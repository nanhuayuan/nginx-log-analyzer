import gc
import numpy as np
import pandas as pd
from openpyxl import Workbook

from self_00_04_excel_processor import add_dataframe_to_excel_with_grouped_headers, format_excel_sheet
from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD, PERCENTILES

# 优化后的时间指标配置
TIME_METRICS = [
    'total_request_duration',  # 请求总时长
    'upstream_response_time',  # 后端响应时长
    'upstream_header_time',  # 后端处理时长
    'upstream_connect_time',  # 后端连接时长
    'backend_connect_phase',  # 后端连接阶段
    'backend_process_phase',  # 后端处理阶段
    'backend_transfer_phase',  # 后端传输阶段
    'nginx_transfer_phase',  # Nginx传输阶段
    'backend_total_phase',  # 后端总阶段
    'network_phase',  # 网络传输阶段
    'processing_phase',  # 纯处理阶段
    'transfer_phase'  # 纯传输阶段
]

# 大小相关指标（已为KB单位）
SIZE_METRICS = [
    'response_body_size_kb',  # 响应体大小(KB)
    'total_bytes_sent_kb'  # 总发送字节(KB)
]

# 效率比率指标（百分比）
RATIO_METRICS = [
    'backend_efficiency',  # 后端处理效率(%)
    'network_overhead',  # 网络开销占比(%)
    'transfer_ratio',  # 传输时间占比(%)
    'connection_cost_ratio',  # 连接成本占比(%)
    'processing_efficiency_index'  # 处理效率指数
]

# 速度相关指标
SPEED_METRICS = [
    'response_transfer_speed',  # 响应传输速度(KB/s)
    'total_transfer_speed',  # 总传输速度(KB/s)
    'nginx_transfer_speed'  # Nginx传输速度(KB/s)
]

# 中文名称映射
METRICS_MAPPING = {
    # 时间指标
    'total_request_duration': '请求总时长',
    'upstream_response_time': '后端响应时长',
    'upstream_header_time': '后端处理时长',
    'upstream_connect_time': '后端连接时长',
    'backend_connect_phase': '后端连接阶段',
    'backend_process_phase': '后端处理阶段',
    'backend_transfer_phase': '后端传输阶段',
    'nginx_transfer_phase': 'Nginx传输阶段',
    'backend_total_phase': '后端总阶段',
    'network_phase': '网络传输阶段',
    'processing_phase': '纯处理阶段',
    'transfer_phase': '纯传输阶段',

    # 大小指标
    'response_body_size_kb': '响应体大小',
    'total_bytes_sent_kb': '总发送字节',

    # 比率指标
    'backend_efficiency': '后端处理效率',
    'network_overhead': '网络开销占比',
    'transfer_ratio': '传输时间占比',
    'connection_cost_ratio': '连接成本占比',
    'processing_efficiency_index': '处理效率指数',

    # 速度指标
    'response_transfer_speed': '响应传输速度',
    'total_transfer_speed': '总传输速度',
    'nginx_transfer_speed': 'Nginx传输速度'
}


def analyze_service_performance(csv_path, output_path, success_codes=None, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """分析服务性能主函数"""
    if success_codes is None:
        from self_00_01_constants import DEFAULT_SUCCESS_CODES
        success_codes = DEFAULT_SUCCESS_CODES
    success_codes = [str(code) for code in success_codes]

    # 初始化统计数据
    service_stats = {}
    app_stats = {}
    total_requests = 0
    success_requests = 0

    # 分块处理CSV文件
    for chunk_idx, chunk in enumerate(pd.read_csv(csv_path, chunksize=DEFAULT_CHUNK_SIZE)):
        total_requests += len(chunk)

        # 处理总请求统计
        _process_total_requests(chunk, service_stats, app_stats)

        # 处理成功请求统计
        successful_requests = chunk[chunk['response_status_code'].astype(str).isin(success_codes)]
        success_requests += len(successful_requests)

        if not successful_requests.empty:
            _process_successful_requests(successful_requests, service_stats, app_stats, slow_threshold)

        # 内存管理
        if (chunk_idx + 1) % 5 == 0:
            del chunk
            gc.collect()

    # 计算最终结果
    service_results = _calculate_final_results(service_stats, success_requests, slow_threshold, 'service')
    app_results = _calculate_final_results(app_stats, success_requests, slow_threshold, 'app')

    # 创建Excel报告
    _create_performance_excel(service_results, app_results, output_path, total_requests, success_requests)

    return service_results.head(5) if not service_results.empty else pd.DataFrame()


def _process_total_requests(chunk, service_stats, app_stats):
    """处理总请求统计"""
    for service, group in chunk.groupby('service_name'):
        service = service if service and service != '' else 'unknown'

        if service not in service_stats:
            _initialize_service_stats(service_stats, service, group)
        service_stats[service]['total_requests'] += len(group)

        # 应用级别统计
        app_name = group['application_name'].iloc[0] if not group['application_name'].empty else "unknown"
        if app_name not in app_stats:
            _initialize_app_stats(app_stats, app_name)
        app_stats[app_name]['total_requests'] += len(group)


def _process_successful_requests(successful_requests, service_stats, app_stats, slow_threshold):
    """处理成功请求统计"""
    for service, group in successful_requests.groupby('service_name'):
        service = service if service and service != '' else 'unknown'

        if service not in service_stats:
            _initialize_service_stats(service_stats, service, group)

        # 更新服务统计
        service_stats[service]['success_requests'] += len(group)
        _update_metrics_stats(service_stats[service], group)

        # 慢请求统计
        if 'total_request_duration' in group.columns:
            request_times = pd.to_numeric(group['total_request_duration'], errors='coerce')
            slow_count = (request_times > slow_threshold).sum()
            service_stats[service]['slow_requests_count'] += slow_count

        # 应用级别统计
        app_name = group['application_name'].iloc[0] if not group['application_name'].empty else "unknown"
        if app_name not in app_stats:
            _initialize_app_stats(app_stats, app_name)

        app_stats[app_name]['success_requests'] += len(group)
        _update_metrics_stats(app_stats[app_name], group)
        if 'total_request_duration' in group.columns:
            app_stats[app_name]['slow_requests_count'] += slow_count


def _initialize_service_stats(stats_dict, service_name, group):
    """初始化服务统计数据"""
    app_name = group['application_name'].iloc[0] if not group['application_name'].empty else "unknown"

    stats = {
        'service_name': service_name,
        'app_name': app_name,
        'total_requests': 0,
        'success_requests': 0,
        'slow_requests_count': 0,
    }

    _initialize_metrics_fields(stats)
    stats_dict[service_name] = stats


def _initialize_app_stats(stats_dict, app_name):
    """初始化应用统计数据"""
    stats = {
        'app_name': app_name,
        'total_requests': 0,
        'success_requests': 0,
        'slow_requests_count': 0,
    }

    _initialize_metrics_fields(stats)
    stats_dict[app_name] = stats


def _initialize_metrics_fields(stats):
    """初始化指标字段"""
    all_metrics = TIME_METRICS + SIZE_METRICS + RATIO_METRICS + SPEED_METRICS

    for metric in all_metrics:
        stats[f'{metric}_min'] = float('inf')
        stats[f'{metric}_max'] = 0
        stats[f'{metric}_total'] = 0
        stats[f'{metric}_count'] = 0
        stats[f'{metric}_sum_sq'] = 0  # 用于计算方差
        stats[f'{metric}_samples'] = []  # 保存采样数据，最多1000个


def _update_metrics_stats(stats, group):
    """更新指标统计数据"""
    all_metrics = TIME_METRICS + SIZE_METRICS + RATIO_METRICS + SPEED_METRICS

    for metric in all_metrics:
        if metric in group.columns:
            values = pd.to_numeric(group[metric], errors='coerce').dropna()
            if len(values) > 0:
                _update_metric_values(stats, metric, values)


def _update_metric_values(stats, metric, values):
    """更新单个指标的统计值（优化版）"""
    import random
    
    min_value = values.min()
    max_value = values.max()
    total_value = values.sum()
    count_value = len(values)
    sum_sq_value = (values ** 2).sum()

    stats[f'{metric}_min'] = min(stats[f'{metric}_min'], min_value) if stats[f'{metric}_min'] != float(
        'inf') else min_value
    stats[f'{metric}_max'] = max(stats[f'{metric}_max'], max_value)
    stats[f'{metric}_total'] += total_value
    stats[f'{metric}_count'] += count_value
    stats[f'{metric}_sum_sq'] += sum_sq_value
    
    # 使用采样策略保存数据，避免内存溢出
    samples = stats[f'{metric}_samples']
    MAX_SAMPLES = 1000
    
    if len(samples) < MAX_SAMPLES:
        # 直接添加到采样中
        remaining = MAX_SAMPLES - len(samples)
        if len(values) <= remaining:
            samples.extend(values.tolist())
        else:
            # 随机采样
            sample_indices = random.sample(range(len(values)), remaining)
            samples.extend(values.iloc[sample_indices].tolist())
    else:
        # 使用蓄水池采样算法替换现有样本
        for value in values:
            if random.random() < MAX_SAMPLES / (stats[f'{metric}_count']):
                replace_idx = random.randint(0, MAX_SAMPLES - 1)
                samples[replace_idx] = value


def _calculate_final_results(stats_dict, success_requests, slow_threshold, stats_type):
    """计算最终统计结果"""
    results = []

    for key, stats in stats_dict.items():
        if stats['success_requests'] == 0:
            continue

        result = _build_basic_result(stats, key, stats_type, success_requests)
        _add_metrics_to_result(result, stats)

        # 清理内存
        for field_key in list(stats.keys()):
            if field_key.endswith('_samples'):
                stats[field_key] = None

        results.append(result)

    if not results:
        return pd.DataFrame()

    results_df = pd.DataFrame(results)

    # 按平均请求时长排序
    sort_column = 'avg_total_request_duration'
    if sort_column in results_df.columns:
        results_df = results_df.sort_values(by=sort_column, ascending=False)

    # 重命名列名
    column_mapping = _create_column_mapping(stats_type)
    results_df = results_df.rename(columns=column_mapping)

    return results_df


def _build_basic_result(stats, key, stats_type, success_requests):
    """构建基础结果数据"""
    result = {f'{stats_type}_name': key}

    if stats_type == 'service':
        result['app_name'] = stats['app_name']

    result.update({
        'total_requests': stats['total_requests'],
        'request_count': stats['success_requests'],
        'request_percentage': round(stats['success_requests'] / success_requests * 100,
                                    2) if success_requests > 0 else 0,
        'success_rate': round(stats['success_requests'] / stats['total_requests'] * 100, 2) if stats[
                                                                                                   'total_requests'] > 0 else 0,
        'slow_requests_count': stats['slow_requests_count'],
        'slow_requests_percentage': round(stats['slow_requests_count'] / stats['success_requests'] * 100, 2) if stats[
                                                                                                                    'success_requests'] > 0 else 0,
    })

    return result


def _add_metrics_to_result(result, stats):
    """将指标数据添加到结果中（优化版）"""
    all_metrics = TIME_METRICS + SIZE_METRICS + RATIO_METRICS + SPEED_METRICS

    for metric in all_metrics:
        # 使用采样数据计算百分位数
        samples = stats.get(f'{metric}_samples', [])
        count = stats.get(f'{metric}_count', 0)
        total = stats.get(f'{metric}_total', 0)

        if count > 0:
            result[f'avg_{metric}'] = total / count  # 使用精确的平均值
            result[f'min_{metric}'] = stats[f'{metric}_min'] if stats[f'{metric}_min'] != float('inf') else 0
            result[f'max_{metric}'] = stats[f'{metric}_max']
            
            # 使用采样数据计算百分位数
            if len(samples) > 0:
                values = np.array(samples)
                result[f'median_{metric}'] = np.median(values)
                for percentile in PERCENTILES:
                    result[f'p{percentile}_{metric}'] = np.percentile(values, percentile)
            else:
                # 如果没有采样数据，使用平均值作为近似值
                avg_val = total / count
                result[f'median_{metric}'] = avg_val
                for percentile in PERCENTILES:
                    result[f'p{percentile}_{metric}'] = avg_val
        else:
            result[f'avg_{metric}'] = 0
            result[f'min_{metric}'] = 0
            result[f'max_{metric}'] = 0
            result[f'median_{metric}'] = 0

            for percentile in PERCENTILES:
                result[f'p{percentile}_{metric}'] = 0


def _create_column_mapping(stats_type):
    """创建列名映射"""
    column_mapping = {
        f'{stats_type}_name': '服务名称' if stats_type == 'service' else '应用名称',
        'app_name': '应用名称',
        'total_requests': '接口请求总数',
        'request_count': '成功请求数',
        'request_percentage': '占总请求比例(%)',
        'success_rate': '成功率(%)',
        'slow_requests_count': '慢请求数',
        'slow_requests_percentage': '慢请求占比(%)',
    }

    # 时间指标映射
    for metric in TIME_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        column_mapping[f'avg_{metric}'] = f'平均{display_name}(秒)'
        column_mapping[f'min_{metric}'] = f'最小{display_name}(秒)'
        column_mapping[f'max_{metric}'] = f'最大{display_name}(秒)'
        column_mapping[f'median_{metric}'] = f'中位数{display_name}(秒)'

        for percentile in PERCENTILES:
            column_mapping[f'p{percentile}_{metric}'] = f'P{percentile}{display_name}(秒)'

    # 大小指标映射
    for metric in SIZE_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        column_mapping[f'avg_{metric}'] = f'平均{display_name}(KB)'
        column_mapping[f'min_{metric}'] = f'最小{display_name}(KB)'
        column_mapping[f'max_{metric}'] = f'最大{display_name}(KB)'
        column_mapping[f'median_{metric}'] = f'中位数{display_name}(KB)'

        for percentile in PERCENTILES:
            column_mapping[f'p{percentile}_{metric}'] = f'P{percentile}{display_name}(KB)'

    # 比率指标映射
    for metric in RATIO_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        column_mapping[f'avg_{metric}'] = f'平均{display_name}(%)'
        column_mapping[f'min_{metric}'] = f'最小{display_name}(%)'
        column_mapping[f'max_{metric}'] = f'最大{display_name}(%)'
        column_mapping[f'median_{metric}'] = f'中位数{display_name}(%)'

        for percentile in PERCENTILES:
            column_mapping[f'p{percentile}_{metric}'] = f'P{percentile}{display_name}(%)'

    # 速度指标映射
    for metric in SPEED_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        column_mapping[f'avg_{metric}'] = f'平均{display_name}(KB/s)'
        column_mapping[f'min_{metric}'] = f'最小{display_name}(KB/s)'
        column_mapping[f'max_{metric}'] = f'最大{display_name}(KB/s)'
        column_mapping[f'median_{metric}'] = f'中位数{display_name}(KB/s)'

        for percentile in PERCENTILES:
            column_mapping[f'p{percentile}_{metric}'] = f'P{percentile}{display_name}(KB/s)'

    return column_mapping


def _create_header_groups():
    """创建表头分组"""
    header_groups = {
        '基本信息': ['服务名称', '应用名称'],
        '请求统计': ['接口请求总数', '成功请求数', '占总请求比例(%)', '成功率(%)'],
        '慢请求': ['慢请求数', '慢请求占比(%)'],
    }

    # 时间指标分组
    for metric in TIME_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        header_groups[f'{display_name}(秒)'] = ['平均', '最小', '最大', '中位数']
        for percentile in PERCENTILES:
            header_groups[f'{display_name}(秒)'].append(f'P{percentile}')

    # 大小指标分组
    for metric in SIZE_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        header_groups[f'{display_name}(KB)'] = ['平均', '最小', '最大', '中位数']
        for percentile in PERCENTILES:
            header_groups[f'{display_name}(KB)'].append(f'P{percentile}')

    # 比率指标分组
    for metric in RATIO_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        header_groups[f'{display_name}(%)'] = ['平均', '最小', '最大', '中位数']
        for percentile in PERCENTILES:
            header_groups[f'{display_name}(%)'].append(f'P{percentile}')

    # 速度指标分组
    for metric in SPEED_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        header_groups[f'{display_name}(KB/s)'] = ['平均', '最小', '最大', '中位数']
        for percentile in PERCENTILES:
            header_groups[f'{display_name}(KB/s)'].append(f'P{percentile}')

    return header_groups


def _create_performance_excel(service_results, app_results, output_path, total_requests, success_requests):
    """创建性能分析Excel报告"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 创建服务分析表
    header_groups = _create_header_groups()
    add_dataframe_to_excel_with_grouped_headers(wb, service_results, '服务分析', header_groups)

    # 创建应用分析表
    app_header_groups = _create_header_groups()
    if '基本信息' in app_header_groups:
        app_header_groups['基本信息'] = ['应用名称']
    add_dataframe_to_excel_with_grouped_headers(wb, app_results, '应用分析', app_header_groups)

    # 创建整体分析表
    _create_overall_analysis_sheet(wb, service_results, app_results, total_requests, success_requests)

    wb.save(output_path)


def _create_overall_analysis_sheet(wb, service_results, app_results, total_requests, success_requests):
    """创建整体分析表"""
    ws = wb.create_sheet(title='整体服务请求分析')

    # 计算整体统计数据
    overall_stats = _calculate_overall_stats(service_results, total_requests, success_requests)

    # 写入整体统计
    for row_idx, (label, value) in enumerate(overall_stats, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    # 添加Top服务和应用
    _add_top_services_section(ws, service_results, len(overall_stats) + 2)
    _add_top_apps_section(ws, app_results, len(overall_stats) + 20)

    # 格式化工作表
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    format_excel_sheet(ws)


def _calculate_overall_stats(service_results, total_requests, success_requests):
    """计算整体统计数据"""
    overall_stats = [
        ['总请求数', total_requests],
        ['成功请求数', success_requests],
        ['成功率(%)', round(success_requests / total_requests * 100, 2) if total_requests > 0 else 0],
        ['服务数量', len(service_results)],
    ]

    # 计算平均响应时间
    time_col = '平均请求总时长(秒)'
    if time_col in service_results.columns and not service_results.empty:
        weighted_avg = 0
        total_weight = 0

        for _, row in service_results.iterrows():
            weight = row['成功请求数']
            time_value = row[time_col]
            weighted_avg += weight * time_value
            total_weight += weight

        if total_weight > 0:
            overall_stats.append(['加权平均响应时间(秒)', round(weighted_avg / total_weight, 3)])

    return overall_stats


def _add_top_services_section(ws, service_results, start_row):
    """添加Top服务部分"""
    from openpyxl.styles import Font

    ws.cell(row=start_row, column=1, value="请求量最多的10个服务").font = Font(bold=True)

    headers = ['服务名称', '应用名称', '成功请求数', '平均响应时间(秒)', '慢请求占比(%)']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row + 1, column=col_idx, value=header)

    # 按请求量排序
    top_services = service_results.sort_values('成功请求数', ascending=False).head(10)

    for row_idx, (_, service) in enumerate(top_services.iterrows(), start=start_row + 2):
        ws.cell(row=row_idx, column=1, value=service['服务名称'])
        ws.cell(row=row_idx, column=2, value=service.get('应用名称', ''))
        ws.cell(row=row_idx, column=3, value=service['成功请求数'])

        # 响应时间列
        time_col = '平均请求总时长(秒)'
        if time_col in service.index:
            ws.cell(row=row_idx, column=4, value=round(service[time_col], 3))

        ws.cell(row=row_idx, column=5, value=service['慢请求占比(%)'])


def _add_top_apps_section(ws, app_results, start_row):
    """添加Top应用部分"""
    from openpyxl.styles import Font

    ws.cell(row=start_row, column=1, value="请求量最多的10个应用").font = Font(bold=True)

    headers = ['应用名称', '成功请求数', '平均响应时间(秒)', '慢请求占比(%)']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row + 1, column=col_idx, value=header)

    # 按请求量排序
    top_apps = app_results.sort_values('成功请求数', ascending=False).head(10)

    for row_idx, (_, app) in enumerate(top_apps.iterrows(), start=start_row + 2):
        ws.cell(row=row_idx, column=1, value=app['应用名称'])
        ws.cell(row=row_idx, column=2, value=app['成功请求数'])

        # 响应时间列
        time_col = '平均请求总时长(秒)'
        if time_col in app.index:
            ws.cell(row=row_idx, column=3, value=round(app[time_col], 3))

        ws.cell(row=row_idx, column=4, value=app['慢请求占比(%)'])