import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import gc

from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    add_dataframe_to_excel_with_grouped_headers,
    format_excel_sheet
)
from self_08_create_http_lifecycle_visualization import create_http_lifecycle_visualization
from self_06_performance_stability_analyzer import (
    analyze_service_stability,
    apply_highlighting,
    add_summary_sheet,
)

# HTTP生命周期参数映射表（基于nginx_lifecycle_analysis.txt）
HTTP_LIFECYCLE_METRICS = {
    # 基础时间参数（4个核心指标）
    'request_time': '请求总时长',
    'upstream_response_time': '后端响应时长',
    'upstream_header_time': '后端处理时长',
    'upstream_connect_time': '后端连接时长',

    # 核心阶段参数（4个关键阶段）
    'backend_connect_phase': '后端连接阶段',
    'backend_process_phase': '后端处理阶段',
    'backend_transfer_phase': '后端传输阶段',
    'nginx_transfer_phase': 'Nginx传输阶段',

    # 组合分析参数
    'backend_total_phase': '后端总阶段',
    'network_phase': '网络传输阶段',
    'processing_phase': '纯处理阶段',
    'transfer_phase': '纯传输阶段',

    # 性能比率参数（百分比形式）
    'backend_efficiency': '后端处理效率(%)',
    'network_overhead': '网络开销占比(%)',
    'transfer_ratio': '传输时间占比(%)',

    # 数据传输相关（KB单位）
    'response_body_size_kb': '响应体大小(KB)',
    'total_bytes_sent_kb': '总传输量(KB)',

    # 传输速度相关
    'response_transfer_speed': '响应传输速度(KB/s)',
    'total_transfer_speed': '总传输速度(KB/s)',
    'nginx_transfer_speed': 'Nginx传输速度(KB/s)',

    # 新增性能指标
    'connection_cost_ratio': '连接成本比例(%)',
    'processing_efficiency_index': '处理效率指数'
}


def generate_summary_report(outputs, output_path):
    """生成优化后的综合报告"""
    log_info("开始生成综合报告...", level="INFO")
    wb = Workbook()
    ws = wb['Sheet']
    ws.title = '综合报告'

    # 报告头部
    row = _add_report_header(ws)

    # 高级摘要（包含HTTP生命周期健康评分）
    row = _add_executive_summary_with_lifecycle(ws, outputs, row)

    # 基础统计信息
    row = _add_basic_statistics(ws, outputs, row)

    # 请求头分析
    row = _add_header_analysis(ws, outputs, row)

    # 请求头性能关联分析
    row = _add_header_performance_analysis(ws, outputs, row)

    # HTTP生命周期性能分析
    row = _add_http_lifecycle_analysis(ws, outputs, row)

    # 资源占用分析（优化KB单位显示）
    row = _add_optimized_resource_analysis(ws, outputs, row)

    # 时间趋势分析
    row = _add_trend_analysis_section(ws, outputs, row)

    # 连接效率分析
    row = _add_connection_efficiency_analysis(ws, outputs, row)

    # 性能瓶颈诊断
    row = _add_performance_bottleneck_diagnosis(ws, outputs, row)

    # 优化建议
    row = _add_optimization_suggestions_enhanced(ws, row)

    # 设置列宽并保存
    ws.column_dimensions['A'].width = 80
    create_http_lifecycle_visualization(wb)

    try:
        wb.save(output_path)
        log_info(f"综合报告已保存到: {output_path}", level="INFO")
    except Exception as e:
        log_info(f"保存报告失败: {str(e)}", level="ERROR")

    _cleanup_resources(wb, ws)


def _add_report_header(ws):
    """添加报告头部"""
    ws.cell(row=1, column=1, value="Nginx日志分析综合报告").font = Font(size=14, bold=True)
    ws.cell(row=3, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    return 5


def _add_executive_summary_with_lifecycle(ws, outputs, start_row):
    """添加包含HTTP生命周期健康评分的高级摘要"""
    row = start_row
    ws.cell(row=row, column=1, value="系统健康评分").font = Font(size=12, bold=True)
    row += 1

    ws.cell(row=row, column=1, value="=" * 60)
    row += 1

    # 计算综合健康评分
    health_score = 100
    health_issues = []

    # 错误率评估
    error_rate = _calculate_error_rate(outputs.get('status_stats'))
    if error_rate > 1.0:
        health_score -= min(30, error_rate * 10)
        health_issues.append(f"⚠️ 5xx错误率: {error_rate:.2f}% (建议<1%)")

    # HTTP生命周期性能评估
    lifecycle_score = _evaluate_lifecycle_performance(outputs)
    health_score = min(health_score, lifecycle_score)

    # 连接效率评估
    conn_efficiency = _evaluate_connection_efficiency(outputs)
    if conn_efficiency < 70:
        health_score -= (70 - conn_efficiency) / 2
        health_issues.append(f"⚠️ 连接复用效率低: {conn_efficiency:.1f}分 (建议>70分)")

    # 显示健康问题
    for issue in health_issues:
        ws.cell(row=row, column=1, value=issue)
        row += 1

    # 显示综合评分
    health_status = _get_health_status(health_score)
    score_cell = ws.cell(row=row, column=1, value=f"系统健康评分: {health_score:.1f}/100 - {health_status}")
    score_cell.font = Font(bold=True, size=11,
                           color=_get_score_color(health_score))
    row += 1

    ws.cell(row=row, column=1, value="=" * 60)
    return row + 2


def _add_basic_statistics(ws, outputs, start_row):
    """添加基础统计信息"""
    row = start_row

    # 总请求数
    if 'total_requests' in outputs:
        total_req = _safe_int_convert(outputs['total_requests'])
        ws.cell(row=row, column=1, value=f"总请求数: {total_req:,}")
        row += 2

    # 状态码分布
    row = _add_stats_table(ws, outputs, 'status_stats', "状态码分布:",
                           ['状态码', '请求数', '百分比(%)'], row)

    # 慢API统计
    row = _add_stats_table(ws, outputs, 'slow_apis', "响应时间最长的API:",
                           ['请求URI', '成功请求数', '平均请求时长(秒)'], row)

    # 服务分布
    row = _add_stats_table(ws, outputs, 'service_stats', "服务请求分布:",
                           ['服务名称', '成功请求数', '占总请求比例(%)'], row)

    return row


def _add_header_analysis(ws, outputs, start_row):
    """添加请求头分析"""
    row = start_row
    ws.cell(row=row, column=1, value="请求头分析:").font = Font(bold=True, size=12)
    row += 1
    
    header_analysis = outputs.get('header_analysis')
    if not header_analysis or not isinstance(header_analysis, dict):
        ws.cell(row=row, column=1, value="- 无请求头分析数据")
        return row + 2
    
    # 基础统计
    total_processed = header_analysis.get('total_processed', 0)
    unique_user_agents = header_analysis.get('unique_user_agents', 0)
    unique_referers = header_analysis.get('unique_referers', 0)
    
    ws.cell(row=row, column=1, value=f"- 总处理记录数: {total_processed:,}")
    row += 1
    ws.cell(row=row, column=1, value=f"- 唯一User-Agent: {unique_user_agents:,} 个")
    row += 1
    ws.cell(row=row, column=1, value=f"- 唯一Referer: {unique_referers:,} 个")
    row += 1
    
    # TOP浏览器
    top_browsers = header_analysis.get('top_browsers', {})
    if top_browsers:
        ws.cell(row=row, column=1, value="TOP浏览器:")
        row += 1
        for browser, count in list(top_browsers.items())[:5]:
            ws.cell(row=row, column=1, value=f"  - {browser}: {count:,} 次")
            row += 1
    
    # TOP来源域名
    top_domains = header_analysis.get('top_domains', {})
    if top_domains:
        ws.cell(row=row, column=1, value="TOP来源域名:")
        row += 1
        for domain, count in list(top_domains.items())[:5]:
            ws.cell(row=row, column=1, value=f"  - {domain}: {count:,} 次")
            row += 1
    
    return row + 1


def _add_header_performance_analysis(ws, outputs, start_row):
    """添加请求头性能关联分析"""
    row = start_row
    ws.cell(row=row, column=1, value="请求头性能关联分析:").font = Font(bold=True, size=12)
    row += 1
    
    header_performance = outputs.get('header_performance_analysis')
    if not header_performance or not isinstance(header_performance, dict):
        ws.cell(row=row, column=1, value="- 无请求头性能关联数据")
        return row + 2
    
    # 基础性能指标
    slow_rate_overall = header_performance.get('slow_rate_overall', 0)
    total_slow_requests = header_performance.get('total_slow_requests', 0)
    slow_threshold = header_performance.get('slow_threshold', 3)
    
    ws.cell(row=row, column=1, value=f"- 整体慢请求率: {slow_rate_overall}% (阈值: {slow_threshold}秒)")
    row += 1
    ws.cell(row=row, column=1, value=f"- 慢请求总数: {total_slow_requests:,}")
    row += 1
    
    # 性能最差的浏览器
    worst_browsers = header_performance.get('worst_browsers', [])
    if worst_browsers:
        ws.cell(row=row, column=1, value="性能最差浏览器:")
        row += 1
        for browser_info in worst_browsers[:3]:
            ws.cell(row=row, column=1, value=f"  - {browser_info['browser']}: 慢请求率 {browser_info['slow_rate']}%, 平均响应时间 {browser_info['avg_response_time']:.3f}秒")
            row += 1
    
    # 性能最差的设备类型
    worst_devices = header_performance.get('worst_devices', [])
    if worst_devices:
        ws.cell(row=row, column=1, value="性能最差设备类型:")
        row += 1
        for device_info in worst_devices[:3]:
            ws.cell(row=row, column=1, value=f"  - {device_info['device']}: 慢请求率 {device_info['slow_rate']}%, 平均响应时间 {device_info['avg_response_time']:.3f}秒")
            row += 1
    
    # 性能最差的来源域名
    worst_domains = header_performance.get('worst_domains', [])
    if worst_domains:
        ws.cell(row=row, column=1, value="性能最差来源域名:")
        row += 1
        for domain_info in worst_domains[:3]:
            ws.cell(row=row, column=1, value=f"  - {domain_info['domain']}: 慢请求率 {domain_info['slow_rate']}%, 平均响应时间 {domain_info['avg_response_time']:.3f}秒")
            row += 1
    
    # 优化建议
    recommendations = header_performance.get('performance_recommendations', [])
    if recommendations:
        ws.cell(row=row, column=1, value="性能优化建议:")
        row += 1
        for recommendation in recommendations[:5]:
            ws.cell(row=row, column=1, value=f"  - {recommendation}")
            row += 1
    
    return row + 1


def _add_http_lifecycle_analysis(ws, outputs, start_row):
    """添加HTTP生命周期性能分析"""
    row = start_row
    ws.cell(row=row, column=1, value="HTTP请求生命周期分析:").font = Font(bold=True, size=12)
    row += 1

    # 检查是否有生命周期数据
    slowest_requests = outputs.get('slowest_requests')
    if not isinstance(slowest_requests, pd.DataFrame) or slowest_requests.empty:
        ws.cell(row=row, column=1, value="- 无生命周期数据")
        return row + 2

    # 生命周期各阶段平均时长分析
    lifecycle_metrics = [
        ('backend_connect_phase', '后端连接阶段'),
        ('backend_process_phase', '后端处理阶段'),
        ('backend_transfer_phase', '后端传输阶段'),
        ('nginx_transfer_phase', 'Nginx传输阶段')
    ]

    ws.cell(row=row, column=1, value="各阶段平均时长分析:")
    row += 1

    total_avg = slowest_requests.get('total_request_duration', pd.Series([0.001])).mean()
    if total_avg <= 0:
        total_avg = 0.001

    for metric_key, metric_name in lifecycle_metrics:
        if metric_key in slowest_requests.columns:
            avg_time = slowest_requests[metric_key].mean()
            percentage = (avg_time / total_avg) * 100

            cell_text = f"- {metric_name}: {avg_time:.3f}秒 ({percentage:.1f}%)"
            cell = ws.cell(row=row, column=1, value=cell_text)

            # 根据阈值设置颜色预警
            if _is_phase_problematic(metric_key, percentage):
                cell.font = Font(color="FF0000")

            row += 1

    # 效率指标分析
    row = _add_efficiency_metrics_analysis(ws, slowest_requests, row)

    return row + 1


def _add_efficiency_metrics_analysis(ws, df, start_row):
    """添加效率指标分析"""
    row = start_row
    ws.cell(row=row, column=1, value="效率指标分析:")
    row += 1

    efficiency_metrics = [
        ('backend_efficiency', '后端处理效率', '%'),
        ('network_overhead', '网络开销占比', '%'),
        ('transfer_ratio', '传输时间占比', '%'),
        ('connection_cost_ratio', '连接成本比例', '%'),
        ('processing_efficiency_index', '处理效率指数', '')
    ]

    for metric_key, metric_name, unit in efficiency_metrics:
        if metric_key in df.columns:
            avg_value = df[metric_key].mean()

            if unit == '%':
                cell_text = f"- {metric_name}: {avg_value:.1f}%"
                cell = ws.cell(row=row, column=1, value=cell_text)

                # 效率指标阈值预警
                if _is_efficiency_problematic(metric_key, avg_value):
                    cell.font = Font(color="FF0000")
            else:
                cell_text = f"- {metric_name}: {avg_value:.2f}"
                ws.cell(row=row, column=1, value=cell_text)

            row += 1

    return row


def _add_optimized_resource_analysis(ws, outputs, start_row):
    """添加优化后的资源占用分析（KB单位统一）"""
    row = start_row
    ws.cell(row=row, column=1, value="资源占用分析:").font = Font(bold=True, size=12)
    row += 1

    # 带宽消耗分析（KB统一单位）
    bandwidth_df = outputs.get('resource_usage_bandwidth')
    if isinstance(bandwidth_df, pd.DataFrame) and not bandwidth_df.empty:
        row = _analyze_bandwidth_usage_kb(ws, bandwidth_df, row)

    # 请求频率分析
    freq_df = outputs.get('service_request_frequency')
    if isinstance(freq_df, pd.DataFrame) and not freq_df.empty:
        row = _analyze_request_frequency(ws, freq_df, row)

    return row


def _analyze_bandwidth_usage_kb(ws, df, start_row):
    """分析带宽使用情况（KB单位）"""
    row = start_row
    ws.cell(row=row, column=1, value="带宽消耗分析(KB单位):")
    row += 1

    try:
        # 查找KB单位的带宽列
        kb_columns = [col for col in df.columns if 'kb' in col.lower() or 'KB' in col]
        bandwidth_col = None

        for col in ['总响应大小KB', 'total_bytes_sent_kb', '总带宽消耗KB']:
            if col in df.columns:
                bandwidth_col = col
                break

        if bandwidth_col:
            total_bandwidth_kb = df[bandwidth_col].sum()
            ws.cell(row=row, column=1, value=f"总带宽消耗: {total_bandwidth_kb:,.1f}KB ({total_bandwidth_kb / 1024:.1f}MB)")
            row += 1

            # 显示top5消耗
            top_bandwidth = df.nlargest(5, bandwidth_col)
            for _, data in top_bandwidth.iterrows():
                service = data.get('服务名称', data.get('service_name', '未知服务'))
                path = _get_request_path(data)
                bandwidth = data[bandwidth_col]
                count = _get_request_count(data)

                info = f"- {service}{path}: {bandwidth:,.1f}KB"
                if count:
                    info += f" ({count:,}次请求)"

                ws.cell(row=row, column=1, value=info)
                row += 1

            _cleanup_dataframe(top_bandwidth)

    except Exception as e:
        ws.cell(row=row, column=1, value=f"带宽分析错误: {str(e)}")
        row += 1

    return row + 1


def _add_connection_efficiency_analysis(ws, outputs, start_row):
    """添加连接效率分析"""
    row = start_row
    ws.cell(row=row, column=1, value="连接效率分析:").font = Font(bold=True, size=12)
    row += 1

    # 连接指标分析
    conn_metrics_df = outputs.get('connection_metrics')
    if isinstance(conn_metrics_df, pd.DataFrame) and not conn_metrics_df.empty:
        # 连接效率评分
        conn_summary = outputs.get('connection_summary', {})
        avg_ratio = conn_summary.get('平均连接/请求比例', 0)
        efficiency_score = max(0, 100 - min(avg_ratio * 200, 90))

        score_text = f"连接效率评分: {efficiency_score:.1f}/100"
        score_cell = ws.cell(row=row, column=1, value=score_text)
        score_cell.font = Font(bold=True, color=_get_score_color(efficiency_score))
        row += 1

        # 连接复用率分析
        if 'total_requests' in outputs:
            total_requests = _safe_int_convert(outputs['total_requests'])
            total_connections = conn_metrics_df['新连接数'].sum()
            if total_connections > 0:
                reuse_ratio = total_requests / total_connections
                ws.cell(row=row, column=1,
                        value=f"平均连接复用率: {reuse_ratio:.2f} 请求/连接")
                row += 1

        # 连接类型分布
        row = _analyze_connection_types(ws, outputs, row)

    return row + 1


def _add_performance_bottleneck_diagnosis(ws, outputs, start_row):
    """添加性能瓶颈诊断"""
    row = start_row
    ws.cell(row=row, column=1, value="性能瓶颈诊断:").font = Font(bold=True, size=12)
    row += 1

    slowest_requests = outputs.get('slowest_requests')
    if not isinstance(slowest_requests, pd.DataFrame) or slowest_requests.empty:
        ws.cell(row=row, column=1, value="无性能数据可供诊断")
        return row + 2

    # 识别主要瓶颈
    bottlenecks = _identify_performance_bottlenecks(slowest_requests)

    if bottlenecks:
        ws.cell(row=row, column=1, value="发现的性能瓶颈:").font = Font(bold=True)
        row += 1

        for bottleneck in bottlenecks:
            ws.cell(row=row, column=1, value=f"- {bottleneck}").font = Font(color="FF0000")
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="针对性优化建议:")
        row += 1

        suggestions = _get_bottleneck_suggestions(bottlenecks)
        for suggestion in suggestions:
            ws.cell(row=row, column=1, value=f"  {suggestion}")
            row += 1
    else:
        ws.cell(row=row, column=1, value="未发现明显性能瓶颈")
        row += 1

    return row + 1


def _add_optimization_suggestions_enhanced(ws, start_row):
    """添加增强版优化建议"""
    row = start_row
    ws.cell(row=row, column=1, value="系统优化建议:").font = Font(bold=True, size=12)
    row += 1

    suggestions = [
        "🔧 HTTP生命周期优化:",
        "  • 后端连接阶段优化: 启用连接池，增加keepalive时间",
        "  • 后端处理阶段优化: 优化数据库查询，启用缓存机制",
        "  • 传输阶段优化: 启用gzip压缩，优化响应体大小",
        "",
        "⚡ 连接效率提升:",
        "  • 配置合适的keepalive_timeout参数(建议30-60秒)",
        "  • 启用HTTP/2减少连接开销",
        "  • 优化upstream keepalive连接数",
        "",
        "📊 性能监控建设:",
        "  • 建立HTTP生命周期各阶段监控",
        "  • 设置连接效率指标告警",
        "  • 监控传输速度和响应体大小趋势",
        "",
        "🛠️ 架构优化方向:",
        "  • 实施微服务拆分，减少单服务处理时间",
        "  • 部署CDN加速静态资源传输",
        "  • 优化负载均衡策略，提升整体吞吐"
    ]

    for suggestion in suggestions:
        ws.cell(row=row, column=1, value=suggestion)
        row += 1

    return row


# 辅助函数
def _calculate_error_rate(status_stats):
    """计算错误率"""
    if status_stats.empty:
        return 0

    error_rate = 0
    if isinstance(status_stats, pd.DataFrame):
        for _, row_data in status_stats.iterrows():
            status_code = str(row_data.get('状态码', ''))
            if status_code.startswith('5'):
                error_rate += float(row_data.get('百分比(%)', 0))

    return error_rate


def _evaluate_lifecycle_performance(outputs):
    """评估HTTP生命周期性能"""
    slowest_requests = outputs.get('slowest_requests')
    if not isinstance(slowest_requests, pd.DataFrame) or slowest_requests.empty:
        return 100

    score = 100

    # 检查各阶段平均时长
    phase_thresholds = {
        'backend_connect_phase': 0.1,  # 连接阶段不应超过100ms
        'backend_process_phase': 0.5,  # 处理阶段不应超过500ms
        'nginx_transfer_phase': 0.1  # Nginx传输不应超过100ms
    }

    for phase, threshold in phase_thresholds.items():
        if phase in slowest_requests.columns:
            avg_time = slowest_requests[phase].mean()
            if avg_time > threshold:
                score -= min(20, (avg_time - threshold) * 40)

    return max(0, score)


def _evaluate_connection_efficiency(outputs):
    """评估连接效率"""
    conn_summary = outputs.get('connection_summary', {})
    avg_ratio = conn_summary.get('平均连接/请求比例', 0.5)

    # 连接/请求比例越低，效率越高
    if avg_ratio < 0.2:
        return 90
    elif avg_ratio < 0.3:
        return 80
    elif avg_ratio < 0.5:
        return 70
    else:
        return max(0, 70 - (avg_ratio - 0.5) * 100)


def _get_health_status(score):
    """获取健康状态描述"""
    if score >= 90:
        return "优秀"
    elif score >= 80:
        return "良好"
    elif score >= 70:
        return "一般"
    else:
        return "需要关注"


def _get_score_color(score):
    """获取分数对应的颜色"""
    if score >= 80:
        return "008000"  # 绿色
    elif score >= 70:
        return "FFA500"  # 橙色
    else:
        return "FF0000"  # 红色


def _is_phase_problematic(phase_key, percentage):
    """判断阶段是否有问题"""
    thresholds = {
        'backend_connect_phase': 30,
        'backend_process_phase': 40,
        'backend_transfer_phase': 50,
        'nginx_transfer_phase': 20
    }
    return percentage > thresholds.get(phase_key, 50)


def _is_efficiency_problematic(metric_key, value):
    """判断效率指标是否有问题"""
    problematic_thresholds = {
        'backend_efficiency': (None, 60),  # 低于60%有问题
        'network_overhead': (30, None),  # 高于30%有问题
        'transfer_ratio': (40, None),  # 高于40%有问题
        'connection_cost_ratio': (20, None)  # 高于20%有问题
    }

    if metric_key not in problematic_thresholds:
        return False

    min_threshold, max_threshold = problematic_thresholds[metric_key]

    if min_threshold and value < min_threshold:
        return True
    if max_threshold and value > max_threshold:
        return True

    return False


def _identify_performance_bottlenecks(df):
    """识别性能瓶颈"""
    bottlenecks = []

    total_avg = df['请求总时长(秒)'].mean()
    if total_avg <= 0:
        return bottlenecks

    # 检查各阶段占比
    phases = {
        'backend_connect_phase': ('后端连接阶段(秒)', 30),
        'backend_process_phase': ('后端处理阶段(秒)', 40),
        'backend_transfer_phase': ('后端传输阶段(秒)', 50),
        'nginx_transfer_phase': ('Nginx传输阶段(秒)', 20)
    }

    for phase_key, (phase_name, threshold) in phases.items():
        if phase_key in df.columns:
            avg_time = df[phase_key].mean()
            percentage = (avg_time / total_avg) * 100
            if percentage > threshold:
                bottlenecks.append(f"{phase_name}耗时过长({percentage:.1f}%)")

    return bottlenecks


def _get_bottleneck_suggestions(bottlenecks):
    """获取瓶颈对应的优化建议"""
    suggestions = []
    suggestion_map = {
        "后端连接阶段": "优化连接池配置，启用keepalive长连接",
        "后端处理阶段": "优化业务逻辑，添加缓存层，检查数据库查询",
        "后端传输阶段": "启用响应压缩，优化响应体大小，检查网络带宽",
        "Nginx传输阶段": "检查Nginx配置，优化客户端网络连接"
    }

    for bottleneck in bottlenecks:
        for key, suggestion in suggestion_map.items():
            if key in bottleneck:
                suggestions.append(f"• {suggestion}")
                break

    return suggestions


def _add_stats_table(ws, outputs, key, title, columns, start_row, max_items=5):
    """添加统计表格"""
    row = start_row
    df = outputs.get(key)

    if df is None or (isinstance(df, pd.DataFrame) and df.empty):
        ws.cell(row=row, column=1, value=f"{title} 无数据")
        return row + 2

    ws.cell(row=row, column=1, value=title).font = Font(bold=True)
    row += 1

    try:
        if isinstance(df, dict):
            df = pd.DataFrame([df])

        for i in range(min(max_items, len(df))):
            values = []
            for col in columns:
                values.append(df.iloc[i].get(col, "N/A") if col in df.columns else "N/A")

            if len(values) >= 3 and all(isinstance(v, (int, float)) for v in values[1:3]):
                cell_text = f"- {values[0]}: {values[1]:,} ({values[2]:.2f}%)"
            else:
                cell_text = f"- {' '.join(str(v) for v in values)}"

            ws.cell(row=row, column=1, value=cell_text)
            row += 1

    except Exception as e:
        ws.cell(row=row, column=1, value=f"数据处理错误: {str(e)}")
        row += 1

    return row + 1


def _safe_int_convert(value):
    """安全的整数转换"""
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0


def _get_request_path(data):
    """获取请求路径"""
    path_cols = ['请求路径', 'request_path', '请求方法']
    return next((data.get(col, '') for col in path_cols if col in data), '')


def _get_request_count(data):
    """获取请求数量"""
    count_cols = ['请求次数', '总请求数', 'request_count']
    return next((data.get(col) for col in count_cols if col in data), None)


def _analyze_request_frequency(ws, df, start_row):
    """分析请求频率"""
    row = start_row
    ws.cell(row=row, column=1, value="请求频率分析:")
    row += 1

    try:
        top_freq = df.head(5)
        for _, data in top_freq.iterrows():
            service = data.get('服务名称', '未知服务')
            avg_rpm = data.get('平均每分钟请求数', 0)
            max_rpm = data.get('最大每分钟请求数', 0)
            total = data.get('总请求数', 0)

            ws.cell(row=row, column=1,
                    value=f"- {service}: 平均{avg_rpm:.1f}次/分钟, 峰值{max_rpm:.1f}次/分钟, 总计{total:,}次")
            row += 1

        _cleanup_dataframe(top_freq)
    except Exception:
        ws.cell(row=row, column=1, value="频率分析处理错误")
        row += 1

    return row + 1


def _analyze_connection_types(ws, outputs, start_row):
    """分析连接类型分布"""
    row = start_row
    conn_types = outputs.get('connection_types', {})

    if conn_types:
        ws.cell(row=row, column=1, value="连接类型分布:")
        row += 1

        keep_alive = conn_types.get('keep_alive', 0)
        short_lived = conn_types.get('short_lived', 0)
        total_conn = keep_alive + short_lived

        if total_conn > 0:
            keep_alive_pct = keep_alive / total_conn * 100
            short_pct = short_lived / total_conn * 100

            ws.cell(row=row, column=1,
                    value=f"- 长连接(Keep-Alive): {keep_alive:,} ({keep_alive_pct:.1f}%)")
            row += 1

            ws.cell(row=row, column=1,
                    value=f"- 短连接: {short_lived:,} ({short_pct:.1f}%)")
            row += 1

            # 连接复用率评估
            if keep_alive_pct < 50:
                ws.cell(row=row, column=1,
                        value="⚠️ 长连接占比较低，建议优化keepalive配置").font = Font(color="FFA500")
                row += 1

    return row


def _cleanup_dataframe(df):
    """清理DataFrame内存"""
    if df is not None:
        del df
        gc.collect()


def _cleanup_resources(*resources):
    """清理资源"""
    for resource in resources:
        if resource is not None:
            del resource
    gc.collect()


def _add_trend_analysis_section(ws, outputs, start_row):
    """优化后的时间趋势分析"""
    row = start_row
    ws.cell(row=row, column=1, value="时间趋势分析:").font = Font(bold=True, size=12)
    row += 1

    # 响应时间趋势分析
    row = _analyze_response_time_trend(ws, outputs, row)

    # 并发连接趋势分析
    row = _analyze_concurrency_trend(ws, outputs, row)

    # 状态码趋势分析
    row = _analyze_status_trend(ws, outputs, row)

    return row


def _analyze_response_time_trend(ws, outputs, start_row):
    """分析响应时间趋势"""
    row = start_row
    resp_trend_df = outputs.get('hourly_response_trend')

    if not isinstance(resp_trend_df, pd.DataFrame) or resp_trend_df.empty:
        return row

    ws.cell(row=row, column=1, value="响应时间趋势分析:")
    row += 1

    try:
        # 查找时间和响应时间列
        time_col = _find_column(resp_trend_df, ['小时', 'hour', 'date_hour'])
        resp_time_col = _find_column(resp_trend_df, ['平均响应时间(秒)', 'avg_response_time', 'total_request_duration'])

        if time_col and resp_time_col:
            # 找出响应时间最高的3个时段
            peak_hours = resp_trend_df.nlargest(3, resp_time_col)

            for _, data in peak_hours.iterrows():
                hour = data[time_col]
                avg_time = data[resp_time_col]

                # 格式化时间显示
                time_str = _format_datetime(hour)

                # 获取其他指标
                median_col = _find_column(data, ['中位数响应时间(秒)', 'median_response_time'])
                count_col = _find_column(data, ['请求数', 'request_count'])

                info_parts = [f"高峰时段 {time_str}: 平均{avg_time:.3f}秒"]

                if median_col and median_col in data.index:
                    median_val = data[median_col]
                    info_parts.append(f"中位数{median_val:.3f}秒")

                if count_col and count_col in data.index:
                    count = data[count_col]
                    info_parts.append(f"请求数{count:,}")

                ws.cell(row=row, column=1, value=f"- {', '.join(info_parts)}")
                row += 1

            _cleanup_dataframe(peak_hours)

    except Exception as e:
        ws.cell(row=row, column=1, value=f"响应时间趋势分析错误: {str(e)}")
        row += 1

    finally:
        _cleanup_dataframe(resp_trend_df)

    return row + 1


def _analyze_concurrency_trend(ws, outputs, start_row):
    """分析并发连接趋势"""
    row = start_row
    concurrency_df = outputs.get('concurrency_estimation')

    if not isinstance(concurrency_df, pd.DataFrame) or concurrency_df.empty:
        return row

    ws.cell(row=row, column=1, value="并发连接趋势分析:")
    row += 1

    try:
        # 趋势分析：比较前后半段数据
        if len(concurrency_df) > 1 and '平均并发数' in concurrency_df.columns:
            half_point = len(concurrency_df) // 2
            first_half_avg = concurrency_df.iloc[:half_point]['平均并发数'].mean()
            second_half_avg = concurrency_df.iloc[half_point:]['平均并发数'].mean()

            if second_half_avg > first_half_avg * 1.2:
                ws.cell(row=row, column=1, value="- 并发数呈明显上升趋势，需关注系统容量").font = Font(color="FFA500")
                row += 1
            elif second_half_avg < first_half_avg * 0.8:
                ws.cell(row=row, column=1, value="- 并发数呈下降趋势")
                row += 1

        # 显示高峰并发时段
        if '最大并发数' in concurrency_df.columns:
            peak_concurrency = concurrency_df.nlargest(3, '最大并发数')
            time_col = _find_column(concurrency_df, ['时间段', 'minute', 'date_hour_minute'])

            if time_col:
                for _, data in peak_concurrency.iterrows():
                    time_val = data[time_col]
                    time_str = _format_datetime(time_val)
                    avg_concurrent = data.get('平均并发数', 0)
                    max_concurrent = data.get('最大并发数', 0)

                    ws.cell(row=row, column=1,
                            value=f"- 高峰时段 {time_str}: 平均并发{avg_concurrent:.1f}, 最大并发{max_concurrent:.1f}")
                    row += 1

            # 整体统计
            overall_avg = concurrency_df['平均并发数'].mean()
            overall_max = concurrency_df['最大并发数'].max()
            ws.cell(row=row, column=1, value=f"- 整体统计: 平均并发{overall_avg:.1f}, 最大并发{overall_max:.1f}")
            row += 1

            _cleanup_dataframe(peak_concurrency)

    except Exception as e:
        ws.cell(row=row, column=1, value=f"并发趋势分析错误: {str(e)}")
        row += 1

    finally:
        _cleanup_dataframe(concurrency_df)

    return row + 1


def _analyze_status_trend(ws, outputs, start_row):
    """分析状态码趋势"""
    row = start_row
    status_trend_df = outputs.get('hourly_status_trend')

    if not isinstance(status_trend_df, pd.DataFrame) or status_trend_df.empty:
        return row

    ws.cell(row=row, column=1, value="状态码分布趋势:")
    row += 1

    try:
        if '5xx服务器错误' in status_trend_df.columns:
            # 计算总请求数和错误率
            status_trend_df['总请求数'] = status_trend_df.select_dtypes(include=['number']).sum(axis=1)

            if status_trend_df['总请求数'].sum() > 0:
                status_trend_df['5xx错误率'] = (status_trend_df['5xx服务器错误'] /
                                             status_trend_df['总请求数'] * 100)

                # 找出错误率最高的时段
                error_threshold = 1.0  # 1%错误率阈值
                high_error_periods = status_trend_df[status_trend_df['5xx错误率'] > error_threshold]

                if not high_error_periods.empty:
                    ws.cell(row=row, column=1, value=f"发现{len(high_error_periods)}个高错误率时段(>1%):")
                    row += 1

                    time_col = _find_column(status_trend_df, ['hour_bucket', '小时', '时间', 'date_hour'])
                    peak_errors = high_error_periods.nlargest(3, '5xx错误率')

                    if time_col:
                        for _, data in peak_errors.iterrows():
                            time_val = data[time_col]
                            time_str = _format_datetime(time_val)
                            error_rate = data['5xx错误率']
                            error_count = data['5xx服务器错误']

                            ws.cell(row=row, column=1,
                                    value=f"- {time_str}: 5xx错误率{error_rate:.2f}% ({error_count:.0f}次)").font = Font(
                                color="FF0000")
                            row += 1

                    _cleanup_dataframe(peak_errors)
                else:
                    ws.cell(row=row, column=1, value="- 系统错误率保持在健康水平(<1%)")
                    row += 1

                _cleanup_dataframe(high_error_periods)

    except Exception as e:
        ws.cell(row=row, column=1, value=f"状态码趋势分析错误: {str(e)}")
        row += 1

    finally:
        _cleanup_dataframe(status_trend_df)

    return row + 1


def _find_column(data, possible_names):
    """在数据中查找可能的列名"""
    if isinstance(data, pd.DataFrame):
        columns = data.columns
    elif isinstance(data, pd.Series):
        columns = data.index
    else:
        return None

    for name in possible_names:
        if name in columns:
            return name
    return None


def _format_datetime(dt_value):
    """格式化日期时间显示"""
    if hasattr(dt_value, 'strftime'):
        return dt_value.strftime('%Y-%m-%d %H:%M')
    else:
        return str(dt_value)


# 主要的辅助分析函数
def calculate_http_lifecycle_metrics(df):
    """计算HTTP生命周期指标"""
    if df.empty:
        return {}

    metrics = {}

    # 基础时间指标统计
    basic_metrics = ['total_request_duration', 'upstream_response_time',
                     'upstream_header_time', 'upstream_connect_time']

    for metric in basic_metrics:
        if metric in df.columns:
            metrics[f'{metric}_avg'] = df[metric].mean()
            metrics[f'{metric}_p95'] = df[metric].quantile(0.95)
            metrics[f'{metric}_max'] = df[metric].max()

    # 阶段指标统计
    phase_metrics = ['backend_connect_phase', 'backend_process_phase',
                     'backend_transfer_phase', 'nginx_transfer_phase']

    for metric in phase_metrics:
        if metric in df.columns:
            metrics[f'{metric}_avg'] = df[metric].mean()
            metrics[f'{metric}_p95'] = df[metric].quantile(0.95)

    # 效率指标统计
    efficiency_metrics = ['backend_efficiency', 'network_overhead',
                          'transfer_ratio', 'connection_cost_ratio']

    for metric in efficiency_metrics:
        if metric in df.columns:
            metrics[f'{metric}_avg'] = df[metric].mean()

    # 传输速度指标
    speed_metrics = ['response_transfer_speed', 'total_transfer_speed',
                     'nginx_transfer_speed']

    for metric in speed_metrics:
        if metric in df.columns:
            metrics[f'{metric}_avg'] = df[metric].mean()
            metrics[f'{metric}_max'] = df[metric].max()

    return metrics


def generate_lifecycle_performance_summary(metrics):
    """生成生命周期性能摘要"""
    summary = []

    # 检查关键性能指标
    if 'backend_efficiency_avg' in metrics:
        efficiency = metrics['backend_efficiency_avg']
        if efficiency < 60:
            summary.append(f"后端处理效率偏低({efficiency:.1f}%)，建议优化业务逻辑")
        elif efficiency > 80:
            summary.append(f"后端处理效率良好({efficiency:.1f}%)")

    if 'network_overhead_avg' in metrics:
        overhead = metrics['network_overhead_avg']
        if overhead > 30:
            summary.append(f"网络开销较高({overhead:.1f}%)，建议优化网络配置")
        elif overhead < 20:
            summary.append(f"网络开销控制良好({overhead:.1f}%)")

    if 'connection_cost_ratio_avg' in metrics:
        conn_cost = metrics['connection_cost_ratio_avg']
        if conn_cost > 20:
            summary.append(f"连接成本较高({conn_cost:.1f}%)，建议启用长连接")
        elif conn_cost < 10:
            summary.append(f"连接成本控制良好({conn_cost:.1f}%)")

    # 传输速度评估
    if 'total_transfer_speed_avg' in metrics:
        speed = metrics['total_transfer_speed_avg']
        if speed < 100:  # KB/s
            summary.append(f"传输速度较慢({speed:.1f}KB/s)，建议检查带宽和压缩设置")
        elif speed > 1000:
            summary.append(f"传输速度良好({speed:.1f}KB/s)")

    return summary


def analyze_performance_trends(df, time_column='arrival_date_hour'):
    """分析性能趋势"""
    if df.empty or time_column not in df.columns:
        return {}

    # 按小时分组分析
    hourly_stats = df.groupby(time_column).agg({
        'total_request_duration': ['mean', 'count'],
        'backend_connect_phase': 'mean',
        'backend_process_phase': 'mean',
        'backend_transfer_phase': 'mean',
        'nginx_transfer_phase': 'mean',
        'backend_efficiency': 'mean',
        'network_overhead': 'mean'
    }).round(4)

    # 扁平化列名
    hourly_stats.columns = ['_'.join(col) for col in hourly_stats.columns]

    # 识别性能异常时段
    response_time_mean = hourly_stats['total_request_duration_mean'].mean()
    response_time_std = hourly_stats['total_request_duration_mean'].std()
    threshold = response_time_mean + 2 * response_time_std

    anomaly_hours = hourly_stats[
        hourly_stats['total_request_duration_mean'] > threshold
        ].index.tolist()

    return {
        'hourly_performance': hourly_stats,
        'anomaly_hours': anomaly_hours,
        'performance_baseline': {
            'avg_response_time': response_time_mean,
            'response_time_std': response_time_std,
            'anomaly_threshold': threshold
        }
    }


# 数据验证和清洗函数
def validate_lifecycle_data(df):
    """验证生命周期数据完整性"""
    required_columns = [
        'total_request_duration',
        'upstream_connect_time',
        'upstream_header_time',
        'upstream_response_time'
    ]

    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        return False, f"缺少必要字段: {', '.join(missing_columns)}"

    # 检查数据逻辑性
    invalid_rows = 0

    # 检查时间逻辑关系
    if all(col in df.columns for col in required_columns):
        # upstream_connect_time <= upstream_header_time <= upstream_response_time <= total_request_duration
        logical_errors = (
                (df['upstream_connect_time'] > df['upstream_header_time']) |
                (df['upstream_header_time'] > df['upstream_response_time']) |
                (df['upstream_response_time'] > df['total_request_duration'])
        )
        invalid_rows = logical_errors.sum()

    if invalid_rows > len(df) * 0.1:  # 如果超过10%的数据有逻辑错误
        return False, f"发现{invalid_rows}行数据存在时间逻辑错误"

    return True, "数据验证通过"


def clean_lifecycle_data(df):
    """清洗生命周期数据"""
    original_rows = len(df)

    # 移除明显异常的数据
    # 1. 移除负数时间
    time_columns = [col for col in df.columns if 'time' in col.lower() or 'phase' in col.lower()]
    for col in time_columns:
        if col in df.columns:
            df = df[df[col] >= 0]

    # 2. 移除超长响应时间（超过10分钟的请求通常是异常）
    if 'total_request_duration' in df.columns:
        df = df[df['total_request_duration'] <= 600]

    # 3. 修复逻辑错误（保守处理）
    if all(col in df.columns for col in ['upstream_connect_time', 'upstream_header_time',
                                         'upstream_response_time', 'total_request_duration']):
        # 确保时间递增关系
        df['upstream_header_time'] = df[['upstream_header_time', 'upstream_connect_time']].max(axis=1)
        df['upstream_response_time'] = df[['upstream_response_time', 'upstream_header_time']].max(axis=1)
        df['total_request_duration'] = df[['total_request_duration', 'upstream_response_time']].max(axis=1)

    cleaned_rows = len(df)
    removed_rows = original_rows - cleaned_rows

    log_info(f"数据清洗完成: 原始{original_rows}行, 清洗后{cleaned_rows}行, 移除{removed_rows}行异常数据")

    return df