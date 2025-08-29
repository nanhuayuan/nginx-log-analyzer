import gc
import re
import pandas as pd
from collections import defaultdict, Counter
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font

from self_00_01_constants import DEFAULT_CHUNK_SIZE
from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    format_excel_sheet,
    add_dataframe_to_excel_with_grouped_headers,
    create_pie_chart,
    create_line_chart
)
from self_00_05_sampling_algorithms import HyperLogLog, ReservoirSampler


def analyze_request_headers(csv_path, output_path, top_n=100):
    """åˆ†æžè¯·æ±‚å¤´æ•°æ®ï¼ŒåŒ…æ‹¬User-Agentå’ŒRefereråˆ†æž - å†…å­˜ä¼˜åŒ–ç‰ˆ"""
    log_info("ðŸš€ å¼€å§‹é«˜çº§è¯·æ±‚å¤´åˆ†æžï¼ˆå†…å­˜ä¼˜åŒ–ç‰ˆï¼‰...", show_memory=True)
    
    chunk_size = max(DEFAULT_CHUNK_SIZE // 2, 10000)
    
    # User-Agentåˆ†æžæ•°æ® - ä½¿ç”¨HyperLogLogä¼˜åŒ–å†…å­˜
    user_agent_stats = defaultdict(lambda: {
        'count': 0,
        'unique_ips_hll': HyperLogLog(precision=12),  # æ›¿ä»£set()
        'success_requests': 0,
        'error_requests': 0,
        'total_response_time': 0.0,
        'avg_response_time': 0.0
    })
    
    # Refereråˆ†æžæ•°æ® - ä½¿ç”¨HyperLogLogä¼˜åŒ–å†…å­˜
    referer_stats = defaultdict(lambda: {
        'count': 0,
        'unique_ips_hll': HyperLogLog(precision=12),  # æ›¿ä»£set()
        'success_requests': 0,
        'error_requests': 0,
        'total_response_time': 0.0,
        'avg_response_time': 0.0
    })
    
    # æµè§ˆå™¨ç±»åž‹åˆ†æž
    browser_stats = defaultdict(int)
    os_stats = defaultdict(int)
    device_stats = defaultdict(int)
    
    # æ¥æºåŸŸååˆ†æž
    domain_stats = defaultdict(int)
    
    # æœç´¢å¼•æ“Žåˆ†æž
    search_engine_stats = defaultdict(int)
    
    # ç¤¾äº¤åª’ä½“åˆ†æž
    social_media_stats = defaultdict(int)
    
    # æœºå™¨äºº/çˆ¬è™«åˆ†æž
    bot_stats = defaultdict(int)
    
    total_processed = 0
    
    # ç¬¬ä¸€éï¼šæ”¶é›†ç»Ÿè®¡æ•°æ®
    log_info("å¼€å§‹æ”¶é›†è¯·æ±‚å¤´ç»Ÿè®¡æ•°æ®")
    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        chunk_size_actual = len(chunk)
        total_processed += chunk_size_actual
        
        # å¤„ç†User-Agent
        if 'user_agent_string' in chunk.columns:
            for _, row in chunk.iterrows():
                user_agent = row.get('user_agent_string', '')
                if pd.notna(user_agent) and user_agent != '' and user_agent != '-':
                    # æ¸…ç†User-Agentå­—ç¬¦ä¸²
                    user_agent = str(user_agent).strip()
                    
                    # ç»Ÿè®¡User-Agent
                    stats = user_agent_stats[user_agent]
                    stats['count'] += 1
                    
                    # æ”¶é›†IPåœ°å€ - ä½¿ç”¨HyperLogLog
                    ip = row.get('client_ip_address', '')
                    if pd.notna(ip) and ip != '':
                        stats['unique_ips_hll'].add(str(ip))
                    
                    # ç»Ÿè®¡æˆåŠŸ/å¤±è´¥è¯·æ±‚
                    status = str(row.get('response_status_code', ''))
                    if status.startswith('2') or status.startswith('3'):
                        stats['success_requests'] += 1
                    elif status.startswith('4') or status.startswith('5'):
                        stats['error_requests'] += 1
                    
                    # å“åº”æ—¶é—´ç»Ÿè®¡
                    response_time = row.get('total_request_duration', 0)
                    if pd.notna(response_time) and response_time > 0:
                        stats['total_response_time'] += float(response_time)
                    
                    # åˆ†æžæµè§ˆå™¨ã€æ“ä½œç³»ç»Ÿã€è®¾å¤‡ç±»åž‹
                    browser = extract_browser_info(user_agent)
                    os_info = extract_os_info(user_agent)
                    device = extract_device_info(user_agent)
                    
                    if browser:
                        browser_stats[browser] += 1
                    if os_info:
                        os_stats[os_info] += 1
                    if device:
                        device_stats[device] += 1
                    
                    # æ£€æµ‹æœºå™¨äºº/çˆ¬è™«
                    bot_type = detect_bot_type(user_agent)
                    if bot_type:
                        bot_stats[bot_type] += 1
        
        # å¤„ç†Referer
        if 'referer_url' in chunk.columns:
            for _, row in chunk.iterrows():
                referer = row.get('referer_url', '')
                if pd.notna(referer) and referer != '' and referer != '-':
                    # æ¸…ç†Refererå­—ç¬¦ä¸²
                    referer = str(referer).strip()
                    
                    # ç»Ÿè®¡Referer
                    stats = referer_stats[referer]
                    stats['count'] += 1
                    
                    # æ”¶é›†IPåœ°å€ - ä½¿ç”¨HyperLogLog
                    ip = row.get('client_ip_address', '')
                    if pd.notna(ip) and ip != '':
                        stats['unique_ips_hll'].add(str(ip))
                    
                    # ç»Ÿè®¡æˆåŠŸ/å¤±è´¥è¯·æ±‚
                    status = str(row.get('response_status_code', ''))
                    if status.startswith('2') or status.startswith('3'):
                        stats['success_requests'] += 1
                    elif status.startswith('4') or status.startswith('5'):
                        stats['error_requests'] += 1
                    
                    # å“åº”æ—¶é—´ç»Ÿè®¡
                    response_time = row.get('total_request_duration', 0)
                    if pd.notna(response_time) and response_time > 0:
                        stats['total_response_time'] += float(response_time)
                    
                    # åˆ†æžæ¥æºåŸŸå
                    domain = extract_domain_from_referer(referer)
                    if domain:
                        domain_stats[domain] += 1
                    
                    # æ£€æµ‹æœç´¢å¼•æ“Ž
                    search_engine = detect_search_engine(referer)
                    if search_engine:
                        search_engine_stats[search_engine] += 1
                    
                    # æ£€æµ‹ç¤¾äº¤åª’ä½“
                    social_media = detect_social_media(referer)
                    if social_media:
                        social_media_stats[social_media] += 1
        
        if total_processed % 100000 == 0:
            gc.collect()
            log_info(f"å·²å¤„ç† {total_processed:,} æ¡è®°å½•")
    
    # è®¡ç®—å¹³å‡å“åº”æ—¶é—´
    for stats in user_agent_stats.values():
        if stats['success_requests'] > 0:
            stats['avg_response_time'] = stats['total_response_time'] / stats['success_requests']
    
    for stats in referer_stats.values():
        if stats['success_requests'] > 0:
            stats['avg_response_time'] = stats['total_response_time'] / stats['success_requests']
    
    log_info(f"âœ… è¯·æ±‚å¤´åˆ†æžå®Œæˆï¼šæ€»è®°å½• {total_processed:,}ï¼Œå”¯ä¸€User-Agent {len(user_agent_stats)}ä¸ªï¼Œå”¯ä¸€Referer {len(referer_stats)}ä¸ª")
    
    # ç”Ÿæˆåˆ†æžæŠ¥å‘Š
    analysis_results = {
        'user_agent_stats': user_agent_stats,
        'referer_stats': referer_stats,
        'browser_stats': browser_stats,
        'os_stats': os_stats,
        'device_stats': device_stats,
        'domain_stats': domain_stats,
        'search_engine_stats': search_engine_stats,
        'social_media_stats': social_media_stats,
        'bot_stats': bot_stats
    }
    
    # åˆ›å»ºExcelæŠ¥å‘Š
    create_request_header_excel(analysis_results, output_path, top_n, total_processed)
    
    log_info(f"ðŸŽ‰ è¯·æ±‚å¤´åˆ†æžå®Œæˆï¼ŒæŠ¥å‘Šå·²ç”Ÿæˆï¼š{output_path}", show_memory=True)
    
    # è¿”å›žæ‘˜è¦ä¿¡æ¯
    return {
        'total_processed': total_processed,
        'unique_user_agents': len(user_agent_stats),
        'unique_referers': len(referer_stats),
        'top_browsers': dict(Counter(browser_stats).most_common(5)),
        'top_domains': dict(Counter(domain_stats).most_common(5))
    }


def extract_browser_info(user_agent):
    """ä»ŽUser-Agentä¸­æå–æµè§ˆå™¨ä¿¡æ¯"""
    if not user_agent:
        return "æœªçŸ¥æµè§ˆå™¨"
    
    user_agent = user_agent.lower()
    
    # æµè§ˆå™¨æ£€æµ‹æ¨¡å¼
    browsers = [
        ('chrome', r'chrome/(\d+)'),
        ('firefox', r'firefox/(\d+)'),
        ('safari', r'safari/(\d+)'),
        ('edge', r'edge/(\d+)'),
        ('opera', r'opera/(\d+)'),
        ('internet explorer', r'msie (\d+)'),
        ('qqæµè§ˆå™¨', r'qqbrowser/(\d+)'),
        ('æœç‹—æµè§ˆå™¨', r'se 2\.x metasr'),
        ('360æµè§ˆå™¨', r'360se|qhbrowser'),
        ('å¾®ä¿¡æµè§ˆå™¨', r'micromessenger/(\d+)'),
        ('æ”¯ä»˜å®', r'alipayclient/(\d+)'),
        ('ç™¾åº¦æµè§ˆå™¨', r'baiduboxapp/(\d+)'),
        ('UCæµè§ˆå™¨', r'ucbrowser/(\d+)')
    ]
    
    for browser_name, pattern in browsers:
        if re.search(pattern, user_agent):
            match = re.search(pattern, user_agent)
            if match and match.group(1) if '(' in pattern else True:
                version = match.group(1) if '(' in pattern and match.group(1) else ""
                return f"{browser_name}" + (f" {version}" if version else "")
    
    return "å…¶ä»–æµè§ˆå™¨"


def extract_os_info(user_agent):
    """ä»ŽUser-Agentä¸­æå–æ“ä½œç³»ç»Ÿä¿¡æ¯ - æ”¹è¿›ç‰ˆ"""
    if not user_agent:
        return "æœªçŸ¥ç³»ç»Ÿ"
    
    user_agent_lower = user_agent.lower()
    
    # æ“ä½œç³»ç»Ÿæ£€æµ‹ - æŒ‰ä¼˜å…ˆçº§æŽ’åºï¼ŒSDKåŒ¹é…ä¼˜å…ˆäºŽç³»ç»ŸåŒ¹é…
    os_patterns = [
        # åº”ç”¨SDK - ä¼˜å…ˆåŒ¹é…ï¼Œé¿å…è¢«ç³»ç»ŸåŒ¹é…è¦†ç›–
        ('iOS_SDK', r'(wst-sdk-ios|zgt-ios/)'),                   # iOS SDK
        ('Android_SDK', r'(wst-sdk-android|zgt-android/)'),       # Android SDK
        ('Java_SDK', r'(wst-sdk-java)'),                          # Java SDK
        
        # Windowsç³»åˆ— - ç²¾ç¡®åŒ¹é…
        ('Windows 10', r'windows nt 10\.0'),
        ('Windows 8.1', r'windows nt 6\.3'),
        ('Windows 8', r'windows nt 6\.2'),
        ('Windows 7', r'windows nt 6\.1'),
        ('Windows Vista', r'windows nt 6\.0'),
        ('Windows XP', r'windows nt 5\.1'),
        
        # macOS - æ›´ç²¾ç¡®åŒ¹é…ï¼Œé¿å…åŒ¹é…iOSè®¾å¤‡
        ('macOS', r'macintosh.*mac os x|macos'),
        
        # Androidç³»ç»Ÿ - æ›´å…¨é¢çš„åŒ¹é…è§„åˆ™
        ('Android', r'android \d+\.|android;|linux.*android|dalvik.*android'),  # åŒ…æ‹¬Dalvikè™šæ‹Ÿæœº
        
        # iOSç³»ç»Ÿ - çœŸæ­£çš„iOSè®¾å¤‡
        ('iOS', r'iphone os|ipad|ipod|cpu os.*like mac os x|ios \d+\.|cfnetwork.*darwin'),
        
        # çˆ¬è™«å’Œå…¶ä»–å·¥å…·
        ('Bot/Spider', r'(spider|bot|crawler|curl|wget)'),
        ('HTTP_Client', r'(okhttp|retrofit|volley|alamofire)'),    # HTTPå®¢æˆ·ç«¯åº“
        ('Development_Tool', r'(dart|flutter|postman)'),
        
        # Linuxç³»åˆ—
        ('Ubuntu', r'ubuntu'),
        ('Linux', r'linux')
    ]
    
    for os_name, pattern in os_patterns:
        if re.search(pattern, user_agent_lower):
            return os_name
    
    return "å…¶ä»–ç³»ç»Ÿ"


def extract_device_info(user_agent):
    """ä»ŽUser-Agentä¸­æå–è®¾å¤‡ä¿¡æ¯"""
    if not user_agent:
        return "æœªçŸ¥è®¾å¤‡"
    
    user_agent = user_agent.lower()
    
    # è®¾å¤‡ç±»åž‹æ£€æµ‹
    if re.search(r'mobile|phone|android|iphone', user_agent):
        return "æ‰‹æœº"
    elif re.search(r'tablet|ipad', user_agent):
        return "å¹³æ¿"
    elif re.search(r'smart-tv|smarttv|television', user_agent):
        return "æ™ºèƒ½ç”µè§†"
    elif re.search(r'bot|crawler|spider|scraper', user_agent):
        return "çˆ¬è™«/æœºå™¨äºº"
    else:
        return "æ¡Œé¢è®¾å¤‡"


def detect_bot_type(user_agent):
    """æ£€æµ‹æœºå™¨äºº/çˆ¬è™«ç±»åž‹"""
    if not user_agent:
        return None
    
    user_agent = user_agent.lower()
    
    # å¸¸è§çˆ¬è™«æ£€æµ‹
    bot_patterns = [
        ('Googlebot', r'googlebot'),
        ('Baiduspider', r'baiduspider'),
        ('Bingbot', r'bingbot'),
        ('Yahooçˆ¬è™«', r'yahoo.*slurp'),
        ('æœç‹—çˆ¬è™«', r'sogou.*spider'),
        ('360çˆ¬è™«', r'360spider'),
        ('å¾®ä¿¡çˆ¬è™«', r'wechatbot'),
        ('Facebookçˆ¬è™«', r'facebookexternalhit'),
        ('Twitterçˆ¬è™«', r'twitterbot'),
        ('å…¶ä»–çˆ¬è™«', r'bot|crawler|spider|scraper')
    ]
    
    for bot_name, pattern in bot_patterns:
        if re.search(pattern, user_agent):
            return bot_name
    
    return None


def extract_domain_from_referer(referer):
    """ä»ŽRefererä¸­æå–åŸŸå"""
    if not referer or referer == '-':
        return None
    
    try:
        parsed = urlparse(referer)
        domain = parsed.netloc
        if domain:
            # åŽ»æŽ‰wwwå‰ç¼€
            domain = re.sub(r'^www\.', '', domain)
            return domain
    except:
        pass
    
    return None


def detect_search_engine(referer):
    """æ£€æµ‹æœç´¢å¼•æ“Žæ¥æº"""
    if not referer:
        return None
    
    referer = referer.lower()
    
    # æœç´¢å¼•æ“Žæ£€æµ‹
    search_engines = [
        ('ç™¾åº¦', r'baidu\.com'),
        ('Google', r'google\.com|google\.cn'),
        ('æœç‹—', r'sogou\.com'),
        ('360æœç´¢', r'so\.com'),
        ('å¿…åº”', r'bing\.com'),
        ('é›…è™Ž', r'yahoo\.com'),
        ('ç¥žé©¬æœç´¢', r'sm\.cn'),
        ('å¤´æ¡æœç´¢', r'toutiao\.com')
    ]
    
    for engine_name, pattern in search_engines:
        if re.search(pattern, referer):
            return engine_name
    
    return None


def detect_social_media(referer):
    """æ£€æµ‹ç¤¾äº¤åª’ä½“æ¥æº"""
    if not referer:
        return None
    
    referer = referer.lower()
    
    # ç¤¾äº¤åª’ä½“æ£€æµ‹
    social_medias = [
        ('å¾®ä¿¡', r'weixin\.qq\.com'),
        ('å¾®åš', r'weibo\.com|t\.cn'),
        ('QQç©ºé—´', r'qzone\.qq\.com'),
        ('çŸ¥ä¹Ž', r'zhihu\.com'),
        ('è±†ç“£', r'douban\.com'),
        ('è´´å§', r'tieba\.baidu\.com'),
        ('Facebook', r'facebook\.com'),
        ('Twitter', r'twitter\.com|t\.co'),
        ('LinkedIn', r'linkedin\.com'),
        ('Instagram', r'instagram\.com')
    ]
    
    for social_name, pattern in social_medias:
        if re.search(pattern, referer):
            return social_name
    
    return None


def create_request_header_excel(analysis_results, output_path, top_n, total_processed):
    """åˆ›å»ºè¯·æ±‚å¤´åˆ†æžExcelæŠ¥å‘Š"""
    log_info(f"åˆ›å»ºè¯·æ±‚å¤´åˆ†æžExcelæŠ¥å‘Š: {output_path}")
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 1. æ¦‚è§ˆå·¥ä½œè¡¨
    create_header_overview_sheet(wb, analysis_results, total_processed)
    
    # 2. User-Agentåˆ†æžå·¥ä½œè¡¨
    create_user_agent_sheet(wb, analysis_results['user_agent_stats'], top_n)
    
    # 3. Refereråˆ†æžå·¥ä½œè¡¨
    create_referer_sheet(wb, analysis_results['referer_stats'], top_n)
    
    # 4. æµè§ˆå™¨ç»Ÿè®¡å·¥ä½œè¡¨
    create_browser_stats_sheet(wb, analysis_results['browser_stats'])
    
    # 5. æ“ä½œç³»ç»Ÿç»Ÿè®¡å·¥ä½œè¡¨
    create_os_stats_sheet(wb, analysis_results['os_stats'])
    
    # 6. è®¾å¤‡ç±»åž‹ç»Ÿè®¡å·¥ä½œè¡¨
    create_device_stats_sheet(wb, analysis_results['device_stats'])
    
    # 7. æ¥æºåŸŸåç»Ÿè®¡å·¥ä½œè¡¨
    create_domain_stats_sheet(wb, analysis_results['domain_stats'], top_n)
    
    # 8. æœç´¢å¼•æ“Žç»Ÿè®¡å·¥ä½œè¡¨
    create_search_engine_sheet(wb, analysis_results['search_engine_stats'])
    
    # 9. ç¤¾äº¤åª’ä½“ç»Ÿè®¡å·¥ä½œè¡¨
    create_social_media_sheet(wb, analysis_results['social_media_stats'])
    
    # 10. æœºå™¨äºº/çˆ¬è™«ç»Ÿè®¡å·¥ä½œè¡¨
    create_bot_stats_sheet(wb, analysis_results['bot_stats'])
    
    # ä¿å­˜æ–‡ä»¶
    wb.save(output_path)
    log_info(f"è¯·æ±‚å¤´åˆ†æžExcelæŠ¥å‘Šå·²ä¿å­˜: {output_path}")


def create_header_overview_sheet(wb, analysis_results, total_processed):
    """åˆ›å»ºæ¦‚è§ˆå·¥ä½œè¡¨"""
    ws = wb.create_sheet(title='æ¦‚è§ˆ')
    
    # åŸºç¡€ç»Ÿè®¡
    user_agent_count = len(analysis_results['user_agent_stats'])
    referer_count = len(analysis_results['referer_stats'])
    browser_count = len(analysis_results['browser_stats'])
    os_count = len(analysis_results['os_stats'])
    device_count = len(analysis_results['device_stats'])
    domain_count = len(analysis_results['domain_stats'])
    bot_count = len(analysis_results['bot_stats'])
    
    # æ¦‚è§ˆæ•°æ®
    overview_data = [
        ['=== åŸºç¡€ç»Ÿè®¡ ===', ''],
        ['æ€»å¤„ç†è®°å½•æ•°', total_processed],
        ['å”¯ä¸€User-Agentæ•°é‡', user_agent_count],
        ['å”¯ä¸€Refereræ•°é‡', referer_count],
        ['', ''],
        
        ['=== åˆ†ç±»ç»Ÿè®¡ ===', ''],
        ['æµè§ˆå™¨ç±»åž‹æ•°', browser_count],
        ['æ“ä½œç³»ç»Ÿç±»åž‹æ•°', os_count],
        ['è®¾å¤‡ç±»åž‹æ•°', device_count],
        ['æ¥æºåŸŸåæ•°', domain_count],
        ['æœºå™¨äºº/çˆ¬è™«ç±»åž‹æ•°', bot_count],
        ['', ''],
        
        ['=== TOPæµè§ˆå™¨ ===', ''],
    ]
    
    # æ·»åŠ TOPæµè§ˆå™¨
    for browser, count in Counter(analysis_results['browser_stats']).most_common(5):
        overview_data.append([browser, count])
    
    overview_data.extend([
        ['', ''],
        ['=== TOPæ¥æºåŸŸå ===', ''],
    ])
    
    # æ·»åŠ TOPæ¥æºåŸŸå
    for domain, count in Counter(analysis_results['domain_stats']).most_common(5):
        overview_data.append([domain, count])
    
    # å†™å…¥æ•°æ®
    for row_idx, (label, value) in enumerate(overview_data, start=1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_value = ws.cell(row=row_idx, column=2, value=value)
        
        if str(label).startswith('===') and str(label).endswith('==='):
            cell_label.font = Font(bold=True, size=12)
    
    # è®¾ç½®åˆ—å®½
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    format_excel_sheet(ws)


def create_user_agent_sheet(wb, user_agent_stats, top_n):
    """åˆ›å»ºUser-Agentåˆ†æžå·¥ä½œè¡¨"""
    if not user_agent_stats:
        return
    
    # è½¬æ¢ä¸ºDataFrame
    ua_data = []
    for ua, stats in user_agent_stats.items():
        ua_data.append({
            'User-Agent': ua[:100] + '...' if len(ua) > 100 else ua,  # æˆªæ–­è¿‡é•¿çš„UA
            'è¯·æ±‚æ¬¡æ•°': stats['count'],
            'å”¯ä¸€IPæ•°(ä¼°è®¡)': int(stats['unique_ips_hll'].cardinality()),
            'æˆåŠŸè¯·æ±‚æ•°': stats['success_requests'],
            'é”™è¯¯è¯·æ±‚æ•°': stats['error_requests'],
            'å¹³å‡å“åº”æ—¶é—´(ç§’)': round(stats['avg_response_time'], 3)
        })
    
    ua_df = pd.DataFrame(ua_data)
    ua_df = ua_df.sort_values(by='è¯·æ±‚æ¬¡æ•°', ascending=False).head(top_n)
    
    # è¡¨å¤´åˆ†ç»„
    ua_headers = {
        "åŸºç¡€ä¿¡æ¯": ["User-Agent"],
        "è¯·æ±‚ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å”¯ä¸€IPæ•°(ä¼°è®¡)"],
        "æ€§èƒ½æŒ‡æ ‡": ["æˆåŠŸè¯·æ±‚æ•°", "é”™è¯¯è¯·æ±‚æ•°", "å¹³å‡å“åº”æ—¶é—´(ç§’)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, ua_df, 'User-Agentåˆ†æž', header_groups=ua_headers
    )


def create_referer_sheet(wb, referer_stats, top_n):
    """åˆ›å»ºRefereråˆ†æžå·¥ä½œè¡¨"""
    if not referer_stats:
        return
    
    # è½¬æ¢ä¸ºDataFrame
    referer_data = []
    for referer, stats in referer_stats.items():
        referer_data.append({
            'Referer': referer[:100] + '...' if len(referer) > 100 else referer,
            'è¯·æ±‚æ¬¡æ•°': stats['count'],
            'å”¯ä¸€IPæ•°(ä¼°è®¡)': int(stats['unique_ips_hll'].cardinality()),
            'æˆåŠŸè¯·æ±‚æ•°': stats['success_requests'],
            'é”™è¯¯è¯·æ±‚æ•°': stats['error_requests'],
            'å¹³å‡å“åº”æ—¶é—´(ç§’)': round(stats['avg_response_time'], 3)
        })
    
    referer_df = pd.DataFrame(referer_data)
    referer_df = referer_df.sort_values(by='è¯·æ±‚æ¬¡æ•°', ascending=False).head(top_n)
    
    # è¡¨å¤´åˆ†ç»„
    referer_headers = {
        "åŸºç¡€ä¿¡æ¯": ["Referer"],
        "è¯·æ±‚ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å”¯ä¸€IPæ•°(ä¼°è®¡)"],
        "æ€§èƒ½æŒ‡æ ‡": ["æˆåŠŸè¯·æ±‚æ•°", "é”™è¯¯è¯·æ±‚æ•°", "å¹³å‡å“åº”æ—¶é—´(ç§’)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, referer_df, 'Refereråˆ†æž', header_groups=referer_headers
    )


def create_browser_stats_sheet(wb, browser_stats):
    """åˆ›å»ºæµè§ˆå™¨ç»Ÿè®¡å·¥ä½œè¡¨"""
    if not browser_stats:
        return
    
    ws = wb.create_sheet(title='æµè§ˆå™¨ç»Ÿè®¡')
    
    # è½¬æ¢ä¸ºDataFrame
    browser_data = []
    for browser, count in Counter(browser_stats).most_common():
        browser_data.append({
            'æµè§ˆå™¨': browser,
            'è¯·æ±‚æ¬¡æ•°': count,
            'å æ¯”(%)': round(count / sum(browser_stats.values()) * 100, 2)
        })
    
    browser_df = pd.DataFrame(browser_data)
    
    # è¡¨å¤´åˆ†ç»„
    browser_headers = {
        "åˆ†ç±»": ["æµè§ˆå™¨"],
        "ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å æ¯”(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, browser_df, 'æµè§ˆå™¨ç»Ÿè®¡', header_groups=browser_headers
    )
    
    # æ·»åŠ é¥¼å›¾
    if len(browser_df) > 1:
        chart_start_row = len(browser_df) + 5
        ws.cell(row=chart_start_row, column=1, value="æµè§ˆå™¨åˆ†å¸ƒå›¾").font = Font(bold=True)
        
        create_pie_chart(
            ws, "æµè§ˆå™¨ä½¿ç”¨åˆ†å¸ƒ",
            data_start_row=3, data_end_row=2 + len(browser_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_os_stats_sheet(wb, os_stats):
    """åˆ›å»ºæ“ä½œç³»ç»Ÿç»Ÿè®¡å·¥ä½œè¡¨"""
    if not os_stats:
        return
    
    # è½¬æ¢ä¸ºDataFrame
    os_data = []
    for os_name, count in Counter(os_stats).most_common():
        os_data.append({
            'æ“ä½œç³»ç»Ÿ': os_name,
            'è¯·æ±‚æ¬¡æ•°': count,
            'å æ¯”(%)': round(count / sum(os_stats.values()) * 100, 2)
        })
    
    os_df = pd.DataFrame(os_data)
    
    # è¡¨å¤´åˆ†ç»„
    os_headers = {
        "åˆ†ç±»": ["æ“ä½œç³»ç»Ÿ"],
        "ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å æ¯”(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, os_df, 'æ“ä½œç³»ç»Ÿç»Ÿè®¡', header_groups=os_headers
    )
    
    # æ·»åŠ é¥¼å›¾
    if len(os_df) > 1:
        create_pie_chart(
            ws, "æ“ä½œç³»ç»Ÿåˆ†å¸ƒ",
            data_start_row=3, data_end_row=2 + len(os_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_device_stats_sheet(wb, device_stats):
    """åˆ›å»ºè®¾å¤‡ç±»åž‹ç»Ÿè®¡å·¥ä½œè¡¨"""
    if not device_stats:
        return
    
    # è½¬æ¢ä¸ºDataFrame
    device_data = []
    for device, count in Counter(device_stats).most_common():
        device_data.append({
            'è®¾å¤‡ç±»åž‹': device,
            'è¯·æ±‚æ¬¡æ•°': count,
            'å æ¯”(%)': round(count / sum(device_stats.values()) * 100, 2)
        })
    
    device_df = pd.DataFrame(device_data)
    
    # è¡¨å¤´åˆ†ç»„
    device_headers = {
        "åˆ†ç±»": ["è®¾å¤‡ç±»åž‹"],
        "ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å æ¯”(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, device_df, 'è®¾å¤‡ç±»åž‹ç»Ÿè®¡', header_groups=device_headers
    )
    
    # æ·»åŠ é¥¼å›¾
    if len(device_df) > 1:
        create_pie_chart(
            ws, "è®¾å¤‡ç±»åž‹åˆ†å¸ƒ",
            data_start_row=3, data_end_row=2 + len(device_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_domain_stats_sheet(wb, domain_stats, top_n):
    """åˆ›å»ºæ¥æºåŸŸåç»Ÿè®¡å·¥ä½œè¡¨"""
    if not domain_stats:
        return
    
    # è½¬æ¢ä¸ºDataFrame
    domain_data = []
    for domain, count in Counter(domain_stats).most_common(top_n):
        domain_data.append({
            'æ¥æºåŸŸå': domain,
            'è¯·æ±‚æ¬¡æ•°': count,
            'å æ¯”(%)': round(count / sum(domain_stats.values()) * 100, 2)
        })
    
    domain_df = pd.DataFrame(domain_data)
    
    # è¡¨å¤´åˆ†ç»„
    domain_headers = {
        "åˆ†ç±»": ["æ¥æºåŸŸå"],
        "ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å æ¯”(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, domain_df, 'æ¥æºåŸŸåç»Ÿè®¡', header_groups=domain_headers
    )


def create_search_engine_sheet(wb, search_engine_stats):
    """åˆ›å»ºæœç´¢å¼•æ“Žç»Ÿè®¡å·¥ä½œè¡¨"""
    if not search_engine_stats:
        return
    
    # è½¬æ¢ä¸ºDataFrame
    search_data = []
    for engine, count in Counter(search_engine_stats).most_common():
        search_data.append({
            'æœç´¢å¼•æ“Ž': engine,
            'è¯·æ±‚æ¬¡æ•°': count,
            'å æ¯”(%)': round(count / sum(search_engine_stats.values()) * 100, 2)
        })
    
    search_df = pd.DataFrame(search_data)
    
    # è¡¨å¤´åˆ†ç»„
    search_headers = {
        "åˆ†ç±»": ["æœç´¢å¼•æ“Ž"],
        "ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å æ¯”(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, search_df, 'æœç´¢å¼•æ“Žç»Ÿè®¡', header_groups=search_headers
    )
    
    # æ·»åŠ é¥¼å›¾
    if len(search_df) > 1:
        create_pie_chart(
            ws, "æœç´¢å¼•æ“Žåˆ†å¸ƒ",
            data_start_row=3, data_end_row=2 + len(search_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_social_media_sheet(wb, social_media_stats):
    """åˆ›å»ºç¤¾äº¤åª’ä½“ç»Ÿè®¡å·¥ä½œè¡¨"""
    if not social_media_stats:
        return
    
    # è½¬æ¢ä¸ºDataFrame
    social_data = []
    for social, count in Counter(social_media_stats).most_common():
        social_data.append({
            'ç¤¾äº¤åª’ä½“': social,
            'è¯·æ±‚æ¬¡æ•°': count,
            'å æ¯”(%)': round(count / sum(social_media_stats.values()) * 100, 2)
        })
    
    social_df = pd.DataFrame(social_data)
    
    # è¡¨å¤´åˆ†ç»„
    social_headers = {
        "åˆ†ç±»": ["ç¤¾äº¤åª’ä½“"],
        "ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å æ¯”(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, social_df, 'ç¤¾äº¤åª’ä½“ç»Ÿè®¡', header_groups=social_headers
    )
    
    # æ·»åŠ é¥¼å›¾
    if len(social_df) > 1:
        create_pie_chart(
            ws, "ç¤¾äº¤åª’ä½“åˆ†å¸ƒ",
            data_start_row=3, data_end_row=2 + len(social_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_bot_stats_sheet(wb, bot_stats):
    """åˆ›å»ºæœºå™¨äºº/çˆ¬è™«ç»Ÿè®¡å·¥ä½œè¡¨"""
    if not bot_stats:
        return
    
    # è½¬æ¢ä¸ºDataFrame
    bot_data = []
    for bot, count in Counter(bot_stats).most_common():
        bot_data.append({
            'æœºå™¨äºº/çˆ¬è™«ç±»åž‹': bot,
            'è¯·æ±‚æ¬¡æ•°': count,
            'å æ¯”(%)': round(count / sum(bot_stats.values()) * 100, 2)
        })
    
    bot_df = pd.DataFrame(bot_data)
    
    # è¡¨å¤´åˆ†ç»„
    bot_headers = {
        "åˆ†ç±»": ["æœºå™¨äºº/çˆ¬è™«ç±»åž‹"],
        "ç»Ÿè®¡": ["è¯·æ±‚æ¬¡æ•°", "å æ¯”(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, bot_df, 'æœºå™¨äººçˆ¬è™«ç»Ÿè®¡', header_groups=bot_headers
    )
    
    # æ·»åŠ é¥¼å›¾
    if len(bot_df) > 1:
        create_pie_chart(
            ws, "æœºå™¨äºº/çˆ¬è™«åˆ†å¸ƒ",
            data_start_row=3, data_end_row=2 + len(bot_df),
            labels_col=1, values_col=2,
            position="E5"
        )