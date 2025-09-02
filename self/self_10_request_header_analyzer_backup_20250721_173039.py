import gc
import re
import pandas as pd
from collections import defaultdict, Counter
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font

from self_00_01_constants import DEFAULT_CHUNK_SIZE
from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    format_excel_sheet,
    add_dataframe_to_excel_with_grouped_headers,
    create_pie_chart,
    create_line_chart
)


def analyze_request_headers(csv_path, output_path, top_n=100):
    """分析请求头数据，包括User-Agent和Referer分析"""
    log_info("开始分析请求头数据...", show_memory=True)
    
    chunk_size = max(DEFAULT_CHUNK_SIZE // 2, 10000)
    
    # User-Agent分析数据
    user_agent_stats = defaultdict(lambda: {
        'count': 0,
        'unique_ips': set(),
        'success_requests': 0,
        'error_requests': 0,
        'total_response_time': 0.0,
        'avg_response_time': 0.0
    })
    
    # Referer分析数据
    referer_stats = defaultdict(lambda: {
        'count': 0,
        'unique_ips': set(),
        'success_requests': 0,
        'error_requests': 0,
        'total_response_time': 0.0,
        'avg_response_time': 0.0
    })
    
    # 浏览器类型分析
    browser_stats = defaultdict(int)
    os_stats = defaultdict(int)
    device_stats = defaultdict(int)
    
    # 来源域名分析
    domain_stats = defaultdict(int)
    
    # 搜索引擎分析
    search_engine_stats = defaultdict(int)
    
    # 社交媒体分析
    social_media_stats = defaultdict(int)
    
    # 机器人/爬虫分析
    bot_stats = defaultdict(int)
    
    total_processed = 0
    
    # 第一遍：收集统计数据
    log_info("开始收集请求头统计数据")
    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        chunk_size_actual = len(chunk)
        total_processed += chunk_size_actual
        
        # 处理User-Agent
        if 'user_agent_string' in chunk.columns:
            for _, row in chunk.iterrows():
                user_agent = row.get('user_agent_string', '')
                if pd.notna(user_agent) and user_agent != '' and user_agent != '-':
                    # 清理User-Agent字符串
                    user_agent = str(user_agent).strip()
                    
                    # 统计User-Agent
                    stats = user_agent_stats[user_agent]
                    stats['count'] += 1
                    
                    # 收集IP地址
                    ip = row.get('client_ip_address', '')
                    if pd.notna(ip) and ip != '':
                        stats['unique_ips'].add(ip)
                    
                    # 统计成功/失败请求
                    status = str(row.get('response_status_code', ''))
                    if status.startswith('2') or status.startswith('3'):
                        stats['success_requests'] += 1
                    elif status.startswith('4') or status.startswith('5'):
                        stats['error_requests'] += 1
                    
                    # 响应时间统计
                    response_time = row.get('total_request_duration', 0)
                    if pd.notna(response_time) and response_time > 0:
                        stats['total_response_time'] += float(response_time)
                    
                    # 分析浏览器、操作系统、设备类型
                    browser = extract_browser_info(user_agent)
                    os_info = extract_os_info(user_agent)
                    device = extract_device_info(user_agent)
                    
                    if browser:
                        browser_stats[browser] += 1
                    if os_info:
                        os_stats[os_info] += 1
                    if device:
                        device_stats[device] += 1
                    
                    # 检测机器人/爬虫
                    bot_type = detect_bot_type(user_agent)
                    if bot_type:
                        bot_stats[bot_type] += 1
        
        # 处理Referer
        if 'referer_url' in chunk.columns:
            for _, row in chunk.iterrows():
                referer = row.get('referer_url', '')
                if pd.notna(referer) and referer != '' and referer != '-':
                    # 清理Referer字符串
                    referer = str(referer).strip()
                    
                    # 统计Referer
                    stats = referer_stats[referer]
                    stats['count'] += 1
                    
                    # 收集IP地址
                    ip = row.get('client_ip_address', '')
                    if pd.notna(ip) and ip != '':
                        stats['unique_ips'].add(ip)
                    
                    # 统计成功/失败请求
                    status = str(row.get('response_status_code', ''))
                    if status.startswith('2') or status.startswith('3'):
                        stats['success_requests'] += 1
                    elif status.startswith('4') or status.startswith('5'):
                        stats['error_requests'] += 1
                    
                    # 响应时间统计
                    response_time = row.get('total_request_duration', 0)
                    if pd.notna(response_time) and response_time > 0:
                        stats['total_response_time'] += float(response_time)
                    
                    # 分析来源域名
                    domain = extract_domain_from_referer(referer)
                    if domain:
                        domain_stats[domain] += 1
                    
                    # 检测搜索引擎
                    search_engine = detect_search_engine(referer)
                    if search_engine:
                        search_engine_stats[search_engine] += 1
                    
                    # 检测社交媒体
                    social_media = detect_social_media(referer)
                    if social_media:
                        social_media_stats[social_media] += 1
        
        if total_processed % 100000 == 0:
            gc.collect()
            log_info(f"已处理 {total_processed:,} 条记录")
    
    # 计算平均响应时间
    for stats in user_agent_stats.values():
        if stats['success_requests'] > 0:
            stats['avg_response_time'] = stats['total_response_time'] / stats['success_requests']
    
    for stats in referer_stats.values():
        if stats['success_requests'] > 0:
            stats['avg_response_time'] = stats['total_response_time'] / stats['success_requests']
    
    log_info(f"请求头分析完成：总记录 {total_processed:,}，唯一User-Agent {len(user_agent_stats)}个，唯一Referer {len(referer_stats)}个")
    
    # 生成分析报告
    analysis_results = {
        'user_agent_stats': user_agent_stats,
        'referer_stats': referer_stats,
        'browser_stats': browser_stats,
        'os_stats': os_stats,
        'device_stats': device_stats,
        'domain_stats': domain_stats,
        'search_engine_stats': search_engine_stats,
        'social_media_stats': social_media_stats,
        'bot_stats': bot_stats
    }
    
    # 创建Excel报告
    create_request_header_excel(analysis_results, output_path, top_n, total_processed)
    
    log_info(f"请求头分析完成，报告已生成：{output_path}", show_memory=True)
    
    # 返回摘要信息
    return {
        'total_processed': total_processed,
        'unique_user_agents': len(user_agent_stats),
        'unique_referers': len(referer_stats),
        'top_browsers': dict(Counter(browser_stats).most_common(5)),
        'top_domains': dict(Counter(domain_stats).most_common(5))
    }


def extract_browser_info(user_agent):
    """从User-Agent中提取浏览器信息"""
    if not user_agent:
        return "未知浏览器"
    
    user_agent = user_agent.lower()
    
    # 浏览器检测模式
    browsers = [
        ('chrome', r'chrome/(\d+)'),
        ('firefox', r'firefox/(\d+)'),
        ('safari', r'safari/(\d+)'),
        ('edge', r'edge/(\d+)'),
        ('opera', r'opera/(\d+)'),
        ('internet explorer', r'msie (\d+)'),
        ('qq浏览器', r'qqbrowser/(\d+)'),
        ('搜狗浏览器', r'se 2\.x metasr'),
        ('360浏览器', r'360se|qhbrowser'),
        ('微信浏览器', r'micromessenger/(\d+)'),
        ('支付宝', r'alipayclient/(\d+)'),
        ('百度浏览器', r'baiduboxapp/(\d+)'),
        ('UC浏览器', r'ucbrowser/(\d+)')
    ]
    
    for browser_name, pattern in browsers:
        if re.search(pattern, user_agent):
            match = re.search(pattern, user_agent)
            if match and match.group(1) if '(' in pattern else True:
                version = match.group(1) if '(' in pattern and match.group(1) else ""
                return f"{browser_name}" + (f" {version}" if version else "")
    
    return "其他浏览器"


def extract_os_info(user_agent):
    """从User-Agent中提取操作系统信息"""
    if not user_agent:
        return "未知系统"
    
    user_agent = user_agent.lower()
    
    # 操作系统检测
    os_patterns = [
        ('Windows 10', r'windows nt 10\.0'),
        ('Windows 8.1', r'windows nt 6\.3'),
        ('Windows 8', r'windows nt 6\.2'),
        ('Windows 7', r'windows nt 6\.1'),
        ('Windows Vista', r'windows nt 6\.0'),
        ('Windows XP', r'windows nt 5\.1'),
        ('macOS', r'mac os x|macos'),
        ('iOS', r'iphone os|ios'),
        ('Android', r'android'),
        ('Linux', r'linux'),
        ('Ubuntu', r'ubuntu')
    ]
    
    for os_name, pattern in os_patterns:
        if re.search(pattern, user_agent):
            return os_name
    
    return "其他系统"


def extract_device_info(user_agent):
    """从User-Agent中提取设备信息"""
    if not user_agent:
        return "未知设备"
    
    user_agent = user_agent.lower()
    
    # 设备类型检测
    if re.search(r'mobile|phone|android|iphone', user_agent):
        return "手机"
    elif re.search(r'tablet|ipad', user_agent):
        return "平板"
    elif re.search(r'smart-tv|smarttv|television', user_agent):
        return "智能电视"
    elif re.search(r'bot|crawler|spider|scraper', user_agent):
        return "爬虫/机器人"
    else:
        return "桌面设备"


def detect_bot_type(user_agent):
    """检测机器人/爬虫类型"""
    if not user_agent:
        return None
    
    user_agent = user_agent.lower()
    
    # 常见爬虫检测
    bot_patterns = [
        ('Googlebot', r'googlebot'),
        ('Baiduspider', r'baiduspider'),
        ('Bingbot', r'bingbot'),
        ('Yahoo爬虫', r'yahoo.*slurp'),
        ('搜狗爬虫', r'sogou.*spider'),
        ('360爬虫', r'360spider'),
        ('微信爬虫', r'wechatbot'),
        ('Facebook爬虫', r'facebookexternalhit'),
        ('Twitter爬虫', r'twitterbot'),
        ('其他爬虫', r'bot|crawler|spider|scraper')
    ]
    
    for bot_name, pattern in bot_patterns:
        if re.search(pattern, user_agent):
            return bot_name
    
    return None


def extract_domain_from_referer(referer):
    """从Referer中提取域名"""
    if not referer or referer == '-':
        return None
    
    try:
        parsed = urlparse(referer)
        domain = parsed.netloc
        if domain:
            # 去掉www前缀
            domain = re.sub(r'^www\.', '', domain)
            return domain
    except:
        pass
    
    return None


def detect_search_engine(referer):
    """检测搜索引擎来源"""
    if not referer:
        return None
    
    referer = referer.lower()
    
    # 搜索引擎检测
    search_engines = [
        ('百度', r'baidu\.com'),
        ('Google', r'google\.com|google\.cn'),
        ('搜狗', r'sogou\.com'),
        ('360搜索', r'so\.com'),
        ('必应', r'bing\.com'),
        ('雅虎', r'yahoo\.com'),
        ('神马搜索', r'sm\.cn'),
        ('头条搜索', r'toutiao\.com')
    ]
    
    for engine_name, pattern in search_engines:
        if re.search(pattern, referer):
            return engine_name
    
    return None


def detect_social_media(referer):
    """检测社交媒体来源"""
    if not referer:
        return None
    
    referer = referer.lower()
    
    # 社交媒体检测
    social_medias = [
        ('微信', r'weixin\.qq\.com'),
        ('微博', r'weibo\.com|t\.cn'),
        ('QQ空间', r'qzone\.qq\.com'),
        ('知乎', r'zhihu\.com'),
        ('豆瓣', r'douban\.com'),
        ('贴吧', r'tieba\.baidu\.com'),
        ('Facebook', r'facebook\.com'),
        ('Twitter', r'twitter\.com|t\.co'),
        ('LinkedIn', r'linkedin\.com'),
        ('Instagram', r'instagram\.com')
    ]
    
    for social_name, pattern in social_medias:
        if re.search(pattern, referer):
            return social_name
    
    return None


def create_request_header_excel(analysis_results, output_path, top_n, total_processed):
    """创建请求头分析Excel报告"""
    log_info(f"创建请求头分析Excel报告: {output_path}")
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 1. 概览工作表
    create_header_overview_sheet(wb, analysis_results, total_processed)
    
    # 2. User-Agent分析工作表
    create_user_agent_sheet(wb, analysis_results['user_agent_stats'], top_n)
    
    # 3. Referer分析工作表
    create_referer_sheet(wb, analysis_results['referer_stats'], top_n)
    
    # 4. 浏览器统计工作表
    create_browser_stats_sheet(wb, analysis_results['browser_stats'])
    
    # 5. 操作系统统计工作表
    create_os_stats_sheet(wb, analysis_results['os_stats'])
    
    # 6. 设备类型统计工作表
    create_device_stats_sheet(wb, analysis_results['device_stats'])
    
    # 7. 来源域名统计工作表
    create_domain_stats_sheet(wb, analysis_results['domain_stats'], top_n)
    
    # 8. 搜索引擎统计工作表
    create_search_engine_sheet(wb, analysis_results['search_engine_stats'])
    
    # 9. 社交媒体统计工作表
    create_social_media_sheet(wb, analysis_results['social_media_stats'])
    
    # 10. 机器人/爬虫统计工作表
    create_bot_stats_sheet(wb, analysis_results['bot_stats'])
    
    # 保存文件
    wb.save(output_path)
    log_info(f"请求头分析Excel报告已保存: {output_path}")


def create_header_overview_sheet(wb, analysis_results, total_processed):
    """创建概览工作表"""
    ws = wb.create_sheet(title='概览')
    
    # 基础统计
    user_agent_count = len(analysis_results['user_agent_stats'])
    referer_count = len(analysis_results['referer_stats'])
    browser_count = len(analysis_results['browser_stats'])
    os_count = len(analysis_results['os_stats'])
    device_count = len(analysis_results['device_stats'])
    domain_count = len(analysis_results['domain_stats'])
    bot_count = len(analysis_results['bot_stats'])
    
    # 概览数据
    overview_data = [
        ['=== 基础统计 ===', ''],
        ['总处理记录数', total_processed],
        ['唯一User-Agent数量', user_agent_count],
        ['唯一Referer数量', referer_count],
        ['', ''],
        
        ['=== 分类统计 ===', ''],
        ['浏览器类型数', browser_count],
        ['操作系统类型数', os_count],
        ['设备类型数', device_count],
        ['来源域名数', domain_count],
        ['机器人/爬虫类型数', bot_count],
        ['', ''],
        
        ['=== TOP浏览器 ===', ''],
    ]
    
    # 添加TOP浏览器
    for browser, count in Counter(analysis_results['browser_stats']).most_common(5):
        overview_data.append([browser, count])
    
    overview_data.extend([
        ['', ''],
        ['=== TOP来源域名 ===', ''],
    ])
    
    # 添加TOP来源域名
    for domain, count in Counter(analysis_results['domain_stats']).most_common(5):
        overview_data.append([domain, count])
    
    # 写入数据
    for row_idx, (label, value) in enumerate(overview_data, start=1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_value = ws.cell(row=row_idx, column=2, value=value)
        
        if str(label).startswith('===') and str(label).endswith('==='):
            cell_label.font = Font(bold=True, size=12)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    format_excel_sheet(ws)


def create_user_agent_sheet(wb, user_agent_stats, top_n):
    """创建User-Agent分析工作表"""
    if not user_agent_stats:
        return
    
    # 转换为DataFrame
    ua_data = []
    for ua, stats in user_agent_stats.items():
        ua_data.append({
            'User-Agent': ua[:100] + '...' if len(ua) > 100 else ua,  # 截断过长的UA
            '请求次数': stats['count'],
            '唯一IP数': len(stats['unique_ips']),
            '成功请求数': stats['success_requests'],
            '错误请求数': stats['error_requests'],
            '平均响应时间(秒)': round(stats['avg_response_time'], 3)
        })
    
    ua_df = pd.DataFrame(ua_data)
    ua_df = ua_df.sort_values(by='请求次数', ascending=False).head(top_n)
    
    # 表头分组
    ua_headers = {
        "基础信息": ["User-Agent"],
        "请求统计": ["请求次数", "唯一IP数"],
        "性能指标": ["成功请求数", "错误请求数", "平均响应时间(秒)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, ua_df, 'User-Agent分析', header_groups=ua_headers
    )


def create_referer_sheet(wb, referer_stats, top_n):
    """创建Referer分析工作表"""
    if not referer_stats:
        return
    
    # 转换为DataFrame
    referer_data = []
    for referer, stats in referer_stats.items():
        referer_data.append({
            'Referer': referer[:100] + '...' if len(referer) > 100 else referer,
            '请求次数': stats['count'],
            '唯一IP数': len(stats['unique_ips']),
            '成功请求数': stats['success_requests'],
            '错误请求数': stats['error_requests'],
            '平均响应时间(秒)': round(stats['avg_response_time'], 3)
        })
    
    referer_df = pd.DataFrame(referer_data)
    referer_df = referer_df.sort_values(by='请求次数', ascending=False).head(top_n)
    
    # 表头分组
    referer_headers = {
        "基础信息": ["Referer"],
        "请求统计": ["请求次数", "唯一IP数"],
        "性能指标": ["成功请求数", "错误请求数", "平均响应时间(秒)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, referer_df, 'Referer分析', header_groups=referer_headers
    )


def create_browser_stats_sheet(wb, browser_stats):
    """创建浏览器统计工作表"""
    if not browser_stats:
        return
    
    ws = wb.create_sheet(title='浏览器统计')
    
    # 转换为DataFrame
    browser_data = []
    for browser, count in Counter(browser_stats).most_common():
        browser_data.append({
            '浏览器': browser,
            '请求次数': count,
            '占比(%)': round(count / sum(browser_stats.values()) * 100, 2)
        })
    
    browser_df = pd.DataFrame(browser_data)
    
    # 表头分组
    browser_headers = {
        "分类": ["浏览器"],
        "统计": ["请求次数", "占比(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, browser_df, '浏览器统计', header_groups=browser_headers
    )
    
    # 添加饼图
    if len(browser_df) > 1:
        chart_start_row = len(browser_df) + 5
        ws.cell(row=chart_start_row, column=1, value="浏览器分布图").font = Font(bold=True)
        
        create_pie_chart(
            ws, "浏览器使用分布",
            data_start_row=3, data_end_row=2 + len(browser_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_os_stats_sheet(wb, os_stats):
    """创建操作系统统计工作表"""
    if not os_stats:
        return
    
    # 转换为DataFrame
    os_data = []
    for os_name, count in Counter(os_stats).most_common():
        os_data.append({
            '操作系统': os_name,
            '请求次数': count,
            '占比(%)': round(count / sum(os_stats.values()) * 100, 2)
        })
    
    os_df = pd.DataFrame(os_data)
    
    # 表头分组
    os_headers = {
        "分类": ["操作系统"],
        "统计": ["请求次数", "占比(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, os_df, '操作系统统计', header_groups=os_headers
    )
    
    # 添加饼图
    if len(os_df) > 1:
        create_pie_chart(
            ws, "操作系统分布",
            data_start_row=3, data_end_row=2 + len(os_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_device_stats_sheet(wb, device_stats):
    """创建设备类型统计工作表"""
    if not device_stats:
        return
    
    # 转换为DataFrame
    device_data = []
    for device, count in Counter(device_stats).most_common():
        device_data.append({
            '设备类型': device,
            '请求次数': count,
            '占比(%)': round(count / sum(device_stats.values()) * 100, 2)
        })
    
    device_df = pd.DataFrame(device_data)
    
    # 表头分组
    device_headers = {
        "分类": ["设备类型"],
        "统计": ["请求次数", "占比(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, device_df, '设备类型统计', header_groups=device_headers
    )
    
    # 添加饼图
    if len(device_df) > 1:
        create_pie_chart(
            ws, "设备类型分布",
            data_start_row=3, data_end_row=2 + len(device_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_domain_stats_sheet(wb, domain_stats, top_n):
    """创建来源域名统计工作表"""
    if not domain_stats:
        return
    
    # 转换为DataFrame
    domain_data = []
    for domain, count in Counter(domain_stats).most_common(top_n):
        domain_data.append({
            '来源域名': domain,
            '请求次数': count,
            '占比(%)': round(count / sum(domain_stats.values()) * 100, 2)
        })
    
    domain_df = pd.DataFrame(domain_data)
    
    # 表头分组
    domain_headers = {
        "分类": ["来源域名"],
        "统计": ["请求次数", "占比(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, domain_df, '来源域名统计', header_groups=domain_headers
    )


def create_search_engine_sheet(wb, search_engine_stats):
    """创建搜索引擎统计工作表"""
    if not search_engine_stats:
        return
    
    # 转换为DataFrame
    search_data = []
    for engine, count in Counter(search_engine_stats).most_common():
        search_data.append({
            '搜索引擎': engine,
            '请求次数': count,
            '占比(%)': round(count / sum(search_engine_stats.values()) * 100, 2)
        })
    
    search_df = pd.DataFrame(search_data)
    
    # 表头分组
    search_headers = {
        "分类": ["搜索引擎"],
        "统计": ["请求次数", "占比(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, search_df, '搜索引擎统计', header_groups=search_headers
    )
    
    # 添加饼图
    if len(search_df) > 1:
        create_pie_chart(
            ws, "搜索引擎分布",
            data_start_row=3, data_end_row=2 + len(search_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_social_media_sheet(wb, social_media_stats):
    """创建社交媒体统计工作表"""
    if not social_media_stats:
        return
    
    # 转换为DataFrame
    social_data = []
    for social, count in Counter(social_media_stats).most_common():
        social_data.append({
            '社交媒体': social,
            '请求次数': count,
            '占比(%)': round(count / sum(social_media_stats.values()) * 100, 2)
        })
    
    social_df = pd.DataFrame(social_data)
    
    # 表头分组
    social_headers = {
        "分类": ["社交媒体"],
        "统计": ["请求次数", "占比(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, social_df, '社交媒体统计', header_groups=social_headers
    )
    
    # 添加饼图
    if len(social_df) > 1:
        create_pie_chart(
            ws, "社交媒体分布",
            data_start_row=3, data_end_row=2 + len(social_df),
            labels_col=1, values_col=2,
            position="E5"
        )


def create_bot_stats_sheet(wb, bot_stats):
    """创建机器人/爬虫统计工作表"""
    if not bot_stats:
        return
    
    # 转换为DataFrame
    bot_data = []
    for bot, count in Counter(bot_stats).most_common():
        bot_data.append({
            '机器人/爬虫类型': bot,
            '请求次数': count,
            '占比(%)': round(count / sum(bot_stats.values()) * 100, 2)
        })
    
    bot_df = pd.DataFrame(bot_data)
    
    # 表头分组
    bot_headers = {
        "分类": ["机器人/爬虫类型"],
        "统计": ["请求次数", "占比(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, bot_df, '机器人爬虫统计', header_groups=bot_headers
    )
    
    # 添加饼图
    if len(bot_df) > 1:
        create_pie_chart(
            ws, "机器人/爬虫分布",
            data_start_row=3, data_end_row=2 + len(bot_df),
            labels_col=1, values_col=2,
            position="E5"
        )