from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def create_http_lifecycle_visualization(wb):
    """
    创建HTTP请求生命周期分析工作表 - 优化版
    包含时序图和参数映射可视化
    """
    ws = wb.create_sheet(title='HTTP请求生命周期分析')

    # 配置样式
    styles = _setup_styles()

    # 阶段颜色配置
    phase_colors = {
        '后端连接': 'FFE6E6',  # 浅红色
        '后端处理': 'E6FFE6',  # 浅绿色
        '后端传输': 'E6E6FF',  # 浅蓝色
        'Nginx传输': 'FFFFE6'  # 浅黄色
    }

    current_row = 1

    # 标题
    ws.cell(row=current_row, column=1, value="HTTP请求生命周期性能分析").font = styles['header']
    current_row += 1
    ws.cell(row=current_row, column=1, value="基于Nginx日志的完整请求链路分析").font = styles['italic']
    current_row += 3

    # 时序图
    current_row = _create_timeline_diagram(ws, current_row, styles, phase_colors)

    # 参数映射可视化
    current_row = _create_parameter_mapping_visual(ws, current_row, styles, phase_colors)

    # 核心指标解析
    current_row = _create_core_metrics_table(ws, current_row, styles)

    # 性能分析指标
    current_row = _create_performance_metrics_table(ws, current_row, styles)

    # 调优指南
    current_row = _create_tuning_guide(ws, current_row, styles, phase_colors)

    # 应用格式
    _apply_formatting(ws, styles['border'])

    return ws


def _setup_styles():
    """设置样式配置"""
    return {
        'header': Font(size=14, bold=True),
        'subheader': Font(size=12, bold=True),
        'bold': Font(bold=True),
        'italic': Font(italic=True),
        'small': Font(size=10),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    }


def _create_timeline_diagram(ws, start_row, styles, phase_colors):
    """创建时序图 - 展示完整的11个阶段"""
    ws.cell(row=start_row, column=1, value="HTTP请求处理时序图").font = styles['subheader']
    start_row += 2

    # 阶段定义
    stages = [
        "1.TCP连接", "2.发送请求", "3.后端连接", "4.请求转发",
        "5.业务处理", "6.响应头", "7.响应体", "8.响应接收",
        "9.客户端传输", "10.响应完成", "11.连接断开"
    ]

    # 第一行：阶段编号和名称
    for i, stage in enumerate(stages):
        cell = ws.cell(row=start_row, column=i + 1, value=stage)
        cell.font = styles['small']
        cell.alignment = Alignment(horizontal='center', vertical='center')

    start_row += 1

    # 参数时间线可视化
    timeline_params = [
        ("total_request_duration", "请求总时长", 2, 10, 'DDDDDD'),
        ("upstream_response_time", "后端响应时长", 3, 7, 'BBBBBB'),
        ("upstream_header_time", "后端处理时长", 3, 6, '999999'),
        ("upstream_connect_time", "后端连接时长", 3, 3, '666666')
    ]

    for param_name, param_desc, start_stage, end_stage, color in timeline_params:
        # 参数名称
        ws.cell(row=start_row, column=1, value=f"{param_desc}")

        # 时间线可视化
        for col in range(1, 12):  # 11个阶段
            cell = ws.cell(row=start_row, column=col)
            if start_stage <= col <= end_stage:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                if col == start_stage:
                    cell.value = "●"
                elif col == end_stage:
                    cell.value = "●"
                else:
                    cell.value = "━"
            cell.alignment = Alignment(horizontal='center')

        start_row += 1

    return start_row + 2


def _create_parameter_mapping_visual(ws, start_row, styles, phase_colors):
    """创建参数映射可视化"""
    ws.cell(row=start_row, column=1, value="阶段参数映射").font = styles['subheader']
    start_row += 2

    # 阶段参数映射
    stage_mappings = [
        ("backend_connect_phase", "后端连接阶段", 3, 3, phase_colors['后端连接']),
        ("backend_process_phase", "后端处理阶段", 4, 6, phase_colors['后端处理']),
        ("backend_transfer_phase", "后端传输阶段", 7, 7, phase_colors['后端传输']),
        ("nginx_transfer_phase", "Nginx传输阶段", 9, 9, phase_colors['Nginx传输'])
    ]

    # 创建阶段标尺
    for i in range(1, 12):
        cell = ws.cell(row=start_row, column=i, value=str(i))
        cell.font = styles['small']
        cell.alignment = Alignment(horizontal='center')

    start_row += 1

    # 绘制阶段映射
    for param_name, param_desc, start_stage, end_stage, color in stage_mappings:
        ws.cell(row=start_row, column=1, value=param_desc)

        for col in range(1, 12):
            cell = ws.cell(row=start_row, column=col)
            if start_stage <= col <= end_stage:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.value = "█"
            cell.alignment = Alignment(horizontal='center')

        start_row += 1

    return start_row + 2


def _create_core_metrics_table(ws, start_row, styles):
    """创建核心指标表格"""
    ws.cell(row=start_row, column=1, value="核心时间指标").font = styles['subheader']
    start_row += 1

    metrics_data = [
        ['参数名称', '中文描述', '计算公式', '性能含义'],
        ['total_request_duration', '请求总时长', '直接获取', '完整请求处理时间'],
        ['upstream_response_time', '后端响应时长', '直接获取', '后端服务总耗时'],
        ['upstream_header_time', '后端处理时长', '直接获取', '后端业务处理耗时'],
        ['upstream_connect_time', '后端连接时长', '直接获取', '网络连接建立耗时'],
        ['backend_connect_phase', '后端连接阶段', 'upstream_connect_time', '连接性能指标'],
        ['backend_process_phase', '后端处理阶段', 'upstream_header_time - upstream_connect_time', '处理效率指标'],
        ['backend_transfer_phase', '后端传输阶段', 'upstream_response_time - upstream_header_time', '传输效率指标'],
        ['nginx_transfer_phase', 'Nginx传输阶段', 'total_request_duration - upstream_response_time', '代理性能指标']
    ]

    return _create_simple_table(ws, start_row, metrics_data, styles) + 2


def _create_performance_metrics_table(ws, start_row, styles):
    """创建性能分析指标表格"""
    ws.cell(row=start_row, column=1, value="性能分析指标").font = styles['subheader']
    start_row += 1

    performance_data = [
        ['指标类型', '参数名称', '计算公式', '健康范围'],
        ['效率比率', 'backend_efficiency', '(backend_process_phase / upstream_response_time) × 100%', '> 60%'],
        ['网络开销', 'network_overhead', '(network_phase / total_request_duration) × 100%', '< 30%'],
        ['传输占比', 'transfer_ratio', '(transfer_phase / total_request_duration) × 100%', '< 40%'],
        ['连接成本', 'connection_cost_ratio', '(backend_connect_phase / total_request_duration) × 100%', '< 10%'],
        ['传输效率', 'response_transfer_speed', 'response_body_size_kb / backend_transfer_phase', '> 1000 KB/s'],
        ['整体速率', 'total_transfer_speed', 'total_bytes_sent_kb / total_request_duration', '> 500 KB/s']
    ]

    return _create_simple_table(ws, start_row, performance_data, styles) + 2


def _create_tuning_guide(ws, start_row, styles, phase_colors):
    """创建调优指南"""
    ws.cell(row=start_row, column=1, value="性能调优指南").font = styles['subheader']
    start_row += 1

    tuning_data = [
        ['问题阶段', '异常症状', '优化方向'],
        ['后端连接', 'backend_connect_phase > 100ms', '启用连接池、网络优化'],
        ['后端处理', 'backend_efficiency < 60%', '代码优化、添加缓存'],
        ['后端传输', 'response_transfer_speed < 1000KB/s', '数据压缩、分页返回'],
        ['Nginx传输', 'nginx_transfer_speed < 500KB/s', '启用gzip、调整缓冲区']
    ]

    # 创建表格并应用颜色
    for i, row_data in enumerate(tuning_data):
        for j, cell_value in enumerate(row_data):
            cell = ws.cell(row=start_row + i, column=j + 1, value=cell_value)
            if i == 0:
                cell.font = styles['bold']
            elif j == 0 and i > 0:  # 阶段列着色
                for phase_key, color in phase_colors.items():
                    if phase_key in str(cell_value):
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        break

    return start_row + len(tuning_data) + 2


def _create_simple_table(ws, start_row, data, styles):
    """创建简单表格"""
    for i, row_data in enumerate(data):
        for j, cell_value in enumerate(row_data):
            cell = ws.cell(row=start_row + i, column=j + 1, value=cell_value)
            if i == 0:  # 表头
                cell.font = styles['bold']
    return start_row + len(data)


def _apply_formatting(ws, border):
    """应用工作表格式"""
    # 设置列宽
    column_widths = [35, 30, 45, 25, 20, 20, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # 应用边框
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.border = border