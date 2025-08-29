"""
优化版API性能分析器 - 使用先进采样算法
集成T-Digest、蓄水池采样、Count-Min Sketch等算法
提供更准确的分位数估计和更高的内存效率

优化内容：
1. T-Digest算法用于响应时间分位数估计
2. 蓄水池采样用于需要原始数据的指标
3. Count-Min Sketch用于API频率统计
4. HyperLogLog用于独立IP统计
5. 分层采样用于时间维度分析
6. 自适应采样策略

Author: Claude Code (Optimized)
Date: 2025-07-18
"""

import gc
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Any, Optional

from self_00_04_excel_processor import format_excel_sheet, add_dataframe_to_excel_with_grouped_headers
from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD, DEFAULT_SLOW_REQUESTS_THRESHOLD, \
    TIME_METRICS, SIZE_METRICS, HIGHLIGHT_FILL
from self_00_02_utils import log_info, get_distribution_stats, calculate_time_percentages
from self_00_05_sampling_algorithms import (
    TDigest, ReservoirSampler, CountMinSketch, HyperLogLog, 
    StratifiedSampler, AdaptiveSampler
)

# 尝试导入scipy，如果失败则使用近似计算
try:
    from scipy.stats import norm
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False


class AdvancedStreamingApiAnalyzer:
    """
    高级流式API性能分析器
    使用多种采样算法提供准确和高效的分析
    """
    
    def __init__(self, slow_threshold=DEFAULT_SLOW_THRESHOLD):
        """
        初始化分析器
        
        Args:
            slow_threshold: 慢请求阈值（秒）
        """
        self.slow_threshold = slow_threshold
        
        # 每个API的统计信息
        self.api_stats = defaultdict(lambda: {
            'total_requests': 0,
            'success_requests': 0,
            'error_requests': 0,  # 新增：错误请求数
            'slow_requests': 0,
            'app_name': '',
            'service_name': '',
            
            # 使用T-Digest进行响应时间分析
            'response_time_digest': TDigest(compression=100),
            
            # 使用蓄水池采样保存样本
            'response_time_reservoir': ReservoirSampler(500),
            'body_size_reservoir': ReservoirSampler(500),
            'bytes_size_reservoir': ReservoirSampler(500),
            
            # 流式统计（用于精确计算均值、方差）
            'request_time_sum': 0.0,
            'request_time_sum_sq': 0.0,
            'request_time_count': 0,
            
            # 阶段时间统计
            'backend_connect_digest': TDigest(compression=50),
            'backend_process_digest': TDigest(compression=50),
            'backend_transfer_digest': TDigest(compression=50),
            'nginx_transfer_digest': TDigest(compression=50),
            
            # 性能指标
            'transfer_speed_reservoir': ReservoirSampler(300),
            'efficiency_reservoir': ReservoirSampler(300)
        })
        
        # 全局统计
        self.global_stats = {
            'total_requests': 0,
            'success_requests': 0,
            'error_requests': 0,  # 新增：全局错误请求数
            'slow_requests': 0,
            
            # 全局T-Digest
            'global_response_time_digest': TDigest(compression=200),
            'global_body_size_digest': TDigest(compression=200),
            'global_bytes_size_digest': TDigest(compression=200),
            
            # API频率统计
            'api_frequency': CountMinSketch(width=2000, depth=7),
            
            # 独立IP统计
            'unique_ips': HyperLogLog(precision=12),
            
            # 分层采样（按小时）
            'hourly_stratified': StratifiedSampler(samples_per_stratum=200),
            
            # 自适应采样
            'adaptive_sampler': AdaptiveSampler(initial_sample_size=1000, adaptation_threshold=50000)
        }
        
        # 性能监控
        self.processing_stats = {
            'chunks_processed': 0,
            'total_records': 0,
            'memory_peaks': [],
            'processing_times': []
        }
    
    def process_chunk(self, chunk, field_mapping, success_codes):
        """
        处理单个数据块
        
        Args:
            chunk: 数据块
            field_mapping: 字段映射
            success_codes: 成功状态码列表
        """
        start_time = datetime.now()
        chunk_rows = len(chunk)
        
        self.global_stats['total_requests'] += chunk_rows
        self.processing_stats['total_records'] += chunk_rows
        self.processing_stats['chunks_processed'] += 1
        
        # 先按API分组统计所有请求（包括失败请求）
        all_requests_data = self._preprocess_all_requests_data(chunk, field_mapping)
        
        # 按API分组处理所有请求
        for api, group_data in all_requests_data.groupby('uri'):
            # 分别统计成功和失败请求
            success_mask = group_data['status'].astype(str).isin(success_codes)
            successful_group = group_data[success_mask]
            
            # 更新API级别的总请求统计
            api_stats = self.api_stats[api]
            total_count = len(group_data)
            success_count = len(successful_group)
            error_count = total_count - success_count
            
            api_stats['total_requests'] += total_count
            api_stats['success_requests'] += success_count
            api_stats['error_requests'] += error_count
            
            # 设置应用和服务名称
            if not api_stats['app_name'] and 'app' in group_data.columns and not group_data['app'].isna().all():
                api_stats['app_name'] = str(group_data['app'].iloc[0])
            if not api_stats['service_name'] and 'service' in group_data.columns and not group_data['service'].isna().all():
                api_stats['service_name'] = str(group_data['service'].iloc[0])
            
            # 只对成功请求进行性能分析
            if len(successful_group) > 0:
                # 提取时间戳
                if 'timestamp' in successful_group.columns:
                    timestamps = pd.to_datetime(successful_group['timestamp'], errors='coerce').reset_index(drop=True)
                else:
                    timestamps = pd.Series([datetime.now()] * len(successful_group))
                
                # 重置索引并添加时间戳
                successful_group_clean = successful_group.reset_index(drop=True).drop('timestamp', axis=1, errors='ignore')
                successful_group_clean['timestamp'] = timestamps
                
                # 处理成功请求的性能数据
                group_timestamps = successful_group_clean['timestamp']
                performance_data = successful_group_clean.drop('timestamp', axis=1)
                self._process_api_group_advanced(api, performance_data, field_mapping, group_timestamps, update_basic_stats=False)
        
        # 更新全局成功请求统计
        successful_requests = chunk[chunk[field_mapping['status']].astype(str).isin(success_codes)]
        global_success_count = len(successful_requests)
        global_error_count = chunk_rows - global_success_count
        
        self.global_stats['success_requests'] += global_success_count
        self.global_stats['error_requests'] += global_error_count
        
        # 记录处理时间
        processing_time = (datetime.now() - start_time).total_seconds()
        self.processing_stats['processing_times'].append(processing_time)
        
        # 定期垃圾回收
        if self.processing_stats['chunks_processed'] % 50 == 0:
            gc.collect()
    
    def _preprocess_all_requests_data(self, chunk, field_mapping):
        """预处理所有请求数据（包括失败请求）"""
        cols_to_process = {
            'uri': field_mapping['uri'],
            'app': field_mapping['app'],
            'service': field_mapping['service'],
            'status': field_mapping['status'],
            'request_time': field_mapping['request_time'],
            'client_ip': field_mapping.get('client_ip', ''),
            'timestamp': field_mapping.get('timestamp', '')
        }
        
        data = {}
        for key, col in cols_to_process.items():
            # 确保col不为空且存在于chunk中
            if col and str(col).strip() and col in chunk.columns:
                try:
                    if key in ['request_time']:
                        # 数值字段
                        data[key] = pd.to_numeric(chunk[col], errors='coerce').fillna(0)
                    else:
                        # 字符串字段
                        data[key] = chunk[col].fillna('').astype(str)
                except Exception as e:
                    # 如果处理失败，使用默认值
                    if key == 'request_time':
                        data[key] = pd.Series([0.0] * len(chunk))
                    else:
                        data[key] = pd.Series([''] * len(chunk))
            else:
                # 如果字段不存在，提供默认值
                if key == 'request_time':
                    data[key] = pd.Series([0.0] * len(chunk))
                else:
                    data[key] = pd.Series([''] * len(chunk))
        
        return pd.DataFrame(data)
    
    def _preprocess_numeric_data(self, chunk, field_mapping):
        """预处理数字数据"""
        cols_to_process = {
            'uri': field_mapping['uri'],
            'app': field_mapping['app'],
            'service': field_mapping['service'],
            'request_time': field_mapping['request_time'],
            'backend_connect': field_mapping.get('backend_connect_phase', ''),
            'backend_process': field_mapping.get('backend_process_phase', ''),
            'backend_transfer': field_mapping.get('backend_transfer_phase', ''),
            'nginx_transfer': field_mapping.get('nginx_transfer_phase', ''),
            'body_size': field_mapping['body_bytes_kb'],
            'bytes_size': field_mapping['bytes_sent_kb'],
            'transfer_speed': field_mapping.get('response_transfer_speed', ''),
            'efficiency': field_mapping.get('processing_efficiency_index', ''),
            'client_ip': field_mapping.get('client_ip', '')
        }
        
        data = {}
        for key, col in cols_to_process.items():
            # 确保col不为空且存在于chunk中
            if col and str(col).strip() and col in chunk.columns:
                try:
                    if key in ['uri', 'app', 'service', 'client_ip']:
                        data[key] = chunk[col].fillna('').astype(str)
                    else:
                        data[key] = pd.to_numeric(chunk[col], errors='coerce').fillna(0)
                except Exception as e:
                    # 如果处理失败，使用默认值
                    if key in ['uri', 'app', 'service', 'client_ip']:
                        data[key] = pd.Series([''] * len(chunk))
                    else:
                        data[key] = pd.Series([0.0] * len(chunk))
            else:
                # 字段不存在时使用默认值
                if key in ['uri', 'app', 'service', 'client_ip']:
                    data[key] = pd.Series([''] * len(chunk))
                else:
                    data[key] = pd.Series([0.0] * len(chunk))
        
        return pd.DataFrame(data)
    
    def _process_api_group_advanced(self, api, group_data, field_mapping, timestamps, update_basic_stats=True):
        """
        使用高级算法处理API组数据
        
        Args:
            api: API标识
            group_data: 组数据
            field_mapping: 字段映射
            timestamps: 时间戳序列
            update_basic_stats: 是否更新基础统计（默认True，修复后设为False）
        """
        group_size = len(group_data)
        stats = self.api_stats[api]
        
        # 只在旧逻辑中更新基础统计（为了兼容性）
        if update_basic_stats:
            stats['total_requests'] += group_size
            stats['success_requests'] += group_size
        
        # 设置应用和服务名称
        if not stats['app_name'] and not group_data['app'].isna().all():
            stats['app_name'] = str(group_data['app'].iloc[0])
        if not stats['service_name'] and not group_data['service'].isna().all():
            stats['service_name'] = str(group_data['service'].iloc[0])
        
        # 处理响应时间 - 使用T-Digest和蓄水池采样
        request_times = group_data['request_time'].dropna()
        if len(request_times) > 0:
            # T-Digest更新（用于分位数）
            stats['response_time_digest'].add_batch(request_times.tolist())
            self.global_stats['global_response_time_digest'].add_batch(request_times.tolist())
            
            # 蓄水池采样更新（保留原始数据）
            stats['response_time_reservoir'].add_batch(request_times.tolist())
            
            # 流式统计更新（用于精确均值计算）
            stats['request_time_sum'] += request_times.sum()
            stats['request_time_sum_sq'] += (request_times ** 2).sum()
            stats['request_time_count'] += len(request_times)
            
            # 慢请求统计
            slow_count = (request_times > self.slow_threshold).sum()
            stats['slow_requests'] += slow_count
            self.global_stats['slow_requests'] += slow_count
            
            # 自适应采样
            for rt in request_times:
                self.global_stats['adaptive_sampler'].add(rt)
        
        # 处理阶段时间 - 使用T-Digest
        phase_fields = {
            'backend_connect': 'backend_connect_digest',
            'backend_process': 'backend_process_digest', 
            'backend_transfer': 'backend_transfer_digest',
            'nginx_transfer': 'nginx_transfer_digest'
        }
        
        for field, digest_key in phase_fields.items():
            if field in group_data.columns:
                phase_data = group_data[field].dropna()
                if len(phase_data) > 0:
                    stats[digest_key].add_batch(phase_data.tolist())
        
        # 处理大小数据 - 使用T-Digest和蓄水池采样
        size_fields = ['body_size', 'bytes_size']
        for field in size_fields:
            if field in group_data.columns:
                size_data = group_data[field].dropna()
                if len(size_data) > 0:
                    # 蓄水池采样
                    stats[f'{field}_reservoir'].add_batch(size_data.tolist())
                    
                    # 全局T-Digest
                    if field == 'body_size':
                        self.global_stats['global_body_size_digest'].add_batch(size_data.tolist())
                    elif field == 'bytes_size':
                        self.global_stats['global_bytes_size_digest'].add_batch(size_data.tolist())
        
        # 处理性能指标
        perf_fields = ['transfer_speed', 'efficiency']
        for field in perf_fields:
            if field in group_data.columns:
                perf_data = group_data[field].dropna()
                if len(perf_data) > 0:
                    stats[f'{field}_reservoir'].add_batch(perf_data.tolist())
        
        # API频率统计
        self.global_stats['api_frequency'].increment(api, group_size)
        
        # 独立IP统计
        if 'client_ip' in group_data.columns:
            unique_ips = group_data['client_ip'].dropna().unique()
            for ip in unique_ips:
                if ip and str(ip) != '':
                    self.global_stats['unique_ips'].add(str(ip))
        
        # 分层采样（按小时）
        for i, timestamp in enumerate(timestamps):
            if pd.notna(timestamp) and i < len(request_times):
                hour_key = f"{timestamp.hour:02d}"
                self.global_stats['hourly_stratified'].add(
                    request_times.iloc[i], hour_key
                )
    
    def get_analysis_summary(self) -> Dict[str, Any]:
        """获取分析总结"""
        return {
            'total_apis': len(self.api_stats),
            'total_requests': self.global_stats['total_requests'],
            'success_requests': self.global_stats['success_requests'],
            'slow_requests': self.global_stats['slow_requests'],
            'unique_ips_estimate': self.global_stats['unique_ips'].cardinality(),
            'chunks_processed': self.processing_stats['chunks_processed'],
            'avg_processing_time': np.mean(self.processing_stats['processing_times']) if self.processing_stats['processing_times'] else 0,
            'memory_efficiency': self._calculate_memory_efficiency()
        }
    
    def _calculate_memory_efficiency(self) -> Dict[str, float]:
        """计算内存效率指标"""
        total_records = self.processing_stats['total_records']
        if total_records == 0:
            return {'efficiency_score': 0.0}
        
        # 估算传统方法需要的内存
        traditional_memory_mb = total_records * 0.001  # 假设每条记录1KB
        
        # 估算当前方法使用的内存
        current_memory_mb = (
            len(self.api_stats) * 0.05 +  # 每个API约50KB
            0.1 +  # 全局T-Digest
            0.05   # 其他算法
        )
        
        efficiency_ratio = traditional_memory_mb / max(current_memory_mb, 0.001)
        
        return {
            'traditional_memory_mb': traditional_memory_mb,
            'current_memory_mb': current_memory_mb,
            'efficiency_ratio': efficiency_ratio,
            'memory_savings_percent': (1 - current_memory_mb / traditional_memory_mb) * 100 if traditional_memory_mb > 0 else 0
        }


def analyze_api_performance_advanced(csv_path, output_path, success_codes=None, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """
    高级API性能分析函数
    
    Args:
        csv_path: CSV文件路径
        output_path: 输出路径
        success_codes: 成功状态码列表
        slow_threshold: 慢请求阈值
        
    Returns:
        分析结果DataFrame
    """
    log_info(f"开始高级API性能分析: {csv_path}", show_memory=True)
    
    if success_codes is None:
        from self_00_01_constants import DEFAULT_SUCCESS_CODES
        success_codes = DEFAULT_SUCCESS_CODES
    
    # 字段映射
    field_mapping = {
        'uri': 'request_full_uri',
        'app': 'application_name', 
        'service': 'service_name',
        'status': 'response_status_code',
        'request_time': 'total_request_duration',
        'header_time': 'upstream_header_time',
        'connect_time': 'upstream_connect_time',
        'response_time': 'upstream_response_time',
        'body_bytes_kb': 'response_body_size_kb',
        'bytes_sent_kb': 'total_bytes_sent_kb',
        'backend_connect_phase': 'backend_connect_phase',
        'backend_process_phase': 'backend_process_phase',
        'backend_transfer_phase': 'backend_transfer_phase',
        'nginx_transfer_phase': 'nginx_transfer_phase',
        'response_transfer_speed': 'response_transfer_speed',
        'processing_efficiency_index': 'processing_efficiency_index',
        'client_ip': 'client_ip'
    }
    
    # 创建高级分析器
    analyzer = AdvancedStreamingApiAnalyzer(slow_threshold)
    
    # 处理参数
    chunk_size = max(DEFAULT_CHUNK_SIZE, 50000)
    success_codes = [str(code) for code in success_codes]
    start_time = datetime.now()
    
    # 检查CSV文件
    if not os.path.exists(csv_path):
        log_info(f"CSV文件不存在: {csv_path}", level="ERROR")
        return pd.DataFrame()
    
    if os.path.getsize(csv_path) == 0:
        log_info(f"CSV文件为空: {csv_path}", level="ERROR")
        return pd.DataFrame()
    
    # 验证CSV文件是否有有效内容
    try:
        # 尝试读取第一行来验证文件格式
        test_df = pd.read_csv(csv_path, nrows=1)
        if test_df.empty:
            log_info(f"CSV文件没有数据行: {csv_path}", level="ERROR")
            return pd.DataFrame()
    except Exception as e:
        log_info(f"CSV文件格式错误: {csv_path}, 错误: {str(e)}", level="ERROR")
        return pd.DataFrame()
    
    # 流式处理数据
    try:
        for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
            analyzer.process_chunk(chunk, field_mapping, success_codes)
            
            # 定期报告进度
            if analyzer.processing_stats['chunks_processed'] % 10 == 0:
                elapsed = (datetime.now() - start_time).total_seconds()
                log_info(
                    f"已处理 {analyzer.processing_stats['chunks_processed']} 个数据块, "
                    f"{analyzer.global_stats['total_requests']} 条记录, "
                    f"耗时: {elapsed:.2f}秒", 
                    show_memory=True
                )
    
    except Exception as e:
        log_info(f"数据处理出错: {e}")
        raise
    
    # 获取分析总结
    summary = analyzer.get_analysis_summary()
    log_info(f"分析完成: {summary}", show_memory=True)
    
    # 生成统计报告
    results = generate_advanced_api_statistics(analyzer)
    
    if results:
        results_df = pd.DataFrame(results)
        if not results_df.empty and '平均请求时长(秒)' in results_df.columns:
            results_df = results_df.sort_values(by='平均请求时长(秒)', ascending=False)
        
        # 创建Excel报告
        create_advanced_api_performance_excel(results_df, output_path, analyzer)
        
        log_info(f"高级API性能分析报告已生成: {output_path}", show_memory=True)
        return results_df.head(5)
    else:
        log_info("没有找到任何API数据，返回空DataFrame", show_memory=True)
        return pd.DataFrame()


def generate_advanced_api_statistics(analyzer):
    """
    生成高级API统计报告
    
    Args:
        analyzer: 高级分析器实例
        
    Returns:
        统计结果列表
    """
    results = []
    api_stats = analyzer.api_stats
    global_stats = analyzer.global_stats
    
    def safe_percentile_tdigest(digest, percentile):
        """使用T-Digest安全计算百分位数"""
        try:
            return round(digest.percentile(percentile), 3)
        except:
            return 0.0
    
    def safe_percentile_reservoir(reservoir, percentile):
        """使用蓄水池采样安全计算百分位数"""
        try:
            return round(reservoir.percentile(percentile), 3)
        except:
            return 0.0
    
    def safe_avg(total, count):
        """安全的平均值计算"""
        return round(total / count, 3) if count > 0 else 0
    
    for api, stats in api_stats.items():
        # 基础指标
        total_requests = stats['total_requests']
        success_requests = stats['success_requests']
        slow_requests = stats['slow_requests']
        
        if success_requests == 0:
            continue
        
        # 计算比例
        success_rate = round(success_requests / total_requests * 100, 2) if total_requests > 0 else 0
        slow_ratio = round(slow_requests / success_requests * 100, 2) if success_requests > 0 else 0
        global_slow_ratio = round(slow_requests / global_stats['slow_requests'] * 100, 2) if global_stats['slow_requests'] > 0 else 0
        global_request_ratio = round(success_requests / global_stats['success_requests'] * 100, 2) if global_stats['success_requests'] > 0 else 0
        
        # 响应时间统计（使用T-Digest）
        avg_request_time = safe_avg(stats['request_time_sum'], stats['request_time_count'])
        is_slow_api = "Y" if (avg_request_time > analyzer.slow_threshold or slow_ratio > DEFAULT_SLOW_REQUESTS_THRESHOLD * 100) else "N"
        
        # T-Digest分位数
        p50_tdigest = safe_percentile_tdigest(stats['response_time_digest'], 50)
        p90_tdigest = safe_percentile_tdigest(stats['response_time_digest'], 90)
        p95_tdigest = safe_percentile_tdigest(stats['response_time_digest'], 95)
        p99_tdigest = safe_percentile_tdigest(stats['response_time_digest'], 99)
        
        # 蓄水池采样分位数（作为对比）
        p50_reservoir = safe_percentile_reservoir(stats['response_time_reservoir'], 50)
        p95_reservoir = safe_percentile_reservoir(stats['response_time_reservoir'], 95)
        
        # 阶段时间统计（使用T-Digest）
        backend_connect_p50 = safe_percentile_tdigest(stats['backend_connect_digest'], 50)
        backend_process_p50 = safe_percentile_tdigest(stats['backend_process_digest'], 50)
        backend_transfer_p50 = safe_percentile_tdigest(stats['backend_transfer_digest'], 50)
        nginx_transfer_p50 = safe_percentile_tdigest(stats['nginx_transfer_digest'], 50)
        
        # 计算阶段占比
        total_phase_time = backend_connect_p50 + backend_process_p50 + backend_transfer_p50 + nginx_transfer_p50
        if total_phase_time > 0:
            connect_ratio = round(backend_connect_p50 / total_phase_time * 100, 2)
            process_ratio = round(backend_process_p50 / total_phase_time * 100, 2)
            transfer_ratio = round(backend_transfer_p50 / total_phase_time * 100, 2)
            nginx_ratio = round(nginx_transfer_p50 / total_phase_time * 100, 2)
        else:
            connect_ratio = process_ratio = transfer_ratio = nginx_ratio = 0
        
        # 大小统计（使用蓄水池采样）
        body_avg = stats['body_size_reservoir'].mean()
        body_p95 = safe_percentile_reservoir(stats['body_size_reservoir'], 95)
        bytes_avg = stats['bytes_size_reservoir'].mean()
        bytes_p95 = safe_percentile_reservoir(stats['bytes_size_reservoir'], 95)
        
        # 性能指标
        transfer_speed_avg = stats['transfer_speed_reservoir'].mean()
        efficiency_avg = stats['efficiency_reservoir'].mean()
        
        # API频率估计
        api_frequency_estimate = global_stats['api_frequency'].estimate(api)
        
        # 数据质量指标
        tdigest_samples = stats['response_time_digest'].count
        reservoir_samples = len(stats['response_time_reservoir'].samples)
        data_quality = round(reservoir_samples / success_requests * 100, 1) if success_requests > 0 else 0
        
        # 构建结果
        result = {
            # 基础信息
            '请求URI': api,
            '应用名称': stats['app_name'],
            '服务名称': stats['service_name'],
            
            # 请求统计
            '请求总数': total_requests,
            '成功请求数': success_requests,
            '错误请求数': stats['error_requests'],  # 新增
            '占总请求比例(%)': global_request_ratio,
            '频率估计': api_frequency_estimate,
            
            # 成功率统计
            '成功率(%)': success_rate,
            '错误率(%)': round((stats['error_requests'] / total_requests * 100), 2) if total_requests > 0 else 0,  # 新增
            '全局错误占比(%)': round((stats['error_requests'] / global_stats['error_requests'] * 100), 2) if global_stats['error_requests'] > 0 else 0,  # 新增
            
            # 慢请求统计
            '慢请求数': slow_requests,
            '慢请求比例(%)': slow_ratio,
            '全局慢请求占比(%)': global_slow_ratio,
            '是否慢接口': is_slow_api,
            
            # 响应时间统计（T-Digest）
            '平均请求时长(秒)': avg_request_time,
            'T-Digest中位数(秒)': p50_tdigest,
            'T-Digest P90(秒)': p90_tdigest,
            'T-Digest P95(秒)': p95_tdigest,
            'T-Digest P99(秒)': p99_tdigest,
            
            # 蓄水池采样对比
            '蓄水池中位数(秒)': p50_reservoir,
            '蓄水池P95(秒)': p95_reservoir,
            
            # 阶段时间统计
            '后端连接时长(秒)': backend_connect_p50,
            '后端处理时长(秒)': backend_process_p50,
            '后端传输时长(秒)': backend_transfer_p50,
            'Nginx传输时长(秒)': nginx_transfer_p50,
            
            # 阶段占比
            '后端连接占比(%)': connect_ratio,
            '后端处理占比(%)': process_ratio,
            '后端传输占比(%)': transfer_ratio,
            'Nginx传输占比(%)': nginx_ratio,
            
            # 响应大小统计
            '平均响应体大小(KB)': round(body_avg, 2),
            'P95响应体大小(KB)': round(body_p95, 2),
            '平均传输大小(KB)': round(bytes_avg, 2),
            'P95传输大小(KB)': round(bytes_p95, 2),
            
            # 性能指标
            '平均传输速度(KB/s)': round(transfer_speed_avg, 2),
            '平均处理效率指数': round(efficiency_avg, 3),
            
            # 数据质量指标
            'T-Digest样本数': tdigest_samples,
            '蓄水池样本数': reservoir_samples,
            '数据质量(%)': data_quality,
            '算法精度': 'T-Digest高精度' if tdigest_samples > 1000 else 'T-Digest中精度' if tdigest_samples > 100 else 'T-Digest低精度'
        }
        
        results.append(result)
    
    log_info(f"已生成 {len(results)} 个API的高级统计报告", show_memory=True)
    return results


def create_advanced_api_performance_excel(results_df, output_path, analyzer):
    """
    创建高级API性能分析Excel报告
    
    Args:
        results_df: 结果DataFrame
        output_path: 输出路径
        analyzer: 分析器实例
    """
    log_info(f"开始创建高级Excel报告: {output_path}", show_memory=True)
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 主要统计表的分组表头
    main_headers = {
        "请求URI": ["请求URI"],
        "应用信息": ["应用名称", "服务名称"],
        "请求统计": ["请求总数", "成功请求数", "错误请求数", "占总请求比例(%)", "频率估计"],
        "成功率统计": ["成功率(%)", "错误率(%)", "全局错误占比(%)"],
        "慢请求统计": ["慢请求数", "慢请求比例(%)", "全局慢请求占比(%)", "是否慢接口"],
        "T-Digest时间分析(秒)": ["平均", "中位数", "P90", "P95", "P99"],
        "蓄水池对比(秒)": ["中位数", "P95"],
        "阶段时间(秒)": ["后端连接", "后端处理", "后端传输", "Nginx传输"],
        "阶段占比(%)": ["后端连接", "后端处理", "后端传输", "Nginx传输"],
        "响应大小(KB)": ["平均响应体", "P95响应体", "平均传输", "P95传输"],
        "性能指标": ["平均传输速度(KB/s)", "平均处理效率指数"],
        "数据质量": ["T-Digest样本数", "蓄水池样本数", "数据质量(%)", "算法精度"]
    }
    
    # 列名映射
    column_mapping = {
        '请求URI': '请求URI',
        '应用名称': '应用名称',
        '服务名称': '服务名称',
        '请求总数': '请求总数',
        '成功请求数': '成功请求数',
        '错误请求数': '错误请求数',  # 新增
        '占总请求比例(%)': '占总请求比例(%)',
        '成功率(%)': '成功率(%)',
        '错误率(%)': '错误率(%)',  # 新增
        '全局错误占比(%)': '全局错误占比(%)',  # 新增
        '频率估计': '频率估计',
        '慢请求数': '慢请求数',
        '慢请求比例(%)': '慢请求比例(%)',
        '全局慢请求占比(%)': '全局慢请求占比(%)',
        '是否慢接口': '是否慢接口',
        '平均请求时长(秒)': '平均',
        'T-Digest中位数(秒)': '中位数',
        'T-Digest P90(秒)': 'P90',
        'T-Digest P95(秒)': 'P95',
        'T-Digest P99(秒)': 'P99',
        '蓄水池中位数(秒)': '中位数',
        '蓄水池P95(秒)': 'P95',
        '后端连接时长(秒)': '后端连接',
        '后端处理时长(秒)': '后端处理',
        '后端传输时长(秒)': '后端传输',
        'Nginx传输时长(秒)': 'Nginx传输',
        '后端连接占比(%)': '后端连接',
        '后端处理占比(%)': '后端处理',
        '后端传输占比(%)': '后端传输',
        'Nginx传输占比(%)': 'Nginx传输',
        '平均响应体大小(KB)': '平均响应体',
        'P95响应体大小(KB)': 'P95响应体',
        '平均传输大小(KB)': '平均传输',
        'P95传输大小(KB)': 'P95传输',
        '平均传输速度(KB/s)': '平均传输速度(KB/s)',
        '平均处理效率指数': '平均处理效率指数',
        'T-Digest样本数': 'T-Digest样本数',
        '蓄水池样本数': '蓄水池样本数',
        '数据质量(%)': '数据质量(%)',
        '算法精度': '算法精度'
    }
    
    # 重命名列
    renamed_df = results_df.copy()
    renamed_df.columns = [column_mapping.get(col, col) for col in results_df.columns]
    
    # 创建主要统计表
    ws1 = add_dataframe_to_excel_with_grouped_headers(wb, renamed_df, 'API性能统计(高级)', header_groups=main_headers)
    
    # 高亮慢接口
    try:
        slow_api_col = renamed_df.columns.get_loc('是否慢接口') + 1
        for row_idx in range(3, len(renamed_df) + 3):
            if ws1.cell(row=row_idx, column=slow_api_col).value == 'Y':
                for col_idx in range(1, len(renamed_df.columns) + 1):
                    ws1.cell(row=row_idx, column=col_idx).fill = HIGHLIGHT_FILL
    except Exception as e:
        log_info(f"高亮慢接口失败: {e}")
    
    # 创建算法对比分析工作表
    create_algorithm_comparison_sheet(wb, analyzer, results_df)
    
    # 创建全局分析工作表  
    create_global_analysis_sheet(wb, analyzer)
    
    # 创建性能优化建议工作表
    create_optimization_recommendations_sheet(wb, results_df, analyzer)
    
    # 格式化工作表
    format_excel_sheet(ws1)
    
    log_info(f"高级Excel报告格式化完成，准备保存", show_memory=True)
    wb.save(output_path)
    log_info(f"高级Excel报告已保存: {output_path}", show_memory=True)


def create_algorithm_comparison_sheet(wb, analyzer, results_df):
    """创建算法对比分析工作表"""
    ws = wb.create_sheet(title='算法对比分析')
    
    current_row = 1
    
    # 标题
    ws.cell(row=current_row, column=1, value='采样算法对比分析').font = Font(bold=True, size=14)
    current_row += 3
    
    # T-Digest vs 蓄水池采样对比
    comparison_data = []
    for _, row in results_df.iterrows():
        if row['T-Digest样本数'] > 0 and row['蓄水池样本数'] > 0:
            tdigest_p95 = row.get('T-Digest P95(秒)', 0)
            reservoir_p95 = row.get('蓄水池P95(秒)', 0)
            diff_percent = abs(tdigest_p95 - reservoir_p95) / max(reservoir_p95, 0.001) * 100
            
            comparison_data.append([
                row['请求URI'][:50] + '...' if len(str(row['请求URI'])) > 50 else row['请求URI'],
                row['T-Digest样本数'],
                row['蓄水池样本数'], 
                tdigest_p95,
                reservoir_p95,
                round(diff_percent, 2)
            ])
    
    # 写入对比数据
    headers = ['API', 'T-Digest样本', '蓄水池样本', 'T-Digest P95', '蓄水池 P95', '差异(%)']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=current_row, column=col_idx, value=header).font = Font(bold=True)
    current_row += 1
    
    for data_row in comparison_data[:20]:  # 显示前20个
        for col_idx, value in enumerate(data_row, start=1):
            ws.cell(row=current_row, column=col_idx, value=value)
        current_row += 1
    
    current_row += 3
    
    # 算法性能总结
    ws.cell(row=current_row, column=1, value='算法性能总结').font = Font(bold=True, size=12)
    current_row += 2
    
    summary_stats = analyzer.get_analysis_summary()
    memory_stats = summary_stats.get('memory_efficiency', {})
    
    summary_data = [
        ['总API数量', summary_stats.get('total_apis', 0)],
        ['总请求数', summary_stats.get('total_requests', 0)],
        ['成功请求数', summary_stats.get('success_requests', 0)],
        ['独立IP估计', summary_stats.get('unique_ips_estimate', 0)],
        ['处理数据块数', summary_stats.get('chunks_processed', 0)],
        ['平均处理时间(秒)', round(summary_stats.get('avg_processing_time', 0), 3)],
        ['', ''],
        ['内存效率分析', ''],
        ['传统方法内存(MB)', round(memory_stats.get('traditional_memory_mb', 0), 2)],
        ['当前方法内存(MB)', round(memory_stats.get('current_memory_mb', 0), 2)],
        ['内存节省(%)', round(memory_stats.get('memory_savings_percent', 0), 2)],
        ['效率提升倍数', round(memory_stats.get('efficiency_ratio', 0), 2)]
    ]
    
    for label, value in summary_data:
        ws.cell(row=current_row, column=1, value=label).font = Font(bold=True) if label and not value else None
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    
    # 设置列宽
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20 if col == 1 else 15
    
    format_excel_sheet(ws)


def create_global_analysis_sheet(wb, analyzer):
    """创建全局分析工作表"""
    ws = wb.create_sheet(title='全局分析概览')
    
    current_row = 1
    
    # 全局T-Digest分析
    global_digest = analyzer.global_stats['global_response_time_digest']
    
    global_stats = [
        ['=== 全局响应时间分析(T-Digest) ===', ''],
        ['总样本数', global_digest.count],
        ['最小值(秒)', round(global_digest.min_value, 3) if global_digest.min_value != float('inf') else 0],
        ['最大值(秒)', round(global_digest.max_value, 3) if global_digest.max_value != float('-inf') else 0],
        ['P50(秒)', round(global_digest.percentile(50), 3)],
        ['P90(秒)', round(global_digest.percentile(90), 3)],
        ['P95(秒)', round(global_digest.percentile(95), 3)],
        ['P99(秒)', round(global_digest.percentile(99), 3)],
        ['P99.9(秒)', round(global_digest.percentile(99.9), 3)],
        ['', ''],
        
        ['=== 分层采样分析(按小时) ===', ''],
    ]
    
    # 分层统计
    hourly_stats = analyzer.global_stats['hourly_stratified'].get_strata_stats()
    for hour, stats in sorted(hourly_stats.items()):
        global_stats.extend([
            [f'{hour}时段统计', ''],
            [f'  样本数', stats['count']],
            [f'  平均响应时间(秒)', round(stats['mean'], 3)],
            [f'  P95响应时间(秒)', round(stats['p95'], 3)],
            [f'  P99响应时间(秒)', round(stats['p99'], 3)]
        ])
    
    global_stats.extend([
        ['', ''],
        ['=== 自适应采样分析 ===', ''],
        ['自适应样本数', len(analyzer.global_stats['adaptive_sampler'].get_samples())],
        ['自适应P95(秒)', round(analyzer.global_stats['adaptive_sampler'].percentile(95), 3)],
        ['自适应P99(秒)', round(analyzer.global_stats['adaptive_sampler'].percentile(99), 3)],
    ])
    
    # 写入数据
    for label, value in global_stats:
        cell_label = ws.cell(row=current_row, column=1, value=label)
        cell_value = ws.cell(row=current_row, column=2, value=value)
        
        if label.startswith('===') and label.endswith('==='):
            cell_label.font = Font(bold=True, size=12)
            cell_value.font = Font(bold=True, size=12)
        elif label.startswith('  '):
            pass  # 缩进项
        elif label and not label.startswith('  '):
            cell_label.font = Font(bold=True)
        
        current_row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    format_excel_sheet(ws)


def create_optimization_recommendations_sheet(wb, results_df, analyzer):
    """创建性能优化建议工作表"""
    ws = wb.create_sheet(title='性能优化建议')
    
    current_row = 1
    
    # 标题
    ws.cell(row=current_row, column=1, value='基于高级分析的性能优化建议').font = Font(bold=True, size=14)
    current_row += 3
    
    # 慢接口优化建议
    slow_apis = results_df[results_df['是否慢接口'] == 'Y'].sort_values('T-Digest P99(秒)', ascending=False)
    
    if not slow_apis.empty:
        ws.cell(row=current_row, column=1, value='1. 慢接口优化建议').font = Font(bold=True, size=12)
        current_row += 2
        
        headers = ['API', 'P99时间(秒)', '主要瓶颈', '优化建议']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col_idx, value=header).font = Font(bold=True)
        current_row += 1
        
        for _, row in slow_apis.head(10).iterrows():
            # 分析主要瓶颈
            phases = {
                '后端连接占比(%)': '网络连接',
                '后端处理占比(%)': '后端处理', 
                '后端传输占比(%)': '数据传输',
                'Nginx传输占比(%)': 'Nginx传输'
            }
            
            max_phase = max(phases.keys(), key=lambda x: row.get(x, 0))
            bottleneck = phases[max_phase]
            
            # 生成优化建议
            if bottleneck == '网络连接':
                suggestion = '优化网络配置、连接池、DNS解析'
            elif bottleneck == '后端处理':
                suggestion = '优化业务逻辑、数据库查询、缓存策略'
            elif bottleneck == '数据传输':
                suggestion = '启用压缩、优化响应体大小、CDN加速'
            else:
                suggestion = '优化Nginx配置、负载均衡策略'
            
            data_row = [
                str(row['请求URI'])[:50] + '...' if len(str(row['请求URI'])) > 50 else str(row['请求URI']),
                row.get('T-Digest P99(秒)', 0),
                f"{bottleneck}({row.get(max_phase, 0):.1f}%)",
                suggestion
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
        
        current_row += 3
    
    # 内存使用优化建议
    ws.cell(row=current_row, column=1, value='2. 内存使用优化效果').font = Font(bold=True, size=12)
    current_row += 2
    
    memory_stats = analyzer._calculate_memory_efficiency()
    
    memory_recommendations = [
        f"✓ 内存使用减少 {memory_stats.get('memory_savings_percent', 0):.1f}%",
        f"✓ 内存效率提升 {memory_stats.get('efficiency_ratio', 0):.1f} 倍",
        f"✓ T-Digest算法提供高精度分位数估计",
        f"✓ 蓄水池采样保证统计代表性",
        f"✓ HyperLogLog实现高效基数统计",
        "✓ 分层采样支持时间维度分析",
        "✓ 自适应采样根据数据分布调整策略"
    ]
    
    for recommendation in memory_recommendations:
        ws.cell(row=current_row, column=1, value=recommendation)
        current_row += 1
    
    current_row += 2
    
    # 算法选择建议
    ws.cell(row=current_row, column=1, value='3. 算法选择建议').font = Font(bold=True, size=12)
    current_row += 2
    
    algorithm_suggestions = [
        "• T-Digest: 适用于响应时间分位数估计，内存占用小，精度高",
        "• 蓄水池采样: 适用于需要原始数据的分析，如异常检测、相关性分析",
        "• Count-Min Sketch: 适用于热点API识别，支持高频更新",
        "• HyperLogLog: 适用于独立用户/IP统计，误差可控",
        "• 分层采样: 适用于时间维度分析，保证各时段代表性",
        "• 自适应采样: 适用于数据分布变化的场景，自动调整策略"
    ]
    
    for suggestion in algorithm_suggestions:
        ws.cell(row=current_row, column=1, value=suggestion)
        current_row += 1
    
    # 设置列宽
    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 30 if col == 1 else 20
    
    format_excel_sheet(ws)


# 保持向后兼容的函数别名
def analyze_api_performance(csv_path, output_path, success_codes=None, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """向后兼容的函数别名"""
    return analyze_api_performance_advanced(csv_path, output_path, success_codes, slow_threshold)