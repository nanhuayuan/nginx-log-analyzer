import gc
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from collections import defaultdict

from self_00_04_excel_processor import format_excel_sheet, add_dataframe_to_excel_with_grouped_headers
from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD, DEFAULT_SLOW_REQUESTS_THRESHOLD, \
    TIME_METRICS, SIZE_METRICS, HIGHLIGHT_FILL
from self_00_02_utils import log_info, get_distribution_stats, calculate_time_percentages

# 尝试导入scipy，如果失败则使用近似计算
try:
    from scipy.stats import norm
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False


class StreamingApiAnalyzer:
    """流式API性能分析器 - 高效处理大数据集"""
    
    def __init__(self, slow_threshold=DEFAULT_SLOW_THRESHOLD):
        self.slow_threshold = slow_threshold
        self.api_stats = defaultdict(lambda: {
            'total_requests': 0,
            'success_requests': 0,
            'slow_requests': 0,
            'app_name': '',
            'service_name': '',
            # 流式统计字段
            'request_time_sum': 0.0,
            'request_time_sum_sq': 0.0,
            'request_time_count': 0,
            'request_time_values': [],  # 仅保存少量样本用于分位数
            
            # 阶段时间统计
            'backend_connect_sum': 0.0,
            'backend_process_sum': 0.0,
            'backend_transfer_sum': 0.0,
            'nginx_transfer_sum': 0.0,
            'backend_connect_count': 0,
            'backend_process_count': 0,
            'backend_transfer_count': 0,
            'nginx_transfer_count': 0,
            
            # 大小统计
            'body_size_sum': 0.0,
            'body_size_count': 0,
            'body_size_values': [],
            'bytes_size_sum': 0.0,
            'bytes_size_count': 0,
            'bytes_size_values': [],
            
            # 性能指标
            'transfer_speed_sum': 0.0,
            'transfer_speed_count': 0,
            'efficiency_sum': 0.0,
            'efficiency_count': 0
        })
        
        # 全局统计
        self.global_stats = {
            'total_requests': 0,
            'success_requests': 0,
            'slow_requests': 0,
            'phase_times': defaultdict(float),
            'global_samples': {
                'response_times': [],
                'body_sizes': [],
                'bytes_sizes': [],
                'transfer_speeds': [],
                'efficiency_scores': []
            }
        }
        
        # 采样配置
        self.MAX_SAMPLES_PER_API = 1000  # 每个API最多保存1000个样本
        self.MAX_GLOBAL_SAMPLES = 10000  # 全局最多保存10000个样本
    
    def process_chunk(self, chunk, field_mapping, success_codes):
        """处理单个数据块"""
        chunk_rows = len(chunk)
        self.global_stats['total_requests'] += chunk_rows
        
        # 筛选成功请求
        successful_requests = chunk[chunk[field_mapping['status']].astype(str).isin(success_codes)]
        success_count = len(successful_requests)
        self.global_stats['success_requests'] += success_count
        
        if success_count == 0:
            return
        
        # 预处理数据 - 向量化转换
        numeric_data = self._preprocess_numeric_data(successful_requests, field_mapping)
        
        # 按API分组并批量处理
        for api, group_data in numeric_data.groupby('uri'):
            self._process_api_group(api, group_data, field_mapping)
    
    def _preprocess_numeric_data(self, chunk, field_mapping):
        """预处理数字数据 - 向量化操作"""
        # 选择需要的列
        cols_to_process = {
            'uri': field_mapping['uri'],
            'app': field_mapping['app'],
            'service': field_mapping['service'],
            'request_time': field_mapping['request_time'],
            'backend_connect': field_mapping.get('backend_connect_phase', ''),
            'backend_process': field_mapping.get('backend_process_phase', ''),
            'backend_transfer': field_mapping.get('backend_transfer_phase', ''),
            'nginx_transfer': field_mapping.get('nginx_transfer_phase', ''),
            'body_size': field_mapping['body_bytes_kb'],
            'bytes_size': field_mapping['bytes_sent_kb'],
            'transfer_speed': field_mapping.get('response_transfer_speed', ''),
            'efficiency': field_mapping.get('processing_efficiency_index', '')
        }
        
        # 创建数据副本
        data = {}
        for key, col in cols_to_process.items():
            if col and col in chunk.columns:
                if key in ['uri', 'app', 'service']:
                    data[key] = chunk[col]
                else:
                    data[key] = pd.to_numeric(chunk[col], errors='coerce')
            else:
                data[key] = pd.Series([None] * len(chunk))
        
        return pd.DataFrame(data)
    
    def _process_api_group(self, api, group_data, field_mapping):
        """处理单个API组的数据"""
        group_size = len(group_data)
        stats = self.api_stats[api]
        
        # 更新基础统计
        stats['total_requests'] += group_size
        stats['success_requests'] += group_size
        
        # 设置应用和服务名称
        if not stats['app_name'] and not group_data['app'].isna().all():
            stats['app_name'] = group_data['app'].iloc[0]
        if not stats['service_name'] and not group_data['service'].isna().all():
            stats['service_name'] = group_data['service'].iloc[0]
        
        # 处理请求时间
        request_times = group_data['request_time'].dropna()
        if len(request_times) > 0:
            # 更新流式统计
            stats['request_time_sum'] += request_times.sum()
            stats['request_time_sum_sq'] += (request_times ** 2).sum()
            stats['request_time_count'] += len(request_times)
            
            # 采样保存用于分位数计算
            self._update_samples(stats['request_time_values'], request_times, self.MAX_SAMPLES_PER_API)
            self._update_samples(self.global_stats['global_samples']['response_times'], request_times, self.MAX_GLOBAL_SAMPLES)
            
            # 统计慢请求
            slow_count = (request_times > self.slow_threshold).sum()
            stats['slow_requests'] += slow_count
            self.global_stats['slow_requests'] += slow_count
        
        # 处理阶段时间
        self._update_phase_stats(stats, group_data, 'backend_connect')
        self._update_phase_stats(stats, group_data, 'backend_process')
        self._update_phase_stats(stats, group_data, 'backend_transfer')
        self._update_phase_stats(stats, group_data, 'nginx_transfer')
        
        # 处理大小统计
        self._update_size_stats(stats, group_data, 'body_size')
        self._update_size_stats(stats, group_data, 'bytes_size')
        
        # 处理性能指标
        self._update_performance_stats(stats, group_data)
    
    def _update_samples(self, sample_list, new_data, max_size):
        """更新采样数据"""
        new_samples = new_data.tolist()
        sample_list.extend(new_samples)
        
        # 如果超过最大大小，随机采样
        if len(sample_list) > max_size:
            import random
            sample_list[:] = random.sample(sample_list, max_size)
    
    def _update_phase_stats(self, stats, group_data, phase_key):
        """更新阶段统计"""
        phase_data = group_data[phase_key].dropna()
        if len(phase_data) > 0:
            stats[f'{phase_key}_sum'] += phase_data.sum()
            stats[f'{phase_key}_count'] += len(phase_data)
            self.global_stats['phase_times'][phase_key] += phase_data.sum()
    
    def _update_size_stats(self, stats, group_data, size_key):
        """更新大小统计"""
        size_data = group_data[size_key].dropna()
        if len(size_data) > 0:
            stats[f'{size_key}_sum'] += size_data.sum()
            stats[f'{size_key}_count'] += len(size_data)
            
            # 采样保存
            self._update_samples(stats[f'{size_key}_values'], size_data, self.MAX_SAMPLES_PER_API)
            
            # 更新全局样本
            if size_key == 'body_size':
                self._update_samples(self.global_stats['global_samples']['body_sizes'], size_data, self.MAX_GLOBAL_SAMPLES)
            elif size_key == 'bytes_size':
                self._update_samples(self.global_stats['global_samples']['bytes_sizes'], size_data, self.MAX_GLOBAL_SAMPLES)
    
    def _update_performance_stats(self, stats, group_data):
        """更新性能统计"""
        # 传输速度
        speed_data = group_data['transfer_speed'].dropna()
        if len(speed_data) > 0:
            stats['transfer_speed_sum'] += speed_data.sum()
            stats['transfer_speed_count'] += len(speed_data)
            self._update_samples(self.global_stats['global_samples']['transfer_speeds'], speed_data, self.MAX_GLOBAL_SAMPLES)
        
        # 效率指标
        efficiency_data = group_data['efficiency'].dropna()
        if len(efficiency_data) > 0:
            stats['efficiency_sum'] += efficiency_data.sum()
            stats['efficiency_count'] += len(efficiency_data)
            self._update_samples(self.global_stats['global_samples']['efficiency_scores'], efficiency_data, self.MAX_GLOBAL_SAMPLES)


def analyze_api_performance(csv_path, output_path, success_codes=None, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """分析API性能数据并生成Excel报告 - 优化版本"""
    log_info(f"开始分析API性能数据: {csv_path}", show_memory=True)

    if success_codes is None:
        from self_00_01_constants import DEFAULT_SUCCESS_CODES
        success_codes = DEFAULT_SUCCESS_CODES

    # 字段映射
    field_mapping = {
        'uri': 'request_full_uri',
        'app': 'application_name',
        'service': 'service_name',
        'status': 'response_status_code',
        'request_time': 'total_request_duration',
        'header_time': 'upstream_header_time',
        'connect_time': 'upstream_connect_time',
        'response_time': 'upstream_response_time',
        'body_bytes_kb': 'response_body_size_kb',
        'bytes_sent_kb': 'total_bytes_sent_kb',
        'backend_connect_phase': 'backend_connect_phase',
        'backend_process_phase': 'backend_process_phase',
        'backend_transfer_phase': 'backend_transfer_phase',
        'nginx_transfer_phase': 'nginx_transfer_phase',
        'response_transfer_speed': 'response_transfer_speed',
        'processing_efficiency_index': 'processing_efficiency_index'
    }

    # 创建流式分析器
    analyzer = StreamingApiAnalyzer(slow_threshold)
    
    # 处理参数
    chunk_size = max(DEFAULT_CHUNK_SIZE, 50000)  # 增加chunk大小以提高效率
    success_codes = [str(code) for code in success_codes]
    chunks_processed = 0
    start_time = datetime.now()

    # 流式处理数据
    try:
        for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
            chunks_processed += 1
            
            # 处理数据块
            analyzer.process_chunk(chunk, field_mapping, success_codes)
            
            # 定期报告进度
            if chunks_processed % 10 == 0:
                elapsed = (datetime.now() - start_time).total_seconds()
                log_info(f"已处理 {chunks_processed} 个数据块, {analyzer.global_stats['total_requests']} 条记录, 耗时: {elapsed:.2f}秒", show_memory=True)
                gc.collect()

    except Exception as e:
        log_info(f"数据处理出错: {e}")
        raise

    log_info(f"数据处理完成，共处理 {analyzer.global_stats['total_requests']} 条请求, 成功请求 {analyzer.global_stats['success_requests']} 条", show_memory=True)

    # 生成统计报告
    results = generate_optimized_api_statistics(analyzer)

    if results:
        results_df = pd.DataFrame(results)
        if not results_df.empty and '平均请求时长(秒)' in results_df.columns:
            results_df = results_df.sort_values(by='平均请求时长(秒)', ascending=False)

        # 创建Excel报告
        create_optimized_api_performance_excel(results_df, output_path, analyzer)

        log_info(f"API性能分析报告已生成: {output_path}", show_memory=True)
        return results_df.head(5)
    else:
        log_info("没有找到任何API数据，返回空DataFrame", show_memory=True)
        return pd.DataFrame()


# 注释掉旧的函数，保留以防需要参考
# def initialize_api_stats(api_stats, api, group, field_mapping):
#     """初始化API统计数据结构"""
#     if api in api_stats:
#         return

    app_name = group[field_mapping['app']].iloc[0] if not group[field_mapping['app']].empty else ""
    service_name = group[field_mapping['service']].iloc[0] if not group[field_mapping['service']].empty else ""

    api_stats[api] = {
        'request_uri': api,
        'app_name': app_name,
        'service_name': service_name,
        'total_requests': 0,
        'success_requests': 0,
        'slow_requests_count': 0
    }

    # 时间相关字段
    time_fields = [
        'request_time', 'connect_time', 'header_time', 'response_time',
        'backend_connect_phase', 'backend_process_phase',
        'backend_transfer_phase', 'nginx_transfer_phase'
    ]

    for field in time_fields:
        api_stats[api][f'{field}_min'] = float('inf')
        api_stats[api][f'{field}_max'] = 0
        api_stats[api][f'{field}_total'] = 0
        # 不再存储原始数据，使用流式统计
        api_stats[api][f'{field}_count'] = 0
        api_stats[api][f'{field}_sum'] = 0.0
        api_stats[api][f'{field}_sum_sq'] = 0.0  # 用于计算方差

    # 大小相关字段 (KB单位)
    size_fields = ['body_kb', 'bytes_kb']
    for field in size_fields:
        api_stats[api][f'{field}_min'] = float('inf')
        api_stats[api][f'{field}_max'] = 0
        api_stats[api][f'{field}_total'] = 0
        api_stats[api][f'{field}_count'] = 0
        api_stats[api][f'{field}_sum'] = 0.0
        api_stats[api][f'{field}_sum_sq'] = 0.0

    # 性能指标字段
    performance_fields = ['transfer_speed', 'efficiency_index']
    for field in performance_fields:
        api_stats[api][f'{field}_count'] = 0
        api_stats[api][f'{field}_sum'] = 0.0
        api_stats[api][f'{field}_sum_sq'] = 0.0


# def process_api_group(api_stats, api, group, slow_threshold, field_mapping):
    """处理单个API组的详细统计"""
    if api not in api_stats:
        initialize_api_stats(api_stats, api, group, field_mapping)

    # 时间字段映射
    time_fields = {
        'request_time': field_mapping['request_time'],
        'connect_time': field_mapping['connect_time'],
        'header_time': field_mapping['header_time'],
        'response_time': field_mapping['response_time'],
        'backend_connect_phase': field_mapping['backend_connect_phase'],
        'backend_process_phase': field_mapping['backend_process_phase'],
        'backend_transfer_phase': field_mapping['backend_transfer_phase'],
        'nginx_transfer_phase': field_mapping['nginx_transfer_phase']
    }

    # 转换时间数据
    numeric_data = {}
    for field_key, field_name in time_fields.items():
        if field_name in group.columns:
            numeric_data[field_key] = pd.to_numeric(group[field_name], errors='coerce')

    # 转换大小数据 (KB单位)
    body_sizes_kb = pd.to_numeric(group[field_mapping['body_bytes_kb']], errors='coerce')
    bytes_sizes_kb = pd.to_numeric(group[field_mapping['bytes_sent_kb']], errors='coerce')

    # 转换性能指标
    transfer_speeds = pd.to_numeric(group[field_mapping['response_transfer_speed']], errors='coerce') if field_mapping[
                                                                                                             'response_transfer_speed'] in group.columns else pd.Series(
        [])
    efficiency_scores = pd.to_numeric(group[field_mapping['processing_efficiency_index']], errors='coerce') if \
    field_mapping['processing_efficiency_index'] in group.columns else pd.Series([])

    # 统计慢请求
    slow_requests_count = (numeric_data.get('request_time', pd.Series([])) > slow_threshold).sum()

    # 更新统计数据
    api_stats[api]['success_requests'] += len(group)
    api_stats[api]['slow_requests_count'] += slow_requests_count

    # 更新时间统计
    for field, series in numeric_data.items():
        update_stats(api_stats[api], field, series)

    # 更新大小统计
    update_stats(api_stats[api], 'body_kb', body_sizes_kb)
    update_stats(api_stats[api], 'bytes_kb', bytes_sizes_kb)

    # 更新性能指标统计（使用增量统计）
    if len(transfer_speeds) > 0:
        update_incremental_stats(api_stats[api], 'transfer_speed', transfer_speeds.dropna())
    if len(efficiency_scores) > 0:
        update_incremental_stats(api_stats[api], 'efficiency_index', efficiency_scores.dropna())


def update_stats(stats_dict, field, series):
    """更新统计数据的通用函数（优化版）"""
    valid_values = series.dropna()
    if len(valid_values) == 0:
        return

    min_field = f"{field}_min"
    max_field = f"{field}_max"
    total_field = f"{field}_total"
    count_field = f"{field}_count"
    sum_field = f"{field}_sum"
    sum_sq_field = f"{field}_sum_sq"

    if all(key in stats_dict for key in [min_field, max_field, total_field]):
        min_val = valid_values.min()
        max_val = valid_values.max()
        total_val = valid_values.sum()
        count_val = len(valid_values)
        sum_sq_val = (valid_values ** 2).sum()

        stats_dict[min_field] = min(stats_dict[min_field], min_val) if stats_dict[min_field] != float(
            'inf') else min_val
        stats_dict[max_field] = max(stats_dict[max_field], max_val)
        stats_dict[total_field] += total_val
        
        # 增量统计
        if count_field in stats_dict:
            stats_dict[count_field] += count_val
        if sum_field in stats_dict:
            stats_dict[sum_field] += total_val
        if sum_sq_field in stats_dict:
            stats_dict[sum_sq_field] += sum_sq_val


def update_incremental_stats(stats_dict, field, series):
    """更新增量统计数据"""
    valid_values = series.dropna()
    if len(valid_values) == 0:
        return
    
    count_field = f"{field}_count"
    sum_field = f"{field}_sum"
    sum_sq_field = f"{field}_sum_sq"
    
    count_val = len(valid_values)
    sum_val = valid_values.sum()
    sum_sq_val = (valid_values ** 2).sum()
    
    if count_field in stats_dict:
        stats_dict[count_field] += count_val
    if sum_field in stats_dict:
        stats_dict[sum_field] += sum_val
    if sum_sq_field in stats_dict:
        stats_dict[sum_sq_field] += sum_sq_val


def generate_optimized_api_statistics(analyzer):
    """生成优化的API统计报告"""
    results = []
    api_stats = analyzer.api_stats
    global_stats = analyzer.global_stats
    
    def safe_percentile(values, percentile):
        """安全的百分位数计算"""
        if not values or len(values) == 0:
            return 0
        return round(np.percentile(values, percentile), 3)
    
    def safe_avg(total, count):
        """安全的平均值计算"""
        return round(total / count, 3) if count > 0 else 0
    
    for api, stats in api_stats.items():
        # 计算基础指标
        total_requests = stats['total_requests']
        success_requests = stats['success_requests']
        slow_requests = stats['slow_requests']
        
        if success_requests == 0:
            continue
        
        # 计算比例
        success_rate = round(success_requests / total_requests * 100, 2) if total_requests > 0 else 0
        slow_ratio = round(slow_requests / success_requests * 100, 2) if success_requests > 0 else 0
        global_slow_ratio = round(slow_requests / global_stats['slow_requests'] * 100, 2) if global_stats['slow_requests'] > 0 else 0
        global_request_ratio = round(success_requests / global_stats['success_requests'] * 100, 2) if global_stats['success_requests'] > 0 else 0
        
        # 计算平均时间
        avg_request_time = safe_avg(stats['request_time_sum'], stats['request_time_count'])
        is_slow_api = "Y" if (avg_request_time > analyzer.slow_threshold or slow_ratio > DEFAULT_SLOW_REQUESTS_THRESHOLD * 100) else "N"
        
        # 计算阶段平均时间
        backend_connect_avg = safe_avg(stats['backend_connect_sum'], stats['backend_connect_count'])
        backend_process_avg = safe_avg(stats['backend_process_sum'], stats['backend_process_count'])
        backend_transfer_avg = safe_avg(stats['backend_transfer_sum'], stats['backend_transfer_count'])
        nginx_transfer_avg = safe_avg(stats['nginx_transfer_sum'], stats['nginx_transfer_count'])
        
        # 计算阶段占比
        total_phase_time = backend_connect_avg + backend_process_avg + backend_transfer_avg + nginx_transfer_avg
        if total_phase_time > 0:
            connect_ratio = round(backend_connect_avg / total_phase_time * 100, 2)
            process_ratio = round(backend_process_avg / total_phase_time * 100, 2)
            transfer_ratio = round(backend_transfer_avg / total_phase_time * 100, 2)
            nginx_ratio = round(nginx_transfer_avg / total_phase_time * 100, 2)
        else:
            connect_ratio = process_ratio = transfer_ratio = nginx_ratio = 0
        
        # 计算大小指标
        body_avg = safe_avg(stats['body_size_sum'], stats['body_size_count'])
        bytes_avg = safe_avg(stats['bytes_size_sum'], stats['bytes_size_count'])
        
        # 计算性能指标
        transfer_speed_avg = safe_avg(stats['transfer_speed_sum'], stats['transfer_speed_count'])
        efficiency_avg = safe_avg(stats['efficiency_sum'], stats['efficiency_count'])
        
        # 计算百分位数(使用采样数据)
        request_time_values = stats['request_time_values']
        body_size_values = stats['body_size_values']
        bytes_size_values = stats['bytes_size_values']
        
        # 数据质量指标
        sample_count = len(request_time_values)
        data_quality = round(sample_count / success_requests * 100, 1) if success_requests > 0 else 0
        
        # 构建结果字典
        result = {
            # 基础信息
            '请求URI': api,
            '应用名称': stats['app_name'],
            '服务名称': stats['service_name'],
            
            # 请求统计
            '请求总数': total_requests,
            '成功请求数': success_requests,
            '占总请求比例(%)': global_request_ratio,
            '成功率(%)': success_rate,
            
            # 慢请求统计
            '慢请求数': slow_requests,
            '慢请求比例(%)': slow_ratio,
            '全局慢请求占比(%)': global_slow_ratio,
            '是否慢接口': is_slow_api,
            
            # 请求时间统计
            '平均请求时长(秒)': avg_request_time,
            '请求时长中位数(秒)': safe_percentile(request_time_values, 50),
            'P90请求时长(秒)': safe_percentile(request_time_values, 90),
            'P95请求时长(秒)': safe_percentile(request_time_values, 95),
            'P99请求时长(秒)': safe_percentile(request_time_values, 99),
            
            # 阶段时间统计
            '后端连接时长(秒)': backend_connect_avg,
            '后端处理时长(秒)': backend_process_avg,
            '后端传输时长(秒)': backend_transfer_avg,
            'Nginx传输时长(秒)': nginx_transfer_avg,
            
            # 阶段占比统计
            '后端连接占比(%)': connect_ratio,
            '后端处理占比(%)': process_ratio,
            '后端传输占比(%)': transfer_ratio,
            'Nginx传输占比(%)': nginx_ratio,
            
            # 响应体大小统计
            '平均响应体大小(KB)': round(body_avg, 2),
            '响应体大小中位数(KB)': round(safe_percentile(body_size_values, 50), 2),
            'P90响应体大小(KB)': round(safe_percentile(body_size_values, 90), 2),
            'P95响应体大小(KB)': round(safe_percentile(body_size_values, 95), 2),
            'P99响应体大小(KB)': round(safe_percentile(body_size_values, 99), 2),
            
            # 传输大小统计
            '平均传输大小(KB)': round(bytes_avg, 2),
            '传输大小中位数(KB)': round(safe_percentile(bytes_size_values, 50), 2),
            'P90传输大小(KB)': round(safe_percentile(bytes_size_values, 90), 2),
            'P95传输大小(KB)': round(safe_percentile(bytes_size_values, 95), 2),
            'P99传输大小(KB)': round(safe_percentile(bytes_size_values, 99), 2),
            
            # 性能指标
            '平均传输速度(KB/s)': round(transfer_speed_avg, 2),
            '平均处理效率指数': round(efficiency_avg, 3),
            
            # 数据质量指标
            '样本数量': sample_count,
            '数据质量(%)': data_quality,
            '计算精度': '高' if data_quality >= 80 else '中' if data_quality >= 50 else '低'
        }
        
        results.append(result)
    
    log_info(f"已生成 {len(results)} 个API的统计报告", show_memory=True)
    return results


def create_optimized_api_performance_excel(results_df, output_path, analyzer):
    """创建优化的API性能分析Excel报告"""
    log_info(f"开始创建Excel报告: {output_path}", show_memory=True)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 定义优化后的分组表头
    main_headers = {
        "请求URI": ["请求URI"],
        "应用信息": ["应用名称", "服务名称"],
        "请求统计": ["请求总数", "成功请求数", "占总请求比例(%)", "成功率(%)"],
        "慢请求统计": ["慢请求数", "慢请求比例(%)", "全局慢请求占比(%)", "是否慢接口"],
        "请求时间分析(秒)": ["平均", "中位数", "P90", "P95", "P99"],
        "阶段时间(秒)": ["后端连接", "后端处理", "后端传输", "Nginx传输"],
        "阶段占比(%)": ["后端连接", "后端处理", "后端传输", "Nginx传输"],
        "响应体大小(KB)": ["平均", "中位数", "P90", "P95", "P99"],
        "传输大小(KB)": ["平均", "中位数", "P90", "P95", "P99"],
        "性能指标": ["平均传输速度(KB/s)", "平均处理效率指数"],
        "数据质量": ["样本数量", "数据质量(%)", "计算精度"]
    }

    # 列名映射 - 简化版本
    column_mapping = {
        '请求URI': '请求URI',
        '应用名称': '应用名称',
        '服务名称': '服务名称',
        '请求总数': '请求总数',
        '成功请求数': '成功请求数',
        '占总请求比例(%)': '占总请求比例(%)',
        '成功率(%)': '成功率(%)',
        '慢请求数': '慢请求数',
        '慢请求比例(%)': '慢请求比例(%)',
        '全局慢请求占比(%)': '全局慢请求占比(%)',
        '是否慢接口': '是否慢接口',
        '平均请求时长(秒)': '平均',
        '请求时长中位数(秒)': '中位数',
        'P90请求时长(秒)': 'P90',
        'P95请求时长(秒)': 'P95',
        'P99请求时长(秒)': 'P99',
        '后端连接时长(秒)': '后端连接',
        '后端处理时长(秒)': '后端处理',
        '后端传输时长(秒)': '后端传输',
        'Nginx传输时长(秒)': 'Nginx传输',
        '后端连接占比(%)': '后端连接',
        '后端处理占比(%)': '后端处理',
        '后端传输占比(%)': '后端传输',
        'Nginx传输占比(%)': 'Nginx传输',
        '平均响应体大小(KB)': '平均',
        '响应体大小中位数(KB)': '中位数',
        'P90响应体大小(KB)': 'P90',
        'P95响应体大小(KB)': 'P95',
        'P99响应体大小(KB)': 'P99',
        '平均传输大小(KB)': '平均',
        '传输大小中位数(KB)': '中位数',
        'P90传输大小(KB)': 'P90',
        'P95传输大小(KB)': 'P95',
        'P99传输大小(KB)': 'P99',
        '平均传输速度(KB/s)': '平均传输速度(KB/s)',
        '平均处理效率指数': '平均处理效率指数',
        '样本数量': '样本数量',
        '数据质量(%)': '数据质量(%)',
        '计算精度': '计算精度'
    }

    # 重命名列
    renamed_df = results_df.copy()
    renamed_df.columns = [column_mapping.get(col, col) for col in results_df.columns]

    # 创建主要统计表
    ws1 = add_dataframe_to_excel_with_grouped_headers(wb, renamed_df, 'API性能统计', header_groups=main_headers)

    # 高亮慢接口行
    try:
        slow_api_col = renamed_df.columns.get_loc('是否慢接口') + 1
        for row_idx in range(3, len(renamed_df) + 3):  # 调整行索引以匹配分组表头
            if ws1.cell(row=row_idx, column=slow_api_col).value == 'Y':
                for col_idx in range(1, len(renamed_df.columns) + 1):
                    ws1.cell(row=row_idx, column=col_idx).fill = HIGHLIGHT_FILL
    except Exception as e:
        log_info(f"高亮慢接口失败: {e}")

    # 创建整体分析工作表
    create_optimized_overview_sheet(wb, analyzer, results_df)

    # 添加性能分析工作表
    create_performance_analysis_sheet(wb, results_df)

    # 格式化工作表
    format_excel_sheet(ws1)
    
    log_info(f"Excel报告格式化完成，准备保存", show_memory=True)
    wb.save(output_path)
    log_info(f"Excel报告已保存: {output_path}", show_memory=True)


def create_optimized_overview_sheet(wb, analyzer, results_df):
    """创建优化的概览工作表"""
    ws = wb.create_sheet(title='整体分析概览')
    global_stats = analyzer.global_stats
    
    # 获取全局样本数据
    global_samples = global_stats['global_samples']
    response_times = global_samples['response_times']
    body_sizes = global_samples['body_sizes']
    bytes_sizes = global_samples['bytes_sizes']
    transfer_speeds = global_samples['transfer_speeds']
    efficiency_scores = global_samples['efficiency_scores']
    
    # 安全的统计计算
    def safe_stat(data, func, default=0):
        try:
            return round(func(data), 3) if data and len(data) > 0 else default
        except:
            return default
    
    # 整体统计数据
    overview_stats = [
        ['=== 基础统计 ===', ''],
        ['总请求数', global_stats['total_requests']],
        ['成功请求数', global_stats['success_requests']],
        ['成功率(%)', round(global_stats['success_requests'] / global_stats['total_requests'] * 100, 2) if global_stats['total_requests'] > 0 else 0],
        ['慢请求数', global_stats['slow_requests']],
        ['慢请求占比(%)', round(global_stats['slow_requests'] / global_stats['success_requests'] * 100, 2) if global_stats['success_requests'] > 0 else 0],
        ['API数量', len(results_df)],
        ['', ''],
        
        ['=== 响应时间分析 ===', ''],
        ['平均响应时间(秒)', safe_stat(response_times, np.mean)],
        ['中位数响应时间(秒)', safe_stat(response_times, np.median)],
        ['P90响应时间(秒)', safe_stat(response_times, lambda x: np.percentile(x, 90))],
        ['P95响应时间(秒)', safe_stat(response_times, lambda x: np.percentile(x, 95))],
        ['P99响应时间(秒)', safe_stat(response_times, lambda x: np.percentile(x, 99))],
        ['样本数量', len(response_times)],
        ['', ''],
        
        ['=== 数据传输分析 ===', ''],
        ['平均响应体大小(KB)', safe_stat(body_sizes, np.mean, 0)],
        ['P90响应体大小(KB)', safe_stat(body_sizes, lambda x: np.percentile(x, 90), 0)],
        ['P99响应体大小(KB)', safe_stat(body_sizes, lambda x: np.percentile(x, 99), 0)],
        ['平均传输大小(KB)', safe_stat(bytes_sizes, np.mean, 0)],
        ['P90传输大小(KB)', safe_stat(bytes_sizes, lambda x: np.percentile(x, 90), 0)],
        ['P99传输大小(KB)', safe_stat(bytes_sizes, lambda x: np.percentile(x, 99), 0)],
        ['', ''],
        
        ['=== 性能指标分析 ===', ''],
        ['平均传输速度(KB/s)', safe_stat(transfer_speeds, np.mean, 0)],
        ['P90传输速度(KB/s)', safe_stat(transfer_speeds, lambda x: np.percentile(x, 90), 0)],
        ['平均处理效率指数', safe_stat(efficiency_scores, np.mean, 0)],
        ['P90处理效率指数', safe_stat(efficiency_scores, lambda x: np.percentile(x, 90), 0)],
        ['', ''],
        
        ['=== 性能阈值设置 ===', ''],
        ['慢请求阈值(秒)', analyzer.slow_threshold],
        ['慢请求占比阈值(%)', DEFAULT_SLOW_REQUESTS_THRESHOLD * 100],
    ]
    
    # 写入数据
    for row_idx, (label, value) in enumerate(overview_stats, start=1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_value = ws.cell(row=row_idx, column=2, value=value)
        
        # 设置标题行格式
        if label.startswith('===') and label.endswith('==='):
            cell_label.font = Font(bold=True, size=12)
            cell_value.font = Font(bold=True, size=12)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # 添加TOP API分析
    next_row = len(overview_stats) + 3
    add_top_apis_analysis(ws, results_df, next_row)
    
    format_excel_sheet(ws)


def add_top_apis_analysis(ws, results_df, start_row):
    """添加TOP API分析"""
    sections = [
        ('请求量最多的10个API', '成功请求数', False,
         ['API', '成功请求数', '占比(%)', '平均时间(秒)'],
         ['请求URI', '成功请求数', '占总请求比例(%)', '平均请求时长(秒)']),
        
        ('响应时间最长的10个API', '平均请求时长(秒)', False,
         ['API', '平均时间(秒)', 'P99时间(秒)', '成功请求数'],
         ['请求URI', '平均请求时长(秒)', 'P99请求时长(秒)', '成功请求数']),
        
        ('响应体最大的10个API', '平均响应体大小(KB)', False,
         ['API', '平均响应体(KB)', 'P99响应体(KB)', '成功请求数'],
         ['请求URI', '平均响应体大小(KB)', 'P99响应体大小(KB)', '成功请求数'])
    ]
    
    current_row = start_row
    for title, sort_col, ascending, display_headers, data_cols in sections:
        # 添加标题
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=12)
        current_row += 2
        
        # 添加表头
        for col_idx, header in enumerate(display_headers, start=1):
            ws.cell(row=current_row, column=col_idx, value=header).font = Font(bold=True)
        current_row += 1
        
        # 添加数据
        top_data = results_df.sort_values(by=sort_col, ascending=ascending).head(10)
        for _, row_data in top_data.iterrows():
            for col_idx, col_name in enumerate(data_cols, start=1):
                value = row_data.get(col_name, '')
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                # 高亮慢请求
                if col_name == '平均请求时长(秒)' and value > DEFAULT_SLOW_THRESHOLD:
                    cell.fill = HIGHLIGHT_FILL
            current_row += 1
        
        current_row += 2  # 空行分隔


def create_performance_analysis_sheet(wb, results_df):
    """创建性能分析工作表"""
    ws = wb.create_sheet(title='性能瓶颈分析')
    
    # 分析各类性能问题
    analysis_sections = [
        ('高耗时接口 (P99>5秒)', 'P99请求时长(秒)', 5,
         ['API', 'P99时间(秒)', '平均时间(秒)', '成功请求数'],
         ['请求URI', 'P99请求时长(秒)', '平均请求时长(秒)', '成功请求数']),
        
        ('后端连接耗时过长接口 (连接占比>30%)', '后端连接占比(%)', 30,
         ['API', '连接占比(%)', '连接时间(秒)', '成功请求数'],
         ['请求URI', '后端连接占比(%)', '后端连接时长(秒)', '成功请求数']),
        
        ('后端处理缓慢接口 (处理占比>60%)', '后端处理占比(%)', 60,
         ['API', '处理占比(%)', '处理时间(秒)', '成功请求数'],
         ['请求URI', '后端处理占比(%)', '后端处理时长(秒)', '成功请求数']),
        
        ('大数据传输接口 (P95>1MB)', 'P95响应体大小(KB)', 1024,
         ['API', 'P95响应体(KB)', '平均响应体(KB)', '成功请求数'],
         ['请求URI', 'P95响应体大小(KB)', '平均响应体大小(KB)', '成功请求数'])
    ]
    
    current_row = 1
    for title, filter_col, threshold, display_headers, data_cols in analysis_sections:
        # 筛选数据
        filtered_data = results_df[results_df[filter_col] > threshold].sort_values(by=filter_col, ascending=False)
        
        if filtered_data.empty:
            continue
        
        # 添加标题
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=12)
        current_row += 2
        
        # 添加表头
        for col_idx, header in enumerate(display_headers, start=1):
            ws.cell(row=current_row, column=col_idx, value=header).font = Font(bold=True)
        current_row += 1
        
        # 添加数据
        for _, row_data in filtered_data.head(10).iterrows():
            for col_idx, col_name in enumerate(data_cols, start=1):
                value = row_data.get(col_name, '')
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
        
        current_row += 2  # 空行分隔
    
    # 设置列宽
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 25 if col == 1 else 15
    
    format_excel_sheet(ws)


def add_performance_bottleneck_analysis(workbook, results_df):
    """添加性能瓶颈分析工作表"""
    ws = workbook.create_sheet(title='性能瓶颈分析')

    # 分析各类性能问题
    analysis_sections = []

    # 1. 网络连接问题API
    high_connect_apis = results_df[results_df['后端连接占比(%)'] > 20].sort_values(by='后端连接占比(%)', ascending=False)
    if not high_connect_apis.empty:
        analysis_sections.append({
            'title': '网络连接耗时过高的API (连接占比>20%)',
            'data': high_connect_apis[['请求URI', '后端连接时长(秒)', '后端连接占比(%)', '成功请求数']].head(10),
            'headers': ['API', '连接时长(秒)', '连接占比(%)', '请求数']
        })

    # 2. 后端处理缓慢API
    slow_process_apis = results_df[results_df['后端处理占比(%)'] > 60].sort_values(by='后端处理占比(%)', ascending=False)
    if not slow_process_apis.empty:
        analysis_sections.append({
            'title': '后端处理耗时过高的API (处理占比>60%)',
            'data': slow_process_apis[['请求URI', '后端处理时长(秒)', '后端处理占比(%)', '成功请求数']].head(10),
            'headers': ['API', '处理时长(秒)', '处理占比(%)', '请求数']
        })

    # 3. 数据传输问题API
    high_transfer_apis = results_df[results_df['后端传输占比(%)'] > 30].sort_values(by='后端传输占比(%)', ascending=False)
    if not high_transfer_apis.empty:
        analysis_sections.append({
            'title': '数据传输耗时过高的API (传输占比>30%)',
            'data': high_transfer_apis[['请求URI', '后端传输时长(秒)', '后端传输占比(%)', '平均响应体大小(KB)', '成功请求数']].head(10),
            'headers': ['API', '传输时长(秒)', '传输占比(%)', '响应体大小(KB)', '请求数']
        })

    # 4. 低效传输API
    if '平均传输速度(KB/s)' in results_df.columns:
        low_speed_apis = results_df[results_df['平均传输速度(KB/s)'] < 1000].sort_values(by='平均传输速度(KB/s)', ascending=True)
        if not low_speed_apis.empty:
            analysis_sections.append({
                'title': '传输速度过低的API (速度<1000KB/s)',
                'data': low_speed_apis[['请求URI', '平均传输速度(KB/s)', '平均响应体大小(KB)', '成功请求数']].head(10),
                'headers': ['API', '传输速度(KB/s)', '响应体大小(KB)', '请求数']
            })

    # 写入分析结果
    current_row = 1
    for section in analysis_sections:
        # 写入标题
        title_cell = ws.cell(row=current_row, column=1, value=section['title'])
        title_cell.font = Font(bold=True, size=12)
        current_row += 2

        # 写入表头
        for col_idx, header in enumerate(section['headers'], start=1):
            header_cell = ws.cell(row=current_row, column=col_idx, value=header)
            header_cell.font = Font(bold=True)
        current_row += 1

        # 写入数据
        for _, row_data in section['data'].iterrows():
            for col_idx, value in enumerate(row_data.values, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

        current_row += 2  # 空行分隔

    # 设置列宽
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 25 if col == 1 else 15

    format_excel_sheet(ws)

def add_top_apis_section(worksheet, df, start_row, title, display_headers, data_columns):
    """添加TOP API分析区域"""
    worksheet.cell(row=start_row, column=1, value=title).font = Font(bold=True)

    # 获取前10个API数据
    top_apis = df.head(10)[data_columns].copy()
    top_apis.columns = display_headers

    # 添加表头
    for col_idx, header in enumerate(display_headers, start=1):
        worksheet.cell(row=start_row + 1, column=col_idx, value=header).font = Font(bold=True)
        worksheet.column_dimensions[chr(64 + col_idx)].width = 20 if col_idx == 1 else 15

    # 添加数据行
    for row_idx, (_, row) in enumerate(top_apis.iterrows(), start=start_row + 2):
        for col_idx, col_name in enumerate(display_headers, start=1):
            value = row[col_name] if col_name in row.index else ""
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            # 高亮慢请求
            if col_name == '平均响应时间(秒)' and value > DEFAULT_SLOW_THRESHOLD:
                cell.fill = HIGHLIGHT_FILL

def add_phase_time_analysis(workbook, total_phase_times, response_times):
    """添加阶段时间分析工作表"""
    if not total_phase_times:
        return

    # 阶段定义映射（基于新的参数体系）
    phases = {
        'backend_connect_phase': '后端连接阶段',
        'backend_process_phase': '后端处理阶段',
        'backend_transfer_phase': '后端传输阶段',
        'nginx_transfer_phase': 'Nginx传输阶段'
    }

    total_time = sum(response_times) / len(response_times) if response_times and len(response_times) > 0 else 0

    data = []
    for phase_key, phase_name in phases.items():
        percentage = total_phase_times.get(phase_key, 0)
        time_value = total_time * percentage / 100 if percentage > 0 else 0
        data.append([phase_name, round(percentage, 2), round(time_value, 3)])

    # 添加总计行
    total_percent = sum(percentage for phase_key, percentage in total_phase_times.items())
    data.append(['总计', round(total_percent, 2), round(total_time, 3)])

    phase_df = pd.DataFrame(data, columns=['请求阶段', '耗时占比(%)', '平均耗时(秒)'])
    ws = add_dataframe_to_excel_with_grouped_headers(workbook, phase_df, '阶段耗时分析')

    # 添加说明注释
    note_row = len(data) + 3
    note = ws.cell(row=note_row, column=1, value='注: 阶段耗时分析基于成功请求的平均耗时')
    note.font = Font(italic=True)

    format_excel_sheet(ws)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def add_transfer_performance_analysis(workbook, results_df):
    """添加传输性能专项分析（基于新增参数）"""
    ws = workbook.create_sheet(title='传输性能分析')

    current_row = 1

    # 1. 传输速度分析
    ws.cell(row=current_row, column=1, value='=== 传输速度综合分析 ===').font = Font(bold=True, size=12)
    current_row += 2

    # 响应传输速度分析
    if 'response_transfer_speed' in results_df.columns:
        speed_stats = [
            ['响应传输速度统计', ''],
            ['平均响应传输速度(KB/s)', round(results_df['response_transfer_speed'].mean(), 2)],
            ['最低响应传输速度(KB/s)', round(results_df['response_transfer_speed'].min(), 2)],
            ['最高响应传输速度(KB/s)', round(results_df['response_transfer_speed'].max(), 2)],
            ['P50响应传输速度(KB/s)', round(results_df['response_transfer_speed'].median(), 2)],
            ['P90响应传输速度(KB/s)', round(results_df['response_transfer_speed'].quantile(0.9), 2)],
            ['P95响应传输速度(KB/s)', round(results_df['response_transfer_speed'].quantile(0.95), 2)],
            ['', '']
        ]

        for label, value in speed_stats:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True) if label and not value else None
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

    # 总传输速度分析
    if 'total_transfer_speed' in results_df.columns:
        total_speed_stats = [
            ['总传输速度统计', ''],
            ['平均总传输速度(KB/s)', round(results_df['total_transfer_speed'].mean(), 2)],
            ['最低总传输速度(KB/s)', round(results_df['total_transfer_speed'].min(), 2)],
            ['最高总传输速度(KB/s)', round(results_df['total_transfer_speed'].max(), 2)],
            ['P50总传输速度(KB/s)', round(results_df['total_transfer_speed'].median(), 2)],
            ['P90总传输速度(KB/s)', round(results_df['total_transfer_speed'].quantile(0.9), 2)],
            ['P95总传输速度(KB/s)', round(results_df['total_transfer_speed'].quantile(0.95), 2)],
            ['', '']
        ]

        for label, value in total_speed_stats:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True) if label and not value else None
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

    # 2. 连接成本分析
    current_row += 2
    ws.cell(row=current_row, column=1, value='=== 连接成本分析 ===').font = Font(bold=True, size=12)
    current_row += 2

    if 'connection_cost_ratio' in results_df.columns:
        connection_stats = [
            ['连接成本比例统计', ''],
            ['平均连接成本比例(%)', round(results_df['connection_cost_ratio'].mean(), 2)],
            ['最低连接成本比例(%)', round(results_df['connection_cost_ratio'].min(), 2)],
            ['最高连接成本比例(%)', round(results_df['connection_cost_ratio'].max(), 2)],
            ['P90连接成本比例(%)', round(results_df['connection_cost_ratio'].quantile(0.9), 2)],
            ['', ''],
            ['连接成本过高的API数量(>20%)', len(results_df[results_df['connection_cost_ratio'] > 20])],
            ['连接成本异常的API数量(>50%)', len(results_df[results_df['connection_cost_ratio'] > 50])],
            ['', '']
        ]

        for label, value in connection_stats:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True) if label and not value else None
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

    # 3. 处理效率指数分析
    current_row += 2
    ws.cell(row=current_row, column=1, value='=== 处理效率指数分析 ===').font = Font(bold=True, size=12)
    current_row += 2

    if 'processing_efficiency_index' in results_df.columns:
        efficiency_stats = [
            ['处理效率指数统计', ''],
            ['平均处理效率指数', round(results_df['processing_efficiency_index'].mean(), 3)],
            ['最低处理效率指数', round(results_df['processing_efficiency_index'].min(), 3)],
            ['最高处理效率指数', round(results_df['processing_efficiency_index'].max(), 3)],
            ['P10处理效率指数', round(results_df['processing_efficiency_index'].quantile(0.1), 3)],
            ['P50处理效率指数', round(results_df['processing_efficiency_index'].quantile(0.5), 3)],
            ['P90处理效率指数', round(results_df['processing_efficiency_index'].quantile(0.9), 3)],
            ['', ''],
            ['低效率API数量(<0.5)', len(results_df[results_df['processing_efficiency_index'] < 0.5])],
            ['高效率API数量(>0.8)', len(results_df[results_df['processing_efficiency_index'] > 0.8])],
            ['', '']
        ]

        for label, value in efficiency_stats:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True) if label and not value else None
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1

    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # 添加传输性能问题API列表
    current_row += 3
    add_transfer_problem_apis(ws, current_row, results_df)

    format_excel_sheet(ws)

def add_transfer_problem_apis(worksheet, start_row, results_df):
    """添加传输性能问题API列表"""

    # 1. 传输速度过慢的API
    if 'response_transfer_speed' in results_df.columns:
        worksheet.cell(row=start_row, column=1, value='响应传输速度过慢的API (< 500KB/s)').font = Font(bold=True)
        start_row += 1

        slow_speed_apis = results_df[results_df['response_transfer_speed'] < 500].sort_values(
            by='response_transfer_speed', ascending=True).head(10)

        headers = ['API', '响应传输速度(KB/s)', '响应体大小(KB)', '后端传输时长(秒)', '请求数']
        columns = ['请求URI', 'response_transfer_speed', 'response_body_size_kb', 'backend_transfer_phase', '成功请求数']

        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=start_row, column=col_idx, value=header).font = Font(bold=True)
        start_row += 1

        for _, row in slow_speed_apis.iterrows():
            for col_idx, col_name in enumerate(columns, start=1):
                value = row.get(col_name, 0)
                worksheet.cell(row=start_row, column=col_idx, value=value)
            start_row += 1

        start_row += 2

    # 2. 连接成本过高的API
    if 'connection_cost_ratio' in results_df.columns:
        worksheet.cell(row=start_row, column=1, value='连接成本过高的API (> 30%)').font = Font(bold=True)
        start_row += 1

        high_cost_apis = results_df[results_df['connection_cost_ratio'] > 30].sort_values(
            by='connection_cost_ratio', ascending=False).head(10)

        headers = ['API', '连接成本比例(%)', '后端连接时长(秒)', '总请求时长(秒)', '请求数']
        columns = ['请求URI', 'connection_cost_ratio', 'backend_connect_phase', 'total_request_duration', '成功请求数']

        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=start_row, column=col_idx, value=header).font = Font(bold=True)
        start_row += 1

        for _, row in high_cost_apis.iterrows():
            for col_idx, col_name in enumerate(columns, start=1):
                value = row.get(col_name, 0)
                worksheet.cell(row=start_row, column=col_idx, value=value)
            start_row += 1

        start_row += 2

    # 3. 处理效率过低的API
    if 'processing_efficiency_index' in results_df.columns:
        worksheet.cell(row=start_row, column=1, value='处理效率过低的API (< 0.3)').font = Font(bold=True)
        start_row += 1

        low_efficiency_apis = results_df[results_df['processing_efficiency_index'] < 0.3].sort_values(
            by='processing_efficiency_index', ascending=True).head(10)

        headers = ['API', '处理效率指数', '后端处理时长(秒)', '总请求时长(秒)', '请求数']
        columns = ['请求URI', 'processing_efficiency_index', 'backend_process_phase', 'total_request_duration',
                   '成功请求数']

        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=start_row, column=col_idx, value=header).font = Font(bold=True)
        start_row += 1

        for _, row in low_efficiency_apis.iterrows():
            for col_idx, col_name in enumerate(columns, start=1):
                value = row.get(col_name, 0)
                worksheet.cell(row=start_row, column=col_idx, value=value)
            start_row += 1

def update_field_mapping_with_unified_units():
    """更新字段映射以适配统一的KB单位"""
    field_mapping = {
        'uri': 'request_full_uri',
        'app': 'application_name',
        'service': 'service_name',
        'status': 'response_status_code',
        'request_time': 'total_request_duration',
        'header_time': 'upstream_header_time',
        'connect_time': 'upstream_connect_time',
        'response_time': 'upstream_response_time',

        # 统一为KB单位的大小字段
        'body_bytes_kb': 'response_body_size_kb',
        'bytes_sent_kb': 'total_bytes_sent_kb',

        # 基于文档的新参数体系
        'backend_connect_phase': 'backend_connect_phase',
        'backend_process_phase': 'backend_process_phase',
        'backend_transfer_phase': 'backend_transfer_phase',
        'nginx_transfer_phase': 'nginx_transfer_phase',
        'backend_efficiency': 'backend_efficiency',
        'network_overhead': 'network_overhead',
        'transfer_ratio': 'transfer_ratio',

        # 新增的传输性能参数
        'response_transfer_speed': 'response_transfer_speed',
        'total_transfer_speed': 'total_transfer_speed',
        'nginx_transfer_speed': 'nginx_transfer_speed',
        'connection_cost_ratio': 'connection_cost_ratio',
        'processing_efficiency_index': 'processing_efficiency_index'
    }
    return field_mapping

def enhance_api_statistics_with_new_metrics(api_stats, group, field_mapping):
    """使用新指标增强API统计信息"""

    # 传输速度相关指标
    response_transfer_speeds = pd.to_numeric(group[field_mapping['response_transfer_speed']], errors='coerce') \
        if field_mapping['response_transfer_speed'] in group.columns else pd.Series([])
    total_transfer_speeds = pd.to_numeric(group[field_mapping['total_transfer_speed']], errors='coerce') \
        if field_mapping['total_transfer_speed'] in group.columns else pd.Series([])
    nginx_transfer_speeds = pd.to_numeric(group[field_mapping['nginx_transfer_speed']], errors='coerce') \
        if field_mapping['nginx_transfer_speed'] in group.columns else pd.Series([])

    # 效率指标
    connection_cost_ratios = pd.to_numeric(group[field_mapping['connection_cost_ratio']], errors='coerce') \
        if field_mapping['connection_cost_ratio'] in group.columns else pd.Series([])
    processing_efficiency_indices = pd.to_numeric(group[field_mapping['processing_efficiency_index']],
                                                  errors='coerce') \
        if field_mapping['processing_efficiency_index'] in group.columns else pd.Series([])

    api = group[field_mapping['uri']].iloc[0]

    # 将新指标添加到统计中
    if len(response_transfer_speeds) > 0:
        api_stats[api]['response_transfer_speed_values'] = response_transfer_speeds.dropna().tolist()
    if len(total_transfer_speeds) > 0:
        api_stats[api]['total_transfer_speed_values'] = total_transfer_speeds.dropna().tolist()
    if len(nginx_transfer_speeds) > 0:
        api_stats[api]['nginx_transfer_speed_values'] = nginx_transfer_speeds.dropna().tolist()
    if len(connection_cost_ratios) > 0:
        api_stats[api]['connection_cost_ratio_values'] = connection_cost_ratios.dropna().tolist()
    if len(processing_efficiency_indices) > 0:
        api_stats[api]['processing_efficiency_index_values'] = processing_efficiency_indices.dropna().tolist()

def generate_enhanced_api_statistics_with_new_metrics(api_stats, success_requests, total_slow_requests):
    """生成包含新指标的增强API统计信息"""
    results = []

    for api, stats in api_stats.items():
        # ... 原有的基础统计计算代码保持不变 ...

        # 新增传输速度指标统计
        response_speeds = np.array(stats.get('response_transfer_speed_values', []))
        total_speeds = np.array(stats.get('total_transfer_speed_values', []))
        nginx_speeds = np.array(stats.get('nginx_transfer_speed_values', []))
        connection_costs = np.array(stats.get('connection_cost_ratio_values', []))
        efficiency_indices = np.array(stats.get('processing_efficiency_index_values', []))

        # 添加新的统计字段到结果中
        enhanced_metrics = {}

        if len(response_speeds) > 0:
            enhanced_metrics.update({
                '平均响应传输速度(KB/s)': round(np.mean(response_speeds), 2),
                'P10响应传输速度(KB/s)': round(np.percentile(response_speeds, 10), 2),
                'P90响应传输速度(KB/s)': round(np.percentile(response_speeds, 90), 2),
            })

        if len(total_speeds) > 0:
            enhanced_metrics.update({
                '平均总传输速度(KB/s)': round(np.mean(total_speeds), 2),
                'P10总传输速度(KB/s)': round(np.percentile(total_speeds, 10), 2),
                'P90总传输速度(KB/s)': round(np.percentile(total_speeds, 90), 2),
            })

        if len(nginx_speeds) > 0:
            enhanced_metrics.update({
                '平均Nginx传输速度(KB/s)': round(np.mean(nginx_speeds), 2),
                'P90Nginx传输速度(KB/s)': round(np.percentile(nginx_speeds, 90), 2),
            })

        if len(connection_costs) > 0:
            enhanced_metrics.update({
                '平均连接成本比例(%)': round(np.mean(connection_costs), 2),
                'P90连接成本比例(%)': round(np.percentile(connection_costs, 90), 2),
                '连接成本过高(>30%)': '是' if np.mean(connection_costs) > 30 else '否',
            })

        if len(efficiency_indices) > 0:
            enhanced_metrics.update({
                '平均处理效率指数': round(np.mean(efficiency_indices), 3),
                'P10处理效率指数': round(np.percentile(efficiency_indices, 10), 3),
                'P90处理效率指数': round(np.percentile(efficiency_indices, 90), 3),
                '处理效率等级': classify_processing_efficiency(np.mean(efficiency_indices)),
            })

        # 将增强指标合并到结果中
        result.update(enhanced_metrics)
        results.append(result)

    return results

def classify_processing_efficiency(efficiency_index):
    """根据处理效率指数进行分级"""
    if efficiency_index >= 0.8:
        return '优秀'
    elif efficiency_index >= 0.6:
        return '良好'
    elif efficiency_index >= 0.4:
        return '一般'
    elif efficiency_index >= 0.2:
        return '较差'
    else:
        return '极差'