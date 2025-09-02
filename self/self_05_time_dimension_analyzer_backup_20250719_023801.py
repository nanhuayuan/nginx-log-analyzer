import gc
import os
import time
from collections import defaultdict
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook

# 导入Excel处理模块
from self_00_04_excel_processor import (
    add_dataframe_to_excel_with_grouped_headers,
    create_line_chart,
    create_pie_chart
)
from self_00_02_utils import log_info

# 常量
DEFAULT_CHUNK_SIZE = 10000
DEFAULT_SLOW_THRESHOLD = 5.0


class StreamingTimeAnalyzer:
    """流式时间维度分析器 - 一次遍历完成所有统计"""

    def __init__(self, slow_threshold=DEFAULT_SLOW_THRESHOLD):
        self.slow_threshold = slow_threshold
        self.stats = {
            'daily': {},
            'hourly': {},
            'minute': {},
            'second': {}
        }
        # 指标统计
        self.metrics_sum = {
            'daily': defaultdict(lambda: defaultdict(float)),
            'hourly': defaultdict(lambda: defaultdict(float)),
            'minute': defaultdict(lambda: defaultdict(float)),
            'second': defaultdict(lambda: defaultdict(float))
        }
        self.metrics_count = {
            'daily': defaultdict(lambda: defaultdict(int)),
            'hourly': defaultdict(lambda: defaultdict(int)),
            'minute': defaultdict(lambda: defaultdict(int)),
            'second': defaultdict(lambda: defaultdict(int))
        }

        # 预定义指标列表
        self.time_metrics = [
            'total_request_duration', 'upstream_response_time',
            'upstream_header_time', 'upstream_connect_time'
        ]

    def extract_time_keys(self, timestamp):
        """提取时间维度键"""
        if pd.isna(timestamp):
            return None

        try:
            dt = pd.to_datetime(timestamp)
            if pd.isna(dt):
                return None

            return {
                'daily': dt.strftime('%Y-%m-%d'),
                'hourly': dt.strftime('%Y-%m-%d %H:00'),
                'minute': dt.strftime('%Y-%m-%d %H:%M'),
                'second': dt.strftime('%Y-%m-%d %H:%M:%S')
            }
        except (ValueError, TypeError):
            return None

    def process_single_record(self, record):
        """处理单条记录"""
        time_keys = self.extract_time_keys(record.get('arrival_time'))
        if not time_keys:
            return

        # 获取关键指标
        status_code = record.get('response_status_code')
        duration = record.get('total_request_duration')

        # 判断请求状态
        is_success = (200 <= status_code < 400) if pd.notna(status_code) else False
        is_slow = (duration > self.slow_threshold) if pd.notna(duration) else False

        # 更新所有维度统计
        for dimension, time_key in time_keys.items():
            if time_key not in self.stats[dimension]:
                self.stats[dimension][time_key] = {
                    'total_requests': 0,
                    'success_requests': 0,
                    'slow_requests': 0
                }

            # 更新基础统计
            stats = self.stats[dimension][time_key]
            stats['total_requests'] += 1
            if is_success:
                stats['success_requests'] += 1
            if is_slow:
                stats['slow_requests'] += 1

            # 更新指标统计
            self._update_metrics(dimension, time_key, record)

    def _update_metrics(self, dimension, time_key, record):
        """更新指标统计"""
        for metric in self.time_metrics:
            value = record.get(metric)
            if pd.notna(value) and value > 0:
                self.metrics_sum[dimension][time_key][metric] += float(value)
                self.metrics_count[dimension][time_key][metric] += 1

    def calculate_derived_metrics(self):
        """计算衍生指标"""
        # 计算平均值
        avg_metrics = {}
        for dimension in self.stats.keys():
            avg_metrics[dimension] = {}
            for time_key in self.stats[dimension].keys():
                avg_metrics[dimension][time_key] = {}
                for metric in self.time_metrics:
                    total = self.metrics_sum[dimension][time_key][metric]
                    count = self.metrics_count[dimension][time_key][metric]
                    avg_metrics[dimension][time_key][metric] = total / count if count > 0 else 0

        # 计算QPS
        window_seconds = {'daily': 86400, 'hourly': 3600, 'minute': 60, 'second': 1}
        for dimension, seconds in window_seconds.items():
            for time_key, stats in self.stats[dimension].items():
                success_requests = stats.get('success_requests', 0)
                stats['qps'] = success_requests / seconds

        return avg_metrics


def read_and_process_data_streaming(csv_path, specific_uri_list=None):
    """流式处理数据"""
    analyzer = StreamingTimeAnalyzer()
    total_records = 0
    processed_records = 0
    error_records = 0
    uri_filtered_records = 0
    chunk_size = 1000

    # 处理URI过滤
    uri_set = None
    if specific_uri_list:
        if isinstance(specific_uri_list, str):
            uri_set = {specific_uri_list}
            log_info(f"分析指定URI: {specific_uri_list}")
        elif isinstance(specific_uri_list, list):
            uri_set = set(specific_uri_list)
            log_info(f"分析 {len(specific_uri_list)} 个URI")
        log_info("开始分析所有请求")

    start_time = time.time()

    try:
        for chunk_idx, chunk in enumerate(pd.read_csv(csv_path, chunksize=chunk_size)):
            total_records += len(chunk)

            # URI过滤
            if uri_set and 'request_full_uri' in chunk.columns:
                before_filter = len(chunk)
                chunk = chunk[chunk['request_full_uri'].isin(uri_set)].copy()
                uri_filtered_records += (before_filter - len(chunk))
                if chunk.empty:
                    continue

            # 预处理数据
            try:
                chunk = _preprocess_chunk_fast(chunk)
            except Exception as e:
                log_info(f"Chunk {chunk_idx} 预处理失败: {e}")
                error_records += len(chunk)
                continue

            # 处理每条记录
            for _, record in chunk.iterrows():
                try:
                    analyzer.process_single_record(record)
                    processed_records += 1
                except Exception as e:
                    error_records += 1
                    if error_records <= 10:
                        log_info(f"记录处理错误: {e}")

            if total_records % 10000 == 0:
                log_info(f"已处理 {total_records} 条记录...")

    except Exception as e:
        log_info(f"数据读取错误: {e}")
        raise

    # 计算衍生指标
    avg_metrics = analyzer.calculate_derived_metrics()

    elapsed = time.time() - start_time
    log_info(f"处理完成 - 总记录: {total_records}, 成功: {processed_records}, 错误: {error_records}")
    log_info(f"耗时: {elapsed:.2f}秒, 速度: {processed_records / elapsed:.0f} 记录/秒")

    return analyzer.stats, avg_metrics, processed_records


def _preprocess_chunk_fast(chunk):
    """快速预处理chunk数据"""
    required_cols = ['arrival_time']
    missing_cols = [col for col in required_cols if col not in chunk.columns]
    if missing_cols:
        raise ValueError(f"缺少必要列: {missing_cols}")

    chunk = chunk.copy()

    # 数据类型转换
    type_mapping = {
        'arrival_time': 'datetime64[ns]',
        'response_status_code': 'int32',
        'total_request_duration': 'float32',
        'upstream_response_time': 'float32',
        'upstream_header_time': 'float32',
        'upstream_connect_time': 'float32'
    }

    for col, dtype in type_mapping.items():
        if col in chunk.columns:
            try:
                if dtype == 'datetime64[ns]':
                    chunk.loc[:, col] = pd.to_datetime(chunk[col], errors='coerce')
                else:
                    chunk.loc[:, col] = pd.to_numeric(chunk[col], errors='coerce').astype(dtype)
            except Exception as e:
                log_info(f"列 {col} 转换失败: {e}")

    return chunk


def create_time_dimension_excel_optimized(output_path, stats_containers, avg_metrics, total_records, peak_hour,
                                          peak_minute):
    """创建优化的Excel报告"""
    wb = Workbook()

    # 移除默认工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 创建概览页
    create_overview_sheet(wb, stats_containers, total_records, peak_hour)

    # 创建各维度分析页
    dimensions = [
        ('日期维度分析', 'daily', '日期'),
        ('小时维度分析', 'hourly', '小时'),
        ('分钟维度分析', 'minute', '分钟'),
        ('秒级维度分析', 'second', '秒')
    ]

    for sheet_name, dimension, time_label in dimensions:
        if stats_containers[dimension]:  # 只创建有数据的维度
            create_dimension_sheet_with_formatting(wb, sheet_name, stats_containers[dimension],
                                                   avg_metrics[dimension], time_label)

    wb.save(output_path)
    wb.close()


def create_overview_sheet(wb, stats_containers, total_records, peak_hour):
    """创建概览页"""
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet(title="概览")

    # 计算总体统计
    daily_stats = stats_containers['daily']
    total_success = sum(stats.get('success_requests', 0) for stats in daily_stats.values())
    total_slow = sum(stats.get('slow_requests', 0) for stats in daily_stats.values())

    # 标题
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "HTTP请求生命周期分析报告"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')

    # 统计数据
    overview_data = [
        ["指标名称", "数值", "单位", "备注"],
        ["总请求数", total_records, "个", "所有HTTP请求"],
        ["成功请求数", total_success, "个", "状态码2xx-3xx"],
        ["成功率", round(total_success / total_records * 100, 2) if total_records > 0 else 0, "%", "成功请求占比"],
        ["慢请求数", total_slow, "个", f"响应时间>{DEFAULT_SLOW_THRESHOLD}s"],
        ["慢请求率", round(total_slow / total_records * 100, 2) if total_records > 0 else 0, "%", "慢请求占比"],
        ["峰值时段", peak_hour or "无数据", "", "请求量最高时段"]
    ]

    # 写入数据到工作表
    for row_idx, row_data in enumerate(overview_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # 表头
                cell.font = Font(bold=True)

    # 调用格式化函数
    from self_00_04_excel_processor import format_excel_sheet
    format_excel_sheet(ws, has_grouped_header=False, header_end_row=3)


def create_dimension_sheet_with_formatting(wb, sheet_name, stats_dict, avg_metrics, time_label):
    """创建带格式化的维度分析页"""
    if not stats_dict:
        return

    # 准备DataFrame数据
    df_data = []
    for time_key in sorted(stats_dict.keys()):
        stats = stats_dict[time_key]
        avg_stats = avg_metrics.get(time_key, {})

        total = stats.get('total_requests', 0)
        success = stats.get('success_requests', 0)
        slow = stats.get('slow_requests', 0)

        # 计算比率和平均值 - 返回数值而不是字符串
        success_rate = (success / total * 100) if total > 0 else 0
        slow_rate = (slow / total * 100) if total > 0 else 0
        qps = stats.get('qps', 0)

        # 平均响应时间指标
        avg_duration = avg_stats.get('total_request_duration', 0)
        avg_upstream_time = avg_stats.get('upstream_response_time', 0)
        avg_header_time = avg_stats.get('upstream_header_time', 0)
        avg_connect_time = avg_stats.get('upstream_connect_time', 0)

        row = [
            time_key,  # 时间
            total,  # 总请求数
            success,  # 成功请求数
            success_rate,  # 成功率(%) - 数值
            slow,  # 慢请求数
            slow_rate,  # 慢请求率(%) - 数值
            qps,  # QPS - 数值
            avg_duration,  # 平均响应时间(s) - 数值
            avg_upstream_time,  # 平均上游响应时间(s) - 数值
            avg_header_time,  # 平均响应头时间(s) - 数值
            avg_connect_time  # 平均连接时间(s) - 数值
        ]
        df_data.append(row)

    # 创建DataFrame
    columns = [
        time_label, '总请求数', '成功请求数', '成功率(%)',
        '慢请求数', '慢请求率(%)', 'QPS',
        '平均响应时间(s)', '平均上游响应时间(s)',
        '平均响应头时间(s)', '平均连接时间(s)'
    ]

    df = pd.DataFrame(df_data, columns=columns)

    # 定义分组表头
    header_groups = {
        '时间维度': [time_label],
        '请求统计': ['总请求数', '成功请求数', '成功率(%)'],
        '性能统计': ['慢请求数', '慢请求率(%)', 'QPS'],
        '响应时间分析': ['平均响应时间(s)', '平均上游响应时间(s)', '平均响应头时间(s)', '平均连接时间(s)']
    }

    # 使用Excel处理模块添加工作表
    ws = add_dataframe_to_excel_with_grouped_headers(wb, df, sheet_name, header_groups)

    # 添加图表（如果数据不为空且不超过图表限制）
    if len(df) > 1 and len(df) <= 100:  # 限制图表数据点
        try:
            # QPS趋势图
            create_line_chart(
                ws,
                min_row=4,  # 从数据行开始（跳过双行表头）
                max_row=3 + len(df),
                title=f"{time_label}QPS趋势",
                x_title=time_label,
                y_title="QPS",
                y_cols=[7],  # QPS列
                chart_position="N5"
            )

            # 平均响应时间趋势图
            create_line_chart(
                ws,
                min_row=4,
                max_row=3 + len(df),
                title=f"{time_label}平均响应时间趋势",
                x_title=time_label,
                y_title="响应时间(秒)",
                y_cols=[8, 9, 10, 11],  # 各种响应时间列
                series_names=['总响应时间', '上游响应时间', '响应头时间', '连接时间'],
                chart_position="N20"
            )

        except Exception as e:
            log_info(f"创建图表失败: {e}")


def identify_peak_period(stats_dict, metric_key):
    """识别峰值时段"""
    if not stats_dict:
        return None
    return max(stats_dict.items(), key=lambda x: x[1].get(metric_key, 0))[0]


def prepare_output_filename(output_path, specific_uri_list):
    """准备输出文件名"""
    if not specific_uri_list:
        return output_path

    base_name = os.path.basename(output_path)
    name_parts = os.path.splitext(base_name)

    if isinstance(specific_uri_list, list):
        uri_identifier = specific_uri_list[0].replace('/', '_').replace('-', '_')
    else:
        uri_identifier = specific_uri_list.replace('/', '_').replace('-', '_')

    if len(uri_identifier) > 30:
        uri_identifier = uri_identifier[:30]

    new_filename = f"{name_parts[0]}_{uri_identifier}{name_parts[1]}"
    return os.path.join(os.path.dirname(output_path), new_filename)


def analyze_time_dimension_optimized(csv_path, output_path, specific_uri_list=None):
    """优化后的时间维度分析主函数"""
    output_filename = prepare_output_filename(output_path, specific_uri_list)

    # 流式处理数据
    stats_containers, avg_metrics, total_records = read_and_process_data_streaming(csv_path, specific_uri_list)

    # 识别峰值时段
    peak_hour = identify_peak_period(stats_containers['hourly'], 'success_requests')
    peak_minute = identify_peak_period(stats_containers['minute'], 'success_requests')

    # 生成Excel报告
    create_time_dimension_excel_optimized(
        output_filename,
        stats_containers,
        avg_metrics,
        total_records,
        peak_hour,
        peak_minute
    )

    log_info(f"时间维度分析完成，报告已生成：{output_filename}")
    return output_filename


# 主函数 - 保持向后兼容
def analyze_time_dimension(csv_path, output_path, specific_uri_list=None):
    """主分析函数 - 兼容原接口"""
    return analyze_time_dimension_optimized(csv_path, output_path, specific_uri_list)