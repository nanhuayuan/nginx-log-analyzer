import gc
import pandas as pd
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font

from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD
from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    format_excel_sheet,
    add_dataframe_to_excel_with_grouped_headers,
    create_pie_chart,
    create_line_chart
)
from self_10_request_header_analyzer import (
    extract_browser_info, 
    extract_os_info, 
    extract_device_info,
    detect_bot_type,
    extract_domain_from_referer,
    detect_search_engine,
    detect_social_media
)


def analyze_header_performance_correlation(csv_path, output_path, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """分析请求头与性能的关联性"""
    log_info("开始分析请求头与性能的关联性...", show_memory=True)
    
    chunk_size = max(DEFAULT_CHUNK_SIZE // 2, 10000)
    
    # 性能统计数据结构
    browser_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times': [],
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    os_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times': [],
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    device_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times': [],
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    domain_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times': [],
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    search_engine_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times': [],
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    bot_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times': [],
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    # 慢请求详细分析
    slow_request_details = []
    
    total_processed = 0
    total_slow_requests = 0
    
    # 第一遍：收集性能关联数据
    log_info("开始收集请求头性能关联数据")
    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        chunk_size_actual = len(chunk)
        total_processed += chunk_size_actual
        
        # 确保必要的列存在
        required_columns = ['user_agent_string', 'referer_url', 'total_request_duration', 'response_status_code']
        missing_columns = [col for col in required_columns if col not in chunk.columns]
        if missing_columns:
            log_info(f"警告: 缺少必要列 {missing_columns}", level="WARNING")
            continue
        
        # 处理数据类型
        chunk['total_request_duration'] = pd.to_numeric(chunk['total_request_duration'], errors='coerce')
        chunk['response_status_code'] = chunk['response_status_code'].astype(str)
        
        if 'response_body_size_kb' in chunk.columns:
            chunk['response_body_size_kb'] = pd.to_numeric(chunk['response_body_size_kb'], errors='coerce')
        
        for _, row in chunk.iterrows():
            user_agent = row.get('user_agent_string', '')
            referer = row.get('referer_url', '')
            response_time = row.get('total_request_duration', 0)
            status_code = str(row.get('response_status_code', ''))
            data_size = row.get('response_body_size_kb', 0) or 0
            
            # 跳过无效数据
            if pd.isna(response_time) or response_time <= 0:
                continue
            
            is_slow = response_time > slow_threshold
            is_error = status_code.startswith('4') or status_code.startswith('5')
            
            if is_slow:
                total_slow_requests += 1
            
            # 分析User-Agent
            if pd.notna(user_agent) and user_agent != '' and user_agent != '-':
                browser = extract_browser_info(user_agent)
                os_info = extract_os_info(user_agent)
                device = extract_device_info(user_agent)
                bot_type = detect_bot_type(user_agent)
                
                # 更新浏览器性能统计
                update_performance_stats(browser_performance[browser], response_time, is_slow, is_error, data_size)
                
                # 更新操作系统性能统计  
                update_performance_stats(os_performance[os_info], response_time, is_slow, is_error, data_size)
                
                # 更新设备类型性能统计
                update_performance_stats(device_performance[device], response_time, is_slow, is_error, data_size)
                
                # 更新机器人性能统计
                if bot_type:
                    update_performance_stats(bot_performance[bot_type], response_time, is_slow, is_error, data_size)
            
            # 分析Referer
            if pd.notna(referer) and referer != '' and referer != '-':
                domain = extract_domain_from_referer(referer)
                search_engine = detect_search_engine(referer)
                
                # 更新域名性能统计
                if domain:
                    update_performance_stats(domain_performance[domain], response_time, is_slow, is_error, data_size)
                
                # 更新搜索引擎性能统计
                if search_engine:
                    update_performance_stats(search_engine_performance[search_engine], response_time, is_slow, is_error, data_size)
            
            # 收集慢请求详细信息
            if is_slow and len(slow_request_details) < 10000:  # 限制详细记录数量
                slow_detail = {
                    '请求时间': row.get('raw_time', ''),
                    '请求URI': row.get('request_full_uri', ''),
                    '响应时间(秒)': round(response_time, 3),
                    '状态码': status_code,
                    '浏览器': extract_browser_info(user_agent) if pd.notna(user_agent) else '未知',
                    '操作系统': extract_os_info(user_agent) if pd.notna(user_agent) else '未知',
                    '设备类型': extract_device_info(user_agent) if pd.notna(user_agent) else '未知',
                    '来源域名': extract_domain_from_referer(referer) if pd.notna(referer) else '直接访问',
                    'User-Agent': (user_agent[:100] + '...') if len(str(user_agent)) > 100 else user_agent,
                    'Referer': (referer[:100] + '...') if len(str(referer)) > 100 else referer
                }
                slow_request_details.append(slow_detail)
        
        if total_processed % 100000 == 0:
            gc.collect()
            log_info(f"已处理 {total_processed:,} 条记录，发现 {total_slow_requests:,} 条慢请求")
    
    log_info(f"性能关联分析完成：总记录 {total_processed:,}，慢请求 {total_slow_requests:,}")
    
    # 生成分析结果
    analysis_results = {
        'browser_performance': calculate_performance_metrics(browser_performance),
        'os_performance': calculate_performance_metrics(os_performance),
        'device_performance': calculate_performance_metrics(device_performance),
        'domain_performance': calculate_performance_metrics(domain_performance),
        'search_engine_performance': calculate_performance_metrics(search_engine_performance),
        'bot_performance': calculate_performance_metrics(bot_performance),
        'slow_request_details': slow_request_details,
        'total_processed': total_processed,
        'total_slow_requests': total_slow_requests
    }
    
    # 创建Excel报告
    create_header_performance_excel(analysis_results, output_path, slow_threshold)
    
    log_info(f"请求头性能关联分析完成，报告已生成：{output_path}", show_memory=True)
    
    # 返回关键洞察
    insights = generate_performance_insights(analysis_results, slow_threshold)
    return insights


def update_performance_stats(stats, response_time, is_slow, is_error, data_size):
    """更新性能统计数据"""
    stats['total_requests'] += 1
    stats['total_response_time'] += response_time
    
    if is_slow:
        stats['slow_requests'] += 1
    
    if is_error:
        stats['error_requests'] += 1
    
    stats['data_transferred'] += data_size
    
    # 采样保存响应时间（避免内存问题）
    if len(stats['response_times']) < 1000:
        stats['response_times'].append(response_time)


def calculate_performance_metrics(performance_data):
    """计算性能指标"""
    results = []
    
    for category, stats in performance_data.items():
        if stats['total_requests'] == 0:
            continue
        
        total_requests = stats['total_requests']
        slow_requests = stats['slow_requests']
        error_requests = stats['error_requests']
        total_response_time = stats['total_response_time']
        data_transferred = stats['data_transferred']
        response_times = stats['response_times']
        
        # 计算关键指标
        slow_rate = (slow_requests / total_requests * 100) if total_requests > 0 else 0
        error_rate = (error_requests / total_requests * 100) if total_requests > 0 else 0
        avg_response_time = total_response_time / total_requests if total_requests > 0 else 0
        avg_data_size = data_transferred / total_requests if total_requests > 0 else 0
        
        # 计算百分位数
        if response_times:
            import numpy as np
            p50 = np.percentile(response_times, 50)
            p95 = np.percentile(response_times, 95)
            p99 = np.percentile(response_times, 99)
        else:
            p50 = p95 = p99 = 0
        
        result = {
            '类别': category,
            '总请求数': total_requests,
            '慢请求数': slow_requests,
            '慢请求率(%)': round(slow_rate, 2),
            '错误请求数': error_requests,
            '错误率(%)': round(error_rate, 2),
            '平均响应时间(秒)': round(avg_response_time, 3),
            'P50响应时间(秒)': round(p50, 3),
            'P95响应时间(秒)': round(p95, 3),
            'P99响应时间(秒)': round(p99, 3),
            '平均数据传输(KB)': round(avg_data_size, 2),
            '总数据传输(MB)': round(data_transferred / 1024, 2)
        }
        
        results.append(result)
    
    # 按慢请求率排序
    results.sort(key=lambda x: x['慢请求率(%)'], reverse=True)
    return pd.DataFrame(results)


def generate_performance_insights(analysis_results, slow_threshold):
    """生成性能洞察"""
    insights = {
        'slow_threshold': slow_threshold,
        'total_processed': analysis_results['total_processed'],
        'total_slow_requests': analysis_results['total_slow_requests'],
        'slow_rate_overall': round(analysis_results['total_slow_requests'] / analysis_results['total_processed'] * 100, 2),
        'worst_browsers': [],
        'worst_os': [],
        'worst_devices': [],
        'worst_domains': [],
        'performance_recommendations': []
    }
    
    # 找出性能最差的浏览器
    browser_df = analysis_results['browser_performance']
    if not browser_df.empty:
        worst_browsers = browser_df.head(3)
        for _, row in worst_browsers.iterrows():
            insights['worst_browsers'].append({
                'browser': row['类别'],
                'slow_rate': row['慢请求率(%)'],
                'avg_response_time': row['平均响应时间(秒)']
            })
    
    # 找出性能最差的操作系统
    os_df = analysis_results['os_performance']
    if not os_df.empty:
        worst_os = os_df.head(3)
        for _, row in worst_os.iterrows():
            insights['worst_os'].append({
                'os': row['类别'],
                'slow_rate': row['慢请求率(%)'],
                'avg_response_time': row['平均响应时间(秒)']
            })
    
    # 找出性能最差的设备类型
    device_df = analysis_results['device_performance']
    if not device_df.empty:
        worst_devices = device_df.head(3)
        for _, row in worst_devices.iterrows():
            insights['worst_devices'].append({
                'device': row['类别'],
                'slow_rate': row['慢请求率(%)'],
                'avg_response_time': row['平均响应时间(秒)']
            })
    
    # 找出性能最差的来源域名
    domain_df = analysis_results['domain_performance']
    if not domain_df.empty:
        worst_domains = domain_df.head(5)
        for _, row in worst_domains.iterrows():
            insights['worst_domains'].append({
                'domain': row['类别'],
                'slow_rate': row['慢请求率(%)'],
                'avg_response_time': row['平均响应时间(秒)']
            })
    
    # 生成优化建议
    insights['performance_recommendations'] = generate_optimization_recommendations(analysis_results)
    
    return insights


def generate_optimization_recommendations(analysis_results):
    """生成性能优化建议"""
    recommendations = []
    
    # 分析浏览器性能
    browser_df = analysis_results['browser_performance']
    if not browser_df.empty and len(browser_df) > 0:
        worst_browser = browser_df.iloc[0]
        if worst_browser['慢请求率(%)'] > 10:
            recommendations.append(f"浏览器优化: {worst_browser['类别']} 的慢请求率达到 {worst_browser['慢请求率(%)']}%，建议针对该浏览器进行专项优化")
    
    # 分析设备性能
    device_df = analysis_results['device_performance']
    if not device_df.empty and len(device_df) > 0:
        for _, row in device_df.iterrows():
            if row['类别'] == '手机' and row['慢请求率(%)'] > 15:
                recommendations.append(f"移动端优化: 手机端慢请求率为 {row['慢请求率(%)']}%，建议优化移动端性能")
            elif row['类别'] == '爬虫/机器人' and row['总请求数'] > 1000:
                recommendations.append(f"机器人流量: 检测到大量机器人请求({row['总请求数']}次)，建议进行流量控制")
    
    # 分析来源域名性能
    domain_df = analysis_results['domain_performance']
    if not domain_df.empty and len(domain_df) > 0:
        high_traffic_slow_domains = domain_df[
            (domain_df['总请求数'] > 100) & (domain_df['慢请求率(%)'] > 20)
        ]
        if not high_traffic_slow_domains.empty:
            for _, row in high_traffic_slow_domains.head(3).iterrows():
                recommendations.append(f"来源优化: 来自 {row['类别']} 的请求慢请求率为 {row['慢请求率(%)']}%，建议检查相关链接或缓存策略")
    
    # 分析机器人性能
    bot_df = analysis_results['bot_performance']
    if not bot_df.empty and len(bot_df) > 0:
        for _, row in bot_df.iterrows():
            if row['总请求数'] > 500 and row['慢请求率(%)'] > 30:
                recommendations.append(f"爬虫优化: {row['类别']} 产生了 {row['总请求数']} 次请求，慢请求率 {row['慢请求率(%)']}%，建议优化爬虫响应或限制频率")
    
    return recommendations


def create_header_performance_excel(analysis_results, output_path, slow_threshold):
    """创建请求头性能关联Excel报告"""
    log_info(f"创建请求头性能关联Excel报告: {output_path}")
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 1. 概览工作表
    create_performance_overview_sheet(wb, analysis_results, slow_threshold)
    
    # 2. 浏览器性能分析
    create_performance_analysis_sheet(wb, analysis_results['browser_performance'], '浏览器性能分析', '浏览器')
    
    # 3. 操作系统性能分析
    create_performance_analysis_sheet(wb, analysis_results['os_performance'], '操作系统性能分析', '操作系统')
    
    # 4. 设备类型性能分析
    create_performance_analysis_sheet(wb, analysis_results['device_performance'], '设备类型性能分析', '设备类型')
    
    # 5. 来源域名性能分析
    create_performance_analysis_sheet(wb, analysis_results['domain_performance'], '来源域名性能分析', '来源域名')
    
    # 6. 搜索引擎性能分析
    create_performance_analysis_sheet(wb, analysis_results['search_engine_performance'], '搜索引擎性能分析', '搜索引擎')
    
    # 7. 机器人性能分析
    create_performance_analysis_sheet(wb, analysis_results['bot_performance'], '机器人性能分析', '机器人类型')
    
    # 8. 慢请求详细列表
    create_slow_requests_detail_sheet(wb, analysis_results['slow_request_details'])
    
    # 保存文件
    wb.save(output_path)
    log_info(f"请求头性能关联Excel报告已保存: {output_path}")


def create_performance_overview_sheet(wb, analysis_results, slow_threshold):
    """创建性能概览工作表"""
    ws = wb.create_sheet(title='性能关联概览')
    
    # 移动到第一个位置
    wb.move_sheet(ws, -(len(wb.worksheets) - 1))
    
    total_processed = analysis_results['total_processed']
    total_slow_requests = analysis_results['total_slow_requests']
    overall_slow_rate = round(total_slow_requests / total_processed * 100, 2) if total_processed > 0 else 0
    
    # 概览数据
    overview_data = [
        ['=== 性能关联分析概览 ===', ''],
        ['分析阈值', f'{slow_threshold} 秒'],
        ['总处理记录数', total_processed],
        ['慢请求总数', total_slow_requests],
        ['整体慢请求率', f'{overall_slow_rate}%'],
        ['', ''],
        
        ['=== 性能表现最差TOP3 ===', ''],
        ['浏览器:', ''],
    ]
    
    # 添加性能最差的浏览器
    browser_df = analysis_results['browser_performance']
    if not browser_df.empty:
        for i, (_, row) in enumerate(browser_df.head(3).iterrows(), 1):
            overview_data.append([f'  {i}. {row["类别"]}', f'慢请求率: {row["慢请求率(%)"]}%'])
    
    overview_data.extend([
        ['', ''],
        ['操作系统:', ''],
    ])
    
    # 添加性能最差的操作系统
    os_df = analysis_results['os_performance']
    if not os_df.empty:
        for i, (_, row) in enumerate(os_df.head(3).iterrows(), 1):
            overview_data.append([f'  {i}. {row["类别"]}', f'慢请求率: {row["慢请求率(%)"]}%'])
    
    overview_data.extend([
        ['', ''],
        ['设备类型:', ''],
    ])
    
    # 添加性能最差的设备类型
    device_df = analysis_results['device_performance']
    if not device_df.empty:
        for i, (_, row) in enumerate(device_df.head(3).iterrows(), 1):
            overview_data.append([f'  {i}. {row["类别"]}', f'慢请求率: {row["慢请求率(%)"]}%'])
    
    # 写入数据
    for row_idx, (label, value) in enumerate(overview_data, start=1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_value = ws.cell(row=row_idx, column=2, value=value)
        
        if str(label).startswith('===') and str(label).endswith('==='):
            cell_label.font = Font(bold=True, size=12)
        elif str(label).endswith(':'):
            cell_label.font = Font(bold=True)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    
    format_excel_sheet(ws)


def create_performance_analysis_sheet(wb, performance_df, sheet_name, category_name):
    """创建性能分析工作表"""
    if performance_df.empty:
        return
    
    # 表头分组
    headers = {
        "分类信息": ["类别"],
        "请求统计": ["总请求数", "慢请求数", "错误请求数"],
        "性能指标": ["慢请求率(%)", "错误率(%)", "平均响应时间(秒)"],
        "响应时间分布": ["P50响应时间(秒)", "P95响应时间(秒)", "P99响应时间(秒)"],
        "数据传输": ["平均数据传输(KB)", "总数据传输(MB)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, performance_df, sheet_name, header_groups=headers
    )
    
    # 添加慢请求率分布图
    if len(performance_df) > 1:
        chart_start_row = len(performance_df) + 5
        ws.cell(row=chart_start_row, column=1, value=f"{category_name}慢请求率分布").font = Font(bold=True)
        
        create_pie_chart(
            ws, f"{category_name}慢请求率分布",
            data_start_row=3, data_end_row=2 + len(performance_df),
            labels_col=1, values_col=4,  # 慢请求率列
            position="H3"
        )


def create_slow_requests_detail_sheet(wb, slow_request_details):
    """创建慢请求详细列表工作表"""
    if not slow_request_details:
        return
    
    slow_df = pd.DataFrame(slow_request_details)
    
    # 表头分组
    headers = {
        "基本信息": ["请求时间", "请求URI", "响应时间(秒)", "状态码"],
        "客户端信息": ["浏览器", "操作系统", "设备类型"],
        "来源信息": ["来源域名"],
        "详细信息": ["User-Agent", "Referer"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, slow_df, '慢请求详细列表', header_groups=headers
    )