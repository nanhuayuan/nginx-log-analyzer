"""
Excel处理模块 - 负责创建和处理Excel文件
"""

import gc
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, PieChart
from openpyxl.chart import Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from self_00_02_utils import log_info
from self_00_01_constants import EXCEL_MAX_ROWS, CHART_MAX_POINTS, HEADER_FILL, DEFAULT_SLOW_THRESHOLD, HIGHLIGHT_FILL


def save_dataframe_to_excel(df, output_path, sheet_name='数据', max_rows=EXCEL_MAX_ROWS):
    """将DataFrame保存为Excel文件，如果数据量过大会自动分页"""
    if df.empty:
        log_info(f"警告: 数据为空, 跳过保存 {output_path}", level="WARNING")
        return

    total_rows = len(df)
    sheets_needed = (total_rows // max_rows) + (1 if total_rows % max_rows > 0 else 0)
    start_time = datetime.now()

    log_info(f"开始保存Excel: {output_path} (总行数: {total_rows:,}, 需创建工作表: {sheets_needed})")

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    for i in range(sheets_needed):
        start_idx = i * max_rows
        end_idx = min((i + 1) * max_rows, total_rows)
        chunk_size = end_idx - start_idx

        current_df = df.iloc[start_idx:end_idx]
        current_sheet_name = f"{sheet_name}_{i + 1}" if sheets_needed > 1 else sheet_name

        log_info(f"正在处理第 {i + 1}/{sheets_needed} 个工作表 '{current_sheet_name}' (行数: {chunk_size:,})")
        ws = wb.create_sheet(title=current_sheet_name)

        # 写入表头
        headers = list(current_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # 写入数据
        row_count = 0
        for row_idx, row in enumerate(current_df.itertuples(), start=2):
            for col_idx, value in enumerate(row[1:], start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

            row_count += 1
            if row_count % 10000 == 0:
                log_info(
                    f"表'{current_sheet_name}'已写入 {row_count:,}/{chunk_size:,} 行 ({row_count / chunk_size * 100:.1f}%)",
                    show_memory=(row_count % 50000 == 0))
                gc.collect()

        log_info(f"正在格式化工作表 '{current_sheet_name}'...")
        format_excel_sheet(ws)
        current_df = None
        gc.collect()

    log_info(f"正在保存Excel文件...")
    wb.save(output_path)

    elapsed = (datetime.now() - start_time).total_seconds()
    log_info(f"Excel保存完成: {output_path} (总行数: {total_rows:,}, 耗时: {elapsed:.2f} 秒)")

def format_excel_sheet(sheet, has_grouped_header=False, header_end_row=1):
    """
    格式化Excel工作表，设置列宽、字体、对齐方式等

    Args:
        sheet: Excel工作表对象
        has_grouped_header: 是否有分组表头(双行表头)
        header_end_row: 表头结束行号(单行表头为1，双行表头为2)
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_NUMBER_00, FORMAT_PERCENTAGE_00
    from openpyxl.utils import get_column_letter

    # 调整列宽_
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # 创建表头样式
    header_font = Font(bold=True, name='等线')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用表头样式 - 处理单行或双行表头
    for row_idx in range(1, header_end_row + 1):
        for cell in sheet[row_idx]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = HEADER_FILL  # 从常量导入的表头填充色

    # 设置冻结窗格，从表头之后的第一行开始
    sheet.freeze_panes = f'A{header_end_row + 1}'

    # 为数值单元格设置样式和格式化
    number_font = Font(name='Consolas')

    # 遍历数据行
    for row in sheet.iter_rows(min_row=header_end_row + 1):
        for cell in row:
            # 设置边框
            cell.border = thin_border

            # 数值格式化
            if isinstance(cell.value, (int, float)):
                cell.font = number_font
                cell.alignment = Alignment(horizontal='right')

                # 根据值类型设置不同的数值格式
                if isinstance(cell.value, int):
                    # 整数格式
                    cell.number_format = FORMAT_NUMBER
                elif cell.column_letter.startswith(('百分比', '比例', '占比')) or (
                        cell.value <= 1 and cell.value >= 0 and '率' in str(
                        sheet.cell(row=1, column=cell.column).value)):
                    # 百分比格式 (0-1之间的小数 或 列名包含"百分比"/"比例"/"占比"/"率")
                    cell.number_format = FORMAT_PERCENTAGE_00  # 百分比显示2位小数
                else:
                    # 浮点数格式 (显示3位小数)
                    cell.number_format = '0.000'

                # 高亮显示慢响应时间
                if cell.column_letter.isalpha() and cell.column_letter >= 'I' and cell.column_letter <= 'P' and isinstance(
                        cell.value, float) and cell.value > DEFAULT_SLOW_THRESHOLD:
                    cell.fill = HIGHLIGHT_FILL  # 从常量导入的高亮填充色


def process_grouped_headers(ws, header_groups):
    """
    处理分组表头，将多级表头写入Excel工作表

    参数:
    ws - Excel工作表对象
    header_groups - 表头分组字典，格式为 {大类名称: [子类名称列表]}

    返回:
    header_end_row - 表头结束的行号
    """
    # 第一行为大分组名称
    current_col = 1
    for group_name, subheaders in header_groups.items():
        if not subheaders:  # 跳过空的子表头列表
            continue

        # 合并单元格用于大分组名称
        if len(subheaders) > 1:
            ws.merge_cells(start_row=1, start_column=current_col,
                           end_row=1, end_column=current_col + len(subheaders) - 1)

        # 写入大分组名称
        cell = ws.cell(row=1, column=current_col, value=group_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # 写入子表头
        for i, subheader in enumerate(subheaders):
            cell = ws.cell(row=2, column=current_col + i, value=subheader)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        current_col += len(subheaders)

    return 2  # 表头结束行为第2行

def add_dataframe_to_excel_with_grouped_headers(wb, df, sheet_name, header_groups=None):
    """
    将DataFrame添加到Excel工作簿，支持分组表头和普通表头。

    参数:
    wb - Excel工作簿对象
    df - 要保存的DataFrame数据
    sheet_name - 工作表名称
    header_groups - 表头分组字典，格式为 {大类名称: [子类名称列表]}，若为None则使用普通表头
    """
    log_info(f"添加工作表: {sheet_name} (行数: {len(df)})")
    ws = wb.create_sheet(title=sheet_name)

    if df.empty:
        log_info(f"警告: '{sheet_name}'工作表数据为空", level="WARNING")
        ws.cell(row=1, column=1, value="无数据")
        return

    has_grouped_header = bool(header_groups)
    header_end_row = 1  # 默认表头为单行

    if not has_grouped_header:
        # 普通表头
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    else:
        # 分组表头处理
        header_end_row = process_grouped_headers(ws, header_groups)

    # 数据从header_end_row后一行开始填充
    # 写入数据行(从表头结束后的下一行开始)
    start_row = header_end_row + 1
    row_count = 0

    for row_idx, row in enumerate(df.itertuples(), start=start_row):
        for col_idx, value in enumerate(row[1:], start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        row_count += 1
        if len(df) > 1000 and row_count % 1000 == 0:
            log_info(f"工作表'{sheet_name}'已写入 {row_count:,}/{len(df):,} 行 ({row_count / len(df) * 100:.1f}%)")

    format_excel_sheet(ws, has_grouped_header=has_grouped_header, header_end_row=header_end_row)
    return ws


def create_line_chart(ws, min_row, max_row, title, x_title, y_title, y_cols=None, series_names=None,
                     chart_position="H1", apply_peak_highlighting=False, peak_time=None):
    """
    创建折线图

    Args:
        ws: Excel工作表
        min_row: 起始行
        max_row: 结束行
        title: 图表标题
        x_title: X轴标题
        y_title: Y轴标题
        y_cols: Y轴数据列索引列表
        series_names: 系列名称列表
        chart_position: 图表位置
        apply_peak_highlighting: 是否突出显示峰值
        peak_time: 峰值时间
    """
    if max_row <= min_row:
        return

    # 限制数据点数量
    max_row = min(max_row, min_row + CHART_MAX_POINTS - 1)

    chart = LineChart()
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title

    # 获取X轴类别
    cats = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)

    if y_cols is None:
        y_cols = [2]

    for i, col_idx in enumerate(y_cols):
        data = Reference(ws, min_col=col_idx, max_col=col_idx,
                         min_row=min_row - 1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)

        if series_names and i < len(series_names):
            chart.series[i].title = SeriesLabel(v=series_names[i])
    chart.set_categories(cats)

    # 如果需要突出显示峰值区域
    if apply_peak_highlighting and peak_time:
        pass  # 这里可以添加峰值突出显示的功能

    # 将图表添加到工作表
    ws.add_chart(chart, chart_position)

def create_pie_chart(ws, title, data_start_row, data_end_row, labels_col=1, values_col=2, position="D1"):
    """
    创建饼图并插入到指定工作表

    Args:
        ws: openpyxl 工作表对象
        title: 图表标题
        data_start_row: 数据起始行（包含标签和数值）
        data_end_row: 数据结束行
        labels_col: 标签所在列（默认第1列）
        values_col: 数值所在列（默认第2列）
        position: 图表插入位置（默认 D1）
    """
    chart = PieChart()
    chart.title = title

    # 提取数值和标签区域
    data = Reference(ws, min_col=values_col, min_row=data_start_row, max_row=data_end_row)
    labels = Reference(ws, min_col=labels_col, min_row=data_start_row, max_row=data_end_row)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    # 插入图表
    ws.add_chart(chart, position)

