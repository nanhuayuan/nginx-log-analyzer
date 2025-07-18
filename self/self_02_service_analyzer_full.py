import gc
import numpy as np
import pandas as pd
from openpyxl import Workbook
from collections import defaultdict
from datetime import datetime

from self_00_04_excel_processor import add_dataframe_to_excel_with_grouped_headers, format_excel_sheet
from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD, PERCENTILES
from self_00_02_utils import log_info

# 核心指标配置（精简版）
CORE_TIME_METRICS = [
    'total_request_duration',  # 请求总时长
    'upstream_response_time',  # 后端响应时长
    'upstream_header_time',    # 后端处理时长
    'upstream_connect_time',   # 后端连接时长
]

CORE_SIZE_METRICS = [
    'response_body_size_kb',   # 响应体大小(KB)
    'total_bytes_sent_kb'      # 总发送字节(KB)
]

CORE_RATIO_METRICS = [
    'backend_efficiency',      # 后端处理效率(%)
    'processing_efficiency_index'  # 处理效率指数
]

# 中文名称映射
METRICS_MAPPING = {
    'total_request_duration': '请求总时长',
    'upstream_response_time': '后端响应时长',
    'upstream_header_time': '后端处理时长',
    'upstream_connect_time': '后端连接时长',
    'response_body_size_kb': '响应体大小',
    'total_bytes_sent_kb': '总发送字节',
    'backend_efficiency': '后端处理效率',
    'processing_efficiency_index': '处理效率指数'
}


class FullPrecisionServiceAnalyzer:
    """全精度服务分析器 - 使用全量数据计算精确分位数"""
    
    def __init__(self, slow_threshold=DEFAULT_SLOW_THRESHOLD):
        self.slow_threshold = slow_threshold
        self.service_stats = defaultdict(lambda: self._init_service_stats())
        self.app_stats = defaultdict(lambda: self._init_app_stats())
        self.global_stats = {
            'total_requests': 0,
            'success_requests': 0,
            'slow_requests': 0
        }
        # 全量数据存储（分批管理）
        self.full_data_buffers = defaultdict(lambda: defaultdict(list))
        self.batch_size = 1000000  # 100万条数据为一批
        
    def _init_service_stats(self):
        return {
            'service_name': '',
            'app_name': '',
            'total_requests': 0,
            'success_requests': 0,
            'slow_requests': 0,
            # 流式统计（用于平均值等）
            'request_time_sum': 0.0,
            'request_time_count': 0,
            'metrics_sum': defaultdict(float),
            'metrics_count': defaultdict(int)
        }
    
    def _init_app_stats(self):
        return {
            'app_name': '',
            'total_requests': 0,
            'success_requests': 0,
            'slow_requests': 0,
            'request_time_sum': 0.0,
            'request_time_count': 0,
            'metrics_sum': defaultdict(float),
            'metrics_count': defaultdict(int)
        }
    
    def process_chunk(self, chunk, success_codes):
        """处理数据块"""
        chunk_size = len(chunk)
        self.global_stats['total_requests'] += chunk_size
        
        # 处理总请求统计
        self._process_total_requests(chunk)
        
        # 处理成功请求统计
        successful_requests = chunk[chunk['response_status_code'].astype(str).isin(success_codes)]
        success_count = len(successful_requests)
        self.global_stats['success_requests'] += success_count
        
        if success_count > 0:
            self._process_successful_requests(successful_requests)
    
    def _process_total_requests(self, chunk):
        """处理总请求统计"""
        for service_name, group in chunk.groupby('service_name'):
            service_name = service_name if service_name and service_name != '' else 'unknown'
            self.service_stats[service_name]['total_requests'] += len(group)
            
            # 设置服务名称和应用名称
            if not self.service_stats[service_name]['service_name']:
                self.service_stats[service_name]['service_name'] = service_name
                app_name = group['application_name'].iloc[0] if not group['application_name'].empty else 'unknown'
                self.service_stats[service_name]['app_name'] = app_name
            
            # 应用统计
            app_name = group['application_name'].iloc[0] if not group['application_name'].empty else 'unknown'
            self.app_stats[app_name]['total_requests'] += len(group)
            if not self.app_stats[app_name]['app_name']:
                self.app_stats[app_name]['app_name'] = app_name
    
    def _process_successful_requests(self, successful_requests):
        """处理成功请求统计"""
        # 预处理数据
        processed_data = self._preprocess_data(successful_requests)
        
        # 服务级别处理
        for service_name, group_data in processed_data.groupby('service_name'):
            self._update_service_stats(service_name, group_data)
        
        # 应用级别处理
        for app_name, group_data in processed_data.groupby('app_name'):
            self._update_app_stats(app_name, group_data)
    
    def _preprocess_data(self, chunk):
        """预处理数据"""
        # 创建数据副本
        data = chunk.copy()
        
        # 填充缺失值
        data['service_name'] = data['service_name'].fillna('unknown')
        data['application_name'] = data['application_name'].fillna('unknown')
        
        # 数据类型转换
        numeric_columns = CORE_TIME_METRICS + CORE_SIZE_METRICS + CORE_RATIO_METRICS
        for col in numeric_columns:
            if col in data.columns:
                data[col] = pd.to_numeric(data[col], errors='coerce')
        
        return data
    
    def _update_service_stats(self, service_name, group_data):
        """更新服务统计"""
        stats = self.service_stats[service_name]
        group_size = len(group_data)
        
        stats['success_requests'] += group_size
        
        # 处理请求时间
        if 'total_request_duration' in group_data.columns:
            request_times = group_data['total_request_duration'].dropna()
            if len(request_times) > 0:
                stats['request_time_sum'] += request_times.sum()
                stats['request_time_count'] += len(request_times)
                
                # 慢请求统计
                slow_count = (request_times > self.slow_threshold).sum()
                stats['slow_requests'] += slow_count
                self.global_stats['slow_requests'] += slow_count
                
                # 存储全量数据用于分位数计算
                self.full_data_buffers[f'service_{service_name}']['total_request_duration'].extend(request_times.tolist())
        
        # 处理其他指标
        all_metrics = CORE_TIME_METRICS + CORE_SIZE_METRICS + CORE_RATIO_METRICS
        for metric in all_metrics:
            if metric in group_data.columns:
                values = group_data[metric].dropna()
                if len(values) > 0:
                    stats['metrics_sum'][metric] += values.sum()
                    stats['metrics_count'][metric] += len(values)
                    
                    # 存储全量数据
                    self.full_data_buffers[f'service_{service_name}'][metric].extend(values.tolist())
        
        # 批量处理检查
        self._check_batch_processing(f'service_{service_name}')
    
    def _update_app_stats(self, app_name, group_data):
        """更新应用统计"""
        stats = self.app_stats[app_name]
        group_size = len(group_data)
        
        stats['success_requests'] += group_size
        
        # 处理请求时间
        if 'total_request_duration' in group_data.columns:
            request_times = group_data['total_request_duration'].dropna()
            if len(request_times) > 0:
                stats['request_time_sum'] += request_times.sum()
                stats['request_time_count'] += len(request_times)
                
                # 慢请求统计
                slow_count = (request_times > self.slow_threshold).sum()
                stats['slow_requests'] += slow_count
                
                # 存储全量数据
                self.full_data_buffers[f'app_{app_name}']['total_request_duration'].extend(request_times.tolist())
        
        # 处理其他指标
        all_metrics = CORE_TIME_METRICS + CORE_SIZE_METRICS + CORE_RATIO_METRICS
        for metric in all_metrics:
            if metric in group_data.columns:
                values = group_data[metric].dropna()
                if len(values) > 0:
                    stats['metrics_sum'][metric] += values.sum()
                    stats['metrics_count'][metric] += len(values)
                    
                    # 存储全量数据
                    self.full_data_buffers[f'app_{app_name}'][metric].extend(values.tolist())
        
        # 批量处理检查
        self._check_batch_processing(f'app_{app_name}')
    
    def _check_batch_processing(self, entity_key):
        """检查是否需要批量处理（控制内存使用）"""
        total_size = sum(len(values) for values in self.full_data_buffers[entity_key].values())
        if total_size > self.batch_size:
            # 计算当前批次的分位数并清理内存
            self._process_batch_percentiles(entity_key)
    
    def _process_batch_percentiles(self, entity_key):
        """处理批次分位数计算"""
        # 这里可以实现增量分位数计算
        # 为了简化，当前版本在最终计算时处理
        pass
    
    def generate_final_results(self):
        """生成最终结果"""
        log_info("开始生成服务分析结果...", show_memory=True)
        
        # 生成服务结果
        service_results = self._generate_service_results()
        
        # 生成应用结果
        app_results = self._generate_app_results()
        
        # 清理内存
        self._cleanup_memory()
        
        return service_results, app_results
    
    def _generate_service_results(self):
        """生成服务结果"""
        results = []
        
        for service_name, stats in self.service_stats.items():
            if stats['success_requests'] == 0:
                continue
            
            result = self._build_service_result(service_name, stats)
            results.append(result)
        
        if not results:
            return pd.DataFrame()
        
        results_df = pd.DataFrame(results)
        
        # 排序
        if '平均请求总时长(秒)' in results_df.columns:
            results_df = results_df.sort_values(by='平均请求总时长(秒)', ascending=False)
        
        return results_df
    
    def _generate_app_results(self):
        """生成应用结果"""
        results = []
        
        for app_name, stats in self.app_stats.items():
            if stats['success_requests'] == 0:
                continue
            
            result = self._build_app_result(app_name, stats)
            results.append(result)
        
        if not results:
            return pd.DataFrame()
        
        results_df = pd.DataFrame(results)
        
        # 排序
        if '平均请求总时长(秒)' in results_df.columns:
            results_df = results_df.sort_values(by='平均请求总时长(秒)', ascending=False)
        
        return results_df
    
    def _build_service_result(self, service_name, stats):
        """构建服务结果"""
        result = {
            '服务名称': service_name,
            '应用名称': stats['app_name'],
            '接口请求总数': stats['total_requests'],
            '成功请求数': stats['success_requests'],
            '占总请求比例(%)': round(stats['success_requests'] / self.global_stats['success_requests'] * 100, 2) if self.global_stats['success_requests'] > 0 else 0,
            '成功率(%)': round(stats['success_requests'] / stats['total_requests'] * 100, 2) if stats['total_requests'] > 0 else 0,
            '慢请求数': stats['slow_requests'],
            '慢请求占比(%)': round(stats['slow_requests'] / stats['success_requests'] * 100, 2) if stats['success_requests'] > 0 else 0,
        }
        
        # 添加指标统计
        self._add_metrics_to_result(result, f'service_{service_name}', stats)
        
        return result
    
    def _build_app_result(self, app_name, stats):
        """构建应用结果"""
        result = {
            '应用名称': app_name,
            '接口请求总数': stats['total_requests'],
            '成功请求数': stats['success_requests'],
            '占总请求比例(%)': round(stats['success_requests'] / self.global_stats['success_requests'] * 100, 2) if self.global_stats['success_requests'] > 0 else 0,
            '成功率(%)': round(stats['success_requests'] / stats['total_requests'] * 100, 2) if stats['total_requests'] > 0 else 0,
            '慢请求数': stats['slow_requests'],
            '慢请求占比(%)': round(stats['slow_requests'] / stats['success_requests'] * 100, 2) if stats['success_requests'] > 0 else 0,
        }
        
        # 添加指标统计
        self._add_metrics_to_result(result, f'app_{app_name}', stats)
        
        return result
    
    def _add_metrics_to_result(self, result, entity_key, stats):
        """添加指标统计到结果"""
        all_metrics = CORE_TIME_METRICS + CORE_SIZE_METRICS + CORE_RATIO_METRICS
        
        for metric in all_metrics:
            display_name = METRICS_MAPPING.get(metric, metric)
            
            # 计算平均值
            if stats['metrics_count'][metric] > 0:
                avg_value = stats['metrics_sum'][metric] / stats['metrics_count'][metric]
            else:
                avg_value = 0
            
            # 获取单位
            if metric in CORE_TIME_METRICS:
                unit = '(秒)'
            elif metric in CORE_SIZE_METRICS:
                unit = '(KB)'
            else:
                unit = '(%)'
            
            result[f'平均{display_name}{unit}'] = round(avg_value, 3)
            
            # 计算分位数（全精度）
            full_data = self.full_data_buffers[entity_key][metric]
            if len(full_data) > 0:
                data_array = np.array(full_data)
                
                # 计算分位数
                result[f'中位数{display_name}{unit}'] = round(np.median(data_array), 3)
                result[f'P90{display_name}{unit}'] = round(np.percentile(data_array, 90), 3)
                result[f'P95{display_name}{unit}'] = round(np.percentile(data_array, 95), 3)
                result[f'P99{display_name}{unit}'] = round(np.percentile(data_array, 99), 3)
                
                # 数据质量指标
                result[f'{display_name}样本数'] = len(full_data)
                result[f'{display_name}计算精度'] = '精确'
            else:
                result[f'中位数{display_name}{unit}'] = 0
                result[f'P90{display_name}{unit}'] = 0
                result[f'P95{display_name}{unit}'] = 0
                result[f'P99{display_name}{unit}'] = 0
                result[f'{display_name}样本数'] = 0
                result[f'{display_name}计算精度'] = '无数据'
    
    def _cleanup_memory(self):
        """清理内存"""
        self.full_data_buffers.clear()
        gc.collect()
        log_info("内存清理完成", show_memory=True)


def analyze_service_performance_full(csv_path, output_path, success_codes=None, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """全精度服务性能分析主函数"""
    log_info(f"开始全精度服务性能分析: {csv_path}", show_memory=True)
    
    if success_codes is None:
        from self_00_01_constants import DEFAULT_SUCCESS_CODES
        success_codes = DEFAULT_SUCCESS_CODES
    success_codes = [str(code) for code in success_codes]
    
    # 创建分析器
    analyzer = FullPrecisionServiceAnalyzer(slow_threshold)
    
    # 处理参数
    chunk_size = max(DEFAULT_CHUNK_SIZE, 50000)
    chunks_processed = 0
    start_time = datetime.now()
    
    try:
        # 分块处理数据
        for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
            chunks_processed += 1
            
            # 处理数据块
            analyzer.process_chunk(chunk, success_codes)
            
            # 定期报告进度
            if chunks_processed % 10 == 0:
                elapsed = (datetime.now() - start_time).total_seconds()
                log_info(f"已处理 {chunks_processed} 个数据块, {analyzer.global_stats['total_requests']} 条记录, 耗时: {elapsed:.2f}秒", show_memory=True)
                gc.collect()
        
        # 生成最终结果
        service_results, app_results = analyzer.generate_final_results()
        
        # 创建Excel报告
        create_service_performance_excel_full(service_results, app_results, output_path, analyzer.global_stats)
        
        log_info(f"全精度服务性能分析完成，报告已生成: {output_path}", show_memory=True)
        return service_results.head(5) if not service_results.empty else pd.DataFrame()
        
    except Exception as e:
        log_info(f"全精度分析出错: {e}")
        raise


def create_service_performance_excel_full(service_results, app_results, output_path, global_stats):
    """创建全精度服务性能分析Excel报告"""
    log_info(f"开始创建全精度Excel报告: {output_path}", show_memory=True)
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 创建服务分析表
    if not service_results.empty:
        header_groups = create_service_header_groups()
        add_dataframe_to_excel_with_grouped_headers(wb, service_results, '服务性能分析', header_groups)
    
    # 创建应用分析表
    if not app_results.empty:
        app_header_groups = create_app_header_groups()
        add_dataframe_to_excel_with_grouped_headers(wb, app_results, '应用性能分析', app_header_groups)
    
    # 创建整体分析表
    create_overall_analysis_sheet_full(wb, service_results, app_results, global_stats)
    
    wb.save(output_path)
    log_info(f"全精度Excel报告已保存: {output_path}", show_memory=True)


def create_service_header_groups():
    """创建服务表头分组"""
    header_groups = {
        '基本信息': ['服务名称', '应用名称'],
        '请求统计': ['接口请求总数', '成功请求数', '占总请求比例(%)', '成功率(%)'],
        '慢请求统计': ['慢请求数', '慢请求占比(%)'],
    }
    
    # 为每个核心指标创建分组
    for metric in CORE_TIME_METRICS + CORE_SIZE_METRICS + CORE_RATIO_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        if metric in CORE_TIME_METRICS:
            unit = '(秒)'
        elif metric in CORE_SIZE_METRICS:
            unit = '(KB)'
        else:
            unit = '(%)'
        
        header_groups[f'{display_name}{unit}'] = ['平均', '中位数', 'P90', 'P95', 'P99', '样本数', '计算精度']
    
    return header_groups


def create_app_header_groups():
    """创建应用表头分组"""
    header_groups = {
        '基本信息': ['应用名称'],
        '请求统计': ['接口请求总数', '成功请求数', '占总请求比例(%)', '成功率(%)'],
        '慢请求统计': ['慢请求数', '慢请求占比(%)'],
    }
    
    # 为每个核心指标创建分组
    for metric in CORE_TIME_METRICS + CORE_SIZE_METRICS + CORE_RATIO_METRICS:
        display_name = METRICS_MAPPING.get(metric, metric)
        if metric in CORE_TIME_METRICS:
            unit = '(秒)'
        elif metric in CORE_SIZE_METRICS:
            unit = '(KB)'
        else:
            unit = '(%)'
        
        header_groups[f'{display_name}{unit}'] = ['平均', '中位数', 'P90', 'P95', 'P99', '样本数', '计算精度']
    
    return header_groups


def create_overall_analysis_sheet_full(wb, service_results, app_results, global_stats):
    """创建整体分析表"""
    ws = wb.create_sheet(title='整体分析概览')
    
    # 整体统计
    overall_stats = [
        ['=== 基础统计 ===', ''],
        ['总请求数', global_stats['total_requests']],
        ['成功请求数', global_stats['success_requests']],
        ['成功率(%)', round(global_stats['success_requests'] / global_stats['total_requests'] * 100, 2) if global_stats['total_requests'] > 0 else 0],
        ['慢请求数', global_stats['slow_requests']],
        ['慢请求占比(%)', round(global_stats['slow_requests'] / global_stats['success_requests'] * 100, 2) if global_stats['success_requests'] > 0 else 0],
        ['服务数量', len(service_results)],
        ['应用数量', len(app_results)],
        ['', ''],
        ['=== 分析精度 ===', ''],
        ['计算方法', '全量数据精确计算'],
        ['分位数精度', '100%准确'],
        ['样本覆盖', '完整数据集'],
    ]
    
    # 写入数据
    for row_idx, (label, value) in enumerate(overall_stats, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
    
    # 格式化工作表
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    format_excel_sheet(ws)


# 主函数别名，保持向后兼容
def analyze_service_performance(csv_path, output_path, success_codes=None, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """主分析函数 - 兼容原接口"""
    return analyze_service_performance_full(csv_path, output_path, success_codes, slow_threshold)