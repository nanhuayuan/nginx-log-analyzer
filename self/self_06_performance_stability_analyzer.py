import os
import gc
import numpy as np
import openpyxl
import pandas as pd
from datetime import datetime, timedelta

from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    add_dataframe_to_excel_with_grouped_headers,
    format_excel_sheet
)


def analyze_service_stability(csv_path, output_path, threshold=None):
    """分析服务稳定性指标"""
    log_info("开始分析服务稳定性指标...", show_memory=True)

    # 默认阈值配置
    if threshold is None:
        threshold = {
            'success_rate': 99.0,  # 成功率阈值
            'response_time': 0.5,  # 响应时间阈值（秒）
            'error_rate': 1.0,  # 错误率阈值
            'backend_efficiency': 60.0,  # 后端处理效率阈值
            'network_overhead': 30.0,  # 网络开销阈值
            'transfer_speed': 1000.0  # 传输速度阈值（KB/s）
        }

    # 初始化数据收集器
    data_collectors = {
        'success_rate': {},
        'response_time': {},
        'resource_usage': {},
        'request_frequency': {},
        'concurrency': [],
        'connection': {},
        'backend_performance': {},  # 新增：后端性能分析
        'transfer_performance': {},  # 新增：传输性能分析
        'nginx_lifecycle': {}  # 新增：Nginx生命周期分析
    }

    # 分块处理数据
    chunk_size = 100000
    total_records = process_data_chunks(csv_path, chunk_size, data_collectors)

    # 生成最终分析结果
    outputs = generate_analysis_results(data_collectors, threshold)

    # 保存到Excel
    save_performance_analysis_to_excel(outputs, output_path)

    log_info(f"服务稳定性分析完成，共处理 {total_records} 条记录", show_memory=True)
    return outputs


def process_data_chunks(csv_path, chunk_size, data_collectors):
    """分块处理数据文件"""
    log_info("正在分块读取数据文件...", show_memory=True)

    chunks_processed = 0
    total_records = 0
    start_time = datetime.now()

    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        chunks_processed += 1
        chunk_records = len(chunk)
        total_records += chunk_records

        # 预处理时间戳
        preprocess_timestamps(chunk)

        # 处理各类指标
        process_success_rate_chunk(chunk, data_collectors['success_rate'])
        process_response_time_chunk(chunk, data_collectors['response_time'])
        process_resource_usage_chunk(chunk, data_collectors['resource_usage'])
        process_request_frequency_chunk(chunk, data_collectors['request_frequency'])
        process_concurrency_chunk(chunk, data_collectors['concurrency'])
        process_connection_chunk(chunk, data_collectors['connection'])

        # 新增的性能分析
        process_backend_performance_chunk(chunk, data_collectors['backend_performance'])
        process_transfer_performance_chunk(chunk, data_collectors['transfer_performance'])
        process_nginx_lifecycle_chunk(chunk, data_collectors['nginx_lifecycle'])

        # 清理内存
        del chunk
        gc.collect()

        # 进度日志
        elapsed = (datetime.now() - start_time).total_seconds()
        log_info(f"已处理 {chunks_processed} 个数据块, {total_records} 条记录, 耗时: {elapsed:.2f}秒", show_memory=True)

    return total_records


def preprocess_timestamps(chunk):
    """预处理时间戳字段"""
    # 处理原始时间字段
    if 'raw_time' in chunk.columns:
        chunk['time'] = pd.to_datetime(chunk['raw_time'])
        chunk['hour_bucket'] = chunk['time'].dt.floor('H')
        chunk['minute_bucket'] = chunk['time'].dt.floor('min')

    # 处理到达时间字段
    if 'arrival_time' in chunk.columns:
        chunk['arrival_time'] = pd.to_datetime(chunk['arrival_time'], errors='coerce')
    elif 'arrival_timestamp' in chunk.columns:
        chunk['arrival_timestamp'] = pd.to_numeric(chunk['arrival_timestamp'], errors='coerce')
        chunk['arrival_time'] = pd.to_datetime(chunk['arrival_timestamp'], unit='s', errors='coerce')


def process_success_rate_chunk(chunk, success_rate_data):
    """处理成功率数据块"""
    for (hour, service), group in chunk.groupby(['hour_bucket', 'service_name']):
        # 统计2xx状态码为成功
        success_count = group['response_status_code'].astype(str).str.startswith('2').sum()
        total_count = len(group)

        if (hour, service) not in success_rate_data:
            success_rate_data[(hour, service)] = {'success': 0, 'total': 0}

        success_rate_data[(hour, service)]['success'] += success_count
        success_rate_data[(hour, service)]['total'] += total_count


def process_response_time_chunk(chunk, response_time_data):
    """处理响应时间数据块"""
    if 'total_request_duration' not in chunk.columns:
        return

    for (hour, service), group in chunk.groupby(['hour_bucket', 'service_name']):
        if (hour, service) not in response_time_data:
            response_time_data[(hour, service)] = {
                'sum': 0, 'count': 0, 'min': float('inf'), 'max': float('-inf'),
                'values': []  # 用于计算百分位数
            }

        data = response_time_data[(hour, service)]
        request_times = group['total_request_duration'].dropna()

        if len(request_times) > 0:
            data['sum'] += request_times.sum()
            data['count'] += len(request_times)
            data['min'] = min(data['min'], request_times.min())
            data['max'] = max(data['max'], request_times.max())
            # 为了节省内存，只保留部分值用于百分位数计算
            if len(data['values']) < 10000:
                data['values'].extend(request_times.tolist())


def process_resource_usage_chunk(chunk, resource_usage_data):
    """处理资源使用数据块 - 使用KB单位"""
    required_cols = ['response_body_size_kb', 'total_bytes_sent_kb', 'http_method']
    if not all(col in chunk.columns for col in required_cols):
        return

    for (service, method), group in chunk.groupby(['service_name', 'http_method']):
        if (service, method) not in resource_usage_data:
            resource_usage_data[(service, method)] = {
                'response_kb_sum': 0, 'total_kb_sum': 0, 'count': 0
            }

        data = resource_usage_data[(service, method)]
        data['response_kb_sum'] += group['response_body_size_kb'].sum()
        data['total_kb_sum'] += group['total_bytes_sent_kb'].sum()
        data['count'] += len(group)


def process_request_frequency_chunk(chunk, request_frequency_data):
    """处理请求频率数据块"""
    for (minute, service), count in chunk.groupby(['minute_bucket', 'service_name']).size().items():
        if service not in request_frequency_data:
            request_frequency_data[service] = {'counts': [], 'total': 0}

        request_frequency_data[service]['counts'].append(count)
        request_frequency_data[service]['total'] += count


def process_concurrency_chunk(chunk, concurrency_data):
    """处理并发数据块"""
    valid_requests = chunk.dropna(subset=['arrival_time', 'total_request_duration'])

    for _, row in valid_requests.iterrows():
        arrival_ts = row['arrival_time']
        duration = row['total_request_duration']
        end_ts = arrival_ts + pd.Timedelta(seconds=duration)

        if pd.notna(arrival_ts) and pd.notna(end_ts):
            concurrency_data.append((arrival_ts, 1))  # 请求开始
            concurrency_data.append((end_ts, -1))  # 请求结束


def process_connection_chunk(chunk, connection_data):
    """处理连接数据块"""
    if 'connection_cost_ratio' not in chunk.columns:
        return

    for minute, group in chunk.groupby('minute_bucket'):
        if minute not in connection_data:
            connection_data[minute] = {
                'request_count': 0,
                'connection_cost_sum': 0.0,
                'avg_connection_cost': 0.0
            }

        data = connection_data[minute]
        data['request_count'] += len(group)
        data['connection_cost_sum'] += group['connection_cost_ratio'].sum()


def process_backend_performance_chunk(chunk, backend_performance_data):
    """处理后端性能数据块 - 基于nginx生命周期分析"""
    required_cols = ['backend_efficiency', 'processing_efficiency_index',
                     'backend_connect_phase', 'backend_process_phase', 'backend_transfer_phase']

    if not all(col in chunk.columns for col in required_cols):
        return

    for (hour, service), group in chunk.groupby(['hour_bucket', 'service_name']):
        if (hour, service) not in backend_performance_data:
            backend_performance_data[(hour, service)] = {
                'efficiency_sum': 0, 'processing_index_sum': 0,
                'connect_time_sum': 0, 'process_time_sum': 0, 'transfer_time_sum': 0,
                'count': 0
            }

        data = backend_performance_data[(hour, service)]
        valid_group = group.dropna(subset=required_cols)

        if len(valid_group) > 0:
            data['efficiency_sum'] += valid_group['backend_efficiency'].sum()
            data['processing_index_sum'] += valid_group['processing_efficiency_index'].sum()
            data['connect_time_sum'] += valid_group['backend_connect_phase'].sum()
            data['process_time_sum'] += valid_group['backend_process_phase'].sum()
            data['transfer_time_sum'] += valid_group['backend_transfer_phase'].sum()
            data['count'] += len(valid_group)


def process_transfer_performance_chunk(chunk, transfer_performance_data):
    """处理传输性能数据块"""
    required_cols = ['response_transfer_speed', 'total_transfer_speed', 'nginx_transfer_speed']

    if not all(col in chunk.columns for col in required_cols):
        return

    for (hour, service), group in chunk.groupby(['hour_bucket', 'service_name']):
        if (hour, service) not in transfer_performance_data:
            transfer_performance_data[(hour, service)] = {
                'response_speed_sum': 0, 'total_speed_sum': 0, 'nginx_speed_sum': 0,
                'count': 0
            }

        data = transfer_performance_data[(hour, service)]
        valid_group = group.dropna(subset=required_cols)

        if len(valid_group) > 0:
            data['response_speed_sum'] += valid_group['response_transfer_speed'].sum()
            data['total_speed_sum'] += valid_group['total_transfer_speed'].sum()
            data['nginx_speed_sum'] += valid_group['nginx_transfer_speed'].sum()
            data['count'] += len(valid_group)


def process_nginx_lifecycle_chunk(chunk, nginx_lifecycle_data):
    """处理Nginx生命周期数据块"""
    required_cols = ['network_overhead', 'transfer_ratio', 'nginx_transfer_phase']

    if not all(col in chunk.columns for col in required_cols):
        return

    for (hour, service), group in chunk.groupby(['hour_bucket', 'service_name']):
        if (hour, service) not in nginx_lifecycle_data:
            nginx_lifecycle_data[(hour, service)] = {
                'network_overhead_sum': 0, 'transfer_ratio_sum': 0,
                'nginx_phase_sum': 0, 'count': 0
            }

        data = nginx_lifecycle_data[(hour, service)]
        valid_group = group.dropna(subset=required_cols)

        if len(valid_group) > 0:
            data['network_overhead_sum'] += valid_group['network_overhead'].sum()
            data['transfer_ratio_sum'] += valid_group['transfer_ratio'].sum()
            data['nginx_phase_sum'] += valid_group['nginx_transfer_phase'].sum()
            data['count'] += len(valid_group)


def generate_analysis_results(data_collectors, threshold):
    """生成最终分析结果"""
    log_info("正在计算最终分析结果...", show_memory=True)

    outputs = {}

    # 基础稳定性指标
    outputs['服务成功率稳定性'] = finalize_success_rate_analysis(
        data_collectors['success_rate'], threshold)
    outputs['服务响应时间稳定性'] = finalize_response_time_analysis(
        data_collectors['response_time'], threshold)
    outputs['资源使用和带宽'] = finalize_resource_usage_analysis(
        data_collectors['resource_usage'])
    outputs['服务请求频率'] = finalize_request_frequency_analysis(
        data_collectors['request_frequency'])

    # 高级分析指标
    if data_collectors['concurrency']:
        outputs['并发连接估算'] = finalize_concurrency_analysis(
            data_collectors['concurrency'])

    if data_collectors['connection']:
        connection_metrics, connection_summary = finalize_connections_analysis(
            data_collectors['connection'])
        outputs['连接性能指标'] = connection_metrics
        outputs['连接性能摘要'] = connection_summary

    # 新增的性能分析
    outputs['后端处理性能'] = finalize_backend_performance_analysis(
        data_collectors['backend_performance'], threshold)
    outputs['数据传输性能'] = finalize_transfer_performance_analysis(
        data_collectors['transfer_performance'], threshold)
    outputs['Nginx生命周期分析'] = finalize_nginx_lifecycle_analysis(
        data_collectors['nginx_lifecycle'], threshold)

    return outputs


def finalize_success_rate_analysis(success_rate_data, threshold):
    """完成成功率分析"""
    hourly_rates = []
    for (hour, service), data in success_rate_data.items():
        success_rate = (data['success'] / data['total'] * 100) if data['total'] > 0 else 0
        hourly_rates.append({
            '时间': hour,
            '服务名称': service,
            '成功率(%)': success_rate,
            '成功请求数': data['success'],
            '总请求数': data['total']
        })

    if not hourly_rates:
        return pd.DataFrame()

    hourly_df = pd.DataFrame(hourly_rates)

    # 按服务聚合统计
    stability_stats = []
    for service, service_data in hourly_df.groupby('服务名称'):
        mean_rate = service_data['成功率(%)'].mean()
        std_rate = service_data['成功率(%)'].std() if len(service_data) > 1 else 0
        min_rate = service_data['成功率(%)'].min()
        max_rate = service_data['成功率(%)'].max()

        # 异常判断
        status = '正常'
        if mean_rate < threshold['success_rate']:
            status = '成功率低'
        elif std_rate > 5.0:
            status = '波动较大'

        stability_stats.append({
            '服务名称': service,
            '平均成功率(%)': round(mean_rate, 2),
            '成功率波动(标准差)': round(std_rate, 2),
            '最低成功率(%)': round(min_rate, 2),
            '最高成功率(%)': round(max_rate, 2),
            '总请求数': service_data['总请求数'].sum(),
            '异常状态': status
        })

    del hourly_df
    gc.collect()

    return pd.DataFrame(stability_stats).sort_values('成功率波动(标准差)', ascending=False)


def finalize_response_time_analysis(response_time_data, threshold):
    """完成响应时间分析"""
    response_stats = []

    for (hour, service), data in response_time_data.items():
        if data['count'] > 0:
            mean_time = data['sum'] / data['count']

            # 计算百分位数（如果有足够数据）
            p95_time = p99_time = mean_time
            if data['values']:
                p95_time = np.percentile(data['values'], 95)
                p99_time = np.percentile(data['values'], 99)

            response_stats.append({
                '时间': hour,
                '服务名称': service,
                '平均响应时间(秒)': round(mean_time, 3),
                '最低响应时间(秒)': round(data['min'], 3),
                '最高响应时间(秒)': round(data['max'], 3),
                'P95响应时间(秒)': round(p95_time, 3),
                'P99响应时间(秒)': round(p99_time, 3),
                '请求数量': data['count']
            })

    if not response_stats:
        return pd.DataFrame()

    hourly_df = pd.DataFrame(response_stats)

    # 按服务聚合
    service_stats = []
    for service, service_data in hourly_df.groupby('服务名称'):
        mean_time = service_data['平均响应时间(秒)'].mean()
        std_time = service_data['平均响应时间(秒)'].std() if len(service_data) > 1 else 0
        min_time = service_data['最低响应时间(秒)'].min()
        max_time = service_data['最高响应时间(秒)'].max()

        # 异常判断
        status = '正常'
        if mean_time > threshold['response_time']:
            status = '响应时间长'
        elif std_time > mean_time * 0.5:  # 波动超过均值的50%
            status = '响应不稳定'

        service_stats.append({
            '服务名称': service,
            '平均响应时间(秒)': round(mean_time, 3),
            '响应时间波动(标准差)': round(std_time, 3),
            '最低响应时间(秒)': round(min_time, 3),
            '最高响应时间(秒)': round(max_time, 3),
            'P95响应时间(秒)': round(service_data['P95响应时间(秒)'].mean(), 3),
            'P99响应时间(秒)': round(service_data['P99响应时间(秒)'].mean(), 3),
            '总请求数': service_data['请求数量'].sum(),
            '异常状态': status
        })

    del hourly_df
    gc.collect()

    return pd.DataFrame(service_stats).sort_values('平均响应时间(秒)', ascending=False)


def finalize_resource_usage_analysis(resource_usage_data):
    """完成资源使用分析 - 基于KB单位"""
    resource_results = []

    for (service, method), data in resource_usage_data.items():
        avg_response_kb = data['response_kb_sum'] / data['count'] if data['count'] > 0 else 0
        avg_total_kb = data['total_kb_sum'] / data['count'] if data['count'] > 0 else 0
        total_response_mb = data['response_kb_sum'] / 1024  # 转换为MB
        total_transfer_mb = data['total_kb_sum'] / 1024  # 转换为MB

        resource_results.append({
            '服务名称': service,
            '请求方法': method,
            '平均响应大小(KB)': round(avg_response_kb, 2),
            '平均传输大小(KB)': round(avg_total_kb, 2),
            '总响应流量(MB)': round(total_response_mb, 2),
            '总传输流量(MB)': round(total_transfer_mb, 2),
            '请求次数': data['count']
        })

    return pd.DataFrame(resource_results).sort_values('总传输流量(MB)', ascending=False)


def finalize_request_frequency_analysis(request_frequency_data):
    """完成请求频率分析"""
    frequency_results = []

    for service, data in request_frequency_data.items():
        counts = data['counts']
        if counts:
            frequency_results.append({
                '服务名称': service,
                '平均每分钟请求数(QPS)': round(np.mean(counts), 2),
                '最大每分钟请求数(峰值QPS)': np.max(counts),
                '最小每分钟请求数': np.min(counts),
                '请求频率波动(标准差)': round(np.std(counts), 2),
                '总请求数': data['total']
            })

    return pd.DataFrame(frequency_results).sort_values('平均每分钟请求数(QPS)', ascending=False)


def finalize_concurrency_analysis(concurrency_data):
    """完成并发分析"""
    # 按时间排序并计算并发数
    concurrency_data.sort(key=lambda x: x[0])

    concurrent_requests = []
    current_count = 0

    for ts, event in concurrency_data:
        current_count += event
        concurrent_requests.append((ts, current_count))

    concurrent_df = pd.DataFrame(concurrent_requests, columns=['时间戳', '并发数'])
    concurrent_df['分钟时间段'] = concurrent_df['时间戳'].dt.floor('min')

    # 按分钟聚合统计
    concurrency_stats = concurrent_df.groupby('分钟时间段').agg(
        平均并发数=('并发数', 'mean'),
        最大并发数=('并发数', 'max'),
        最小并发数=('并发数', 'min')
    ).reset_index()

    concurrency_stats.rename(columns={'分钟时间段': '时间段'}, inplace=True)

    # 四舍五入
    concurrency_stats['平均并发数'] = concurrency_stats['平均并发数'].round(2)

    del concurrent_df, concurrent_requests
    gc.collect()

    return concurrency_stats


def finalize_connections_analysis(connection_data):
    """完成连接分析"""
    connection_metrics = []

    for minute, data in connection_data.items():
        avg_connection_cost = (data['connection_cost_sum'] / data['request_count']
                               if data['request_count'] > 0 else 0)

        connection_metrics.append({
            '时间': minute,
            '请求数量': data['request_count'],
            '平均连接成本比率': round(avg_connection_cost, 4),
            '总连接成本': round(data['connection_cost_sum'], 2)
        })

    connection_df = pd.DataFrame(connection_metrics)

    # 生成摘要统计
    connection_summary = {
        '平均每分钟请求数': round(connection_df['请求数量'].mean(), 2),
        '最大每分钟请求数': connection_df['请求数量'].max(),
        '平均连接成本比率': round(connection_df['平均连接成本比率'].mean(), 4),
        '最高连接成本比率': round(connection_df['平均连接成本比率'].max(), 4),
        '连接成本波动(标准差)': round(connection_df['平均连接成本比率'].std(), 4)
    }

    return connection_df, connection_summary


def finalize_backend_performance_analysis(backend_performance_data, threshold):
    """完成后端性能分析"""
    backend_stats = []

    for (hour, service), data in backend_performance_data.items():
        if data['count'] > 0:
            avg_efficiency = data['efficiency_sum'] / data['count']
            avg_processing_index = data['processing_index_sum'] / data['count']
            avg_connect_time = data['connect_time_sum'] / data['count']
            avg_process_time = data['process_time_sum'] / data['count']
            avg_transfer_time = data['transfer_time_sum'] / data['count']

            # 性能状态判断
            status = '正常'
            if avg_efficiency < threshold['backend_efficiency']:
                status = '处理效率低'
            elif avg_connect_time > 0.1:  # 连接时间超过100ms
                status = '连接延迟高'

            backend_stats.append({
                '时间': hour,
                '服务名称': service,
                '后端处理效率(%)': round(avg_efficiency, 2),
                '处理效率指数': round(avg_processing_index, 3),
                '平均连接时间(秒)': round(avg_connect_time, 3),
                '平均处理时间(秒)': round(avg_process_time, 3),
                '平均传输时间(秒)': round(avg_transfer_time, 3),
                '请求数量': data['count'],
                '性能状态': status
            })

    return pd.DataFrame(backend_stats).sort_values('后端处理效率(%)', ascending=True)


def finalize_transfer_performance_analysis(transfer_performance_data, threshold):
    """完成传输性能分析"""
    transfer_stats = []

    for (hour, service), data in transfer_performance_data.items():
        if data['count'] > 0:
            avg_response_speed = data['response_speed_sum'] / data['count']
            avg_total_speed = data['total_speed_sum'] / data['count']
            avg_nginx_speed = data['nginx_speed_sum'] / data['count']

            # 传输性能状态判断
            status = '正常'
            if avg_total_speed < threshold['transfer_speed']:
                status = '传输速度慢'
            elif avg_nginx_speed < avg_response_speed * 0.8:
                status = 'Nginx传输瓶颈'

            transfer_stats.append({
                '时间': hour,
                '服务名称': service,
                '响应传输速度(KB/s)': round(avg_response_speed, 2),
                '总传输速度(KB/s)': round(avg_total_speed, 2),
                'Nginx传输速度(KB/s)': round(avg_nginx_speed, 2),
                '请求数量': data['count'],
                '传输状态': status
            })

    return pd.DataFrame(transfer_stats).sort_values('总传输速度(KB/s)', ascending=True)


def finalize_nginx_lifecycle_analysis(nginx_lifecycle_data, threshold):
    """完成Nginx生命周期分析"""
    lifecycle_stats = []

    for (hour, service), data in nginx_lifecycle_data.items():
        if data['count'] > 0:
            avg_network_overhead = data['network_overhead_sum'] / data['count']
            avg_transfer_ratio = data['transfer_ratio_sum'] / data['count']
            avg_nginx_phase = data['nginx_phase_sum'] / data['count']

            # 生命周期状态判断
            status = '正常'
            if avg_network_overhead > threshold['network_overhead']:
                status = '网络开销高'
            elif avg_transfer_ratio > 60.0:  # 传输时间占比超过60%
                status = '传输时间占比高'

            lifecycle_stats.append({
                '时间': hour,
                '服务名称': service,
                '网络开销占比(%)': round(avg_network_overhead, 2),
                '传输时间占比(%)': round(avg_transfer_ratio, 2),
                '平均Nginx传输阶段(秒)': round(avg_nginx_phase, 3),
                '请求数量': data['count'],
                '生命周期状态': status
            })

    return pd.DataFrame(lifecycle_stats).sort_values('网络开销占比(%)', ascending=False)


def save_performance_analysis_to_excel(outputs, output_path):
    log_info(f"正在保存服务稳定性分析到Excel: {output_path}", show_memory=True)

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 定义工作表信息和高亮规则
    sheet_info = {
        '服务成功率稳定性': {
            'data': outputs.get('服务成功率稳定性'),
            'highlight_column': '异常状态',
            'highlight_values': {'成功率低': 'FF0000', '波动较大': 'FFA500'}
        },
        '服务响应时间稳定性': {
            'data': outputs.get('服务响应时间稳定性'),
            'highlight_column': '异常状态',
            'highlight_values': {'响应时间长': 'FF0000', '响应不稳定': 'FFA500'}
        },
        '资源使用和带宽': {
            'data': outputs.get('资源使用和带宽')
        },
        '服务请求频率': {
            'data': outputs.get('服务请求频率')
        },
        '后端处理性能': {
            'data': outputs.get('后端处理性能'),
            'highlight_column': '性能状态',
            'highlight_values': {'处理效率低': 'FF0000', '连接延迟高': 'FFA500'}
        },
        '数据传输性能': {
            'data': outputs.get('数据传输性能'),
            'highlight_column': '传输状态',
            'highlight_values': {'传输速度慢': 'FF0000', 'Nginx传输瓶颈': 'FFA500'}
        },
        'Nginx生命周期分析': {
            'data': outputs.get('Nginx生命周期分析'),
            'highlight_column': '生命周期状态',
            'highlight_values': {'网络开销高': 'FF0000', '传输时间占比高': 'FFA500'}
        }
    }

    # 添加并发连接估算（如果存在）
    if outputs.get('并发连接估算') is not None:
        sheet_info['并发连接估算'] = {'data': outputs['并发连接估算']}

    # 添加连接性能指标（如果存在）
    if outputs.get('连接性能指标') is not None:
        sheet_info['连接性能指标'] = {'data': outputs['连接性能指标']}

    # 创建各个工作表
    for sheet_name, info in sheet_info.items():
        if info['data'] is not None and not info['data'].empty:
            ws = add_dataframe_to_excel_with_grouped_headers(wb, info['data'], sheet_name)

            # 应用条件格式高亮
            if 'highlight_column' in info and 'highlight_values' in info:
                apply_highlighting(ws, info['data'], info['highlight_column'], info['highlight_values'])

            # 格式化工作表
            format_excel_sheet(ws)
            gc.collect()

    # 添加连接性能摘要（如果存在）
    if outputs.get('连接性能摘要') is not None:
        add_summary_sheet(wb, outputs['连接性能摘要'], '连接性能摘要')

    # 添加整体性能摘要
    add_overall_performance_summary(wb, outputs)

    wb.save(output_path)
    log_info(f"服务稳定性分析已保存到: {output_path}", show_memory=True)


def apply_highlighting(ws, df, highlight_column, highlight_values):
    """应用条件格式高亮显示"""
    from openpyxl.styles import Font, PatternFill

    if highlight_column not in df.columns:
        return

    col_idx = list(df.columns).index(highlight_column) + 1

    for r, row in enumerate(df.itertuples(index=False), start=2):
        try:
            # 处理列名中的特殊字符
            clean_column = highlight_column.replace(' ', '_').replace('(', '').replace(')', '').replace('%', '')
            cell_value = getattr(row, clean_column, None)

            if cell_value in highlight_values:
                cell = ws.cell(row=r, column=col_idx)
                # 使用背景色而不是字体色，更明显
                cell.fill = PatternFill(start_color=highlight_values[cell_value],
                                        end_color=highlight_values[cell_value],
                                        fill_type='solid')
                cell.font = Font(bold=True)
        except AttributeError:
            continue


def add_summary_sheet(wb, summary_data, sheet_name):
    """添加摘要工作表"""
    from openpyxl.styles import Font, Alignment

    ws = wb.create_sheet(title=sheet_name)

    # 设置标题行
    ws.cell(row=1, column=1, value='性能指标').font = Font(bold=True, size=12)
    ws.cell(row=1, column=2, value='数值').font = Font(bold=True, size=12)

    # 填充数据
    for r, (metric, value) in enumerate(summary_data.items(), start=2):
        ws.cell(row=r, column=1, value=metric)
        ws.cell(row=r, column=2, value=value)

        # 设置对齐方式
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='right')

    # 调整列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    format_excel_sheet(ws)


def add_overall_performance_summary(wb, outputs):
    """添加整体性能摘要工作表"""
    from openpyxl.styles import Font, Alignment, PatternFill

    ws = wb.create_sheet(title='整体性能摘要')

    # 标题
    ws.cell(row=1, column=1, value='Nginx服务性能分析摘要').font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    current_row = 3

    # 1. 服务成功率摘要
    if outputs.get('服务成功率稳定性') is not None and not outputs['服务成功率稳定性'].empty:
        success_df = outputs['服务成功率稳定性']
        ws.cell(row=current_row, column=1, value='📊 服务成功率概览').font = Font(bold=True, size=12)
        current_row += 1

        avg_success_rate = success_df['平均成功率(%)'].mean()
        min_success_rate = success_df['平均成功率(%)'].min()
        max_volatility = success_df['成功率波动(标准差)'].max()
        total_requests = success_df['总请求数'].sum()

        summary_data = [
            ('平均成功率', f"{avg_success_rate:.2f}%"),
            ('最低成功率', f"{min_success_rate:.2f}%"),
            ('最大波动性', f"{max_volatility:.2f}%"),
            ('总请求数', f"{total_requests:,}")
        ]

        for metric, value in summary_data:
            ws.cell(row=current_row, column=2, value=metric)
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1
        current_row += 1

    # 2. 响应时间摘要
    if outputs.get('服务响应时间稳定性') is not None and not outputs['服务响应时间稳定性'].empty:
        response_df = outputs['服务响应时间稳定性']
        ws.cell(row=current_row, column=1, value='⏱️ 响应时间概览').font = Font(bold=True, size=12)
        current_row += 1

        avg_response_time = response_df['平均响应时间(秒)'].mean()
        max_response_time = response_df['最高响应时间(秒)'].max()
        avg_p95 = response_df['P95响应时间(秒)'].mean()
        avg_p99 = response_df['P99响应时间(秒)'].mean()

        summary_data = [
            ('平均响应时间', f"{avg_response_time:.3f}秒"),
            ('最高响应时间', f"{max_response_time:.3f}秒"),
            ('平均P95响应时间', f"{avg_p95:.3f}秒"),
            ('平均P99响应时间', f"{avg_p99:.3f}秒")
        ]

        for metric, value in summary_data:
            ws.cell(row=current_row, column=2, value=metric)
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1
        current_row += 1

    # 3. 后端性能摘要
    if outputs.get('后端处理性能') is not None and not outputs['后端处理性能'].empty:
        backend_df = outputs['后端处理性能']
        ws.cell(row=current_row, column=1, value='🔧 后端处理性能概览').font = Font(bold=True, size=12)
        current_row += 1

        avg_efficiency = backend_df['后端处理效率(%)'].mean()
        avg_connect_time = backend_df['平均连接时间(秒)'].mean()
        avg_process_time = backend_df['平均处理时间(秒)'].mean()
        avg_transfer_time = backend_df['平均传输时间(秒)'].mean()

        summary_data = [
            ('平均后端效率', f"{avg_efficiency:.2f}%"),
            ('平均连接时间', f"{avg_connect_time:.3f}秒"),
            ('平均处理时间', f"{avg_process_time:.3f}秒"),
            ('平均传输时间', f"{avg_transfer_time:.3f}秒")
        ]

        for metric, value in summary_data:
            ws.cell(row=current_row, column=2, value=metric)
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1
        current_row += 1

    # 4. 数据传输性能摘要
    if outputs.get('数据传输性能') is not None and not outputs['数据传输性能'].empty:
        transfer_df = outputs['数据传输性能']
        ws.cell(row=current_row, column=1, value='📡 数据传输性能概览').font = Font(bold=True, size=12)
        current_row += 1

        avg_response_speed = transfer_df['响应传输速度(KB/s)'].mean()
        avg_total_speed = transfer_df['总传输速度(KB/s)'].mean()
        avg_nginx_speed = transfer_df['Nginx传输速度(KB/s)'].mean()

        summary_data = [
            ('平均响应传输速度', f"{avg_response_speed:.2f} KB/s"),
            ('平均总传输速度', f"{avg_total_speed:.2f} KB/s"),
            ('平均Nginx传输速度', f"{avg_nginx_speed:.2f} KB/s")
        ]

        for metric, value in summary_data:
            ws.cell(row=current_row, column=2, value=metric)
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1
        current_row += 1

    # 5. Nginx生命周期摘要
    if outputs.get('Nginx生命周期分析') is not None and not outputs['Nginx生命周期分析'].empty:
        lifecycle_df = outputs['Nginx生命周期分析']
        ws.cell(row=current_row, column=1, value='🔄 Nginx生命周期概览').font = Font(bold=True, size=12)
        current_row += 1

        avg_network_overhead = lifecycle_df['网络开销占比(%)'].mean()
        avg_transfer_ratio = lifecycle_df['传输时间占比(%)'].mean()
        avg_nginx_phase = lifecycle_df['平均Nginx传输阶段(秒)'].mean()

        summary_data = [
            ('平均网络开销占比', f"{avg_network_overhead:.2f}%"),
            ('平均传输时间占比', f"{avg_transfer_ratio:.2f}%"),
            ('平均Nginx传输阶段', f"{avg_nginx_phase:.3f}秒")
        ]

        for metric, value in summary_data:
            ws.cell(row=current_row, column=2, value=metric)
            ws.cell(row=current_row, column=3, value=value)
            current_row += 1

    # 设置列宽和样式
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    # 设置对齐方式
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    format_excel_sheet(ws)