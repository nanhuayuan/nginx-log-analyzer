import gc
import numpy as np
import pandas as pd
# ipaddress模块在Python 3.3+中可用
try:
    import ipaddress
    IPADDRESS_AVAILABLE = True
except ImportError:
    IPADDRESS_AVAILABLE = False
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
import math

from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD
from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    format_excel_sheet,
    add_dataframe_to_excel_with_grouped_headers,
    create_pie_chart,
    create_line_chart
)
from self_00_05_sampling_algorithms import (
    TDigest, HyperLogLog, ReservoirSampler, StratifiedSampler
)


class AdvancedIPAnalyzer:
    """高级IP分析器 - 使用流式算法优化内存使用"""
    
    def __init__(self):
        # IP统计数据结构 - 使用流式算法
        self.ip_stats = {}
        self.total_processed = 0
        
        # 全局时间分布统计
        self.global_hourly_distribution = defaultdict(int)
        
        # 配置参数
        self.max_sample_size = 1000  # 限制样本大小
        self.compression = 100       # T-Digest压缩参数
        self.hll_precision = 12      # HyperLogLog精度
        
    def _init_ip_stats(self, ip):
        """初始化单个IP的统计结构"""
        return {
            # 基础计数
            'total_requests': 0,
            'success_requests': 0,
            'error_requests': 0,
            'slow_requests': 0,
            
            # 流式统计算法
            'response_time_digest': TDigest(compression=self.compression),
            'data_size_digest': TDigest(compression=self.compression),
            'unique_apis_hll': HyperLogLog(precision=self.hll_precision),
            'user_agents_sampler': ReservoirSampler(self.max_sample_size),
            
            # 累计统计（内存可控）
            'total_response_time': 0.0,
            'total_data_size': 0.0,
            'status_codes': defaultdict(int),
            'hourly_distribution': defaultdict(int),
            
            # 限制大小的集合
            'sample_request_times': [],
            'sample_apis': set()
        }
    
    def analyze_ip_sources(self, csv_path, output_path, top_n=100):
        """分析来源IP，包括请求分布、地理位置、异常检测等 - 优化版"""
        log_info("🚀 开始高级IP分析（内存优化版）...", show_memory=True)
        
        chunk_size = max(DEFAULT_CHUNK_SIZE // 2, 10000)
        
        # 第一遍：收集IP统计数据
        log_info("📊 第一遍扫描：收集IP统计数据")
        for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
            self._process_chunk(chunk)
            
            if self.total_processed % 100000 == 0:
                gc.collect()
                log_info(f"已处理 {self.total_processed:,} 条记录，发现 {len(self.ip_stats)} 个唯一IP")
        
        total_unique_ips = len(self.ip_stats)
        log_info(f"✅ IP统计完成：总记录 {self.total_processed:,}，唯一IP {total_unique_ips:,}")
        
        if total_unique_ips == 0:
            log_info("⚠️ 未找到有效的IP数据", level="WARNING")
            return pd.DataFrame()
        
        # 生成高级IP分析报告
        ip_analysis_results = self._generate_advanced_ip_analysis_report(top_n)
        
        # 创建高级Excel报告
        self._create_advanced_ip_analysis_excel(ip_analysis_results, output_path)
        
        log_info(f"🎉 高级IP分析完成，报告已生成：{output_path}", show_memory=True)
        return ip_analysis_results.head(10)
    
    def _process_chunk(self, chunk):
        """处理数据块"""
        chunk_size_actual = len(chunk)
        self.total_processed += chunk_size_actual
        
        # 处理必要的列
        if 'client_ip_address' not in chunk.columns:
            log_info("⚠️ 未找到client_ip_address列，跳过IP分析", level="WARNING")
            return
            
        # 数据类型转换和清洗
        chunk = self._clean_chunk_data(chunk)
        
        # 按IP分组处理
        for ip, group in chunk.groupby('client_ip_address'):
            if pd.isna(ip) or ip == '' or ip == 'unknown':
                continue
            
            # 初始化IP统计（如果不存在）
            if ip not in self.ip_stats:
                self.ip_stats[ip] = self._init_ip_stats(ip)
            
            self._process_ip_group(ip, group)
    
    def _clean_chunk_data(self, chunk):
        """清洗数据块"""
        # 数据类型转换
        if 'total_request_duration' in chunk.columns:
            chunk['total_request_duration'] = pd.to_numeric(chunk['total_request_duration'], errors='coerce')
        
        # 状态码处理 - 更严格的清理
        if 'response_status_code' in chunk.columns:
            # 先转换为字符串，然后清理
            chunk['response_status_code'] = chunk['response_status_code'].astype(str).str.strip()
            # 过滤掉无效的状态码
            chunk['response_status_code'] = chunk['response_status_code'].replace({'nan': None, '': None, '-': None})
        
        if 'response_body_size_kb' in chunk.columns:
            chunk['response_body_size_kb'] = pd.to_numeric(chunk['response_body_size_kb'], errors='coerce')
        
        return chunk
    
    def _process_ip_group(self, ip, group):
        """处理单个IP的分组数据"""
        stats = self.ip_stats[ip]
        group_size = len(group)
        
        # 基础统计
        stats['total_requests'] += group_size
        
        # 成功和错误请求统计
        if 'response_status_code' in group.columns:
            self._process_status_codes(stats, group)
        
        # 响应时间统计 - 使用T-Digest流式算法
        if 'total_request_duration' in group.columns:
            self._process_response_times(stats, group)
        
        # 数据大小统计 - 使用T-Digest
        if 'response_body_size_kb' in group.columns:
            self._process_data_sizes(stats, group)
        
        # API统计 - 使用HyperLogLog
        if 'request_full_uri' in group.columns:
            self._process_apis(stats, group)
        
        # 时间分布统计
        if 'hour' in group.columns:
            self._process_time_distribution(stats, group)
        
        # User Agent统计 - 使用蓄水池采样
        if 'user_agent_string' in group.columns:
            self._process_user_agents(stats, group)
    
    def _process_status_codes(self, stats, group):
        """处理状态码统计"""
        # 过滤掉空值和无效状态码
        valid_status_codes = group['response_status_code'].dropna()
        if valid_status_codes.empty:
            return
            
        status_counts = valid_status_codes.value_counts()
        for status, count in status_counts.items():
            # 确保状态码是字符串格式并清理
            status_str = str(status).strip()
            
            # 跳过无效状态码
            if status_str in ['None', 'nan', '', '-'] or len(status_str) < 3:
                continue
                
            stats['status_codes'][status_str] += count
            
            # 更严格的状态码判断
            if status_str.startswith('2') or status_str.startswith('3'):
                stats['success_requests'] += count
            elif status_str.startswith('4') or status_str.startswith('5'):
                stats['error_requests'] += count
            # 如果状态码不是标准格式（如1xx），记录但不计入成功/错误
    
    def _process_response_times(self, stats, group):
        """处理响应时间统计 - 流式算法"""
        durations = group['total_request_duration'].dropna()
        if durations.empty:
            return
            
        # 累计统计
        stats['total_response_time'] += durations.sum()
        slow_count = (durations > DEFAULT_SLOW_THRESHOLD).sum()
        stats['slow_requests'] += slow_count
        
        # T-Digest流式分位数计算
        for duration in durations:
            if not math.isinf(duration) and not math.isnan(duration):
                stats['response_time_digest'].add(float(duration))
        
        # 保持少量样本用于详细分析（内存可控）
        if len(stats['sample_request_times']) < self.max_sample_size:
            sample_size = min(self.max_sample_size - len(stats['sample_request_times']), len(durations))
            if sample_size > 0:
                sample = durations.sample(sample_size, random_state=42) if len(durations) > sample_size else durations
                stats['sample_request_times'].extend(sample.tolist())
    
    def _process_data_sizes(self, stats, group):
        """处理数据大小统计 - 流式算法"""
        sizes = group['response_body_size_kb'].dropna()
        if sizes.empty:
            return
            
        # 累计统计
        stats['total_data_size'] += sizes.sum()
        
        # T-Digest流式统计
        for size in sizes:
            if not math.isinf(size) and not math.isnan(size) and size >= 0:
                stats['data_size_digest'].add(float(size))
    
    def _process_apis(self, stats, group):
        """处理API统计 - HyperLogLog"""
        apis = group['request_full_uri'].dropna().unique()
        
        # HyperLogLog流式唯一计数
        for api in apis:
            stats['unique_apis_hll'].add(str(api))
        
        # 保持少量样本用于展示
        if len(stats['sample_apis']) < 50:
            remaining_slots = 50 - len(stats['sample_apis'])
            sample_apis = apis[:remaining_slots] if len(apis) > remaining_slots else apis
            stats['sample_apis'].update(sample_apis)
    
    def _process_time_distribution(self, stats, group):
        """处理时间分布统计"""
        hour_counts = group['hour'].value_counts()
        for hour, count in hour_counts.items():
            if pd.notna(hour):
                hour_int = int(hour)
                stats['hourly_distribution'][hour_int] += count
                self.global_hourly_distribution[hour_int] += count
    
    def _process_user_agents(self, stats, group):
        """处理User Agent统计 - 蓄水池采样"""
        agents = group['user_agent_string'].dropna().unique()
        for agent in agents:
            stats['user_agents_sampler'].add({'user_agent': str(agent)})
    
    def _generate_advanced_ip_analysis_report(self, top_n):
        """生成高级IP分析报告"""
        log_info("📋 生成高级IP分析报告...")
        
        results = []
        for ip, stats in self.ip_stats.items():
            result = self._calculate_ip_metrics(ip, stats)
            results.append(result)
        
        # 转换为DataFrame并排序
        df = pd.DataFrame(results)
        df = df.sort_values(by='总请求数', ascending=False).head(top_n)
        
        log_info(f"✅ 生成了 {len(df)} 个IP的高级分析报告")
        return df
    
    def _calculate_ip_metrics(self, ip, stats):
        """计算单个IP的指标"""
        total_requests = stats['total_requests']
        success_requests = stats['success_requests']
        error_requests = stats['error_requests']
        slow_requests = stats['slow_requests']
        
        # 计算比率
        success_rate = (success_requests / total_requests * 100) if total_requests > 0 else 0
        error_rate = (error_requests / total_requests * 100) if total_requests > 0 else 0
        slow_rate = (slow_requests / total_requests * 100) if total_requests > 0 else 0
        
        # 平均响应时间
        avg_response_time = (stats['total_response_time'] / success_requests) if success_requests > 0 else 0
        
        # 高级分位数计算 - 使用T-Digest
        response_time_digest = stats['response_time_digest']
        median_time = response_time_digest.percentile(50) if response_time_digest.count > 0 else 0
        p95_time = response_time_digest.percentile(95) if response_time_digest.count > 0 else 0
        p99_time = response_time_digest.percentile(99) if response_time_digest.count > 0 else 0
        
        # 数据传输统计
        data_size_digest = stats['data_size_digest']
        avg_data_size = (stats['total_data_size'] / total_requests) if total_requests > 0 else 0
        median_data_size = data_size_digest.percentile(50) if data_size_digest.count > 0 else 0
        p95_data_size = data_size_digest.percentile(95) if data_size_digest.count > 0 else 0
        
        # 唯一API数量 - HyperLogLog估计
        unique_api_count = stats['unique_apis_hll'].cardinality()
        
        # IP类型分析
        ip_type = self._classify_ip_type(ip)
        
        # 高级风险评分
        risk_score, risk_factors = self._calculate_advanced_risk_score(stats, total_requests, error_rate, slow_rate, unique_api_count)
        
        # 异常检测评分
        anomaly_score, anomaly_level = self._calculate_anomaly_score(stats, total_requests, error_rate, slow_rate)
        
        # 最常见的状态码和时段
        most_common_status = max(stats['status_codes'].items(), key=lambda x: x[1])[0] if stats['status_codes'] else 'N/A'
        peak_hour = max(stats['hourly_distribution'].items(), key=lambda x: x[1])[0] if stats['hourly_distribution'] else 'N/A'
        
        # User Agent采样数量
        user_agent_count = len(stats['user_agents_sampler'].get_samples())
        
        # 行为模式分析
        behavior_pattern = self._analyze_behavior_pattern(stats, total_requests, unique_api_count, error_rate)
        
        return {
            'IP地址': ip,
            'IP类型': ip_type,
            '总请求数': total_requests,
            '成功请求数': success_requests,
            '错误请求数': error_requests,
            '慢请求数': slow_requests,
            '成功率(%)': round(success_rate, 2),
            '错误率(%)': round(error_rate, 2),
            '慢请求率(%)': round(slow_rate, 2),
            '平均响应时间(秒)': round(avg_response_time, 3),
            '响应时间中位数(秒)': round(median_time, 3),
            'P95响应时间(秒)': round(p95_time, 3),
            'P99响应时间(秒)': round(p99_time, 3),
            '平均数据传输(KB)': round(avg_data_size, 2),
            '数据传输中位数(KB)': round(median_data_size, 2),
            'P95数据传输(KB)': round(p95_data_size, 2),
            '总数据传输(MB)': round(stats['total_data_size'] / 1024, 2),
            '唯一API数(估计)': int(unique_api_count),
            '最常见状态码': most_common_status,
            '活跃时段': f"{peak_hour}:00" if peak_hour != 'N/A' else 'N/A',
            '风险评分': risk_score,
            '风险因子': '; '.join(risk_factors) if risk_factors else '无',
            '异常评分': anomaly_score,
            '异常等级': anomaly_level,
            '行为模式': behavior_pattern,
            'User Agent采样数': user_agent_count
        }
    
    def _classify_ip_type(self, ip_str):
        """分类IP类型"""
        if not IPADDRESS_AVAILABLE:
            # 当ipaddress模块不可用时的简单分类
            if ip_str.startswith(('192.168.', '10.', '172.')):
                return "内网IP"
            elif ip_str.startswith('127.'):
                return "回环IP"
            elif ip_str.startswith(('224.', '225.', '226.', '227.', '228.', '229.', '230.', '231.', '232.', '233.', '234.', '235.', '236.', '237.', '238.', '239.')):
                return "组播IP"
            else:
                return "公网IP"
        
        try:
            ip = ipaddress.ip_address(ip_str)
            if ip.is_private:
                return "内网IP"
            elif ip.is_loopback:
                return "回环IP"
            elif ip.is_multicast:
                return "组播IP"
            elif ip.is_reserved:
                return "保留IP"
            else:
                return "公网IP"
        except ValueError:
            return "无效IP"
    
    def _calculate_advanced_risk_score(self, stats, total_requests, error_rate, slow_rate, unique_api_count):
        """计算高级风险评分（0-100，分数越高风险越大）"""
        risk_score = 0
        risk_factors = []
        
        # 基于请求量的风险（大量请求可能是攻击）
        if total_requests > 50000:
            risk_score += 40
            risk_factors.append('超高请求量')
        elif total_requests > 10000:
            risk_score += 30
            risk_factors.append('高请求量')
        elif total_requests > 1000:
            risk_score += 15
            risk_factors.append('中等请求量')
        elif total_requests > 100:
            risk_score += 5
        
        # 基于错误率的风险
        if error_rate > 50:
            risk_score += 25
            risk_factors.append('极高错误率')
        elif error_rate > 20:
            risk_score += 15
            risk_factors.append('高错误率')
        elif error_rate > 10:
            risk_score += 10
            risk_factors.append('中等错误率')
        
        # 基于慢请求率的风险
        if slow_rate > 30:
            risk_score += 20
            risk_factors.append('极高慢请求率')
        elif slow_rate > 10:
            risk_score += 10
            risk_factors.append('高慢请求率')
        
        # 基于API多样性的风险（访问过多不同API可能是扫描）
        if unique_api_count > 100:
            risk_score += 20
            risk_factors.append('API扫描行为')
        elif unique_api_count > 50:
            risk_score += 15
            risk_factors.append('高API多样性')
        elif unique_api_count > 20:
            risk_score += 10
            risk_factors.append('中等API多样性')
        
        # 基于4xx状态码比例
        status_4xx_count = sum(count for status, count in stats['status_codes'].items() if status.startswith('4'))
        if total_requests > 0:
            status_4xx_rate = status_4xx_count / total_requests * 100
            if status_4xx_rate > 30:
                risk_score += 15
                risk_factors.append('高4xx错误率')
            elif status_4xx_rate > 10:
                risk_score += 10
                risk_factors.append('中等4xx错误率')
        
        # 基于时间分布的风险（非正常时间大量访问）
        if stats['hourly_distribution']:
            night_requests = sum(stats['hourly_distribution'].get(hour, 0) for hour in [0, 1, 2, 3, 4, 5])
            night_ratio = night_requests / total_requests * 100 if total_requests > 0 else 0
            if night_ratio > 50:
                risk_score += 15
                risk_factors.append('深夜异常活跃')
        
        return min(risk_score, 100), risk_factors
    
    def _calculate_anomaly_score(self, stats, total_requests, error_rate, slow_rate):
        """计算异常检测评分"""
        anomaly_score = 0
        
        # 请求量异常
        if total_requests > 20000:
            anomaly_score += 30
        
        # 成功率异常
        success_rate = ((stats['success_requests'] / total_requests * 100) if total_requests > 0 else 0)
        if success_rate < 50:
            anomaly_score += 40
        
        # 响应时间异常
        if stats['response_time_digest'].count > 0:
            p99_time = stats['response_time_digest'].percentile(99)
            if p99_time > 10:  # P99超过10秒
                anomaly_score += 30
        
        # 错误率异常
        if error_rate > 20:
            anomaly_score += 25
        
        # 异常等级分类
        if anomaly_score >= 80:
            anomaly_level = "严重异常"
        elif anomaly_score >= 60:
            anomaly_level = "中度异常"
        elif anomaly_score >= 40:
            anomaly_level = "轻微异常"
        else:
            anomaly_level = "正常"
            
        return anomaly_score, anomaly_level
    
    def _analyze_behavior_pattern(self, stats, total_requests, unique_api_count, error_rate):
        """分析行为模式"""
        if unique_api_count > 50 and total_requests > 1000:
            if error_rate > 30:
                return "疑似恶意扫描"
            else:
                return "深度访问用户"
        elif total_requests > 5000 and unique_api_count < 5:
            return "高频单一访问"
        elif error_rate > 50:
            return "异常访问"
        elif total_requests < 10:
            return "轻度访问"
        else:
            return "正常访问"
    
    def _create_advanced_ip_analysis_excel(self, ip_df, output_path):
        """创建高级IP分析Excel报告"""
        log_info(f"📊 创建高级IP分析Excel报告: {output_path}")
        
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # 主要IP统计表 - 增强版
        header_groups = {
            "基础信息": ["IP地址", "IP类型", "行为模式"],
            "请求统计": ["总请求数", "成功请求数", "错误请求数", "慢请求数"],
            "性能比率": ["成功率(%)", "错误率(%)", "慢请求率(%)"],
            "响应时间分析": ["平均响应时间(秒)", "响应时间中位数(秒)", "P95响应时间(秒)", "P99响应时间(秒)"],
            "数据传输分析": ["平均数据传输(KB)", "数据传输中位数(KB)", "P95数据传输(KB)", "总数据传输(MB)"],
            "风险评估": ["风险评分", "风险因子", "异常评分", "异常等级"],
            "其他指标": ["唯一API数(估计)", "最常见状态码", "活跃时段", "User Agent采样数"]
        }
        
        ws_main = add_dataframe_to_excel_with_grouped_headers(
            wb, ip_df, '高级IP分析统计', header_groups=header_groups
        )
        
        # 高风险IP工作表 - 增强版
        self._create_advanced_high_risk_ip_sheet(wb, ip_df)
        
        # IP类型分布工作表 - 增强版
        self._create_advanced_ip_type_distribution_sheet(wb, ip_df)
        
        # 行为模式分析工作表
        self._create_behavior_pattern_analysis_sheet(wb, ip_df)
        
        # 时间分布分析工作表 - 增强版
        self._create_advanced_time_distribution_sheet(wb)
        
        # 异常检测工作表
        self._create_anomaly_detection_sheet(wb, ip_df)
        
        # 概览工作表 - 增强版
        self._create_advanced_ip_overview_sheet(wb, ip_df)
        
        # 保存文件
        wb.save(output_path)
        log_info(f"✅ 高级IP分析Excel报告已保存: {output_path}")
    
    def _create_advanced_high_risk_ip_sheet(self, wb, ip_df):
        """创建高级高风险IP工作表"""
        ws = wb.create_sheet(title='高风险IP分析')
        
        # 筛选高风险IP（风险评分 > 50或异常评分 > 60）
        high_risk_ips = ip_df[
            (ip_df['风险评分'] > 50) | (ip_df['异常评分'] > 60)
        ].sort_values(by=['风险评分', '异常评分'], ascending=False)
        
        if high_risk_ips.empty:
            ws.cell(row=1, column=1, value="🎉 未发现高风险IP").font = Font(bold=True)
            return
        
        # 高风险IP表头
        high_risk_headers = {
            "基础信息": ["IP地址", "IP类型", "行为模式"],
            "风险指标": ["风险评分", "异常评分", "异常等级"],
            "关键统计": ["总请求数", "错误率(%)", "慢请求率(%)", "唯一API数(估计)"],
            "详细信息": ["风险因子", "最常见状态码", "活跃时段", "总数据传输(MB)"]
        }
        
        risk_columns = [
            "IP地址", "IP类型", "行为模式", "风险评分", "异常评分", "异常等级", 
            "总请求数", "错误率(%)", "慢请求率(%)", "唯一API数(估计)", 
            "风险因子", "最常见状态码", "活跃时段", "总数据传输(MB)"
        ]
        risk_df = high_risk_ips[risk_columns].copy()
        
        # 由于工作表已创建，需要先删除再重新创建
        wb.remove(ws)
        ws = add_dataframe_to_excel_with_grouped_headers(
            wb, risk_df, '高风险IP分析', header_groups=high_risk_headers
        )
        
        # 添加风险分析说明
        note_row = len(risk_df) + 5
        ws.cell(row=note_row, column=1, value="🔍 风险评分说明：").font = Font(bold=True)
        ws.cell(row=note_row + 1, column=1, value="• 风险评分 70-100: 高风险，需要立即关注")
        ws.cell(row=note_row + 2, column=1, value="• 风险评分 50-70: 中等风险，建议监控")
        ws.cell(row=note_row + 3, column=1, value="• 异常评分 80+: 严重异常")
        ws.cell(row=note_row + 4, column=1, value="• 异常评分 60-79: 中度异常")
        
        format_excel_sheet(ws)
    
    def _create_advanced_ip_type_distribution_sheet(self, wb, ip_df):
        """创建高级IP类型分布工作表"""
        ws = wb.create_sheet(title='IP类型分布分析')
        
        # IP类型统计 - 增强版
        ip_type_stats = ip_df.groupby('IP类型').agg({
            'IP地址': 'count',
            '总请求数': ['sum', 'mean'],
            '成功率(%)': 'mean',
            '错误率(%)': 'mean',
            '风险评分': 'mean',
            '异常评分': 'mean',
            '唯一API数(估计)': 'mean'
        }).round(2)
        
        # 展平列名
        ip_type_stats.columns = [
            'IP数量', '总请求数', '平均每IP请求数', '平均成功率(%)', 
            '平均错误率(%)', '平均风险评分', '平均异常评分', '平均API数'
        ]
        ip_type_stats = ip_type_stats.reset_index()
        
        # 添加到工作表
        type_headers = {
            "分类": ["IP类型"],
            "数量统计": ["IP数量", "总请求数", "平均每IP请求数"],
            "性能指标": ["平均成功率(%)", "平均错误率(%)"],
            "风险指标": ["平均风险评分", "平均异常评分", "平均API数"]
        }
        
        ws = add_dataframe_to_excel_with_grouped_headers(
            wb, ip_type_stats, 'IP类型分布分析', header_groups=type_headers
        )
        
        format_excel_sheet(ws)
    
    def _create_behavior_pattern_analysis_sheet(self, wb, ip_df):
        """创建行为模式分析工作表"""
        ws = wb.create_sheet(title='行为模式分析')
        
        # 行为模式统计
        behavior_stats = ip_df.groupby('行为模式').agg({
            'IP地址': 'count',
            '总请求数': ['sum', 'mean'],
            '风险评分': 'mean',
            '异常评分': 'mean',
            '错误率(%)': 'mean'
        }).round(2)
        
        behavior_stats.columns = ['IP数量', '总请求数', '平均每IP请求数', '平均风险评分', '平均异常评分', '平均错误率(%)']
        behavior_stats = behavior_stats.reset_index()
        behavior_stats = behavior_stats.sort_values(by='平均风险评分', ascending=False)
        
        # 添加到工作表
        behavior_headers = {
            "模式": ["行为模式"],
            "数量统计": ["IP数量", "总请求数", "平均每IP请求数"],
            "风险评估": ["平均风险评分", "平均异常评分", "平均错误率(%)"]
        }
        
        ws = add_dataframe_to_excel_with_grouped_headers(
            wb, behavior_stats, '行为模式分析', header_groups=behavior_headers
        )
        
        format_excel_sheet(ws)
    
    def _create_advanced_time_distribution_sheet(self, wb):
        """创建高级时间分布分析工作表"""
        ws = wb.create_sheet(title='时间分布分析')
        
        if not self.global_hourly_distribution:
            ws.cell(row=1, column=1, value="⚠️ 无时间分布数据").font = Font(bold=True)
            return
        
        # 创建小时分布数据 - 增强版
        hours = list(range(24))
        total_requests = sum(self.global_hourly_distribution.values())
        
        time_data = []
        for hour in hours:
            requests = self.global_hourly_distribution.get(hour, 0)
            percentage = round(requests / total_requests * 100, 2) if total_requests > 0 else 0
            
            # 分析时段特征
            if 6 <= hour <= 12:
                period = "上午"
            elif 13 <= hour <= 18:
                period = "下午"
            elif 19 <= hour <= 23:
                period = "晚上"
            else:
                period = "深夜"
            
            time_data.append({
                '小时': f"{hour:02d}:00",
                '时段': period,
                '请求数': requests,
                '占比(%)': percentage,
                '活跃度': '高' if percentage > 6 else ('中' if percentage > 3 else '低')
            })
        
        time_df = pd.DataFrame(time_data)
        
        # 添加到工作表
        time_headers = {
            "时间": ["小时", "时段"],
            "统计": ["请求数", "占比(%)", "活跃度"]
        }
        
        ws = add_dataframe_to_excel_with_grouped_headers(
            wb, time_df, '时间分布分析', header_groups=time_headers
        )
        
        format_excel_sheet(ws)
    
    def _create_anomaly_detection_sheet(self, wb, ip_df):
        """创建异常检测工作表"""
        ws = wb.create_sheet(title='异常检测分析')
        
        # 筛选异常IP
        anomaly_ips = ip_df[ip_df['异常等级'] != '正常'].sort_values(by='异常评分', ascending=False)
        
        if anomaly_ips.empty:
            ws.cell(row=1, column=1, value="🎉 未检测到异常IP").font = Font(bold=True)
            return
        
        # 异常IP表头
        anomaly_headers = {
            "基础信息": ["IP地址", "IP类型", "行为模式"],
            "异常指标": ["异常评分", "异常等级", "风险评分"],
            "性能指标": ["总请求数", "成功率(%)", "P99响应时间(秒)", "唯一API数(估计)"]
        }
        
        anomaly_columns = [
            "IP地址", "IP类型", "行为模式", "异常评分", "异常等级", "风险评分",
            "总请求数", "成功率(%)", "P99响应时间(秒)", "唯一API数(估计)"
        ]
        anomaly_df = anomaly_ips[anomaly_columns].copy()
        
        ws = add_dataframe_to_excel_with_grouped_headers(
            wb, anomaly_df, '异常检测分析', header_groups=anomaly_headers
        )
        
        format_excel_sheet(ws)
    
    def _create_advanced_ip_overview_sheet(self, wb, ip_df):
        """创建高级IP分析概览工作表"""
        ws = wb.create_sheet(title='分析概览')
        
        # 移动到第一个位置
        wb.move_sheet(ws, -(len(wb.worksheets) - 1))
        
        # 总体统计
        total_unique_ips = len(ip_df)
        total_requests = ip_df['总请求数'].sum()
        avg_requests_per_ip = total_requests / total_unique_ips if total_unique_ips > 0 else 0
        
        # 风险统计
        high_risk_count = len(ip_df[ip_df['风险评分'] > 70])
        medium_risk_count = len(ip_df[(ip_df['风险评分'] > 50) & (ip_df['风险评分'] <= 70)])
        low_risk_count = len(ip_df[ip_df['风险评分'] <= 50])
        
        # 异常统计
        severe_anomaly = len(ip_df[ip_df['异常等级'] == '严重异常'])
        moderate_anomaly = len(ip_df[ip_df['异常等级'] == '中度异常'])
        mild_anomaly = len(ip_df[ip_df['异常等级'] == '轻微异常'])
        normal_count = len(ip_df[ip_df['异常等级'] == '正常'])
        
        # IP类型统计
        ip_type_counts = ip_df['IP类型'].value_counts()
        
        # 行为模式统计
        behavior_counts = ip_df['行为模式'].value_counts()
        
        # 性能统计
        avg_success_rate = ip_df['成功率(%)'].mean()
        avg_error_rate = ip_df['错误率(%)'].mean()
        avg_response_time = ip_df['平均响应时间(秒)'].mean()
        
        # 概览数据
        overview_data = [
            ['🚀 === 高级IP分析概览 ===', ''],
            ['', ''],
            
            ['📊 === 基础统计 ===', ''],
            ['总处理记录数', self.total_processed],
            ['唯一IP数量', total_unique_ips],
            ['总请求数', total_requests],
            ['平均每IP请求数', round(avg_requests_per_ip, 2)],
            ['', ''],
            
            ['🔍 === 风险分布 ===', ''],
            ['高风险IP数量 (>70)', high_risk_count],
            ['中等风险IP数量 (50-70)', medium_risk_count],
            ['低风险IP数量 (≤50)', low_risk_count],
            ['', ''],
            
            ['⚠️ === 异常检测 ===', ''],
            ['严重异常IP', severe_anomaly],
            ['中度异常IP', moderate_anomaly],
            ['轻微异常IP', mild_anomaly],
            ['正常IP', normal_count],
            ['', ''],
            
            ['🌐 === IP类型分布 ===', ''],
        ]
        
        # 添加IP类型统计
        for ip_type, count in ip_type_counts.items():
            overview_data.append([f'{ip_type}数量', count])
        
        overview_data.extend([
            ['', ''],
            ['👤 === 行为模式分布 ===', ''],
        ])
        
        # 添加行为模式统计
        for behavior, count in behavior_counts.items():
            overview_data.append([f'{behavior}数量', count])
        
        overview_data.extend([
            ['', ''],
            ['📈 === 性能统计 ===', ''],
            ['平均成功率(%)', round(avg_success_rate, 2)],
            ['平均错误率(%)', round(avg_error_rate, 2)],
            ['平均响应时间(秒)', round(avg_response_time, 3)],
            ['', ''],
            
            ['🏆 === TOP指标 ===', ''],
            ['请求量最大IP', ip_df.iloc[0]['IP地址'] if not ip_df.empty else 'N/A'],
            ['最大请求量', ip_df.iloc[0]['总请求数'] if not ip_df.empty else 0],
            ['最高风险评分IP', ip_df.loc[ip_df['风险评分'].idxmax(), 'IP地址'] if not ip_df.empty else 'N/A'],
            ['最高风险评分', ip_df['风险评分'].max() if not ip_df.empty else 0],
            ['', ''],
            
            ['🔧 === 优化说明 ===', ''],
            ['算法优化', 'T-Digest + HyperLogLog + 蓄水池采样'],
            ['内存优化', '流式算法，支持40G+数据'],
            ['分析增强', '多维风险评分 + 异常检测 + 行为分析'],
        ])
        
        # 写入数据
        for row_idx, (label, value) in enumerate(overview_data, start=1):
            cell_label = ws.cell(row=row_idx, column=1, value=label)
            cell_value = ws.cell(row=row_idx, column=2, value=value)
            
            if label.startswith('===') and label.endswith('==='):
                cell_label.font = Font(bold=True, size=12)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        
        format_excel_sheet(ws)


# 向后兼容的函数接口
def analyze_ip_sources(csv_path, output_path, top_n=100):
    """分析来源IP - 兼容接口，使用高级分析器"""
    analyzer = AdvancedIPAnalyzer()
    return analyzer.analyze_ip_sources(csv_path, output_path, top_n)


if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("使用方法: python self_08_ip_analyzer_advanced.py <csv_path> <output_path>")
        sys.exit(1)
    
    csv_path = sys.argv[1]
    output_path = sys.argv[2]
    
    log_info("🚀 启动高级IP分析器...")
    result = analyze_ip_sources(csv_path, output_path, top_n=100)
    if not result.empty:
        log_info("✅ 高级IP分析完成！")
        print(result.head())
    else:
        log_info("❌ 分析失败或无有效数据")