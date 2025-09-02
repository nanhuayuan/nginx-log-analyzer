import gc
import numpy as np
import pandas as pd
# ipaddress模块在Python 3.3+中可用
try:
    import ipaddress
    IPADDRESS_AVAILABLE = True
except ImportError:
    IPADDRESS_AVAILABLE = False
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font

from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD
from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    format_excel_sheet,
    add_dataframe_to_excel_with_grouped_headers,
    create_pie_chart,
    create_line_chart
)


def analyze_ip_sources(csv_path, output_path, top_n=100):
    """分析来源IP，包括请求分布、地理位置、异常检测等"""
    log_info("开始分析来源IP数据...", show_memory=True)
    
    chunk_size = max(DEFAULT_CHUNK_SIZE // 2, 10000)
    ip_stats = defaultdict(lambda: {
        'total_requests': 0,
        'success_requests': 0,
        'error_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'total_data_size': 0.0,
        'unique_apis': set(),
        'status_codes': defaultdict(int),
        'request_times': [],
        'hourly_distribution': defaultdict(int),
        'user_agents': set()
    })
    
    total_processed = 0
    total_unique_ips = 0
    
    # 第一遍：收集IP统计数据
    log_info("第一遍扫描：收集IP统计数据")
    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        chunk_size_actual = len(chunk)
        total_processed += chunk_size_actual
        
        # 处理必要的列
        if 'client_ip_address' not in chunk.columns:
            log_info("未找到client_ip_address列，跳过IP分析", level="WARNING")
            continue
            
        # 数据类型转换
        if 'total_request_duration' in chunk.columns:
            chunk['total_request_duration'] = pd.to_numeric(chunk['total_request_duration'], errors='coerce')
        if 'response_status_code' in chunk.columns:
            chunk['response_status_code'] = chunk['response_status_code'].astype(str)
        if 'response_body_size_kb' in chunk.columns:
            chunk['response_body_size_kb'] = pd.to_numeric(chunk['response_body_size_kb'], errors='coerce')
            
        # 按IP分组处理
        for ip, group in chunk.groupby('client_ip_address'):
            if pd.isna(ip) or ip == '' or ip == 'unknown':
                continue
                
            stats = ip_stats[ip]
            group_size = len(group)
            
            # 基础统计
            stats['total_requests'] += group_size
            
            # 成功和错误请求统计
            if 'response_status_code' in group.columns:
                status_counts = group['response_status_code'].value_counts()
                for status, count in status_counts.items():
                    stats['status_codes'][status] += count
                    if status.startswith('2') or status.startswith('3'):
                        stats['success_requests'] += count
                    elif status.startswith('4') or status.startswith('5'):
                        stats['error_requests'] += count
            
            # 慢请求统计
            if 'total_request_duration' in group.columns:
                durations = group['total_request_duration'].dropna()
                stats['total_response_time'] += durations.sum()
                slow_count = (durations > DEFAULT_SLOW_THRESHOLD).sum()
                stats['slow_requests'] += slow_count
                
                # 保存少量样本用于分析（避免内存问题）
                if len(stats['request_times']) < 1000:
                    sample_size = min(1000 - len(stats['request_times']), len(durations))
                    if sample_size > 0:
                        sample = durations.sample(sample_size, random_state=42) if len(durations) > sample_size else durations
                        stats['request_times'].extend(sample.tolist())
            
            # 数据大小统计
            if 'response_body_size_kb' in group.columns:
                sizes = group['response_body_size_kb'].dropna()
                stats['total_data_size'] += sizes.sum()
            
            # API统计
            if 'request_full_uri' in group.columns:
                apis = group['request_full_uri'].dropna().unique()
                stats['unique_apis'].update(apis)
            
            # 时间分布统计
            if 'hour' in group.columns:
                hour_counts = group['hour'].value_counts()
                for hour, count in hour_counts.items():
                    if pd.notna(hour):
                        stats['hourly_distribution'][int(hour)] += count
            
            # User Agent统计（取样）
            if 'user_agent_string' in group.columns:
                if len(stats['user_agents']) < 10:
                    agents = group['user_agent_string'].dropna().unique()
                    stats['user_agents'].update(agents[:10 - len(stats['user_agents'])])
        
        if total_processed % 100000 == 0:
            gc.collect()
            log_info(f"已处理 {total_processed:,} 条记录，发现 {len(ip_stats)} 个唯一IP")
    
    total_unique_ips = len(ip_stats)
    log_info(f"IP统计完成：总记录 {total_processed:,}，唯一IP {total_unique_ips:,}")
    
    if total_unique_ips == 0:
        log_info("未找到有效的IP数据", level="WARNING")
        return pd.DataFrame()
    
    # 生成IP分析报告
    ip_analysis_results = generate_ip_analysis_report(ip_stats, top_n)
    
    # 创建Excel报告
    create_ip_analysis_excel(ip_analysis_results, output_path, ip_stats, total_processed)
    
    log_info(f"IP分析完成，报告已生成：{output_path}", show_memory=True)
    return ip_analysis_results.head(10)


def generate_ip_analysis_report(ip_stats, top_n):
    """生成IP分析报告"""
    log_info("生成IP分析报告...")
    
    results = []
    for ip, stats in ip_stats.items():
        # 基础统计
        total_requests = stats['total_requests']
        success_requests = stats['success_requests']
        error_requests = stats['error_requests']
        slow_requests = stats['slow_requests']
        
        # 计算比率
        success_rate = (success_requests / total_requests * 100) if total_requests > 0 else 0
        error_rate = (error_requests / total_requests * 100) if total_requests > 0 else 0
        slow_rate = (slow_requests / total_requests * 100) if total_requests > 0 else 0
        
        # 平均响应时间
        avg_response_time = (stats['total_response_time'] / success_requests) if success_requests > 0 else 0
        
        # 请求时间分析
        request_times = stats['request_times']
        if request_times:
            median_time = np.median(request_times)
            p95_time = np.percentile(request_times, 95)
            p99_time = np.percentile(request_times, 99)
        else:
            median_time = p95_time = p99_time = 0
        
        # 平均数据传输量
        avg_data_size = (stats['total_data_size'] / total_requests) if total_requests > 0 else 0
        
        # 唯一API数量
        unique_api_count = len(stats['unique_apis'])
        
        # IP类型分析
        ip_type = classify_ip_type(ip)
        
        # 风险评分（简单评分系统）
        risk_score = calculate_risk_score(stats, total_requests, error_rate, slow_rate)
        
        # 最常见的状态码
        most_common_status = max(stats['status_codes'].items(), key=lambda x: x[1])[0] if stats['status_codes'] else 'N/A'
        
        # 活跃时段
        peak_hour = max(stats['hourly_distribution'].items(), key=lambda x: x[1])[0] if stats['hourly_distribution'] else 'N/A'
        
        result = {
            'IP地址': ip,
            'IP类型': ip_type,
            '总请求数': total_requests,
            '成功请求数': success_requests,
            '错误请求数': error_requests,
            '慢请求数': slow_requests,
            '成功率(%)': round(success_rate, 2),
            '错误率(%)': round(error_rate, 2),
            '慢请求率(%)': round(slow_rate, 2),
            '平均响应时间(秒)': round(avg_response_time, 3),
            '响应时间中位数(秒)': round(median_time, 3),
            'P95响应时间(秒)': round(p95_time, 3),
            'P99响应时间(秒)': round(p99_time, 3),
            '平均数据传输(KB)': round(avg_data_size, 2),
            '总数据传输(MB)': round(stats['total_data_size'] / 1024, 2),
            '唯一API数': unique_api_count,
            '最常见状态码': most_common_status,
            '活跃时段': f"{peak_hour}:00" if peak_hour != 'N/A' else 'N/A',
            '风险评分': risk_score,
            'User Agent数': len(stats['user_agents'])
        }
        
        results.append(result)
    
    # 转换为DataFrame并排序
    df = pd.DataFrame(results)
    df = df.sort_values(by='总请求数', ascending=False).head(top_n)
    
    log_info(f"生成了 {len(df)} 个IP的分析报告")
    return df


def classify_ip_type(ip_str):
    """分类IP类型"""
    if not IPADDRESS_AVAILABLE:
        # 当ipaddress模块不可用时的简单分类
        if ip_str.startswith(('192.168.', '10.', '172.')):
            return "内网IP"
        elif ip_str.startswith('127.'):
            return "回环IP"
        elif ip_str.startswith(('224.', '225.', '226.', '227.', '228.', '229.', '230.', '231.', '232.', '233.', '234.', '235.', '236.', '237.', '238.', '239.')):
            return "组播IP"
        else:
            return "公网IP"
    
    try:
        ip = ipaddress.ip_address(ip_str)
        if ip.is_private:
            return "内网IP"
        elif ip.is_loopback:
            return "回环IP"
        elif ip.is_multicast:
            return "组播IP"
        elif ip.is_reserved:
            return "保留IP"
        else:
            return "公网IP"
    except ValueError:
        return "无效IP"


def calculate_risk_score(stats, total_requests, error_rate, slow_rate):
    """计算风险评分（0-100，分数越高风险越大）"""
    risk_score = 0
    
    # 基于请求量的风险（大量请求可能是攻击）
    if total_requests > 10000:
        risk_score += 30
    elif total_requests > 1000:
        risk_score += 15
    elif total_requests > 100:
        risk_score += 5
    
    # 基于错误率的风险
    if error_rate > 50:
        risk_score += 25
    elif error_rate > 20:
        risk_score += 15
    elif error_rate > 10:
        risk_score += 10
    
    # 基于慢请求率的风险
    if slow_rate > 30:
        risk_score += 20
    elif slow_rate > 10:
        risk_score += 10
    
    # 基于API多样性的风险（访问过多不同API可能是扫描）
    unique_apis = len(stats['unique_apis'])
    if unique_apis > 50:
        risk_score += 15
    elif unique_apis > 20:
        risk_score += 10
    
    # 基于4xx状态码比例
    status_4xx_count = sum(count for status, count in stats['status_codes'].items() if status.startswith('4'))
    if total_requests > 0:
        status_4xx_rate = status_4xx_count / total_requests * 100
        if status_4xx_rate > 30:
            risk_score += 10
    
    return min(risk_score, 100)


def create_ip_analysis_excel(ip_df, output_path, ip_stats, total_processed):
    """创建IP分析Excel报告"""
    log_info(f"创建IP分析Excel报告: {output_path}")
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 主要IP统计表
    header_groups = {
        "基础信息": ["IP地址", "IP类型"],
        "请求统计": ["总请求数", "成功请求数", "错误请求数", "慢请求数"],
        "性能指标": ["成功率(%)", "错误率(%)", "慢请求率(%)", "风险评分"],
        "响应时间": ["平均响应时间(秒)", "响应时间中位数(秒)", "P95响应时间(秒)", "P99响应时间(秒)"],
        "数据传输": ["平均数据传输(KB)", "总数据传输(MB)"],
        "其他指标": ["唯一API数", "最常见状态码", "活跃时段", "User Agent数"]
    }
    
    ws_main = add_dataframe_to_excel_with_grouped_headers(
        wb, ip_df, 'IP分析统计', header_groups=header_groups
    )
    
    # 高风险IP工作表
    create_high_risk_ip_sheet(wb, ip_df)
    
    # IP类型分布工作表
    create_ip_type_distribution_sheet(wb, ip_df)
    
    # 时间分布分析工作表
    create_time_distribution_sheet(wb, ip_stats)
    
    # 概览工作表
    create_ip_overview_sheet(wb, ip_df, total_processed)
    
    # 保存文件
    wb.save(output_path)
    log_info(f"IP分析Excel报告已保存: {output_path}")


def create_high_risk_ip_sheet(wb, ip_df):
    """创建高风险IP工作表"""
    ws = wb.create_sheet(title='高风险IP')
    
    # 筛选高风险IP（风险评分 > 50）
    high_risk_ips = ip_df[ip_df['风险评分'] > 50].sort_values(by='风险评分', ascending=False)
    
    if high_risk_ips.empty:
        ws.cell(row=1, column=1, value="未发现高风险IP").font = Font(bold=True)
        return
    
    # 高风险IP表头
    high_risk_headers = {
        "基础信息": ["IP地址", "IP类型", "风险评分"],
        "异常指标": ["总请求数", "错误率(%)", "慢请求率(%)", "唯一API数"],
        "详细信息": ["最常见状态码", "活跃时段", "总数据传输(MB)"]
    }
    
    risk_columns = ["IP地址", "IP类型", "风险评分", "总请求数", "错误率(%)", "慢请求率(%)", "唯一API数", "最常见状态码", "活跃时段", "总数据传输(MB)"]
    risk_df = high_risk_ips[risk_columns].copy()
    
    # 由于工作表已创建，需要先删除再重新创建
    wb.remove(ws)
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, risk_df, '高风险IP', header_groups=high_risk_headers
    )
    
    # 添加风险说明
    note_row = len(risk_df) + 5
    ws.cell(row=note_row, column=1, value="风险评分说明：").font = Font(bold=True)
    ws.cell(row=note_row + 1, column=1, value="• 70-100: 高风险，需要立即关注")
    ws.cell(row=note_row + 2, column=1, value="• 50-70: 中等风险，建议监控")
    ws.cell(row=note_row + 3, column=1, value="• 0-50: 低风险")
    
    format_excel_sheet(ws)


def create_ip_type_distribution_sheet(wb, ip_df):
    """创建IP类型分布工作表"""
    ws = wb.create_sheet(title='IP类型分布')
    
    # IP类型统计
    ip_type_stats = ip_df.groupby('IP类型').agg({
        'IP地址': 'count',
        '总请求数': 'sum',
        '成功率(%)': 'mean',
        '错误率(%)': 'mean',
        '风险评分': 'mean'
    }).round(2)
    
    ip_type_stats.columns = ['IP数量', '总请求数', '平均成功率(%)', '平均错误率(%)', '平均风险评分']
    ip_type_stats = ip_type_stats.reset_index()
    
    # 添加到工作表
    type_headers = {
        "分类": ["IP类型"],
        "数量统计": ["IP数量", "总请求数"],
        "性能指标": ["平均成功率(%)", "平均错误率(%)", "平均风险评分"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, ip_type_stats, 'IP类型分布', header_groups=type_headers
    )
    
    # 添加饼图
    if len(ip_type_stats) > 1:
        chart_start_row = len(ip_type_stats) + 5
        ws.cell(row=chart_start_row, column=1, value="IP类型分布图").font = Font(bold=True)
        
        for i, row in ip_type_stats.iterrows():
            ws.cell(row=chart_start_row + 1 + i, column=1, value=row['IP类型'])
            ws.cell(row=chart_start_row + 1 + i, column=2, value=row['IP数量'])
        
        create_pie_chart(
            ws, "IP类型分布",
            data_start_row=chart_start_row + 1,
            data_end_row=chart_start_row + len(ip_type_stats),
            labels_col=1, values_col=2,
            position="D" + str(chart_start_row)
        )
    
    format_excel_sheet(ws)


def create_time_distribution_sheet(wb, ip_stats):
    """创建时间分布分析工作表"""
    ws = wb.create_sheet(title='时间分布分析')
    
    # 聚合所有IP的小时分布
    hourly_total = defaultdict(int)
    for stats in ip_stats.values():
        for hour, count in stats['hourly_distribution'].items():
            hourly_total[hour] += count
    
    if not hourly_total:
        ws.cell(row=1, column=1, value="无时间分布数据").font = Font(bold=True)
        return
    
    # 创建小时分布数据
    hours = list(range(24))
    requests = [hourly_total.get(hour, 0) for hour in hours]
    
    time_data = []
    for hour in hours:
        time_data.append({
            '小时': f"{hour:02d}:00",
            '请求数': hourly_total.get(hour, 0),
            '占比(%)': round(hourly_total.get(hour, 0) / sum(hourly_total.values()) * 100, 2) if sum(hourly_total.values()) > 0 else 0
        })
    
    time_df = pd.DataFrame(time_data)
    
    # 添加到工作表
    time_headers = {
        "时间": ["小时"],
        "统计": ["请求数", "占比(%)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, time_df, '时间分布分析', header_groups=time_headers
    )
    
    # 添加折线图
    if len(time_df) > 1:
        try:
            create_line_chart(
                ws,
                min_row=4, max_row=3 + len(time_df),
                title="24小时请求分布",
                x_title="小时", y_title="请求数",
                y_cols=[2], chart_position="E5"
            )
        except Exception as e:
            log_info(f"创建时间分布图表失败: {e}")
    
    format_excel_sheet(ws)


def create_ip_overview_sheet(wb, ip_df, total_processed):
    """创建IP分析概览工作表"""
    ws = wb.create_sheet(title='概览')
    
    # 移动到第一个位置
    wb.move_sheet(ws, -(len(wb.worksheets) - 1))
    
    # 总体统计
    total_unique_ips = len(ip_df)
    total_requests = ip_df['总请求数'].sum()
    avg_requests_per_ip = total_requests / total_unique_ips if total_unique_ips > 0 else 0
    
    # 风险统计
    high_risk_count = len(ip_df[ip_df['风险评分'] > 70])
    medium_risk_count = len(ip_df[(ip_df['风险评分'] > 50) & (ip_df['风险评分'] <= 70)])
    low_risk_count = len(ip_df[ip_df['风险评分'] <= 50])
    
    # IP类型统计
    ip_type_counts = ip_df['IP类型'].value_counts()
    
    # 概览数据
    overview_data = [
        ['=== 基础统计 ===', ''],
        ['总处理记录数', total_processed],
        ['唯一IP数量', total_unique_ips],
        ['总请求数', total_requests],
        ['平均每IP请求数', round(avg_requests_per_ip, 2)],
        ['', ''],
        
        ['=== 风险分布 ===', ''],
        ['高风险IP数量 (>70)', high_risk_count],
        ['中等风险IP数量 (50-70)', medium_risk_count],
        ['低风险IP数量 (≤50)', low_risk_count],
        ['', ''],
        
        ['=== IP类型分布 ===', ''],
    ]
    
    # 添加IP类型统计
    for ip_type, count in ip_type_counts.items():
        overview_data.append([f'{ip_type}数量', count])
    
    overview_data.extend([
        ['', ''],
        ['=== TOP指标 ===', ''],
        ['请求量最大IP', ip_df.iloc[0]['IP地址'] if not ip_df.empty else 'N/A'],
        ['最大请求量', ip_df.iloc[0]['总请求数'] if not ip_df.empty else 0],
        ['最高风险评分IP', ip_df.loc[ip_df['风险评分'].idxmax(), 'IP地址'] if not ip_df.empty else 'N/A'],
        ['最高风险评分', ip_df['风险评分'].max() if not ip_df.empty else 0],
    ])
    
    # 写入数据
    for row_idx, (label, value) in enumerate(overview_data, start=1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_value = ws.cell(row=row_idx, column=2, value=value)
        
        if label.startswith('===') and label.endswith('==='):
            cell_label.font = Font(bold=True, size=12)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    format_excel_sheet(ws)