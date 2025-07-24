import gc
import pandas as pd
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font

from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD
from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    format_excel_sheet,
    add_dataframe_to_excel_with_grouped_headers,
    create_pie_chart,
    create_line_chart
)
from self_00_05_sampling_algorithms import ReservoirSampler
from self_10_request_header_analyzer import (
    extract_browser_info, 
    extract_os_info, 
    extract_device_info,
    detect_bot_type,
    extract_domain_from_referer,
    detect_search_engine,
    detect_social_media
)


def analyze_header_performance_correlation(csv_path, output_path, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """åˆ†æè¯·æ±‚å¤´ä¸æ€§èƒ½çš„å…³è”æ€§ - å†…å­˜ä¼˜åŒ–ç‰ˆ"""
    log_info("ğŸš€ å¼€å§‹è¯·æ±‚å¤´æ€§èƒ½å…³è”åˆ†æï¼ˆå†…å­˜ä¼˜åŒ–ç‰ˆï¼‰...", show_memory=True)
    
    chunk_size = max(DEFAULT_CHUNK_SIZE // 2, 10000)
    
    # æ€§èƒ½ç»Ÿè®¡æ•°æ®ç»“æ„
    browser_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times_sampler': ReservoirSampler(1000),  # æ›¿ä»£æ— é™åˆ¶æ•°ç»„
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    os_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times_sampler': ReservoirSampler(1000),  # æ›¿ä»£æ— é™åˆ¶æ•°ç»„
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    device_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times_sampler': ReservoirSampler(1000),  # æ›¿ä»£æ— é™åˆ¶æ•°ç»„
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    domain_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times_sampler': ReservoirSampler(1000),  # æ›¿ä»£æ— é™åˆ¶æ•°ç»„
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    search_engine_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times_sampler': ReservoirSampler(1000),  # æ›¿ä»£æ— é™åˆ¶æ•°ç»„
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    bot_performance = defaultdict(lambda: {
        'total_requests': 0,
        'slow_requests': 0,
        'total_response_time': 0.0,
        'response_times_sampler': ReservoirSampler(1000),  # æ›¿ä»£æ— é™åˆ¶æ•°ç»„
        'error_requests': 0,
        'data_transferred': 0.0
    })
    
    # æ…¢è¯·æ±‚è¯¦ç»†åˆ†æ
    slow_request_details = []
    
    total_processed = 0
    total_slow_requests = 0
    
    # ç¬¬ä¸€éï¼šæ”¶é›†æ€§èƒ½å…³è”æ•°æ®
    log_info("å¼€å§‹æ”¶é›†è¯·æ±‚å¤´æ€§èƒ½å…³è”æ•°æ®")
    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        chunk_size_actual = len(chunk)
        total_processed += chunk_size_actual
        
        # ç¡®ä¿å¿…è¦çš„åˆ—å­˜åœ¨
        required_columns = ['user_agent_string', 'referer_url', 'total_request_duration', 'response_status_code']
        missing_columns = [col for col in required_columns if col not in chunk.columns]
        if missing_columns:
            log_info(f"è­¦å‘Š: ç¼ºå°‘å¿…è¦åˆ— {missing_columns}", level="WARNING")
            continue
        
        # å¤„ç†æ•°æ®ç±»å‹
        chunk['total_request_duration'] = pd.to_numeric(chunk['total_request_duration'], errors='coerce')
        chunk['response_status_code'] = chunk['response_status_code'].astype(str)
        
        if 'response_body_size_kb' in chunk.columns:
            chunk['response_body_size_kb'] = pd.to_numeric(chunk['response_body_size_kb'], errors='coerce')
        
        for _, row in chunk.iterrows():
            user_agent = row.get('user_agent_string', '')
            referer = row.get('referer_url', '')
            response_time = row.get('total_request_duration', 0)
            status_code = str(row.get('response_status_code', ''))
            data_size = row.get('response_body_size_kb', 0) or 0
            
            # è·³è¿‡æ— æ•ˆæ•°æ®
            if pd.isna(response_time) or response_time <= 0:
                continue
            
            is_slow = response_time > slow_threshold
            is_error = status_code.startswith('4') or status_code.startswith('5')
            
            if is_slow:
                total_slow_requests += 1
            
            # åˆ†æUser-Agent
            if pd.notna(user_agent) and user_agent != '' and user_agent != '-':
                browser = extract_browser_info(user_agent)
                os_info = extract_os_info(user_agent)
                device = extract_device_info(user_agent)
                bot_type = detect_bot_type(user_agent)
                
                # æ›´æ–°æµè§ˆå™¨æ€§èƒ½ç»Ÿè®¡
                update_performance_stats(browser_performance[browser], response_time, is_slow, is_error, data_size)
                
                # æ›´æ–°æ“ä½œç³»ç»Ÿæ€§èƒ½ç»Ÿè®¡  
                update_performance_stats(os_performance[os_info], response_time, is_slow, is_error, data_size)
                
                # æ›´æ–°è®¾å¤‡ç±»å‹æ€§èƒ½ç»Ÿè®¡
                update_performance_stats(device_performance[device], response_time, is_slow, is_error, data_size)
                
                # æ›´æ–°æœºå™¨äººæ€§èƒ½ç»Ÿè®¡
                if bot_type:
                    update_performance_stats(bot_performance[bot_type], response_time, is_slow, is_error, data_size)
            
            # åˆ†æReferer
            if pd.notna(referer) and referer != '' and referer != '-':
                domain = extract_domain_from_referer(referer)
                search_engine = detect_search_engine(referer)
                
                # æ›´æ–°åŸŸåæ€§èƒ½ç»Ÿè®¡
                if domain:
                    update_performance_stats(domain_performance[domain], response_time, is_slow, is_error, data_size)
                
                # æ›´æ–°æœç´¢å¼•æ“æ€§èƒ½ç»Ÿè®¡
                if search_engine:
                    update_performance_stats(search_engine_performance[search_engine], response_time, is_slow, is_error, data_size)
            
            # æ”¶é›†æ…¢è¯·æ±‚è¯¦ç»†ä¿¡æ¯
            if is_slow and len(slow_request_details) < 10000:  # é™åˆ¶è¯¦ç»†è®°å½•æ•°é‡
                slow_detail = {
                    'è¯·æ±‚æ—¶é—´': row.get('raw_time', ''),
                    'è¯·æ±‚URI': row.get('request_full_uri', ''),
                    'å“åº”æ—¶é—´(ç§’)': round(response_time, 3),
                    'çŠ¶æ€ç ': status_code,
                    'æµè§ˆå™¨': extract_browser_info(user_agent) if pd.notna(user_agent) else 'æœªçŸ¥',
                    'æ“ä½œç³»ç»Ÿ': extract_os_info(user_agent) if pd.notna(user_agent) else 'æœªçŸ¥',
                    'è®¾å¤‡ç±»å‹': extract_device_info(user_agent) if pd.notna(user_agent) else 'æœªçŸ¥',
                    'æ¥æºåŸŸå': extract_domain_from_referer(referer) if pd.notna(referer) else 'ç›´æ¥è®¿é—®',
                    'User-Agent': (user_agent[:100] + '...') if len(str(user_agent)) > 100 else user_agent,
                    'Referer': (referer[:100] + '...') if len(str(referer)) > 100 else referer
                }
                slow_request_details.append(slow_detail)
        
        if total_processed % 100000 == 0:
            gc.collect()
            log_info(f"å·²å¤„ç† {total_processed:,} æ¡è®°å½•ï¼Œå‘ç° {total_slow_requests:,} æ¡æ…¢è¯·æ±‚")
    
    log_info(f"âœ… æ€§èƒ½å…³è”åˆ†æå®Œæˆï¼šæ€»è®°å½• {total_processed:,}ï¼Œæ…¢è¯·æ±‚ {total_slow_requests:,}")
    
    # ç”Ÿæˆåˆ†æç»“æœ
    analysis_results = {
        'browser_performance': calculate_performance_metrics(browser_performance),
        'os_performance': calculate_performance_metrics(os_performance),
        'device_performance': calculate_performance_metrics(device_performance),
        'domain_performance': calculate_performance_metrics(domain_performance),
        'search_engine_performance': calculate_performance_metrics(search_engine_performance),
        'bot_performance': calculate_performance_metrics(bot_performance),
        'slow_request_details': slow_request_details,
        'total_processed': total_processed,
        'total_slow_requests': total_slow_requests
    }
    
    # åˆ›å»ºExcelæŠ¥å‘Š
    create_header_performance_excel(analysis_results, output_path, slow_threshold)
    
    log_info(f"ğŸ‰ è¯·æ±‚å¤´æ€§èƒ½å…³è”åˆ†æå®Œæˆï¼ŒæŠ¥å‘Šå·²ç”Ÿæˆï¼š{output_path}", show_memory=True)
    
    # è¿”å›å…³é”®æ´å¯Ÿ
    insights = generate_performance_insights(analysis_results, slow_threshold)
    return insights


def update_performance_stats(stats, response_time, is_slow, is_error, data_size):
    """æ›´æ–°æ€§èƒ½ç»Ÿè®¡æ•°æ®"""
    stats['total_requests'] += 1
    stats['total_response_time'] += response_time
    
    if is_slow:
        stats['slow_requests'] += 1
    
    if is_error:
        stats['error_requests'] += 1
    
    stats['data_transferred'] += data_size
    
    # ä½¿ç”¨è“„æ°´æ± é‡‡æ ·ä¿å­˜å“åº”æ—¶é—´ï¼ˆé¿å…å†…å­˜é—®é¢˜ï¼‰
    stats['response_times_sampler'].add({'response_time': response_time})


def calculate_performance_metrics(performance_data):
    """è®¡ç®—æ€§èƒ½æŒ‡æ ‡"""
    results = []
    
    for category, stats in performance_data.items():
        if stats['total_requests'] == 0:
            continue
        
        total_requests = stats['total_requests']
        slow_requests = stats['slow_requests']
        error_requests = stats['error_requests']
        total_response_time = stats['total_response_time']
        data_transferred = stats['data_transferred']
        response_times_sample = [item['response_time'] for item in stats['response_times_sampler'].get_samples()]
        
        # è®¡ç®—å…³é”®æŒ‡æ ‡
        slow_rate = (slow_requests / total_requests * 100) if total_requests > 0 else 0
        error_rate = (error_requests / total_requests * 100) if total_requests > 0 else 0
        avg_response_time = total_response_time / total_requests if total_requests > 0 else 0
        avg_data_size = data_transferred / total_requests if total_requests > 0 else 0
        
        # è®¡ç®—ç™¾åˆ†ä½æ•° - åŸºäºé‡‡æ ·æ•°æ®
        if response_times_sample:
            import numpy as np
            p50 = np.percentile(response_times_sample, 50)
            p95 = np.percentile(response_times_sample, 95)
            p99 = np.percentile(response_times_sample, 99)
        else:
            p50 = p95 = p99 = 0
        
        result = {
            'ç±»åˆ«': category,
            'æ€»è¯·æ±‚æ•°': total_requests,
            'æ…¢è¯·æ±‚æ•°': slow_requests,
            'æ…¢è¯·æ±‚ç‡(%)': round(slow_rate, 2),
            'é”™è¯¯è¯·æ±‚æ•°': error_requests,
            'é”™è¯¯ç‡(%)': round(error_rate, 2),
            'å¹³å‡å“åº”æ—¶é—´(ç§’)': round(avg_response_time, 3),
            'P50å“åº”æ—¶é—´(ç§’)': round(p50, 3),
            'P95å“åº”æ—¶é—´(ç§’)': round(p95, 3),
            'P99å“åº”æ—¶é—´(ç§’)': round(p99, 3),
            'å¹³å‡æ•°æ®ä¼ è¾“(KB)': round(avg_data_size, 2),
            'æ€»æ•°æ®ä¼ è¾“(MB)': round(data_transferred / 1024, 2)
        }
        
        results.append(result)
    
    # æŒ‰æ…¢è¯·æ±‚ç‡æ’åº
    results.sort(key=lambda x: x['æ…¢è¯·æ±‚ç‡(%)'], reverse=True)
    return pd.DataFrame(results)


def generate_performance_insights(analysis_results, slow_threshold):
    """ç”Ÿæˆæ€§èƒ½æ´å¯Ÿ"""
    insights = {
        'slow_threshold': slow_threshold,
        'total_processed': analysis_results['total_processed'],
        'total_slow_requests': analysis_results['total_slow_requests'],
        'slow_rate_overall': round(analysis_results['total_slow_requests'] / analysis_results['total_processed'] * 100, 2),
        'worst_browsers': [],
        'worst_os': [],
        'worst_devices': [],
        'worst_domains': [],
        'performance_recommendations': []
    }
    
    # æ‰¾å‡ºæ€§èƒ½æœ€å·®çš„æµè§ˆå™¨
    browser_df = analysis_results['browser_performance']
    if not browser_df.empty:
        worst_browsers = browser_df.head(3)
        for _, row in worst_browsers.iterrows():
            insights['worst_browsers'].append({
                'browser': row['ç±»åˆ«'],
                'slow_rate': row['æ…¢è¯·æ±‚ç‡(%)'],
                'avg_response_time': row['å¹³å‡å“åº”æ—¶é—´(ç§’)']
            })
    
    # æ‰¾å‡ºæ€§èƒ½æœ€å·®çš„æ“ä½œç³»ç»Ÿ
    os_df = analysis_results['os_performance']
    if not os_df.empty:
        worst_os = os_df.head(3)
        for _, row in worst_os.iterrows():
            insights['worst_os'].append({
                'os': row['ç±»åˆ«'],
                'slow_rate': row['æ…¢è¯·æ±‚ç‡(%)'],
                'avg_response_time': row['å¹³å‡å“åº”æ—¶é—´(ç§’)']
            })
    
    # æ‰¾å‡ºæ€§èƒ½æœ€å·®çš„è®¾å¤‡ç±»å‹
    device_df = analysis_results['device_performance']
    if not device_df.empty:
        worst_devices = device_df.head(3)
        for _, row in worst_devices.iterrows():
            insights['worst_devices'].append({
                'device': row['ç±»åˆ«'],
                'slow_rate': row['æ…¢è¯·æ±‚ç‡(%)'],
                'avg_response_time': row['å¹³å‡å“åº”æ—¶é—´(ç§’)']
            })
    
    # æ‰¾å‡ºæ€§èƒ½æœ€å·®çš„æ¥æºåŸŸå
    domain_df = analysis_results['domain_performance']
    if not domain_df.empty:
        worst_domains = domain_df.head(5)
        for _, row in worst_domains.iterrows():
            insights['worst_domains'].append({
                'domain': row['ç±»åˆ«'],
                'slow_rate': row['æ…¢è¯·æ±‚ç‡(%)'],
                'avg_response_time': row['å¹³å‡å“åº”æ—¶é—´(ç§’)']
            })
    
    # ç”Ÿæˆä¼˜åŒ–å»ºè®®
    insights['performance_recommendations'] = generate_optimization_recommendations(analysis_results)
    
    return insights


def generate_optimization_recommendations(analysis_results):
    """ç”Ÿæˆæ€§èƒ½ä¼˜åŒ–å»ºè®®"""
    recommendations = []
    
    # åˆ†ææµè§ˆå™¨æ€§èƒ½
    browser_df = analysis_results['browser_performance']
    if not browser_df.empty and len(browser_df) > 0:
        worst_browser = browser_df.iloc[0]
        if worst_browser['æ…¢è¯·æ±‚ç‡(%)'] > 10:
            recommendations.append(f"æµè§ˆå™¨ä¼˜åŒ–: {worst_browser['ç±»åˆ«']} çš„æ…¢è¯·æ±‚ç‡è¾¾åˆ° {worst_browser['æ…¢è¯·æ±‚ç‡(%)']}%ï¼Œå»ºè®®é’ˆå¯¹è¯¥æµè§ˆå™¨è¿›è¡Œä¸“é¡¹ä¼˜åŒ–")
    
    # åˆ†æè®¾å¤‡æ€§èƒ½
    device_df = analysis_results['device_performance']
    if not device_df.empty and len(device_df) > 0:
        for _, row in device_df.iterrows():
            if row['ç±»åˆ«'] == 'æ‰‹æœº' and row['æ…¢è¯·æ±‚ç‡(%)'] > 15:
                recommendations.append(f"ç§»åŠ¨ç«¯ä¼˜åŒ–: æ‰‹æœºç«¯æ…¢è¯·æ±‚ç‡ä¸º {row['æ…¢è¯·æ±‚ç‡(%)']}%ï¼Œå»ºè®®ä¼˜åŒ–ç§»åŠ¨ç«¯æ€§èƒ½")
            elif row['ç±»åˆ«'] == 'çˆ¬è™«/æœºå™¨äºº' and row['æ€»è¯·æ±‚æ•°'] > 1000:
                recommendations.append(f"æœºå™¨äººæµé‡: æ£€æµ‹åˆ°å¤§é‡æœºå™¨äººè¯·æ±‚({row['æ€»è¯·æ±‚æ•°']}æ¬¡)ï¼Œå»ºè®®è¿›è¡Œæµé‡æ§åˆ¶")
    
    # åˆ†ææ¥æºåŸŸåæ€§èƒ½
    domain_df = analysis_results['domain_performance']
    if not domain_df.empty and len(domain_df) > 0:
        high_traffic_slow_domains = domain_df[
            (domain_df['æ€»è¯·æ±‚æ•°'] > 100) & (domain_df['æ…¢è¯·æ±‚ç‡(%)'] > 20)
        ]
        if not high_traffic_slow_domains.empty:
            for _, row in high_traffic_slow_domains.head(3).iterrows():
                recommendations.append(f"æ¥æºä¼˜åŒ–: æ¥è‡ª {row['ç±»åˆ«']} çš„è¯·æ±‚æ…¢è¯·æ±‚ç‡ä¸º {row['æ…¢è¯·æ±‚ç‡(%)']}%ï¼Œå»ºè®®æ£€æŸ¥ç›¸å…³é“¾æ¥æˆ–ç¼“å­˜ç­–ç•¥")
    
    # åˆ†ææœºå™¨äººæ€§èƒ½
    bot_df = analysis_results['bot_performance']
    if not bot_df.empty and len(bot_df) > 0:
        for _, row in bot_df.iterrows():
            if row['æ€»è¯·æ±‚æ•°'] > 500 and row['æ…¢è¯·æ±‚ç‡(%)'] > 30:
                recommendations.append(f"çˆ¬è™«ä¼˜åŒ–: {row['ç±»åˆ«']} äº§ç”Ÿäº† {row['æ€»è¯·æ±‚æ•°']} æ¬¡è¯·æ±‚ï¼Œæ…¢è¯·æ±‚ç‡ {row['æ…¢è¯·æ±‚ç‡(%)']}%ï¼Œå»ºè®®ä¼˜åŒ–çˆ¬è™«å“åº”æˆ–é™åˆ¶é¢‘ç‡")
    
    return recommendations


def create_header_performance_excel(analysis_results, output_path, slow_threshold):
    """åˆ›å»ºè¯·æ±‚å¤´æ€§èƒ½å…³è”ExcelæŠ¥å‘Š"""
    log_info(f"åˆ›å»ºè¯·æ±‚å¤´æ€§èƒ½å…³è”ExcelæŠ¥å‘Š: {output_path}")
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 1. æ¦‚è§ˆå·¥ä½œè¡¨
    create_performance_overview_sheet(wb, analysis_results, slow_threshold)
    
    # 2. æµè§ˆå™¨æ€§èƒ½åˆ†æ
    create_performance_analysis_sheet(wb, analysis_results['browser_performance'], 'æµè§ˆå™¨æ€§èƒ½åˆ†æ', 'æµè§ˆå™¨')
    
    # 3. æ“ä½œç³»ç»Ÿæ€§èƒ½åˆ†æ
    create_performance_analysis_sheet(wb, analysis_results['os_performance'], 'æ“ä½œç³»ç»Ÿæ€§èƒ½åˆ†æ', 'æ“ä½œç³»ç»Ÿ')
    
    # 4. è®¾å¤‡ç±»å‹æ€§èƒ½åˆ†æ
    create_performance_analysis_sheet(wb, analysis_results['device_performance'], 'è®¾å¤‡ç±»å‹æ€§èƒ½åˆ†æ', 'è®¾å¤‡ç±»å‹')
    
    # 5. æ¥æºåŸŸåæ€§èƒ½åˆ†æ
    create_performance_analysis_sheet(wb, analysis_results['domain_performance'], 'æ¥æºåŸŸåæ€§èƒ½åˆ†æ', 'æ¥æºåŸŸå')
    
    # 6. æœç´¢å¼•æ“æ€§èƒ½åˆ†æ
    create_performance_analysis_sheet(wb, analysis_results['search_engine_performance'], 'æœç´¢å¼•æ“æ€§èƒ½åˆ†æ', 'æœç´¢å¼•æ“')
    
    # 7. æœºå™¨äººæ€§èƒ½åˆ†æ
    create_performance_analysis_sheet(wb, analysis_results['bot_performance'], 'æœºå™¨äººæ€§èƒ½åˆ†æ', 'æœºå™¨äººç±»å‹')
    
    # 8. æ…¢è¯·æ±‚è¯¦ç»†åˆ—è¡¨
    create_slow_requests_detail_sheet(wb, analysis_results['slow_request_details'])
    
    # ä¿å­˜æ–‡ä»¶
    wb.save(output_path)
    log_info(f"è¯·æ±‚å¤´æ€§èƒ½å…³è”ExcelæŠ¥å‘Šå·²ä¿å­˜: {output_path}")


def create_performance_overview_sheet(wb, analysis_results, slow_threshold):
    """åˆ›å»ºæ€§èƒ½æ¦‚è§ˆå·¥ä½œè¡¨"""
    ws = wb.create_sheet(title='æ€§èƒ½å…³è”æ¦‚è§ˆ')
    
    # ç§»åŠ¨åˆ°ç¬¬ä¸€ä¸ªä½ç½®
    wb.move_sheet(ws, -(len(wb.worksheets) - 1))
    
    total_processed = analysis_results['total_processed']
    total_slow_requests = analysis_results['total_slow_requests']
    overall_slow_rate = round(total_slow_requests / total_processed * 100, 2) if total_processed > 0 else 0
    
    # æ¦‚è§ˆæ•°æ®
    overview_data = [
        ['=== æ€§èƒ½å…³è”åˆ†ææ¦‚è§ˆ ===', ''],
        ['åˆ†æé˜ˆå€¼', f'{slow_threshold} ç§’'],
        ['æ€»å¤„ç†è®°å½•æ•°', total_processed],
        ['æ…¢è¯·æ±‚æ€»æ•°', total_slow_requests],
        ['æ•´ä½“æ…¢è¯·æ±‚ç‡', f'{overall_slow_rate}%'],
        ['', ''],
        
        ['=== æ€§èƒ½è¡¨ç°æœ€å·®TOP3 ===', ''],
        ['æµè§ˆå™¨:', ''],
    ]
    
    # æ·»åŠ æ€§èƒ½æœ€å·®çš„æµè§ˆå™¨
    browser_df = analysis_results['browser_performance']
    if not browser_df.empty:
        for i, (_, row) in enumerate(browser_df.head(3).iterrows(), 1):
            overview_data.append([f'  {i}. {row["ç±»åˆ«"]}', f'æ…¢è¯·æ±‚ç‡: {row["æ…¢è¯·æ±‚ç‡(%)"]}%'])
    
    overview_data.extend([
        ['', ''],
        ['æ“ä½œç³»ç»Ÿ:', ''],
    ])
    
    # æ·»åŠ æ€§èƒ½æœ€å·®çš„æ“ä½œç³»ç»Ÿ
    os_df = analysis_results['os_performance']
    if not os_df.empty:
        for i, (_, row) in enumerate(os_df.head(3).iterrows(), 1):
            overview_data.append([f'  {i}. {row["ç±»åˆ«"]}', f'æ…¢è¯·æ±‚ç‡: {row["æ…¢è¯·æ±‚ç‡(%)"]}%'])
    
    overview_data.extend([
        ['', ''],
        ['è®¾å¤‡ç±»å‹:', ''],
    ])
    
    # æ·»åŠ æ€§èƒ½æœ€å·®çš„è®¾å¤‡ç±»å‹
    device_df = analysis_results['device_performance']
    if not device_df.empty:
        for i, (_, row) in enumerate(device_df.head(3).iterrows(), 1):
            overview_data.append([f'  {i}. {row["ç±»åˆ«"]}', f'æ…¢è¯·æ±‚ç‡: {row["æ…¢è¯·æ±‚ç‡(%)"]}%'])
    
    # å†™å…¥æ•°æ®
    for row_idx, (label, value) in enumerate(overview_data, start=1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_value = ws.cell(row=row_idx, column=2, value=value)
        
        if str(label).startswith('===') and str(label).endswith('==='):
            cell_label.font = Font(bold=True, size=12)
        elif str(label).endswith(':'):
            cell_label.font = Font(bold=True)
    
    # è®¾ç½®åˆ—å®½
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    
    format_excel_sheet(ws)


def create_performance_analysis_sheet(wb, performance_df, sheet_name, category_name):
    """åˆ›å»ºæ€§èƒ½åˆ†æå·¥ä½œè¡¨"""
    if performance_df.empty:
        return
    
    # è¡¨å¤´åˆ†ç»„
    headers = {
        "åˆ†ç±»ä¿¡æ¯": ["ç±»åˆ«"],
        "è¯·æ±‚ç»Ÿè®¡": ["æ€»è¯·æ±‚æ•°", "æ…¢è¯·æ±‚æ•°", "é”™è¯¯è¯·æ±‚æ•°"],
        "æ€§èƒ½æŒ‡æ ‡": ["æ…¢è¯·æ±‚ç‡(%)", "é”™è¯¯ç‡(%)", "å¹³å‡å“åº”æ—¶é—´(ç§’)"],
        "å“åº”æ—¶é—´åˆ†å¸ƒ": ["P50å“åº”æ—¶é—´(ç§’)", "P95å“åº”æ—¶é—´(ç§’)", "P99å“åº”æ—¶é—´(ç§’)"],
        "æ•°æ®ä¼ è¾“": ["å¹³å‡æ•°æ®ä¼ è¾“(KB)", "æ€»æ•°æ®ä¼ è¾“(MB)"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, performance_df, sheet_name, header_groups=headers
    )
    
    # æ·»åŠ æ…¢è¯·æ±‚ç‡åˆ†å¸ƒå›¾
    if len(performance_df) > 1:
        chart_start_row = len(performance_df) + 5
        ws.cell(row=chart_start_row, column=1, value=f"{category_name}æ…¢è¯·æ±‚ç‡åˆ†å¸ƒ").font = Font(bold=True)
        
        create_pie_chart(
            ws, f"{category_name}æ…¢è¯·æ±‚ç‡åˆ†å¸ƒ",
            data_start_row=3, data_end_row=2 + len(performance_df),
            labels_col=1, values_col=4,  # æ…¢è¯·æ±‚ç‡åˆ—
            position="H3"
        )


def create_slow_requests_detail_sheet(wb, slow_request_details):
    """åˆ›å»ºæ…¢è¯·æ±‚è¯¦ç»†åˆ—è¡¨å·¥ä½œè¡¨"""
    if not slow_request_details:
        return
    
    slow_df = pd.DataFrame(slow_request_details)
    
    # è¡¨å¤´åˆ†ç»„
    headers = {
        "åŸºæœ¬ä¿¡æ¯": ["è¯·æ±‚æ—¶é—´", "è¯·æ±‚URI", "å“åº”æ—¶é—´(ç§’)", "çŠ¶æ€ç "],
        "å®¢æˆ·ç«¯ä¿¡æ¯": ["æµè§ˆå™¨", "æ“ä½œç³»ç»Ÿ", "è®¾å¤‡ç±»å‹"],
        "æ¥æºä¿¡æ¯": ["æ¥æºåŸŸå"],
        "è¯¦ç»†ä¿¡æ¯": ["User-Agent", "Referer"]
    }
    
    ws = add_dataframe_to_excel_with_grouped_headers(
        wb, slow_df, 'æ…¢è¯·æ±‚è¯¦ç»†åˆ—è¡¨', header_groups=headers
    )