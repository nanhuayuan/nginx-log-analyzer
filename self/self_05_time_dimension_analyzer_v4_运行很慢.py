import gc
import os
import logging
from collections import defaultdict
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from self_00_01_constants import DEFAULT_CHUNK_SIZE, CHART_MAX_POINTS, DEFAULT_SLOW_REQUESTS_THRESHOLD, \
    DEFAULT_SLOW_THRESHOLD
from self_00_02_utils import log_info
from self_00_04_excel_processor import format_excel_sheet, add_dataframe_to_excel_with_grouped_headers, \
    create_line_chart


def analyze_time_dimension(csv_path, output_path, specific_uri_list=None):
    """分析请求时间维度数据并生成Excel报告"""
    output_filename = prepare_output_filename(output_path, specific_uri_list)

    # 定义核心指标组
    metrics_config = get_metrics_configuration()

    stats_containers = initialize_stats_containers()
    total_records, success_requests = read_and_process_data(
        csv_path, specific_uri_list, metrics_config, stats_containers
    )

    calculate_all_time_dimensions(stats_containers)
    avg_metrics = calculate_avg_metrics_for_all_dimensions(stats_containers)
    peak_hour = identify_peak_period(stats_containers['hourly']['stats'], 'success_requests')
    peak_minute = identify_peak_period(stats_containers['minute']['stats'], 'success_requests')

    create_time_dimension_excel(
        output_filename,
        stats_containers['daily']['stats'],
        stats_containers['hourly']['stats'],
        stats_containers['minute']['stats'],
        stats_containers['second']['stats'],
        total_records,
        avg_metrics['daily'],
        avg_metrics['hourly'],
        avg_metrics['minute'],
        avg_metrics['second'],
        peak_hour,
        peak_minute
    )

    log_info(f"报告已生成：{output_filename}", True)
    return output_filename


def get_metrics_configuration():
    """获取指标配置信息"""
    return {
        # 基础时间指标（4个核心指标）
        'basic_time_metrics': [
            'total_request_duration',  # 请求总时长
            'upstream_response_time',  # 后端响应时长
            'upstream_header_time',  # 后端处理时长
            'upstream_connect_time'  # 后端连接时长
        ],

        # 核心阶段指标（4个关键阶段）
        'phase_metrics': [
            'backend_connect_phase',  # 后端连接阶段
            'backend_process_phase',  # 后端处理阶段
            'backend_transfer_phase',  # 后端传输阶段
            'nginx_transfer_phase'  # Nginx传输阶段
        ],

        # 组合分析指标
        'composite_metrics': [
            'backend_total_phase',  # 后端总阶段
            'network_phase',  # 网络传输阶段
            'processing_phase',  # 纯处理阶段
            'transfer_phase'  # 纯传输阶段
        ],

        # 传输大小指标（已转换为KB）
        'size_metrics': [
            'response_body_size_kb',  # 响应体大小(KB)
            'total_bytes_sent_kb'  # 总传输大小(KB)
        ],

        # 性能比率指标（百分比形式）
        'efficiency_metrics': [
            'backend_efficiency',  # 后端处理效率
            'network_overhead',  # 网络开销占比
            'transfer_ratio',  # 传输时间占比
            'connection_cost_ratio',  # 连接成本占比
            'processing_efficiency_index'  # 处理效率指数
        ],

        # 传输速度指标
        'speed_metrics': [
            'response_transfer_speed',  # 响应传输速度
            'total_transfer_speed',  # 总传输速度
            'nginx_transfer_speed'  # Nginx传输速度
        ]
    }


def get_metric_display_names():
    """获取指标的中文显示名称"""
    return {
        # 基础时间指标
        'total_request_duration': '请求总时长',
        'upstream_response_time': '后端响应时长',
        'upstream_header_time': '后端处理时长',
        'upstream_connect_time': '后端连接时长',

        # 阶段指标
        'backend_connect_phase': '后端连接阶段',
        'backend_process_phase': '后端处理阶段',
        'backend_transfer_phase': '后端传输阶段',
        'nginx_transfer_phase': 'Nginx传输阶段',

        # 组合指标
        'backend_total_phase': '后端总阶段',
        'network_phase': '网络传输阶段',
        'processing_phase': '纯处理阶段',
        'transfer_phase': '纯传输阶段',

        # 大小指标
        'response_body_size_kb': '响应体大小',
        'total_bytes_sent_kb': '总传输大小',

        # 效率指标
        'backend_efficiency': '后端处理效率',
        'network_overhead': '网络开销占比',
        'transfer_ratio': '传输时间占比',
        'connection_cost_ratio': '连接成本占比',
        'processing_efficiency_index': '处理效率指数',

        # 速度指标
        'response_transfer_speed': '响应传输速度',
        'total_transfer_speed': '总传输速度',
        'nginx_transfer_speed': 'Nginx传输速度'
    }


def prepare_time_dimension_data(stats_dict, time_label, avg_metrics):
    """准备时间维度数据，优化后的版本"""
    from collections import OrderedDict

    data = OrderedDict()
    data[time_label] = list(stats_dict.keys())

    # 基础统计指标
    basic_stats = [
        ('total_requests', '总请求数'),
        ('success_requests', '成功请求数'),
        ('slow_requests', '慢请求数'),
        ('new_connections', '新建连接数'),
        ('concurrent_connections', '并发连接数'),
        ('active_connections', '活跃连接数'),
        ('qps', 'QPS')
    ]

    # 填充基础统计数据
    for metric_code, metric_name in basic_stats:
        data[metric_name] = [stats_dict[time_key].get(metric_code, 0) for time_key in data[time_label]]

    # 计算慢请求占比
    slow_ratio_data = []
    for i, time_key in enumerate(data[time_label]):
        total = data['总请求数'][i]
        slow = data['慢请求数'][i]
        ratio = (slow / total * 100) if total > 0 and slow is not None else 0
        slow_ratio_data.append(ratio)

    # 构建有序数据结构
    ordered_data = OrderedDict()
    ordered_data[time_label] = data[time_label]
    ordered_data['总请求数'] = data['总请求数']
    ordered_data['成功请求数'] = data['成功请求数']
    ordered_data['慢请求数'] = data['慢请求数']
    ordered_data['慢请求占比(%)'] = slow_ratio_data
    ordered_data['新建连接数'] = data['新建连接数']
    ordered_data['并发连接数'] = data['并发连接数']
    ordered_data['活跃连接数'] = data['活跃连接数']
    ordered_data['QPS'] = data['QPS']

    # 处理平均指标数据
    if avg_metrics:
        add_average_metrics_to_data(ordered_data, avg_metrics, data[time_label])

    df = pd.DataFrame(ordered_data)
    df = df.sort_values(by=time_label)

    # 清理内存
    del data, slow_ratio_data, ordered_data
    gc.collect()

    return df


def add_average_metrics_to_data(ordered_data, avg_metrics, time_keys):
    """添加平均指标数据到有序数据结构中"""
    metrics_config = get_metrics_configuration()
    display_names = get_metric_display_names()

    # 基础时间指标 - 秒为单位
    for metric in metrics_config['basic_time_metrics']:
        display_name = display_names.get(metric, metric)
        ordered_data[f'平均{display_name}(秒)'] = [
            avg_metrics.get(time_key, {}).get(metric, 0) for time_key in time_keys
        ]

    # 阶段时间指标 - 秒为单位
    for metric in metrics_config['phase_metrics']:
        display_name = display_names.get(metric, metric)
        ordered_data[f'平均{display_name}(秒)'] = [
            avg_metrics.get(time_key, {}).get(metric, 0) for time_key in time_keys
        ]

    # 组合时间指标 - 秒为单位
    for metric in metrics_config['composite_metrics']:
        display_name = display_names.get(metric, metric)
        ordered_data[f'平均{display_name}(秒)'] = [
            avg_metrics.get(time_key, {}).get(metric, 0) for time_key in time_keys
        ]

    # 大小指标 - KB为单位（CSV中已经是KB了）
    for metric in metrics_config['size_metrics']:
        display_name = display_names.get(metric, metric)
        ordered_data[f'平均{display_name}(KB)'] = [
            avg_metrics.get(time_key, {}).get(metric, 0) for time_key in time_keys
        ]

    # 效率指标 - 百分比形式
    for metric in metrics_config['efficiency_metrics']:
        display_name = display_names.get(metric, metric)
        ordered_data[f'平均{display_name}(%)'] = [
            avg_metrics.get(time_key, {}).get(metric, 0) for time_key in time_keys
        ]

    # 速度指标 - KB/s为单位
    for metric in metrics_config['speed_metrics']:
        display_name = display_names.get(metric, metric)
        ordered_data[f'平均{display_name}(KB/s)'] = [
            avg_metrics.get(time_key, {}).get(metric, 0) for time_key in time_keys
        ]


def create_header_groups(df, time_label):
    """创建优化后的表头分组"""
    header_groups = {
        time_label: [time_label],
        '请求数量统计': ['总请求数', '成功请求数', '慢请求数', '慢请求占比(%)'],
        '连接统计': ['新建连接数', '并发连接数', '活跃连接数', 'QPS'],
        '基础时间指标': [col for col in df.columns if
                   '平均' in col and any(x in col for x in ['请求总时长', '后端响应时长', '后端处理时长', '后端连接时长']) and '(秒)' in col],
        '阶段时间指标': [col for col in df.columns if
                   '平均' in col and any(x in col for x in ['连接阶段', '处理阶段', '传输阶段']) and '(秒)' in col],
        '组合时间指标': [col for col in df.columns if
                   '平均' in col and any(x in col for x in ['总阶段', '网络传输', '纯处理', '纯传输']) and '(秒)' in col],
        '传输大小指标': [col for col in df.columns if '平均' in col and '大小' in col and '(KB)' in col],
        '性能效率指标': [col for col in df.columns if
                   '平均' in col and ('效率' in col or '占比' in col or '指数' in col) and '(%)' in col],
        '传输速度指标': [col for col in df.columns if '平均' in col and '速度' in col and '(KB/s)' in col]
    }

    # 移除空分组
    header_groups = {k: v for k, v in header_groups.items() if v}
    return header_groups


def add_dimension_charts(ws, df, sheet_name, time_label, header_groups, peak_time):
    """添加优化后的维度图表"""
    start_row = 3
    end_row = start_row + len(df) - 1
    apply_peak_highlight = sheet_name in ['分钟维度', '秒维度'] and peak_time is not None

    # 请求数量分布图表
    request_cols, series_names = get_chart_columns(header_groups, ['总请求数', '成功请求数', '慢请求数'])
    if request_cols:
        create_line_chart(
            ws, start_row, end_row,
            f"{sheet_name}请求数量分布", time_label, "请求数",
            request_cols, series_names,
            "H1", apply_peak_highlight, peak_time
        )

    # 连接数分布图表
    conn_cols, conn_series_names = get_chart_columns(header_groups, ['新建连接数', '并发连接数', '活跃连接数'])
    if conn_cols:
        create_line_chart(
            ws, start_row, end_row,
            f"{sheet_name}连接数分布", time_label, "连接数",
            conn_cols, conn_series_names,
            "H20", apply_peak_highlight, peak_time
        )

    # 性能指标图表
    perf_cols, perf_series_names = get_chart_columns(header_groups, ['QPS', '慢请求占比(%)'])
    if perf_cols:
        create_line_chart(
            ws, start_row, end_row,
            f"{sheet_name}性能指标分布", time_label, "值",
            perf_cols, perf_series_names,
            "H40", apply_peak_highlight, peak_time
        )

    # 基础时间指标图表
    basic_time_cols = header_groups.get('基础时间指标', [])
    if basic_time_cols:
        cols, series_names = get_chart_columns_by_headers(header_groups, basic_time_cols)
        if cols:
            create_line_chart(
                ws, start_row, end_row,
                f"{sheet_name}基础时间指标分布", time_label, "时长(秒)",
                cols, series_names,
                "R1", apply_peak_highlight, peak_time
            )

    # 阶段时间指标图表
    phase_time_cols = header_groups.get('阶段时间指标', [])
    if phase_time_cols:
        cols, series_names = get_chart_columns_by_headers(header_groups, phase_time_cols)
        if cols:
            create_line_chart(
                ws, start_row, end_row,
                f"{sheet_name}阶段时间指标分布", time_label, "时长(秒)",
                cols, series_names,
                "R20", apply_peak_highlight, peak_time
            )

    # 性能效率指标图表
    efficiency_cols = header_groups.get('性能效率指标', [])
    if efficiency_cols:
        cols, series_names = get_chart_columns_by_headers(header_groups, efficiency_cols)
        if cols:
            create_line_chart(
                ws, start_row, end_row,
                f"{sheet_name}性能效率指标分布", time_label, "百分比(%)",
                cols, series_names,
                "R40", apply_peak_highlight, peak_time
            )

    # 传输大小指标图表
    size_cols = header_groups.get('传输大小指标', [])
    if size_cols:
        cols, series_names = get_chart_columns_by_headers(header_groups, size_cols)
        if cols:
            create_line_chart(
                ws, start_row, end_row,
                f"{sheet_name}传输大小指标分布", time_label, "大小(KB)",
                cols, series_names,
                "AB1", apply_peak_highlight, peak_time
            )

    # 传输速度指标图表
    speed_cols = header_groups.get('传输速度指标', [])
    if speed_cols:
        cols, series_names = get_chart_columns_by_headers(header_groups, speed_cols)
        if cols:
            create_line_chart(
                ws, start_row, end_row,
                f"{sheet_name}传输速度指标分布", time_label, "速度(KB/s)",
                cols, series_names,
                "AB20", apply_peak_highlight, peak_time
            )

    # 清理内存
    gc.collect()


def get_chart_columns_by_headers(header_groups, target_headers):
    """根据表头列表获取图表列索引"""
    cols = []
    series_names = []
    col_idx = 1

    for group_name, subheaders in header_groups.items():
        for subheader in subheaders:
            if subheader in target_headers:
                cols.append(col_idx)
                series_names.append(subheader)
            col_idx += 1

    return cols, series_names


def add_time_series_analysis(wb, time_series_metrics, sheet_name):
    """添加优化后的时间序列分析页面"""
    if not time_series_metrics:
        log_info("无时间序列指标数据，跳过时间序列分析页面")
        return

    if '秒维度' in wb.sheetnames:
        second_dimension_sheet = wb['秒维度']
        if second_dimension_sheet.max_row > 10:
            log_info("秒维度页面已包含足够详细的时间序列数据，跳过创建单独的时间序列分析页面")
            return

    # 重新组织指标分组，基于新的参数体系
    metric_groups = {
        '基础时间指标': [
            'total_request_duration',
            'upstream_response_time',
            'upstream_header_time',
            'upstream_connect_time'
        ],
        '阶段时间指标': [
            'backend_connect_phase',
            'backend_process_phase',
            'backend_transfer_phase',
            'nginx_transfer_phase'
        ],
        '组合时间指标': [
            'backend_total_phase',
            'network_phase',
            'processing_phase',
            'transfer_phase'
        ],
        '传输大小指标': [
            'response_body_size_kb',
            'total_bytes_sent_kb'
        ],
        '性能效率指标': [
            'backend_efficiency',
            'network_overhead',
            'transfer_ratio',
            'connection_cost_ratio',
            'processing_efficiency_index'
        ],
        '传输速度指标': [
            'response_transfer_speed',
            'total_transfer_speed',
            'nginx_transfer_speed'
        ]
    }

    ws = wb.create_sheet(title=sheet_name)
    row = 1
    display_names = get_metric_display_names()

    for group_name, metrics in metric_groups.items():
        ws.cell(row=row, column=1, value=group_name).font = Font(bold=True, size=14)
        row += 2

        # 表头
        ws.cell(row=row, column=1, value="指标名称").font = Font(bold=True)
        ws.cell(row=row, column=2, value="指标含义").font = Font(bold=True)
        ws.cell(row=row, column=3, value="单位").font = Font(bold=True)
        ws.cell(row=row, column=4, value="优化建议").font = Font(bold=True)
        row += 1

        # 指标详细信息
        meanings = get_metric_meanings()
        optimization_tips = get_optimization_tips()

        for metric in metrics:
            display_name = display_names.get(metric, metric)
            unit = get_metric_unit(metric)
            meaning = meanings.get(metric, "待定义")
            tip = optimization_tips.get(metric, "持续监控")

            ws.cell(row=row, column=1, value=display_name).font = Font(bold=True)
            ws.cell(row=row, column=2, value=meaning)
            ws.cell(row=row, column=3, value=unit)
            ws.cell(row=row, column=4, value=tip)
            row += 1

        row += 2

    format_excel_sheet(ws)


def get_metric_meanings():
    """获取指标含义说明"""
    return {
        # 基础时间指标
        'total_request_duration': '从接收请求到发送完响应的总时间',
        'upstream_response_time': '从建立连接到接收完整后端响应的时间',
        'upstream_header_time': '从建立连接到接收后端响应头的时间',
        'upstream_connect_time': '与后端服务建立连接所需的时间',

        # 阶段时间指标
        'backend_connect_phase': '后端连接建立阶段耗时',
        'backend_process_phase': '后端业务处理阶段耗时',
        'backend_transfer_phase': '后端响应传输阶段耗时',
        'nginx_transfer_phase': 'Nginx向客户端传输阶段耗时',

        # 组合时间指标
        'backend_total_phase': '后端处理的总时间（连接+处理+传输）',
        'network_phase': '网络相关的总时间（连接+客户端传输）',
        'processing_phase': '纯业务处理时间',
        'transfer_phase': '纯数据传输时间',

        # 传输大小指标
        'response_body_size_kb': '响应体数据大小，单位KB',
        'total_bytes_sent_kb': '发送给客户端的总数据量，单位KB',

        # 性能效率指标
        'backend_efficiency': '后端处理时间占后端总时间的比例',
        'network_overhead': '网络开销占总请求时间的比例',
        'transfer_ratio': '数据传输时间占总请求时间的比例',
        'connection_cost_ratio': '连接成本占总时间的比例',
        'processing_efficiency_index': '处理效率综合指数',

        # 传输速度指标
        'response_transfer_speed': '响应体传输速度',
        'total_transfer_speed': '总数据传输速度',
        'nginx_transfer_speed': 'Nginx传输速度'
    }


def get_optimization_tips():
    """获取优化建议"""
    return {
        # 基础时间指标
        'total_request_duration': '总时长过长时检查各阶段耗时分布',
        'upstream_response_time': '后端响应慢时优化业务逻辑或数据库',
        'upstream_header_time': '处理慢时检查业务逻辑复杂度',
        'upstream_connect_time': '连接慢时检查网络和服务可用性',

        # 阶段时间指标
        'backend_connect_phase': '连接慢时检查网络延迟和服务健康状态',
        'backend_process_phase': '处理慢时优化算法和数据库查询',
        'backend_transfer_phase': '传输慢时检查网络带宽和数据大小',
        'nginx_transfer_phase': 'Nginx传输慢时检查配置和客户端网络',

        # 组合时间指标
        'backend_total_phase': '后端总时间长时重点优化处理逻辑',
        'network_phase': '网络时间长时检查基础设施',
        'processing_phase': '处理时间长时优化业务算法',
        'transfer_phase': '传输时间长时优化数据大小和压缩',

        # 传输大小指标
        'response_body_size_kb': '响应过大时考虑数据压缩和分页',
        'total_bytes_sent_kb': '总传输量大时优化响应头和压缩',

        # 性能效率指标
        'backend_efficiency': '效率低时说明网络或传输占比过高',
        'network_overhead': '开销高时优化网络配置',
        'transfer_ratio': '传输占比高时优化数据大小',
        'connection_cost_ratio': '连接成本高时考虑连接池优化',
        'processing_efficiency_index': '指数低时综合优化各个环节',

        # 传输速度指标
        'response_transfer_speed': '速度慢时检查网络带宽',
        'total_transfer_speed': '总速度慢时优化网络和压缩',
        'nginx_transfer_speed': 'Nginx速度慢时调整配置参数'
    }


def get_metric_unit(metric_name):
    """获取指标单位"""
    if 'size_kb' in metric_name:
        return 'KB'
    elif 'speed' in metric_name:
        return 'KB/s'
    elif any(x in metric_name for x in ['efficiency', 'overhead', 'ratio', 'index']):
        return '%'
    else:
        return '秒'


def process_numeric_metrics(chunk, metrics_config):
    """处理数值指标，基于新的配置结构"""
    # 收集所有需要处理的指标
    all_metrics = []
    for metric_group in metrics_config.values():
        all_metrics.extend(metric_group)

    # 转换为数值类型
    for metric in all_metrics:
        if metric in chunk.columns:
            chunk[metric] = pd.to_numeric(chunk[metric], errors='coerce')

    # 清理内存
    del all_metrics
    gc.collect()

    return chunk


def collect_metric_data(group, metric_stats_dict, time_key):
    """收集指标数据，基于新的指标体系"""
    metrics_config = get_metrics_configuration()

    # 收集所有指标
    all_metrics = []
    for metric_group in metrics_config.values():
        all_metrics.extend(metric_group)

    for metric in all_metrics:
        if metric in group.columns:
            values = group[metric].dropna()
            if len(values) > 0:
                metric_stats_dict[time_key][metric] = values.tolist()
                del values
                gc.collect()

    del all_metrics


# 保持原有的其他函数不变，但进行内存优化
def initialize_stats_containers():
    """初始化统计容器"""
    dimensions = ['daily', 'hourly', 'minute', 'second']
    containers = {}
    for dim in dimensions:
        containers[dim] = {
            'stats': defaultdict(lambda: defaultdict(int)),
            'metric_stats': defaultdict(lambda: defaultdict(list)),
            'requests': defaultdict(list),
            'arrival_requests': defaultdict(list)
        }
    return containers


def calculate_average_metrics(metric_stats):
    """计算平均指标"""
    avg_metrics = {}
    for time_key, metrics in metric_stats.items():
        avg_metrics[time_key] = {
            metric: np.mean(np.array(values))
            for metric, values in metrics.items()
            if values
        }
    return avg_metrics


def calculate_all_time_dimensions(stats_containers):
    """计算所有时间维度的QPS - 使用向量化运算"""
    window_seconds = {'daily': 86400, 'hourly': 3600, 'minute': 60, 'second': 1}
    
    for dimension, seconds in window_seconds.items():
        stats = stats_containers[dimension]['stats']
        for time_key, metrics in stats.items():
            success_requests = metrics.get('success_requests', 0)
            qps = success_requests / seconds
            stats[time_key]['qps'] = qps


def calculate_avg_metrics_for_all_dimensions(stats_containers):
    """计算所有维度的平均指标"""
    avg_metrics = {}
    for dimension in stats_containers:
        avg_metrics[dimension] = calculate_average_metrics(stats_containers[dimension]['metric_stats'])
    return avg_metrics


def identify_peak_period(stats_dict, metric_key):
    """识别峰值时段"""
    if not stats_dict:
        return None
    return max(stats_dict.items(), key=lambda x: x[1].get(metric_key, 0))[0]


def prepare_output_filename(output_path, specific_uri_list):
    """准备输出文件名"""
    if not specific_uri_list:
        return output_path

    base_name = os.path.basename(output_path)
    name_parts = os.path.splitext(base_name)

    if isinstance(specific_uri_list, list):
        uri_identifier = specific_uri_list[0].replace('/', '_').replace('-', '_')
    else:
        uri_identifier = specific_uri_list.replace('/', '_').replace('-', '_')

    if len(uri_identifier) > 30:
        uri_identifier = uri_identifier[:30]

    new_filename = f"{name_parts[0]}_{uri_identifier}{name_parts[1]}"
    return os.path.join(os.path.dirname(output_path), new_filename)


def calculate_qps(stats_dict, seconds_per_unit):
    """计算QPS"""
    for time_key, metrics in stats_dict.items():
        success_requests = metrics.get('success_requests', 0)
        qps = success_requests / seconds_per_unit
        stats_dict[time_key]['qps'] = qps


def create_time_dimension_excel(
        output_path, daily_stats, hourly_stats, minute_stats, second_stats,
        total_records, daily_avg_metrics, hourly_avg_metrics,
        minute_avg_metrics, second_avg_metrics, peak_hour=None, peak_minute=None
):
    """创建时间维度Excel报告 - 采用流式保存方式"""
    wb = Workbook(write_only=True)  # 使用只写模式降低内存占用

    # 创建概览页面
    ws_overview = wb.create_sheet(title="概览")
    add_overview_content(ws_overview, daily_stats, hourly_stats, total_records, peak_hour)

    # 创建各维度页面
    dimensions = [
        ('日期维度', daily_stats, '日期', daily_avg_metrics, None),
        ('小时维度', hourly_stats, '时间', hourly_avg_metrics, None),
        ('分钟维度', minute_stats, '时间', minute_avg_metrics, peak_minute),
        ('秒维度', second_stats, '时间', second_avg_metrics, peak_minute)
    ]

    for sheet_name, stats, time_label, avg_metrics, peak_time in dimensions:
        ws = wb.create_sheet(title=sheet_name)
        data = prepare_time_dimension_data(stats, time_label, avg_metrics)
        add_dataframe_with_paging(ws, data)
        add_charts(ws, data, time_label, peak_time)

    # 流式保存文件
    wb.save(output_path)
    wb.close()


def add_time_dimension_sheet_with_paging(wb, sheet_name, stats_dict, time_label, avg_metrics=None, peak_time=None):
    """添加时间维度页面（带分页处理）"""
    log_info(f"处理{sheet_name}数据", True)

    df = prepare_time_dimension_data(stats_dict, time_label, avg_metrics)
    header_groups = create_header_groups(df, time_label)

    ws = add_dataframe_to_excel_with_grouped_headers(wb, df, sheet_name, header_groups)

    apply_conditional_formatting(ws, df, header_groups, 2)

    add_dimension_charts(ws, df, sheet_name, time_label, header_groups, peak_time)

    # 清理内存
    del df, header_groups
    gc.collect()

    return ws


def get_chart_columns(header_groups, target_headers):
    """获取图表列索引"""
    cols = []
    series_names = []
    col_idx = 1

    for group_name, subheaders in header_groups.items():
        for subheader in subheaders:
            if subheader in target_headers:
                cols.append(col_idx)
                series_names.append(subheader)
            col_idx += 1

    return cols, series_names


def get_chart_columns_by_pattern(header_groups, pattern):
    """根据模式获取图表列索引"""
    cols = []
    series_names = []
    col_idx = 1

    for group_name, subheaders in header_groups.items():
        for subheader in subheaders:
            if pattern(subheader):
                cols.append(col_idx)
                series_names.append(subheader)
            col_idx += 1

    return cols, series_names


def add_overview_sheet(wb, daily_stats, hourly_stats, active_total, peak_hour):
    """添加概览页面"""
    ws = wb.create_sheet(title="概览")

    # 计算总体统计
    total_requests = sum(stats['total_requests'] for stats in daily_stats.values())
    success_requests = sum(stats['success_requests'] for stats in daily_stats.values())
    total_slow_requests = sum(stats['slow_requests'] for stats in daily_stats.values())
    success_ratio = (success_requests / total_requests * 100) if total_requests > 0 else 0
    slow_ratio = (total_slow_requests / total_requests * 100) if total_requests > 0 else 0

    max_concurrent = max(
        [stats.get('concurrent_connections', 0) for stats in hourly_stats.values()]) if hourly_stats else 0
    busiest_hour = max(hourly_stats.items(), key=lambda x: x[1]['success_requests']) if hourly_stats else (
        "N/A", {'success_requests': 0})

    # 计算最差时段
    hourly_slow_ratios = {}
    for hour, stats in hourly_stats.items():
        total = stats.get('total_requests', 0)
        slow = stats.get('slow_requests', 0)
        if total > 0:
            hourly_slow_ratios[hour] = (slow / total) * 100
    worst_hour = max(hourly_slow_ratios.items(), key=lambda x: x[1]) if hourly_slow_ratios else ("N/A", 0)

    # 清理内存
    del hourly_slow_ratios
    gc.collect()

    # 添加标题和说明
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = ws.cell(row=1, column=1,
                         value=f"Nginx HTTP请求生命周期分析报告 - 新参数体系")
    title_cell.font = Font(bold=True, size=14, color="FF0000")
    title_cell.fill = PatternFill("solid", fgColor="FFFF00")

    # 添加参数体系说明
    row = 3
    ws.cell(row=row, column=1, value="新参数体系说明").font = Font(bold=True, size=12)
    row += 1

    # 参数分类说明
    param_categories = [
        ("基础时间参数", "请求总时长、后端响应时长、后端处理时长、后端连接时长"),
        ("阶段时间参数", "后端连接阶段、后端处理阶段、后端传输阶段、Nginx传输阶段"),
        ("组合分析参数", "后端总阶段、网络传输阶段、纯处理阶段、纯传输阶段"),
        ("性能比率参数", "后端处理效率、网络开销占比、传输时间占比、连接成本占比"),
        ("传输分析参数", "响应体大小(KB)、总传输大小(KB)、各阶段传输速度(KB/s)")
    ]

    for category, description in param_categories:
        ws.cell(row=row, column=1, value=category).font = Font(bold=True)
        ws.cell(row=row, column=2, value=description)
        row += 1

    row += 2

    # 总体统计
    ws.cell(row=row, column=1, value="总体统计").font = Font(bold=True, size=14)
    row += 2

    stats_data = [
        ("总请求数", total_requests),
        ("成功请求数", success_requests),
        ("成功请求占比(%)", f"{success_ratio:.2f}"),
        (f"慢请求数 (>{DEFAULT_SLOW_THRESHOLD}秒)", total_slow_requests),
        ("慢请求占比(%)", f"{slow_ratio:.2f}"),
        ("最大并发连接数", max_concurrent),
        ("活跃连接总数", active_total)
    ]

    for label, value in stats_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # 清理内存
    del stats_data
    gc.collect()

    # 峰值时段信息
    row += 2
    ws.cell(row=row, column=1, value="最繁忙时段").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="时间")
    ws.cell(row=row, column=2, value="请求数")
    row += 1
    ws.cell(row=row, column=1, value=busiest_hour[0])
    ws.cell(row=row, column=2, value=busiest_hour[1]['success_requests'])

    row += 2
    ws.cell(row=row, column=1, value="慢请求占比最高时段").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="时间")
    ws.cell(row=row, column=2, value="慢请求占比(%)")
    row += 1
    ws.cell(row=row, column=1, value=worst_hour[0])
    ws.cell(row=row, column=2, value=f"{worst_hour[1]:.2f}")

    # 日期分布详情
    if daily_stats:
        row += 2
        ws.cell(row=row, column=1, value="日期请求量分布").font = Font(bold=True)
        row += 1

        headers = ["日期", "总请求数", "成功请求数", "慢请求数", "慢请求占比(%)", "QPS"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)

        row += 1
        for date in sorted(daily_stats.keys()):
            stats = daily_stats[date]
            total = stats.get('total_requests', 0)
            success = stats.get('success_requests', 0)
            slow = stats.get('slow_requests', 0)
            slow_ratio = (slow / total * 100) if total > 0 else 0
            qps = stats.get('qps', 0)

            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=success)
            ws.cell(row=row, column=4, value=slow)
            ws.cell(row=row, column=5, value=f"{slow_ratio:.2f}")
            ws.cell(row=row, column=6, value=f"{qps:.2f}")
            row += 1

    format_excel_sheet(ws)


def apply_conditional_formatting(ws, df, header_groups, header_start_row):
    """应用条件格式"""
    from openpyxl.utils import get_column_letter

    # 慢请求占比条件格式
    slow_ratio_col = None
    col_idx = 1

    for group_name, subheaders in header_groups.items():
        for subheader in subheaders:
            if subheader == '慢请求占比(%)':
                slow_ratio_col = col_idx
            col_idx += 1

    if slow_ratio_col:
        col_letter = get_column_letter(slow_ratio_col)
        start_row = header_start_row + 1
        end_row = start_row + len(df) - 1

        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type='min', start_color='00FF00',
                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                end_type='max', end_color='FF0000'
            )
        )

    # QPS条件格式
    qps_col = None
    col_idx = 1
    for group_name, subheaders in header_groups.items():
        for subheader in subheaders:
            if subheader == 'QPS':
                qps_col = col_idx
            col_idx += 1

    if qps_col:
        col_letter = get_column_letter(qps_col)
        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type='min', start_color='FFFFFF',
                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                end_type='max', end_color='FF9900'
            )
        )
    # 总请求时长条件格式
    time_cols = []
    col_idx = 1
    for group_name, subheaders in header_groups.items():
        for subheader in subheaders:
            if '平均' in subheader and '请求总时长' in subheader:
                time_cols.append(col_idx)
            col_idx += 1

    for time_col in time_cols:
        col_letter = get_column_letter(time_col)
        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type='min', start_color='00FF00',
                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                end_type='max', end_color='FF0000'
            )
        )

    # 清理内存
    del time_cols
    gc.collect()


def preprocess_chunk_data(chunk):
    """预处理数据块"""
    # 转换时间戳列
    timestamp_cols = ['raw_timestamp', 'arrival_timestamp']
    for col in timestamp_cols:
        if col in chunk.columns:
            chunk[col] = pd.to_numeric(chunk[col], errors='coerce')

    # 转换时间列
    time_cols = ['raw_time', 'arrival_time']
    for col in time_cols:
        if col in chunk.columns:
            chunk[col] = pd.to_datetime(chunk[col], errors='coerce')

    return chunk


def store_dimensional_data(chunk, stats_containers, success_mask, slow_mask):
    """存储维度数据"""
    dimension_mappings = [
        ('date', 'arrival_date', 'daily'),
        ('date_hour', 'arrival_date_hour', 'hourly'),
        ('date_hour_minute', 'arrival_date_hour_minute', 'minute'),
        ('date_hour_minute_second', 'arrival_date_hour_minute_second', 'second')
    ]

    window_seconds = {'daily': 86400, 'hourly': 3600, 'minute': 60, 'second': 1}

    for completion_col, arrival_col, container_key in dimension_mappings:
        seconds = window_seconds[container_key]

        # 处理到达时间分组
        if arrival_col in chunk.columns:
            arrival_groups = chunk.groupby(arrival_col, sort=False)
            for time_key, group in arrival_groups:
                if pd.notna(time_key):
                    stats_containers[container_key]['stats'][time_key]['new_connections'] = len(group)
            del arrival_groups
            gc.collect()

        # 处理完成时间分组
        if completion_col in chunk.columns:
            completion_groups = chunk.groupby(completion_col, sort=False)
            for time_key, group in completion_groups:
                if pd.notna(time_key):
                    group_indices = group.index
                    total_count = len(group)
                    success_count = success_mask[group_indices].sum()
                    slow_count = slow_mask[group_indices].sum()

                    stats_dict = stats_containers[container_key]['stats']
                    stats_dict[time_key]['total_requests'] = total_count
                    stats_dict[time_key]['success_requests'] = success_count
                    stats_dict[time_key]['slow_requests'] = slow_count
                    stats_dict[time_key]['slow_ratio'] = (slow_count / total_count) * 100 if total_count > 0 else 0

                    calculate_connection_metrics(group, stats_dict[time_key], time_key, seconds)
                    collect_metric_data(group, stats_containers[container_key]['metric_stats'], time_key)

                    del group_indices
                    gc.collect()

            del completion_groups
            gc.collect()

        log_info(f"{container_key} 维度数据处理完成")


def calculate_connection_metrics(group, stats_dict, time_key, window_seconds):
    """计算连接指标"""
    window_start = pd.to_datetime(time_key)
    window_end = window_start + pd.Timedelta(seconds=window_seconds)
    window_start_ts = window_start.timestamp()
    window_end_ts = window_end.timestamp()

    concurrent_mask = (group['arrival_timestamp'] < window_end_ts) & (group['raw_timestamp'] >= window_end_ts)
    active_mask = (group['arrival_timestamp'] < window_end_ts) & (group['raw_timestamp'] < window_start_ts)

    stats_dict['concurrent_connections'] = concurrent_mask.sum()
    stats_dict['active_connections'] = active_mask.sum()

    del concurrent_mask, active_mask
    gc.collect()


def read_and_process_data(csv_path, specific_uri_list, metrics_config, stats_containers):
    """读取和处理数据 - 优化版本"""
    chunk_size = DEFAULT_CHUNK_SIZE * 2  # 增加块大小以提高吞吐量
    total_records = 0
    success_requests = 0
    chunks_processed = 0
    slow_threshold = DEFAULT_SLOW_THRESHOLD
    all_metrics = _get_all_metrics(metrics_config)
    dimension_config = _get_dimension_config()

    log_info(f"开始分析{'所有请求' if specific_uri_list is None else f'特定URI: {specific_uri_list}'}")

    import time
    start_time = time.time()
    chunk_generator = pd.read_csv(csv_path, chunksize=chunk_size)

    for chunk in chunk_generator:
        chunks_processed += 1

        # URI过滤
        if specific_uri_list:
            chunk = _filter_chunk_by_uri(chunk, specific_uri_list)
            if chunk.empty:
                continue

        # 预处理数据
        chunk = _batch_preprocess_chunk(chunk, all_metrics, slow_threshold)
        # 重置索引确保连续性
        chunk = chunk.reset_index(drop=True)

        # 创建mask（直接使用numpy数组）
        success_mask = ((chunk['response_status_code'] >= 200) & (chunk['response_status_code'] < 400)).values
        slow_mask = (chunk['total_request_duration'] > slow_threshold).values

        chunk_record_count = len(chunk)
        chunk_success_count = np.sum(success_mask)

        total_records += chunk_record_count
        success_requests += chunk_success_count

        if chunk_record_count > 0:
            _store_dimensional_data_optimized(chunk, stats_containers, success_mask, slow_mask, dimension_config)

        # 内存管理
        del chunk, success_mask, slow_mask
        gc.collect()

        if chunks_processed % 50 == 0:  # 调整GC频率，减少性能开销
            elapsed = time.time() - start_time
            log_info(f"已处理 {chunks_processed} 个数据块, {total_records} 条记录, 耗时: {elapsed:.2f}秒")

    elapsed = time.time() - start_time
    log_info(
        f"数据处理完成，共处理 {chunks_processed} 个数据块, {total_records} 条记录, 成功请求 {success_requests} 条, 总耗时: {elapsed:.2f}秒"
    )

    return total_records, success_requests


def _get_all_metrics(metrics_config):
    """获取所有指标"""
    all_metrics = []
    for metric_group in metrics_config.values():
        all_metrics.extend(metric_group)
    return all_metrics


def _get_dimension_config():
    """获取维度配置"""
    return {
        'daily': 'arrival_date',
        'hourly': 'arrival_date_hour',
        'minute': 'arrival_date_hour_minute',
        'second': 'arrival_date_hour_minute_second'
    }


def _batch_preprocess_chunk(chunk, all_metrics, slow_threshold):
    """批量预处理数据块 - 优化版本"""
    if chunk.empty:
        return chunk

    # 转换时间戳列
    timestamp_cols = ['raw_timestamp', 'arrival_timestamp']
    for col in timestamp_cols:
        if col in chunk.columns:
            chunk[col] = pd.to_numeric(chunk[col], errors='coerce')

    # 转换时间列
    time_cols = ['raw_time', 'arrival_time']
    for col in time_cols:
        if col in chunk.columns:
            chunk[col] = pd.to_datetime(chunk[col], errors='coerce').dt.tz_localize(None)  # 移除时区信息

    # 转换指标列并转换为category类型
    for metric in all_metrics:
        if metric in chunk.columns:
            chunk[metric] = pd.to_numeric(chunk[metric], errors='coerce')

    # 删除全空列
    chunk = chunk.dropna(axis=1, how='all')

    return chunk


def _filter_chunk_by_uri(chunk, specific_uri_list):
    """按URI过滤数据块 - 优化版本"""
    if chunk.empty or 'request_full_uri' not in chunk.columns:
        return pd.DataFrame()

    uri_set = set(specific_uri_list) if isinstance(specific_uri_list, list) else {specific_uri_list}
    filtered = chunk[chunk['request_full_uri'].isin(uri_set)]
    return filtered


def _store_dimensional_data_optimized(chunk, stats_containers, success_mask, slow_mask, dimension_config):
    """存储维度数据 - 优化版本"""
    for dimension, arrival_col in dimension_config.items():
        if arrival_col in chunk.columns:
            stats = stats_containers[dimension]['stats']
            metric_stats = stats_containers[dimension]['metric_stats']

            # 使用groupby进行分组
            grouped = chunk.groupby(arrival_col, sort=False)

            for time_key, group in grouped:
                if pd.notna(time_key):
                    group_indices = group.index
                    total_count = len(group)
                    success_count = success_mask[group_indices].sum()
                    slow_count = slow_mask[group_indices].sum()

                    stats[time_key]['total_requests'] = total_count
                    stats[time_key]['success_requests'] = success_count
                    stats[time_key]['slow_requests'] = slow_count
                    stats[time_key]['slow_ratio'] = (slow_count / total_count) * 100 if total_count > 0 else 0

                    calculate_connection_metrics(group, stats[time_key], time_key, 86400)
                    collect_metric_data(group, metric_stats, time_key)

                    del group_indices
                    gc.collect()

            del grouped
            gc.collect()


def add_overview_content(ws, daily_stats, hourly_stats, total_records, peak_hour):
    """添加概览页面内容"""
    # 添加表头
    ws.append([
        '统计维度', '成功请求数', '总请求数', '慢请求数', 'QPS'
    ])

    # 添加每日数据
    for date_key, stats in daily_stats.items():
        ws.append([
            date_key,
            stats.get('success_requests', 0),
            stats.get('total_requests', 0),
            stats.get('slow_requests', 0),
            stats.get('qps', 0)
        ])

    # 添加汇总信息
    ws.append([
        '总计',
        sum(stats.get('success_requests', 0) for stats in daily_stats.values()),
        total_records,
        sum(stats.get('slow_requests', 0) for stats in daily_stats.values()),
        sum(stats.get('qps', 0) for stats in daily_stats.values()) / len(daily_stats) if daily_stats else 0
    ])


def add_dataframe_with_paging(ws, data):
    """分页添加数据到工作表"""
    from openpyxl.utils.dataframe import dataframe_to_rows  # 导入缺失的依赖

    page_size = 1048576  # Excel最大行数限制
    total_rows = len(data)

    for start_row in range(0, total_rows, page_size):
        end_row = min(start_row + page_size, total_rows)
        page_data = data.iloc[start_row:end_row]

        for row in dataframe_to_rows(page_data, index=False, header=(start_row == 0)):
            ws.append(row)


def add_charts(ws, data, time_label, peak_time):
    """为工作表添加图表"""
    from openpyxl.chart import BarChart, LineChart, Reference

    chart = BarChart()
    chart.title = f"{time_label} 统计分析"
    chart.style = 10
    chart.x_axis.title = time_label
    chart.y_axis.title = '请求量'

    data_ref = Reference(ws, min_col=2, min_row=2, max_col=5, max_row=len(data) + 1)
    categories_ref = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)

    ws.add_chart(chart, "F2")
