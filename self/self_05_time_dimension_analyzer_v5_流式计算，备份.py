import gc
import os
import logging
from collections import defaultdict, Counter
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# 假设这些常量已定义
DEFAULT_CHUNK_SIZE = 10000
DEFAULT_SLOW_THRESHOLD = 5.0

class StreamingTimeAnalyzer:
    """流式时间维度分析器 - 一次遍历完成所有统计"""
    
    def __init__(self, slow_threshold=DEFAULT_SLOW_THRESHOLD):
        self.slow_threshold = slow_threshold
        self.stats = {
            'daily': {},
            'hourly': {},
            'minute': {},
            'second': {}
        }
        # 使用更高效的数据结构存储指标
        self.metrics_sum = {
            'daily': defaultdict(lambda: defaultdict(float)),
            'hourly': defaultdict(lambda: defaultdict(float)),
            'minute': defaultdict(lambda: defaultdict(float)),
            'second': defaultdict(lambda: defaultdict(float))
        }
        self.metrics_count = {
            'daily': defaultdict(lambda: defaultdict(int)),
            'hourly': defaultdict(lambda: defaultdict(int)),
            'minute': defaultdict(lambda: defaultdict(int)),
            'second': defaultdict(lambda: defaultdict(int))
        }
        
        # 预定义指标列表
        self.time_metrics = [
            'total_request_duration', 'upstream_response_time', 
            'upstream_header_time', 'upstream_connect_time',
            'backend_connect_phase', 'backend_process_phase',
            'backend_transfer_phase', 'nginx_transfer_phase'
        ]
        
        self.size_metrics = [
            'response_body_size_kb', 'total_bytes_sent_kb'
        ]

    def extract_time_keys(self, timestamp):
        """一次性提取所有时间维度的键，处理无效时间戳"""
        if pd.isna(timestamp):
            return None

        try:
            dt = pd.to_datetime(timestamp)
            if pd.isna(dt):  # 检查转换后是否为NaT
                return None

            return {
                'daily': dt.strftime('%Y-%m-%d'),
                'hourly': dt.strftime('%Y-%m-%d %H:00'),
                'minute': dt.strftime('%Y-%m-%d %H:%M'),
                'second': dt.strftime('%Y-%m-%d %H:%M:%S')
            }
        except (ValueError, TypeError):
            return None

    def process_single_record(self, record):
        """处理单条记录，更新所有维度统计"""
        # 提取时间键，跳过无效时间戳
        time_keys = self.extract_time_keys(record.get('arrival_time'))
        if not time_keys:
            return  # 跳过无效记录

        # 安全获取数值，避免NaN
        status_code = record.get('response_status_code')
        duration = record.get('total_request_duration')

        # 判断请求状态
        is_success = (200 <= status_code < 400) if pd.notna(status_code) else False
        is_slow = (duration > self.slow_threshold) if pd.notna(duration) else False

        # 更新所有维度的统计
        for dimension, time_key in time_keys.items():
            stats = self.stats[dimension]

            # 初始化统计结构
            if time_key not in stats:
                stats[time_key] = {
                    'total_requests': 0,
                    'success_requests': 0,
                    'slow_requests': 0,
                    'new_connections': 0
                }

            # 更新基础统计
            stats[time_key]['total_requests'] += 1
            if is_success:
                stats[time_key]['success_requests'] += 1
            if is_slow:
                stats[time_key]['slow_requests'] += 1

            # 更新指标统计
            self._update_metrics(dimension, time_key, record)

    def _update_metrics(self, dimension, time_key, record):
        """更新指标统计，安全处理数值"""
        # 时间指标
        for metric in self.time_metrics:
            value = record.get(metric)
            if pd.notna(value) and value > 0:  # 同时检查非空和正值
                self.metrics_sum[dimension][time_key][metric] += float(value)
                self.metrics_count[dimension][time_key][metric] += 1

        # 大小指标
        for metric in self.size_metrics:
            value = record.get(metric)
            if pd.notna(value) and value >= 0:  # 大小可以为0
                self.metrics_sum[dimension][time_key][metric] += float(value)
                self.metrics_count[dimension][time_key][metric] += 1

    def calculate_averages(self):
        """计算平均值"""
        avg_metrics = {}
        for dimension in self.stats.keys():
            avg_metrics[dimension] = {}
            for time_key in self.stats[dimension].keys():
                avg_metrics[dimension][time_key] = {}
                for metric in self.time_metrics + self.size_metrics:
                    total = self.metrics_sum[dimension][time_key][metric]
                    count = self.metrics_count[dimension][time_key][metric]
                    avg_metrics[dimension][time_key][metric] = total / count if count > 0 else 0
        return avg_metrics

    def calculate_qps(self):
        """计算QPS"""
        window_seconds = {'daily': 86400, 'hourly': 3600, 'minute': 60, 'second': 1}
        
        for dimension, seconds in window_seconds.items():
            for time_key, stats in self.stats[dimension].items():
                success_requests = stats.get('success_requests', 0)
                stats['qps'] = success_requests / seconds


def read_and_process_data_streaming(csv_path, specific_uri_list=None):
    """流式处理数据，增强错误处理"""
    analyzer = StreamingTimeAnalyzer()
    total_records = 0
    processed_records = 0
    error_records = 0

    chunk_size = 1000
    log_info(f"开始流式分析{'所有请求' if specific_uri_list is None else f'特定URI: {specific_uri_list}'}")

    import time
    start_time = time.time()

    try:
        # 流式读取并处理
        for chunk_idx, chunk in enumerate(pd.read_csv(csv_path, chunksize=chunk_size)):
            total_records += len(chunk)

            # URI过滤
            if specific_uri_list:
                if 'request_full_uri' in chunk.columns:
                    uri_set = set(specific_uri_list) if isinstance(specific_uri_list, list) else {specific_uri_list}
                    chunk = chunk[chunk['request_full_uri'].isin(uri_set)]
                if chunk.empty:
                    continue

            # 预处理数据类型
            try:
                chunk = _preprocess_chunk_fast(chunk)
            except Exception as e:
                log_info(f"Chunk {chunk_idx} 预处理失败: {e}")
                error_records += len(chunk)
                continue

            # 逐行处理
            for _, record in chunk.iterrows():
                try:
                    analyzer.process_single_record(record)
                    processed_records += 1
                except Exception as e:
                    error_records += 1
                    if error_records <= 10:  # 只记录前10个错误
                        log_info(f"记录处理错误: {e}")

            # 每处理10000条记录输出进度
            if total_records % 10000 == 0:
                log_info(f"已处理 {total_records} 条记录...")

    except Exception as e:
        log_info(f"数据读取错误: {e}")
        raise

    # 计算衍生指标
    analyzer.calculate_qps()
    avg_metrics = analyzer.calculate_averages()

    elapsed = time.time() - start_time
    log_info(f"流式处理完成")
    log_info(f"总记录数: {total_records}, 成功处理: {processed_records}, 错误记录: {error_records}")
    log_info(f"耗时: {elapsed:.2f}秒, 处理速度: {processed_records / elapsed:.0f} 记录/秒")

    return analyzer.stats, avg_metrics, processed_records


def _preprocess_chunk_fast(chunk):
    """快速预处理chunk数据，增强数据验证"""
    # 检查必要列是否存在
    required_cols = ['arrival_time']
    missing_cols = [col for col in required_cols if col not in chunk.columns]
    if missing_cols:
        raise ValueError(f"缺少必要列: {missing_cols}")

    # 数据类型转换配置
    essential_cols = {
        'arrival_time': 'datetime64[ns]',
        'response_status_code': 'int32',
        'total_request_duration': 'float32',
        'upstream_response_time': 'float32',
        'upstream_header_time': 'float32',
        'upstream_connect_time': 'float32',
        'response_body_size': 'float32',
        'total_bytes_sent': 'float32'
    }

    for col, dtype in essential_cols.items():
        if col in chunk.columns:
            try:
                if dtype == 'datetime64[ns]':
                    chunk[col] = pd.to_datetime(chunk[col], errors='coerce')
                else:
                    chunk[col] = pd.to_numeric(chunk[col], errors='coerce').astype(dtype)
            except Exception as e:
                log_info(f"列 {col} 转换失败: {e}")
                continue

    return chunk

def analyze_time_dimension_optimized(csv_path, output_path, specific_uri_list=None):
    """优化后的时间维度分析主函数"""
    output_filename = prepare_output_filename(output_path, specific_uri_list)
    
    # 使用流式处理
    stats_containers, avg_metrics, total_records = read_and_process_data_streaming(csv_path, specific_uri_list)
    
    # 识别峰值时段
    peak_hour = identify_peak_period(stats_containers['hourly'], 'success_requests')
    peak_minute = identify_peak_period(stats_containers['minute'], 'success_requests')
    
    # 生成Excel报告
    create_time_dimension_excel_optimized(
        output_filename,
        stats_containers,
        avg_metrics,
        total_records,
        peak_hour,
        peak_minute
    )
    
    log_info(f"报告已生成：{output_filename}")
    return output_filename

def identify_peak_period(stats_dict, metric_key):
    """识别峰值时段"""
    if not stats_dict:
        return None
    return max(stats_dict.items(), key=lambda x: x[1].get(metric_key, 0))[0]

def prepare_output_filename(output_path, specific_uri_list):
    """准备输出文件名"""
    if not specific_uri_list:
        return output_path
    
    base_name = os.path.basename(output_path)
    name_parts = os.path.splitext(base_name)
    
    if isinstance(specific_uri_list, list):
        uri_identifier = specific_uri_list[0].replace('/', '_').replace('-', '_')
    else:
        uri_identifier = specific_uri_list.replace('/', '_').replace('-', '_')
    
    if len(uri_identifier) > 30:
        uri_identifier = uri_identifier[:30]
    
    new_filename = f"{name_parts[0]}_{uri_identifier}{name_parts[1]}"
    return os.path.join(os.path.dirname(output_path), new_filename)

def create_time_dimension_excel_optimized(output_path, stats_containers, avg_metrics, total_records, peak_hour, peak_minute):
    """优化后的Excel生成"""
    wb = Workbook()
    
    # 移除默认工作表
    wb.remove(wb.active)
    
    # 创建概览页
    create_overview_sheet(wb, stats_containers, total_records, peak_hour)
    
    # 创建各维度页面
    dimensions = [
        ('日期维度', 'daily', '日期'),
        ('小时维度', 'hourly', '时间'),
        ('分钟维度', 'minute', '时间'),
        ('秒维度', 'second', '时间')
    ]
    
    for sheet_name, dimension, time_label in dimensions:
        create_dimension_sheet(wb, sheet_name, stats_containers[dimension], avg_metrics[dimension], time_label)
    
    wb.save(output_path)
    wb.close()

def create_overview_sheet(wb, stats_containers, total_records, peak_hour):
    """创建概览页"""
    ws = wb.create_sheet(title="概览")
    
    # 计算总体统计
    daily_stats = stats_containers['daily']
    total_success = sum(stats.get('success_requests', 0) for stats in daily_stats.values())
    total_slow = sum(stats.get('slow_requests', 0) for stats in daily_stats.values())
    
    # 添加标题
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Nginx HTTP请求生命周期分析报告 - 优化版"
    title_cell.font = Font(bold=True, size=14, color="FF0000")
    title_cell.fill = PatternFill("solid", fgColor="FFFF00")
    
    # 添加统计数据
    row = 3
    stats_data = [
        ("总请求数", total_records),
        ("成功请求数", total_success),
        ("成功率(%)", f"{total_success/total_records*100:.2f}" if total_records > 0 else "0"),
        ("慢请求数", total_slow),
        ("慢请求率(%)", f"{total_slow/total_records*100:.2f}" if total_records > 0 else "0"),
        ("峰值时段", peak_hour or "N/A")
    ]
    
    for label, value in stats_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

def create_dimension_sheet(wb, sheet_name, stats_dict, avg_metrics, time_label):
    """创建维度分析页"""
    ws = wb.create_sheet(title=sheet_name)
    
    # 准备数据
    data = prepare_dimension_data(stats_dict, avg_metrics, time_label)
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 格式化
    if data:
        # 设置表头格式
        for col in range(1, len(data[0]) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

def prepare_dimension_data(stats_dict, avg_metrics, time_label):
    """准备维度数据"""
    if not stats_dict:
        return []
    
    # 表头
    headers = [
        time_label, '总请求数', '成功请求数', '慢请求数', '慢请求率(%)', 'QPS',
        '平均响应时间(s)', '平均处理时间(s)', '平均连接时间(s)'
    ]
    
    data = [headers]
    
    # 数据行
    for time_key in sorted(stats_dict.keys()):
        stats = stats_dict[time_key]
        avg_stats = avg_metrics.get(time_key, {})
        
        total = stats.get('total_requests', 0)
        slow = stats.get('slow_requests', 0)
        slow_rate = (slow / total * 100) if total > 0 else 0
        
        row = [
            time_key,
            total,
            stats.get('success_requests', 0),
            slow,
            f"{slow_rate:.2f}",
            f"{stats.get('qps', 0):.2f}",
            f"{avg_stats.get('total_request_duration', 0):.3f}",
            f"{avg_stats.get('upstream_header_time', 0):.3f}",
            f"{avg_stats.get('upstream_connect_time', 0):.3f}"
        ]
        data.append(row)
    
    return data

def log_info(message, force_print=False):
    """日志输出"""
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [INFO] {message}")

# 主函数保持兼容性
def analyze_time_dimension(csv_path, output_path, specific_uri_list=None):
    """主分析函数 - 兼容原接口"""
    return analyze_time_dimension_optimized(csv_path, output_path, specific_uri_list)