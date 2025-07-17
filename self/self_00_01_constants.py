from openpyxl.styles import PatternFill
import os

DEFAULT_LOG_DIR = "../data/demo/自研Ng2025.05.09日志-样例"
#DEFAULT_LOG_DIR = "D:/temp/gaokao0625"
DEFAULT_CHUNK_SIZE = 50000  # 减小chunk大小以优化内存使用
DEFAULT_BATCH_SIZE = 5000   # 减小批处理大小
DEFAULT_SLOW_THRESHOLD = 3 # 慢接口时间阈值
DEFAULT_SLOW_REQUESTS_THRESHOLD = 0.1 # 默认接口超时数量比例，即为慢接口
DEFAULT_SUCCESS_CODES = ['200']
DEFAULT_COLUMN_API = ["scmp-gateway/wxxcx/prod-api/portals/affairsGuide/search", "/scmp-gateway/column/getColumnSource1", "/scmp-gateway/gxrz-rest/newUser/pwdFreeReg", "/scmp-gateway/gxrz-rest/newUser/loginByVerificationCode", "/scmp-gateway/gxrz-rest/newUser/queryByCardOrPhoneAes", "/scmp-gateway/gxrz-rest/user/getUserInfoByToken"]
#DEFAULT_COLUMN_API = ["/api/zgt-exam/exam/gxeea_cjcx", "/api/zgt-exam/exam/checkAndSend", "/exam/ptgkcjcx"]
#EFAULT_COLUMN_API = ["/api/zgt-exam/exam/checkAndSend", "/exam/ptgkcjcx"]
EXCEL_MAX_ROWS = 1000000
CHART_MAX_POINTS = 1000

# 日志类型
LOG_TYPE_SELF_DEVELOPED = "自研"
LOG_TYPE_BASE = "底座"
LOG_TYPE_AUTO = "自动识别"

# 默认日期范围过滤（None表示不过滤）
DEFAULT_START_DATE = None  # 格式: "2023-05-17 14:30:25"
DEFAULT_END_DATE = None    # 格式: "2023-05-17 14:30:25"

# Excel样式
HEADER_FILL = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid")
REDIRECT_FILL = PatternFill(start_color="D9E6FF", end_color="D9E6FF", fill_type="solid")

# 度量指标
TIME_METRICS = [
    'request_time',
    'upstream_connect_time',
    'upstream_header_time',
    'upstream_response_time',
    'upstream_connect_phase',
    'upstream_header_phase',
    'upstream_body_phase',
    'client_transfer_phase'
]
SIZE_METRICS = [
    'body_bytes_sent',
    'bytes_sent'
]
TIME_METRICS_MAPPING = {
    'request_time': '请求耗时',
    'upstream_connect_time': '上游连接耗时',
    'upstream_header_time': '上游头部耗时',
    'upstream_response_time': '上游响应耗时',
    'upstream_connect_phase': '上游连接阶段',
    'upstream_header_phase': '上游头部阶段',
    'upstream_body_phase': '上游响应体阶段',
    'client_transfer_phase': '客户端传输阶段'
}
PERCENTILES = [50, 90, 95, 99]
STATUS_CATEGORIES = {
    'success': ['2'],
    'redirect': ['3'],
    'client_error': ['4'],
    'server_error': ['5']
}
STATUS_DESCRIPTIONS = {
    '200': '请求成功',
    '201': '已创建',
    '204': '无内容',
    '301': '永久重定向',
    '302': '临时重定向',
    '304': '未修改',
    '400': '错误请求',
    '401': '未授权',
    '403': '禁止访问',
    '404': '未找到',
    '405': '方法不允许',
    '408': '请求超时',
    '429': '请求过多',
    '500': '服务器错误',
    '502': '网关错误',
    '503': '服务不可用',
    '504': '网关超时'
}

# 预估的HTTP头部大小（用于底座日志中计算bytes_sent）
ESTIMATED_HEADER_SIZE = 200

# 新旧字段映射
FIELD_MAPPING = {
    'log_source_file': 'source',
    'application_name': 'app_name',
    'raw_time': 'time',
    'raw_timestamp': 'timestamp',
    'total_request_duration': 'request_time',
    'client_ip_address': 'client_ip',
    'client_port_number': 'client_port',
    'http_method': 'request_method',
    'request_full_uri': 'request_uri',
    'request_path': 'request_path',
    'query_parameters': 'query_string',
    'http_protocol_version': 'request_protocol',
    'response_status_code': 'status',
    'response_body_size': 'body_bytes_sent',
    'total_bytes_sent': 'bytes_sent',
    'response_content_type': 'content_type',
    'upstream_connect_time': 'upstream_connect_time',
    'upstream_header_time': 'upstream_header_time',
    'upstream_response_time': 'upstream_response_time',
    'upstream_server_address': 'upstream_addr',
    'upstream_status_code': 'upstream_status',
    'server_name': 'server_name',
    'host_header': 'host',
    'user_agent_string': 'user_agent',
    'referer_url': 'referer',
    'service_name': 'service_name',
    'phase_upstream_connect': 'upstream_connect_phase',
    'phase_upstream_header': 'upstream_header_phase',
    'phase_upstream_body': 'upstream_body_phase',
    'phase_client_transfer': 'client_transfer_phase'
}

# 反向映射，便于代码中从新字段名查找对应的旧字段名
REVERSE_FIELD_MAPPING = {v: k for k, v in FIELD_MAPPING.items()}

# 时间指标映射
TIME_METRICS_NEW = [
    'total_request_duration',
    'upstream_connect_time',
    'upstream_header_time',
    'upstream_response_time',
    'phase_upstream_connect',
    'phase_upstream_header',
    'phase_upstream_body',
    'phase_client_transfer'
]

# 大小指标映射
SIZE_METRICS_NEW = [
    'response_body_size',
    'total_bytes_sent'
]