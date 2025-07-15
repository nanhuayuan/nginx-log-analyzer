import gc
import os
import pandas as pd
import numpy as np
from datetime import datetime
from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook
from openpyxl.chart import PieChart, BarChart, Series, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from self_00_01_constants import (
    DEFAULT_CHUNK_SIZE, DEFAULT_SUCCESS_CODES, EXCEL_MAX_ROWS,
    HEADER_FILL, ERROR_FILL, WARNING_FILL, SUCCESS_FILL, REDIRECT_FILL,
    STATUS_CATEGORIES, STATUS_DESCRIPTIONS
)
from self_00_02_utils import log_info
from self_00_04_excel_processor import save_dataframe_to_excel, format_excel_sheet, \
    add_dataframe_to_excel_with_grouped_headers


def analyze_status_codes(csv_path, output_path, slow_request_threshold=3.0):
    """
    优化后的状态码分析函数
    - 适配新的CSV字段结构
    - 整合nginx生命周期分析参数
    - 简化代码逻辑，提高可读性
    """

    # ==================== 字段映射配置 ====================
    # 基础字段映射 - 适配新的CSV结构
    field_mapping = {
        'status': 'response_status_code',
        'app_name': 'application_name',
        'service_name': 'service_name',
        'time': 'raw_time',
        'timestamp': 'raw_timestamp',
        'date': 'date',
        'hour': 'hour',
        'method': 'http_method',
        'request_time': 'total_request_duration',
        'client_ip': 'client_ip_address',
        'request_path': 'request_path',
        'upstream_addr': 'upstream_server_address',
        'upstream_connect_time': 'upstream_connect_time',
        'upstream_header_time': 'upstream_header_time',
        'upstream_response_time': 'upstream_response_time',
        'body_bytes_sent': 'response_body_size_kb',  # 已转换为KB
        'bytes_sent': 'total_bytes_sent_kb'  # 已转换为KB
    }

    # nginx生命周期分析参数映射
    lifecycle_fields = {
        # 核心阶段指标
        'backend_connect_phase': 'backend_connect_phase',
        'backend_process_phase': 'backend_process_phase',
        'backend_transfer_phase': 'backend_transfer_phase',
        'nginx_transfer_phase': 'nginx_transfer_phase',

        # 组合分析指标
        'backend_total_phase': 'backend_total_phase',
        'network_phase': 'network_phase',
        'processing_phase': 'processing_phase',
        'transfer_phase': 'transfer_phase',

        # 性能比率指标(%)
        'backend_efficiency': 'backend_efficiency',
        'network_overhead': 'network_overhead',
        'transfer_ratio': 'transfer_ratio',

        # 传输速度指标
        'response_transfer_speed': 'response_transfer_speed',
        'total_transfer_speed': 'total_transfer_speed',
        'nginx_transfer_speed': 'nginx_transfer_speed',

        # 效率指标
        'connection_cost_ratio': 'connection_cost_ratio',
        'processing_efficiency_index': 'processing_efficiency_index'
    }

    # ==================== 数据收集器初始化 ====================
    collectors = {
        'status_counts': Counter(),
        'app_status_counts': defaultdict(Counter),
        'service_status_counts': defaultdict(Counter),
        'status_time_stats': defaultdict(list),
        'status_lifecycle_stats': defaultdict(lambda: defaultdict(list)),
        'error_details': defaultdict(list),
        'status_by_hour': defaultdict(Counter),
        'status_by_date': defaultdict(Counter),
        'status_method_counts': defaultdict(Counter),
        'slow_requests_by_status': defaultdict(int)
    }

    # 统计变量
    stats = {
        'total_requests': 0,
        'total_slow_requests': 0,
        'chunks_processed': 0
    }

    log_info("开始分析状态码...", True)
    start_time = datetime.now()

    # ==================== 数据处理主循环 ====================
    chunk_size = DEFAULT_CHUNK_SIZE
    reader = pd.read_csv(csv_path, chunksize=chunk_size)

    for chunk in reader:
        stats['chunks_processed'] += 1
        chunk_rows = len(chunk)
        stats['total_requests'] += chunk_rows

        # 处理当前数据块
        _process_chunk(chunk, collectors, stats, field_mapping, lifecycle_fields, slow_request_threshold)

        # 内存清理和进度报告
        if stats['chunks_processed'] % 5 == 0:
            elapsed = (datetime.now() - start_time).total_seconds()
            log_info(f"已处理 {stats['chunks_processed']} 个数据块, {stats['total_requests']} 条记录, 耗时: {elapsed:.2f}秒",
                     show_memory=True)
            del chunk
            gc.collect()

    # ==================== 生成分析报告 ====================
    log_info("生成状态码统计报告...", True)

    # 生成各类分析表
    dataframes = _generate_analysis_dataframes(collectors, stats, slow_request_threshold, start_time)

    # 创建Excel报告
    _create_excel_report(output_path, dataframes)

    # 生成图表
    _create_charts(output_path, dataframes)

    end_time = datetime.now()
    log_info(f"状态码分析完成，耗时: {(end_time - start_time).total_seconds():.2f} 秒", True)
    log_info(f"分析报告已保存至: {output_path}", True)

    return dataframes['status_df'].head(10)


def _process_chunk(chunk, collectors, stats, field_mapping, lifecycle_fields, slow_request_threshold):
    """处理单个数据块"""
    status_field = field_mapping['status']
    app_field = field_mapping['app_name']
    request_time_field = field_mapping['request_time']

    # 基础状态码统计
    collectors['status_counts'].update(chunk[status_field].value_counts().to_dict())

    # 应用级状态码统计
    if app_field in chunk.columns:
        for app_name, app_group in chunk.groupby(app_field):
            collectors['app_status_counts'][app_name].update(
                app_group[status_field].value_counts().to_dict()
            )

    # 服务级状态码统计
    if field_mapping['service_name'] in chunk.columns:
        for service_name, service_group in chunk.groupby(field_mapping['service_name']):
            collectors['service_status_counts'][service_name].update(
                service_group[status_field].value_counts().to_dict()
            )

    # 慢请求统计
    if request_time_field in chunk.columns:
        slow_mask = chunk[request_time_field].astype(float) > slow_request_threshold
        chunk_slow_requests = chunk[slow_mask]
        stats['total_slow_requests'] += len(chunk_slow_requests)

        for status, group in chunk_slow_requests.groupby(status_field):
            collectors['slow_requests_by_status'][status] += len(group)

    # 时间和生命周期统计
    _collect_time_and_lifecycle_stats(chunk, collectors, field_mapping, lifecycle_fields, status_field)

    # 错误详情收集
    _collect_error_details(chunk, collectors, field_mapping, status_field)

    # 时间维度统计
    _collect_time_dimension_stats(chunk, collectors, field_mapping, status_field)


def _collect_time_and_lifecycle_stats(chunk, collectors, field_mapping, lifecycle_fields, status_field):
    """收集时间和生命周期统计数据"""
    request_time_field = field_mapping['request_time']

    for status, status_group in chunk.groupby(status_field):
        # 基础响应时间统计
        if request_time_field in chunk.columns:
            valid_times = status_group[request_time_field].dropna().astype(float).tolist()
            collectors['status_time_stats'][status].extend(valid_times)

        # nginx生命周期各阶段统计
        for phase_name, phase_field in lifecycle_fields.items():
            if phase_field in chunk.columns:
                phase_values = status_group[phase_field].dropna().astype(float).tolist()
                collectors['status_lifecycle_stats'][status][phase_name].extend(phase_values)

        # 传输大小统计(已经是KB单位)
        for size_field in ['response_body_size_kb', 'total_bytes_sent_kb']:
            if size_field in chunk.columns:
                size_values = status_group[size_field].dropna().astype(float).tolist()
                collectors['status_lifecycle_stats'][status][size_field].extend(size_values)


def _collect_error_details(chunk, collectors, field_mapping, status_field):
    """收集错误状态码详情"""
    error_mask = chunk[status_field].astype(str).str.startswith(('4', '5'))
    error_codes = chunk[error_mask]

    if not error_codes.empty:
        error_detail_fields = [
            field_mapping.get('time', 'raw_time'),
            field_mapping.get('app_name', 'application_name'),
            field_mapping.get('service_name', 'service_name'),
            field_mapping.get('method', 'http_method'),
            field_mapping.get('request_path', 'request_path'),
            field_mapping.get('request_time', 'total_request_duration'),
            field_mapping.get('client_ip', 'client_ip_address'),
            field_mapping.get('upstream_addr', 'upstream_server_address')
        ]

        available_fields = [field for field in error_detail_fields if field in error_codes.columns]

        for status, group in error_codes.groupby(status_field):
            if len(collectors['error_details'][status]) >= 1000:
                continue

            for _, row in group.iterrows():
                if len(collectors['error_details'][status]) >= 1000:
                    break
                error_info = {field: row.get(field, '') for field in available_fields}
                collectors['error_details'][status].append(error_info)


def _collect_time_dimension_stats(chunk, collectors, field_mapping, status_field):
    """收集时间维度统计数据"""
    # 小时维度统计
    hour_field = 'date_hour' if 'date_hour' in chunk.columns else field_mapping['hour']
    if hour_field in chunk.columns:
        for hour, hour_group in chunk.groupby(hour_field):
            collectors['status_by_hour'][hour].update(
                hour_group[status_field].value_counts().to_dict()
            )

    # 日期维度统计
    if field_mapping['date'] in chunk.columns:
        for date, date_group in chunk.groupby(field_mapping['date']):
            collectors['status_by_date'][date].update(
                date_group[status_field].value_counts().to_dict()
            )

    # HTTP方法维度统计
    method_field = field_mapping['method']
    if method_field in chunk.columns:
        for method, method_group in chunk.groupby(method_field):
            collectors['status_method_counts'][method].update(
                method_group[status_field].value_counts().to_dict()
            )


def _generate_analysis_dataframes(collectors, stats, slow_request_threshold, start_time):
    """生成各类分析数据表"""
    dataframes = {}

    # 计算类别统计
    category_counts = defaultdict(int)
    for code, count in collectors['status_counts'].items():
        category = get_status_category(str(code))
        category_counts[category] += count

    # 摘要信息
    dataframes['summary_df'] = _create_summary_dataframe(
        stats, category_counts, slow_request_threshold, start_time
    )

    # 详细状态码统计(包含生命周期分析)
    dataframes['status_df'] = _create_detailed_status_dataframe(
        collectors, stats['total_requests']
    )

    # 应用状态码统计
    dataframes['app_status_df'] = _create_app_status_dataframe(collectors['app_status_counts'])

    # 服务状态码统计
    dataframes['service_status_df'] = _create_service_status_dataframe(collectors['service_status_counts'])

    # 状态码类别统计
    dataframes['category_df'] = _create_category_dataframe(category_counts, stats['total_requests'])

    # 错误详情
    dataframes['error_details_df'] = _create_error_details_dataframe(collectors['error_details'])

    # 时间趋势统计
    dataframes['hour_status_df'] = _create_hour_status_dataframe(collectors['status_by_hour'])
    dataframes['date_status_df'] = _create_date_status_dataframe(collectors['status_by_date'])

    # HTTP方法统计
    dataframes['method_status_df'] = _create_method_status_dataframe(collectors['status_method_counts'])

    # 生命周期性能分布
    dataframes['lifecycle_performance_df'] = _create_lifecycle_performance_dataframe(
        collectors['status_time_stats'], collectors['status_lifecycle_stats']
    )

    return dataframes


def _create_summary_dataframe(stats, category_counts, slow_request_threshold, start_time):
    """创建摘要信息数据表"""
    total_requests = stats['total_requests']
    slow_request_rate = (stats['total_slow_requests'] / total_requests * 100) if total_requests > 0 else 0

    summary_data = [
        {'指标': '总请求数', '值': total_requests},
        {'指标': '成功率(%)',
         '值': round(category_counts.get('成功(2xx)', 0) / total_requests * 100, 2) if total_requests > 0 else 0},
        {'指标': '客户端错误率(%)',
         '值': round(category_counts.get('客户端错误(4xx)', 0) / total_requests * 100, 2) if total_requests > 0 else 0},
        {'指标': '服务器错误率(%)',
         '值': round(category_counts.get('服务器错误(5xx)', 0) / total_requests * 100, 2) if total_requests > 0 else 0},
        {'指标': '慢请求总数', '值': stats['total_slow_requests']},
        {'指标': f'慢请求阈值(秒)', '值': slow_request_threshold},
        {'指标': '全局慢请求占比(%)', '值': round(slow_request_rate, 2)},
        {'指标': '分析开始时间', '值': start_time.strftime('%Y-%m-%d %H:%M:%S')},
        {'指标': '分析结束时间', '值': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
        {'指标': '分析耗时(秒)', '值': round((datetime.now() - start_time).total_seconds(), 2)}
    ]

    return pd.DataFrame(summary_data)


def _create_detailed_status_dataframe(collectors, total_requests):
    """创建详细状态码统计表(整合生命周期分析)"""
    status_data = []

    for code, count in collectors['status_counts'].items():
        category = get_status_category(str(code))
        description = get_status_description(str(code))
        percentage = (count / total_requests * 100) if total_requests > 0 else 0

        # 慢请求统计
        slow_count = collectors['slow_requests_by_status'].get(code, 0)
        slow_percentage = (slow_count / count * 100) if count > 0 else 0

        # 基础响应时间统计
        request_time_stats = _calculate_time_stats(collectors['status_time_stats'].get(code, []))

        # nginx生命周期各阶段统计
        lifecycle_stats = _calculate_lifecycle_stats(collectors['status_lifecycle_stats'].get(code, {}))

        # 组合数据行
        row_data = {
            '状态码': code,
            '状态描述': description,
            '类别': category,
            '请求数': count,
            '百分比(%)': round(percentage, 2),
            '慢请求数': slow_count,
            '慢请求比例(%)': round(slow_percentage, 2),
        }

        # 添加响应时间统计
        row_data.update(request_time_stats)

        # 添加生命周期统计
        row_data.update(lifecycle_stats)

        status_data.append(row_data)

    status_df = pd.DataFrame(status_data)
    return status_df.sort_values(by='请求数', ascending=False) if not status_df.empty else status_df


def _calculate_time_stats(times_list):
    """计算时间统计指标"""
    if not times_list:
        return {
            '平均响应时间(秒)': None,
            '中位响应时间(秒)': None,
            '最小响应时间(秒)': None,
            '最大响应时间(秒)': None,
            'P95响应时间(秒)': None,
            'P99响应时间(秒)': None
        }

    times_array = np.array(times_list)
    return {
        '平均响应时间(秒)': round(np.mean(times_array), 3),
        '中位响应时间(秒)': round(np.median(times_array), 3),
        '最小响应时间(秒)': round(np.min(times_array), 3),
        '最大响应时间(秒)': round(np.max(times_array), 3),
        'P95响应时间(秒)': round(np.percentile(times_array, 95), 3),
        'P99响应时间(秒)': round(np.percentile(times_array, 99), 3)
    }


def _calculate_lifecycle_stats(lifecycle_dict):
    """计算nginx生命周期各阶段统计"""
    stats = {}

    # 核心阶段统计
    phase_fields = {
        'backend_connect_phase': '后端连接阶段均值(秒)',
        'backend_process_phase': '后端处理阶段均值(秒)',
        'backend_transfer_phase': '后端传输阶段均值(秒)',
        'nginx_transfer_phase': 'Nginx传输阶段均值(秒)',
        'backend_total_phase': '后端总阶段均值(秒)',
        'network_phase': '网络传输阶段均值(秒)',
        'processing_phase': '纯处理阶段均值(秒)',
        'transfer_phase': '纯传输阶段均值(秒)'
    }

    for field, label in phase_fields.items():
        values = lifecycle_dict.get(field, [])
        stats[label] = round(np.mean(values), 3) if values else None

    # 性能比率统计
    ratio_fields = {
        'backend_efficiency': '后端处理效率(%)',
        'network_overhead': '网络开销占比(%)',
        'transfer_ratio': '传输时间占比(%)',
        'connection_cost_ratio': '连接成本比率(%)',
        'processing_efficiency_index': '处理效率指数'
    }

    for field, label in ratio_fields.items():
        values = lifecycle_dict.get(field, [])
        stats[label] = round(np.mean(values), 2) if values else None

    # 传输大小统计(KB)
    size_fields = {
        'response_body_size_kb': '平均响应体大小(KB)',
        'total_bytes_sent_kb': '平均总传输大小(KB)'
    }

    for field, label in size_fields.items():
        values = lifecycle_dict.get(field, [])
        stats[label] = round(np.mean(values), 2) if values else None

    # 传输速度统计
    speed_fields = {
        'response_transfer_speed': '响应传输速度(KB/s)',
        'total_transfer_speed': '总传输速度(KB/s)',
        'nginx_transfer_speed': 'Nginx传输速度(KB/s)'
    }

    for field, label in speed_fields.items():
        values = lifecycle_dict.get(field, [])
        stats[label] = round(np.mean(values), 2) if values else None

    return stats


def _create_app_status_dataframe(app_status_counts):
    """创建应用状态码统计表"""
    app_status_data = []

    for app_name, code_counter in app_status_counts.items():
        app_total = sum(code_counter.values())
        success_count = sum(code_counter[code] for code in code_counter if str(code).startswith('2'))
        client_error_count = sum(code_counter[code] for code in code_counter if str(code).startswith('4'))
        server_error_count = sum(code_counter[code] for code in code_counter if str(code).startswith('5'))

        success_rate = (success_count / app_total * 100) if app_total > 0 else 0
        error_rate = ((client_error_count + server_error_count) / app_total * 100) if app_total > 0 else 0

        app_status_data.append({
            '应用名称': app_name,
            '总请求数': app_total,
            '成功请求数': success_count,
            '客户端错误数': client_error_count,
            '服务器错误数': server_error_count,
            '成功率(%)': round(success_rate, 2),
            '错误率(%)': round(error_rate, 2)
        })

    app_status_df = pd.DataFrame(app_status_data)
    return app_status_df.sort_values(by='总请求数', ascending=False) if not app_status_df.empty else app_status_df


def _create_service_status_dataframe(service_status_counts):
    """创建服务状态码统计表"""
    service_status_data = []

    for service_name, code_counter in service_status_counts.items():
        service_total = sum(code_counter.values())
        success_count = sum(code_counter[code] for code in code_counter if str(code).startswith('2'))
        client_error_count = sum(code_counter[code] for code in code_counter if str(code).startswith('4'))
        server_error_count = sum(code_counter[code] for code in code_counter if str(code).startswith('5'))

        success_rate = (success_count / service_total * 100) if service_total > 0 else 0
        error_rate = ((client_error_count + server_error_count) / service_total * 100) if service_total > 0 else 0

        service_status_data.append({
            '服务名称': service_name,
            '总请求数': service_total,
            '成功请求数': success_count,
            '客户端错误数': client_error_count,
            '服务器错误数': server_error_count,
            '成功率(%)': round(success_rate, 2),
            '错误率(%)': round(error_rate, 2)
        })

    service_status_df = pd.DataFrame(service_status_data)
    return service_status_df.sort_values(by='总请求数',
                                         ascending=False) if not service_status_df.empty else service_status_df


def _create_category_dataframe(category_counts, total_requests):
    """创建状态码类别统计表"""
    category_data = [
        {
            '状态码类别': category,
            '请求数': count,
            '百分比(%)': round((count / total_requests * 100), 2) if total_requests > 0 else 0
        }
        for category, count in category_counts.items()
    ]

    category_df = pd.DataFrame(category_data)
    return category_df.sort_values(by='请求数', ascending=False) if not category_df.empty else category_df


def _create_error_details_dataframe(error_details):
    """创建错误详情数据表"""
    error_details_data = []

    for status, details in error_details.items():
        for detail in details:
            error_details_data.append({
                '状态码': status,
                '时间': detail.get('raw_time', ''),
                '应用': detail.get('application_name', ''),
                '服务': detail.get('service_name', ''),
                '请求方法': detail.get('http_method', ''),
                '请求路径': detail.get('request_path', ''),
                '响应时间(秒)': detail.get('total_request_duration', ''),
                '客户端IP': detail.get('client_ip_address', ''),
                '上游地址': detail.get('upstream_server_address', '')
            })

    error_details_df = pd.DataFrame(error_details_data)
    return error_details_df.sort_values(by=['状态码', '时间'],
                                        ascending=[True, False]) if not error_details_df.empty else error_details_df


def _create_hour_status_dataframe(status_by_hour):
    """创建小时状态码趋势表"""
    hour_status_data = []

    for hour, codes in sorted(status_by_hour.items()):
        hour_total = sum(codes.values())
        success_count = sum(codes[code] for code in codes if str(code).startswith('2'))
        client_error_count = sum(codes[code] for code in codes if str(code).startswith('4'))
        server_error_count = sum(codes[code] for code in codes if str(code).startswith('5'))

        success_rate = (success_count / hour_total * 100) if hour_total > 0 else 0
        error_rate = ((client_error_count + server_error_count) / hour_total * 100) if hour_total > 0 else 0

        hour_status_data.append({
            '小时': hour,
            '总请求数': hour_total,
            '成功请求数': success_count,
            '客户端错误数': client_error_count,
            '服务器错误数': server_error_count,
            '成功率(%)': round(success_rate, 2),
            '错误率(%)': round(error_rate, 2)
        })

    return pd.DataFrame(hour_status_data)


def _create_date_status_dataframe(status_by_date):
    """创建日期状态码趋势表"""
    date_status_data = []

    for date, status_dict in sorted(status_by_date.items()):
        date_total = sum(status_dict.values())
        success_count = sum(status_dict[code] for code in status_dict if str(code).startswith('2'))
        client_error_count = sum(status_dict[code] for code in status_dict if str(code).startswith('4'))
        server_error_count = sum(status_dict[code] for code in status_dict if str(code).startswith('5'))

        success_rate = (success_count / date_total * 100) if date_total > 0 else 0
        error_rate = ((client_error_count + server_error_count) / date_total * 100) if date_total > 0 else 0

        date_status_data.append({
            '日期': date,
            '总请求数': date_total,
            '成功请求数': success_count,
            '客户端错误数': client_error_count,
            '服务器错误数': server_error_count,
            '成功率(%)': round(success_rate, 2),
            '错误率(%)': round(error_rate, 2)
        })

    return pd.DataFrame(date_status_data) if date_status_data else pd.DataFrame()


def _create_method_status_dataframe(status_method_counts):
    """创建HTTP方法状态码交叉分析数据框"""
    method_status_data = []

    for method, status_dict in status_method_counts.items():
        method_total = sum(status_dict.values())
        success_count = sum(status_dict[code] for code in status_dict if str(code).startswith('2'))
        client_error_count = sum(status_dict[code] for code in status_dict if str(code).startswith('4'))
        server_error_count = sum(status_dict[code] for code in status_dict if str(code).startswith('5'))
        redirect_count = sum(status_dict[code] for code in status_dict if str(code).startswith('3'))

        success_rate = (success_count / method_total * 100) if method_total > 0 else 0
        error_rate = ((client_error_count + server_error_count) / method_total * 100) if method_total > 0 else 0

        method_status_data.append({
            'HTTP方法': method,
            '总请求数': method_total,
            '成功请求数': success_count,
            '重定向数': redirect_count,
            '客户端错误数': client_error_count,
            '服务器错误数': server_error_count,
            '成功率(%)': round(success_rate, 2),
            '错误率(%)': round(error_rate, 2)
        })

    method_status_df = pd.DataFrame(method_status_data)
    return method_status_df.sort_values(by='总请求数', ascending=False) if not method_status_df.empty else method_status_df


def _create_lifecycle_performance_dataframe(status_time_stats, status_lifecycle_stats):
    """创建生命周期性能分析数据框"""
    lifecycle_data = []

    for status, lifecycle_dict in status_lifecycle_stats.items():
        if not lifecycle_dict:
            continue

        row_data = {
            '状态码': status,
            '请求数': len(status_time_stats.get(status, [])),
            '状态描述': get_status_description(str(status)),
            '类别': get_status_category(str(status))
        }

        # 基础时间统计
        times_list = status_time_stats.get(status, [])
        if times_list:
            times_array = np.array(times_list)
            row_data.update({
                '平均总时长(秒)': round(np.mean(times_array), 3),
                '中位总时长(秒)': round(np.median(times_array), 3),
                'P95总时长(秒)': round(np.percentile(times_array, 95), 3),
                'P99总时长(秒)': round(np.percentile(times_array, 99), 3)
            })

        # 后端各阶段时长统计
        backend_phases = {
            'backend_connect_phase': '后端连接时长(秒)',
            'backend_process_phase': '后端处理时长(秒)',
            'backend_transfer_phase': '后端传输时长(秒)',
            'nginx_transfer_phase': 'Nginx传输时长(秒)',
            'backend_total_phase': '后端总时长(秒)'
        }

        for phase_field, phase_label in backend_phases.items():
            phase_values = lifecycle_dict.get(phase_field, [])
            if phase_values:
                row_data[phase_label] = round(np.mean(phase_values), 3)

        # 性能效率指标
        efficiency_metrics = {
            'backend_efficiency': '后端处理效率(%)',
            'network_overhead': '网络开销占比(%)',
            'transfer_ratio': '传输时间占比(%)',
            'connection_cost_ratio': '连接成本比率(%)',
            'processing_efficiency_index': '处理效率指数'
        }

        for metric_field, metric_label in efficiency_metrics.items():
            metric_values = lifecycle_dict.get(metric_field, [])
            if metric_values:
                row_data[metric_label] = round(np.mean(metric_values), 2)

        # 传输速度指标
        speed_metrics = {
            'response_transfer_speed': '响应传输速度(KB/s)',
            'total_transfer_speed': '总传输速度(KB/s)',
            'nginx_transfer_speed': 'Nginx传输速度(KB/s)'
        }

        for speed_field, speed_label in speed_metrics.items():
            speed_values = lifecycle_dict.get(speed_field, [])
            if speed_values:
                row_data[speed_label] = round(np.mean(speed_values), 2)

        # 数据大小统计（已调整为KB）
        size_metrics = {
            'response_body_size_kb': '平均响应体大小(KB)',
            'total_bytes_sent_kb': '平均总传输大小(KB)'
        }

        for size_field, size_label in size_metrics.items():
            size_values = lifecycle_dict.get(size_field, [])
            if size_values:
                row_data[size_label] = round(np.mean(size_values), 2)

        lifecycle_data.append(row_data)

    lifecycle_df = pd.DataFrame(lifecycle_data)
    return lifecycle_df.sort_values(by='请求数', ascending=False) if not lifecycle_df.empty else lifecycle_df


def _create_excel_report(output_path, dataframes):
    """创建Excel报告"""
    log_info("创建Excel工作簿...", True)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 摘要信息
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['summary_df'], '摘要信息')

    # 状态码统计 - 使用分组标题
    status_header_groups = {
        '基本信息': ['状态码', '状态描述', '类别', '请求数', '百分比(%)'],
        '慢请求分析': ['慢请求数', '慢请求比例(%)'],
        '响应时间统计(秒)': ['平均响应时间(秒)', '中位响应时间(秒)', '最小响应时间(秒)',
                      '最大响应时间(秒)', 'P95响应时间(秒)', 'P99响应时间(秒)'],
        '后端阶段时长(秒)': ['后端连接阶段均值(秒)', '后端处理阶段均值(秒)', '后端传输阶段均值(秒)', 'Nginx传输阶段均值(秒)'],
        '性能效率指标(%)': ['后端处理效率(%)', '网络开销占比(%)', '传输时间占比(%)', '连接成本比率(%)'],
        '传输性能': ['响应传输速度(KB/s)', '总传输速度(KB/s)', 'Nginx传输速度(KB/s)',
                 '平均响应体大小(KB)', '平均总传输大小(KB)']
    }
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['status_df'], '状态码统计',
                                                header_groups=status_header_groups)

    # 应用和服务状态码
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['app_status_df'], '应用状态码统计')
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['service_status_df'], '服务状态码统计')
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['category_df'], '状态码类别统计')

    # 错误详情
    if not dataframes['error_details_df'].empty:
        add_dataframe_to_excel_with_grouped_headers(wb, dataframes['error_details_df'], '错误状态码详情')

    # 时间维度分析
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['hour_status_df'], '小时状态码趋势')
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['date_status_df'], '日期状态码趋势')
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['method_status_df'], 'HTTP方法状态码')

    # 生命周期性能分析 - 重点新增
    lifecycle_header_groups = {
        '基本信息': ['状态码', '请求数', '状态描述', '类别'],
        '总体时长统计(秒)': ['平均总时长(秒)', '中位总时长(秒)', 'P95总时长(秒)', 'P99总时长(秒)'],
        '后端各阶段时长(秒)': ['后端连接时长(秒)', '后端处理时长(秒)', '后端传输时长(秒)',
                       'Nginx传输时长(秒)', '后端总时长(秒)'],
        '性能效率分析(%)': ['后端处理效率(%)', '网络开销占比(%)', '传输时间占比(%)',
                      '连接成本比率(%)', '处理效率指数'],
        '传输性能指标': ['响应传输速度(KB/s)', '总传输速度(KB/s)', 'Nginx传输速度(KB/s)',
                   '平均响应体大小(KB)', '平均总传输大小(KB)']
    }
    add_dataframe_to_excel_with_grouped_headers(wb, dataframes['lifecycle_performance_df'],
                                                '生命周期性能分析', header_groups=lifecycle_header_groups)

    wb.save(output_path)


def _create_charts(output_path, dataframes):
    """创建图表"""
    wb = load_workbook(output_path)

    # 状态码分布饼图
    if '状态码类别统计' in wb.sheetnames and not dataframes['category_df'].empty:
        ws = wb['状态码类别统计']
        chart_sheet = wb.create_sheet(title="状态码分布图")

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=1 + len(dataframes['category_df']))
        data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(dataframes['category_df']))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "HTTP状态码类别分布"

        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True

        # 设置饼图颜色
        slice_colors = {
            '成功(2xx)': '92D050',
            '重定向(3xx)': '00B0F0',
            '客户端错误(4xx)': 'FFC000',
            '服务器错误(5xx)': 'FF0000',
            '未知': 'A5A5A5'
        }

        for i, category in enumerate([row[0] for row in dataframes['category_df'].values]):
            if category in slice_colors:
                slice = DataPoint(idx=i)
                slice.graphicalProperties.solidFill = slice_colors[category]
                pie.series[0].dPt.append(slice)

        chart_sheet.add_chart(pie, "A1")

    # 小时趋势图
    if '小时状态码趋势' in wb.sheetnames and len(dataframes['hour_status_df']) > 1:
        ws = wb['小时状态码趋势']
        chart_sheet = wb.create_sheet(title="小时趋势图")

        chart = LineChart()
        chart.title = "HTTP状态码小时趋势"
        chart.style = 12
        chart.x_axis.title = "小时"
        chart.y_axis.title = "请求数"

        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(dataframes['hour_status_df']))

        # 成功请求趋势
        success_data = Reference(ws, min_col=3, min_row=1, max_row=1 + len(dataframes['hour_status_df']))
        chart.add_data(success_data, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = "92D050"

        # 客户端错误趋势
        client_error_data = Reference(ws, min_col=4, min_row=1, max_row=1 + len(dataframes['hour_status_df']))
        chart.add_data(client_error_data, titles_from_data=True)
        chart.series[1].graphicalProperties.line.solidFill = "FFC000"

        # 服务器错误趋势
        server_error_data = Reference(ws, min_col=5, min_row=1, max_row=1 + len(dataframes['hour_status_df']))
        chart.add_data(server_error_data, titles_from_data=True)
        chart.series[2].graphicalProperties.line.solidFill = "FF0000"

        chart.set_categories(cats)
        chart_sheet.add_chart(chart, "A1")

    # 日期趋势图
    if '日期状态码趋势' in wb.sheetnames and len(dataframes['date_status_df']) > 1:
        ws = wb['日期状态码趋势']
        date_chart_sheet = wb.create_sheet(title="日期趋势图")

        date_chart = LineChart()
        date_chart.title = "HTTP状态码日期趋势"
        date_chart.style = 12
        date_chart.x_axis.title = "日期"
        date_chart.y_axis.title = "请求数"

        date_cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(dataframes['date_status_df']))

        # 成功请求趋势
        date_success_data = Reference(ws, min_col=3, min_row=1, max_row=1 + len(dataframes['date_status_df']))
        date_chart.add_data(date_success_data, titles_from_data=True)
        date_chart.series[0].graphicalProperties.line.solidFill = "92D050"

        # 客户端错误趋势
        date_client_error_data = Reference(ws, min_col=4, min_row=1, max_row=1 + len(dataframes['date_status_df']))
        date_chart.add_data(date_client_error_data, titles_from_data=True)
        date_chart.series[1].graphicalProperties.line.solidFill = "FFC000"

        # 服务器错误趋势
        date_server_error_data = Reference(ws, min_col=5, min_row=1, max_row=1 + len(dataframes['date_status_df']))
        date_chart.add_data(date_server_error_data, titles_from_data=True)
        date_chart.series[2].graphicalProperties.line.solidFill = "FF0000"

        date_chart.set_categories(date_cats)
        date_chart_sheet.add_chart(date_chart, "A1")

    wb.save(output_path)

def get_status_category(status_code):
    """获取状态码类别"""
    if not status_code or not isinstance(status_code, str):
        return '未知'

    first_digit = status_code[0] if status_code else ''

    if first_digit == '2':
        return '成功(2xx)'
    elif first_digit == '3':
        return '重定向(3xx)'
    elif first_digit == '4':
        return '客户端错误(4xx)'
    elif first_digit == '5':
        return '服务器错误(5xx)'
    else:
        return '未知'

def get_status_description(status_code):
    """获取状态码描述"""
    if not status_code or not isinstance(status_code, str):
        return '未知状态码'

    return STATUS_DESCRIPTIONS.get(status_code, f'未知状态码: {status_code}')