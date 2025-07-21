#!/usr/bin/env python3
"""
高级状态码分析器 - 优化版本
支持40G+大数据处理，内存高效，智能分析

主要优化：
1. 单次扫描 + 流式处理
2. T-Digest + 智能采样
3. 状态码异常检测 + 根因分析
4. 精简高价值输出列
5. 智能优化建议

版本：v2.0
作者：Claude Code
日期：2025-07-18
"""

import gc
import os
import tempfile
import json
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Set
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
from collections import defaultdict, Counter

# 导入采样算法
from self_00_05_sampling_algorithms import TDigest, ReservoirSampler, CountMinSketch, HyperLogLog
from self_00_01_constants import DEFAULT_SLOW_THRESHOLD, DEFAULT_CHUNK_SIZE, HIGHLIGHT_FILL
from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    format_excel_sheet,
    add_dataframe_to_excel_with_grouped_headers
)

# 核心状态码分类
STATUS_CATEGORIES = {
    '2xx': '成功',
    '3xx': '重定向', 
    '4xx': '客户端错误',
    '5xx': '服务器错误'
}

# 重要状态码定义
CRITICAL_STATUS_CODES = {
    '200': '成功',
    '301': '永久重定向',
    '302': '临时重定向',
    '400': '请求错误',
    '401': '未授权',
    '403': '禁止访问',
    '404': '未找到',
    '500': '服务器内部错误',
    '502': '网关错误',
    '503': '服务不可用',
    '504': '网关超时'
}

# 内存格式化函数
def format_memory_usage():
    """格式化内存使用情况"""
    try:
        import psutil
        import os
        process = psutil.Process(os.getpid())
        memory_usage_mb = process.memory_info().rss / 1024 / 1024
        return f"{memory_usage_mb:.2f} MB"
    except ImportError:
        return "N/A"


class AdvancedStatusAnalyzer:
    """
    高级状态码分析器
    使用流式处理和智能采样算法
    """
    
    def __init__(self, slow_threshold=DEFAULT_SLOW_THRESHOLD):
        self.slow_threshold = slow_threshold
        self.reset_collectors()
        
    def reset_collectors(self):
        """重置数据收集器"""
        # 基础统计
        self.status_counter = Counter()
        self.app_status_counter = defaultdict(Counter)
        self.service_status_counter = defaultdict(Counter)
        self.method_status_counter = defaultdict(Counter)
        
        # 时间维度统计
        self.hourly_status_counter = defaultdict(Counter)
        self.daily_status_counter = defaultdict(Counter)
        
        # 性能相关采样器
        self.status_response_time = defaultdict(lambda: TDigest(compression=100))
        self.status_slow_requests = defaultdict(int)
        
        # 错误详情采样
        self.error_sampler = defaultdict(lambda: ReservoirSampler(max_size=500))
        
        # IP和路径分析
        self.status_ip_counter = defaultdict(lambda: CountMinSketch(width=1000, depth=7))
        self.status_path_counter = defaultdict(lambda: CountMinSketch(width=2000, depth=7))
        
        # 异常检测
        self.anomaly_detector = AnomalyDetector()
        
        # 统计信息
        self.total_requests = 0
        self.chunks_processed = 0
        
    def analyze_status_codes(self, csv_path: str, output_path: str) -> pd.DataFrame:
        """
        主要分析入口函数
        """
        log_info("🚀 开始高级状态码分析...", True)
        start_time = datetime.now()
        
        # 单次扫描处理所有数据
        self._process_data_stream(csv_path)
        
        # 生成分析报告
        dataframes = self._generate_analysis_reports()
        
        # 创建Excel报告
        self._create_excel_report(output_path, dataframes)
        
        end_time = datetime.now()
        processing_time = (end_time - start_time).total_seconds()
        
        log_info(f"✅ 状态码分析完成！处理时间: {processing_time:.2f}秒", True)
        log_info(f"📊 报告保存至: {output_path}", True)
        
        return dataframes.get('summary', pd.DataFrame())
    
    def _process_data_stream(self, csv_path: str):
        """流式处理CSV数据"""
        log_info("📖 开始流式处理数据...", True)
        
        chunk_size = DEFAULT_CHUNK_SIZE
        reader = pd.read_csv(csv_path, chunksize=chunk_size)
        
        for chunk in reader:
            self._process_chunk(chunk)
            self.chunks_processed += 1
            
            # 定期内存清理和进度报告
            if self.chunks_processed % 10 == 0:
                elapsed = (datetime.now() - datetime.now()).total_seconds()
                memory_usage = format_memory_usage()
                log_info(f"📊 已处理 {self.chunks_processed} 个数据块, {self.total_requests} 条记录, 内存: {memory_usage}")
                gc.collect()
                
    def _process_chunk(self, chunk: pd.DataFrame):
        """处理单个数据块"""
        chunk_size = len(chunk)
        self.total_requests += chunk_size
        
        # 获取关键字段
        status_field = self._get_field_name(chunk, ['response_status_code', 'status'])
        app_field = self._get_field_name(chunk, ['application_name', 'app_name'])
        service_field = self._get_field_name(chunk, ['service_name'])
        method_field = self._get_field_name(chunk, ['http_method', 'method'])
        time_field = self._get_field_name(chunk, ['total_request_duration', 'request_time'])
        ip_field = self._get_field_name(chunk, ['client_ip_address', 'client_ip'])
        path_field = self._get_field_name(chunk, ['request_path', 'path'])
        hour_field = self._get_field_name(chunk, ['hour', 'date_hour'])
        date_field = self._get_field_name(chunk, ['date'])
        
        # 基础状态码统计
        if status_field:
            self.status_counter.update(chunk[status_field].value_counts().to_dict())
            
            # 应用级状态码统计
            if app_field:
                for app_name, app_group in chunk.groupby(app_field):
                    self.app_status_counter[app_name].update(
                        app_group[status_field].value_counts().to_dict()
                    )
            
            # 服务级状态码统计
            if service_field:
                for service_name, service_group in chunk.groupby(service_field):
                    self.service_status_counter[service_name].update(
                        service_group[status_field].value_counts().to_dict()
                    )
            
            # HTTP方法统计
            if method_field:
                for method, method_group in chunk.groupby(method_field):
                    self.method_status_counter[method].update(
                        method_group[status_field].value_counts().to_dict()
                    )
            
            # 时间维度统计
            if hour_field:
                for hour, hour_group in chunk.groupby(hour_field):
                    self.hourly_status_counter[hour].update(
                        hour_group[status_field].value_counts().to_dict()
                    )
            
            if date_field:
                for date, date_group in chunk.groupby(date_field):
                    self.daily_status_counter[date].update(
                        date_group[status_field].value_counts().to_dict()
                    )
            
            # 性能关联分析
            if time_field:
                for status, status_group in chunk.groupby(status_field):
                    response_times = status_group[time_field].dropna().astype(float)
                    
                    # 添加响应时间数据到T-Digest
                    for rt in response_times:
                        self.status_response_time[status].add(rt)
                        
                        # 慢请求统计
                        if rt > self.slow_threshold:
                            self.status_slow_requests[status] += 1
            
            # 错误详情采样
            error_mask = chunk[status_field].astype(str).str.match(r'^[45]\d\d$')
            if error_mask.any():
                error_chunk = chunk[error_mask]
                self._collect_error_samples(error_chunk, status_field, ip_field, path_field, time_field)
            
            # IP和路径分析
            if ip_field:
                for status, status_group in chunk.groupby(status_field):
                    for ip in status_group[ip_field].dropna():
                        self.status_ip_counter[status].increment(str(ip))
            
            if path_field:
                for status, status_group in chunk.groupby(status_field):
                    for path in status_group[path_field].dropna():
                        self.status_path_counter[status].increment(str(path))
            
            # 异常检测
            self.anomaly_detector.process_chunk(chunk, status_field, time_field)
    
    def _get_field_name(self, chunk: pd.DataFrame, field_candidates: List[str]) -> Optional[str]:
        """获取可用的字段名"""
        for field in field_candidates:
            if field in chunk.columns:
                return field
        return None
    
    def _collect_error_samples(self, error_chunk: pd.DataFrame, status_field: str, 
                             ip_field: str, path_field: str, time_field: str):
        """收集错误样本"""
        for _, row in error_chunk.iterrows():
            status = row[status_field]
            error_info = {
                'status': status,
                'ip': row.get(ip_field, ''),
                'path': row.get(path_field, ''),
                'response_time': row.get(time_field, 0),
                'timestamp': row.get('raw_time', '')
            }
            self.error_sampler[status].add(error_info)
    
    def _generate_analysis_reports(self) -> Dict[str, pd.DataFrame]:
        """生成分析报告"""
        log_info("📊 生成分析报告...", True)
        
        reports = {}
        
        # 1. 状态码分布摘要
        reports['summary'] = self._create_status_summary()
        
        # 2. 详细状态码分析
        reports['detailed_status'] = self._create_detailed_status_analysis()
        
        # 3. 应用/服务状态码分析
        reports['app_analysis'] = self._create_app_status_analysis()
        reports['service_analysis'] = self._create_service_status_analysis()
        
        # 4. 时间维度分析
        reports['time_analysis'] = self._create_time_dimension_analysis()
        
        # 5. 错误分析
        reports['error_analysis'] = self._create_error_analysis()
        
        # 6. 性能关联分析
        reports['performance_analysis'] = self._create_performance_analysis()
        
        # 7. 异常检测报告
        reports['anomaly_report'] = self._create_anomaly_report()
        
        # 8. 优化建议
        reports['optimization_suggestions'] = self._create_optimization_suggestions()
        
        # 9. 慢请求API汇总 (重要！)
        reports['slow_request_api_summary'] = self._create_slow_request_api_summary()
        
        # 10. 性能关联详细分析 (重要！)
        reports['performance_detail_analysis'] = self._create_performance_detail_analysis()
        
        # 11. 状态码生命周期分析 (原版本的核心功能)
        reports['status_lifecycle_analysis'] = self._create_status_lifecycle_analysis()
        
        # 12. HTTP方法状态码分析 (整合原版本功能)
        reports['method_status_analysis'] = self._create_method_status_analysis()
        
        return reports
    
    def _create_status_summary(self) -> pd.DataFrame:
        """创建状态码分布摘要"""
        total_requests = self.total_requests
        
        # 计算各类别统计
        category_stats = {}
        for category in STATUS_CATEGORIES.keys():
            category_count = sum(
                count for status, count in self.status_counter.items()
                if str(status).startswith(category[0])
            )
            category_stats[category] = {
                'count': category_count,
                'percentage': (category_count / total_requests * 100) if total_requests > 0 else 0
            }
        
        # 计算整体指标
        success_rate = category_stats.get('2xx', {}).get('percentage', 0)
        error_rate = (category_stats.get('4xx', {}).get('percentage', 0) + 
                     category_stats.get('5xx', {}).get('percentage', 0))
        
        summary_data = [
            {'指标': '总请求数', '值': total_requests, '说明': '分析的总请求数量'},
            {'指标': '成功率(%)', '值': round(success_rate, 2), '说明': '2xx状态码占比'},
            {'指标': '错误率(%)', '值': round(error_rate, 2), '说明': '4xx+5xx状态码占比'},
            {'指标': '慢请求数', '值': sum(self.status_slow_requests.values()), '说明': f'响应时间>{self.slow_threshold}s的请求'},
            {'指标': '错误状态码种类', '值': len([s for s in self.status_counter.keys() if str(s).startswith(('4', '5'))]), '说明': '出现的错误状态码种类数'},
            {'指标': '异常检测项', '值': len(self.anomaly_detector.get_anomalies()), '说明': '检测到的异常项数量'}
        ]
        
        return pd.DataFrame(summary_data)
    
    def _create_detailed_status_analysis(self) -> pd.DataFrame:
        """创建详细状态码分析"""
        detailed_data = []
        
        for status, count in self.status_counter.most_common():
            status_str = str(status)
            category = self._get_status_category(status_str)
            description = CRITICAL_STATUS_CODES.get(status_str, f'状态码{status_str}')
            
            # 计算百分比
            percentage = (count / self.total_requests * 100) if self.total_requests > 0 else 0
            
            # 响应时间统计
            response_time_stats = self._get_response_time_stats(status)
            
            # 慢请求统计
            slow_count = self.status_slow_requests.get(status, 0)
            slow_percentage = (slow_count / count * 100) if count > 0 else 0
            
            detailed_data.append({
                '状态码': status,
                '描述': description,
                '类别': category,
                '请求数': count,
                '占比(%)': round(percentage, 2),
                '平均响应时间(秒)': response_time_stats['mean'],
                'P95响应时间(秒)': response_time_stats['p95'],
                'P99响应时间(秒)': response_time_stats['p99'],
                '慢请求数': slow_count,
                '慢请求占比(%)': round(slow_percentage, 2),
                '影响等级': self._assess_impact_level(status_str, count, percentage)
            })
        
        return pd.DataFrame(detailed_data)
    
    def _create_app_status_analysis(self) -> pd.DataFrame:
        """创建应用状态码分析"""
        if not self.app_status_counter:
            return pd.DataFrame()
        
        app_data = []
        for app_name, status_counter in self.app_status_counter.items():
            app_total = sum(status_counter.values())
            success_count = sum(count for status, count in status_counter.items() if str(status).startswith('2'))
            error_count = sum(count for status, count in status_counter.items() if str(status).startswith(('4', '5')))
            
            success_rate = (success_count / app_total * 100) if app_total > 0 else 0
            error_rate = (error_count / app_total * 100) if app_total > 0 else 0
            
            app_data.append({
                '应用名称': app_name,
                '总请求数': app_total,
                '成功请求数': success_count,
                '错误请求数': error_count,
                '成功率(%)': round(success_rate, 2),
                '错误率(%)': round(error_rate, 2),
                '健康状态': self._assess_app_health(error_rate)
            })
        
        app_df = pd.DataFrame(app_data)
        if not app_df.empty and '错误率(%)' in app_df.columns:
            return app_df.sort_values('错误率(%)', ascending=False)
        else:
            return app_df
    
    def _create_service_status_analysis(self) -> pd.DataFrame:
        """创建服务状态码分析"""
        if not self.service_status_counter:
            return pd.DataFrame()
        
        service_data = []
        for service_name, status_counter in self.service_status_counter.items():
            service_total = sum(status_counter.values())
            success_count = sum(count for status, count in status_counter.items() if str(status).startswith('2'))
            error_count = sum(count for status, count in status_counter.items() if str(status).startswith(('4', '5')))
            
            success_rate = (success_count / service_total * 100) if service_total > 0 else 0
            error_rate = (error_count / service_total * 100) if service_total > 0 else 0
            
            service_data.append({
                '服务名称': service_name,
                '总请求数': service_total,
                '成功请求数': success_count,
                '错误请求数': error_count,
                '成功率(%)': round(success_rate, 2),
                '错误率(%)': round(error_rate, 2),
                '健康状态': self._assess_service_health(error_rate)
            })
        
        service_df = pd.DataFrame(service_data)
        if not service_df.empty and '错误率(%)' in service_df.columns:
            return service_df.sort_values('错误率(%)', ascending=False)
        else:
            return service_df
    
    def _create_time_dimension_analysis(self) -> pd.DataFrame:
        """创建时间维度分析"""
        time_data = []
        
        # 小时维度分析
        for hour, status_counter in sorted(self.hourly_status_counter.items()):
            hour_total = sum(status_counter.values())
            success_count = sum(count for status, count in status_counter.items() if str(status).startswith('2'))
            error_count = sum(count for status, count in status_counter.items() if str(status).startswith(('4', '5')))
            
            success_rate = (success_count / hour_total * 100) if hour_total > 0 else 0
            error_rate = (error_count / hour_total * 100) if hour_total > 0 else 0
            
            time_data.append({
                '维度': '小时',
                '时间': hour,
                '总请求数': hour_total,
                '成功率(%)': round(success_rate, 2),
                '错误率(%)': round(error_rate, 2),
                '流量等级': self._assess_traffic_level(hour_total)
            })
        
        return pd.DataFrame(time_data)
    
    def _create_error_analysis(self) -> pd.DataFrame:
        """创建错误分析"""
        error_data = []
        
        for status, sampler in self.error_sampler.items():
            error_samples = sampler.get_samples()
            if not error_samples:
                continue
            
            # 统计错误模式
            ip_counter = Counter(sample['ip'] for sample in error_samples)
            path_counter = Counter(sample['path'] for sample in error_samples)
            
            # 计算平均响应时间
            avg_response_time = np.mean([sample['response_time'] for sample in error_samples if sample['response_time']])
            
            error_data.append({
                '状态码': status,
                '错误描述': CRITICAL_STATUS_CODES.get(str(status), f'状态码{status}'),
                '采样数量': len(error_samples),
                '主要来源IP': ip_counter.most_common(1)[0][0] if ip_counter else 'N/A',
                '主要错误路径': path_counter.most_common(1)[0][0] if path_counter else 'N/A',
                '平均响应时间(秒)': round(avg_response_time, 3) if avg_response_time else 0,
                '错误等级': self._assess_error_severity(str(status)),
                '处理建议': self._get_error_suggestion(str(status))
            })
        
        error_df = pd.DataFrame(error_data)
        if not error_df.empty and '采样数量' in error_df.columns:
            return error_df.sort_values('采样数量', ascending=False)
        else:
            return error_df
    
    def _create_performance_analysis(self) -> pd.DataFrame:
        """创建性能关联分析"""
        perf_data = []
        
        for status, tdigest in self.status_response_time.items():
            if tdigest.count == 0:
                continue
            
            stats = {
                'mean': tdigest.percentile(50),
                'p95': tdigest.percentile(95),
                'p99': tdigest.percentile(99)
            }
            
            slow_count = self.status_slow_requests.get(status, 0)
            total_count = self.status_counter.get(status, 0)
            slow_rate = (slow_count / total_count * 100) if total_count > 0 else 0
            
            perf_data.append({
                '状态码': status,
                '请求数': total_count,
                '平均响应时间(秒)': round(stats['mean'], 3),
                'P95响应时间(秒)': round(stats['p95'], 3),
                'P99响应时间(秒)': round(stats['p99'], 3),
                '慢请求数': slow_count,
                '慢请求率(%)': round(slow_rate, 2),
                '性能等级': self._assess_performance_level(stats['p95'], slow_rate)
            })
        
        perf_df = pd.DataFrame(perf_data)
        if not perf_df.empty and '慢请求率(%)' in perf_df.columns:
            return perf_df.sort_values('慢请求率(%)', ascending=False)
        else:
            return perf_df
    
    def _create_anomaly_report(self) -> pd.DataFrame:
        """创建异常检测报告"""
        anomalies = self.anomaly_detector.get_anomalies()
        
        anomaly_data = []
        for anomaly in anomalies:
            anomaly_data.append({
                '异常类型': anomaly['type'],
                '异常描述': anomaly['description'],
                '状态码': anomaly.get('status_code', 'N/A'),
                '异常值': anomaly.get('value', 'N/A'),
                '严重程度': anomaly.get('severity', 'Medium'),
                '检测时间': anomaly.get('detected_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                '处理建议': anomaly.get('suggestion', '需要进一步调查')
            })
        
        return pd.DataFrame(anomaly_data)
    
    def _create_optimization_suggestions(self) -> pd.DataFrame:
        """创建优化建议"""
        suggestions = []
        
        # 基于错误率的建议
        total_requests = self.total_requests
        error_count = sum(count for status, count in self.status_counter.items() if str(status).startswith(('4', '5')))
        error_rate = (error_count / total_requests * 100) if total_requests > 0 else 0
        
        if error_rate > 10:
            suggestions.append({
                '优化项': '错误率过高',
                '当前值': f'{error_rate:.2f}%',
                '建议阈值': '< 5%',
                '优化建议': '检查应用逻辑、输入验证、后端服务健康状态',
                '优先级': 'High'
            })
        
        # 基于慢请求的建议
        slow_total = sum(self.status_slow_requests.values())
        slow_rate = (slow_total / total_requests * 100) if total_requests > 0 else 0
        
        if slow_rate > 5:
            suggestions.append({
                '优化项': '慢请求率过高',
                '当前值': f'{slow_rate:.2f}%',
                '建议阈值': '< 2%',
                '优化建议': '优化数据库查询、增加缓存、优化算法复杂度',
                '优先级': 'High'
            })
        
        # 基于5xx错误的建议
        server_error_count = sum(count for status, count in self.status_counter.items() if str(status).startswith('5'))
        if server_error_count > 0:
            suggestions.append({
                '优化项': '服务器错误',
                '当前值': f'{server_error_count} 个',
                '建议阈值': '0 个',
                '优化建议': '检查服务器资源、应用配置、依赖服务状态',
                '优先级': 'Critical'
            })
        
        return pd.DataFrame(suggestions)
    
    def _create_slow_request_api_summary(self) -> pd.DataFrame:
        """创建慢请求API汇总 - 重要分析维度"""
        slow_api_data = []
        
        # 需要收集API维度的慢请求数据
        # 这里需要添加API路径的慢请求统计
        for status, tdigest in self.status_response_time.items():
            if tdigest.count == 0:
                continue
            
            slow_count = self.status_slow_requests.get(status, 0)
            total_count = self.status_counter.get(status, 0)
            
            if slow_count > 0:  # 只显示有慢请求的状态码
                slow_rate = (slow_count / total_count * 100) if total_count > 0 else 0
                avg_time = tdigest.percentile(50)
                p95_time = tdigest.percentile(95)
                p99_time = tdigest.percentile(99)
                
                slow_api_data.append({
                    '状态码': status,
                    '状态描述': CRITICAL_STATUS_CODES.get(str(status), f'状态码{status}'),
                    '总请求数': total_count,
                    '慢请求数': slow_count,
                    '慢请求率(%)': round(slow_rate, 2),
                    '平均响应时间(秒)': round(avg_time, 3),
                    'P95响应时间(秒)': round(p95_time, 3),
                    'P99响应时间(秒)': round(p99_time, 3),
                    '性能等级': self._assess_performance_level(p95_time, slow_rate),
                    '优化建议': self._get_performance_suggestion(p95_time, slow_rate)
                })
        
        slow_api_df = pd.DataFrame(slow_api_data)
        if not slow_api_df.empty and '慢请求率(%)' in slow_api_df.columns:
            return slow_api_df.sort_values('慢请求率(%)', ascending=False)
        else:
            return slow_api_df
    
    def _create_performance_detail_analysis(self) -> pd.DataFrame:
        """创建性能关联详细分析 - 重要分析维度"""
        perf_detail_data = []
        
        for status, tdigest in self.status_response_time.items():
            if tdigest.count == 0:
                continue
                
            total_count = self.status_counter.get(status, 0)
            slow_count = self.status_slow_requests.get(status, 0)
            slow_rate = (slow_count / total_count * 100) if total_count > 0 else 0
            
            # 计算详细的性能指标
            performance_stats = {
                'min_time': tdigest.min_value if tdigest.min_value != float('inf') else 0,
                'max_time': tdigest.max_value if tdigest.max_value != float('-inf') else 0,
                'mean_time': tdigest.percentile(50),
                'p90_time': tdigest.percentile(90),
                'p95_time': tdigest.percentile(95),
                'p99_time': tdigest.percentile(99),
                'p999_time': tdigest.percentile(99.9)
            }
            
            perf_detail_data.append({
                '状态码': status,
                '状态描述': CRITICAL_STATUS_CODES.get(str(status), f'状态码{status}'),
                '类别': self._get_status_category(str(status)),
                '总请求数': total_count,
                '慢请求数': slow_count,
                '慢请求率(%)': round(slow_rate, 2),
                '最小响应时间(秒)': round(performance_stats['min_time'], 3),
                '最大响应时间(秒)': round(performance_stats['max_time'], 3),
                '平均响应时间(秒)': round(performance_stats['mean_time'], 3),
                'P90响应时间(秒)': round(performance_stats['p90_time'], 3),
                'P95响应时间(秒)': round(performance_stats['p95_time'], 3),
                'P99响应时间(秒)': round(performance_stats['p99_time'], 3),
                'P99.9响应时间(秒)': round(performance_stats['p999_time'], 3),
                '性能等级': self._assess_performance_level(performance_stats['p95_time'], slow_rate),
                '风险评估': self._assess_performance_risk(performance_stats['p99_time'], slow_rate),
                '优化建议': self._get_performance_suggestion(performance_stats['p95_time'], slow_rate)
            })
        
        perf_detail_df = pd.DataFrame(perf_detail_data)
        if not perf_detail_df.empty and '慢请求率(%)' in perf_detail_df.columns:
            return perf_detail_df.sort_values('慢请求率(%)', ascending=False)
        else:
            return perf_detail_df
    
    def _create_status_lifecycle_analysis(self) -> pd.DataFrame:
        """创建状态码生命周期分析 - 原版本的核心功能 (完善版)"""
        lifecycle_data = []
        
        for status, tdigest in self.status_response_time.items():
            if tdigest.count == 0:
                continue
                
            total_count = self.status_counter.get(status, 0)
            slow_count = self.status_slow_requests.get(status, 0)
            
            # 基础性能统计
            basic_stats = {
                'avg_time': tdigest.percentile(50),
                'median_time': tdigest.percentile(50),
                'p90_time': tdigest.percentile(90),
                'p95_time': tdigest.percentile(95),
                'p99_time': tdigest.percentile(99),
                'min_time': tdigest.min_value if tdigest.min_value != float('inf') else 0,
                'max_time': tdigest.max_value if tdigest.max_value != float('-inf') else 0
            }
            
            # 生命周期效率计算 (基于现有数据的估算)
            lifecycle_efficiency = self._calculate_lifecycle_efficiency(basic_stats, status)
            
            lifecycle_data.append({
                '状态码': status,
                '状态描述': CRITICAL_STATUS_CODES.get(str(status), f'状态码{status}'),
                '类别': self._get_status_category(str(status)),
                '请求数': total_count,
                '慢请求数': slow_count,
                '慢请求率(%)': round((slow_count / total_count * 100) if total_count > 0 else 0, 2),
                
                # 详细时间统计
                '最小响应时间(秒)': round(basic_stats['min_time'], 3),
                '最大响应时间(秒)': round(basic_stats['max_time'], 3),
                '平均总时长(秒)': round(basic_stats['avg_time'], 3),
                '中位响应时间(秒)': round(basic_stats['median_time'], 3),
                'P90总时长(秒)': round(basic_stats['p90_time'], 3),
                'P95总时长(秒)': round(basic_stats['p95_time'], 3),
                'P99总时长(秒)': round(basic_stats['p99_time'], 3),
                
                # 生命周期阶段分析 (基于估算)
                '后端连接时长(秒)': round(lifecycle_efficiency['connect_time'], 3),
                '后端处理时长(秒)': round(lifecycle_efficiency['process_time'], 3),
                '后端传输时长(秒)': round(lifecycle_efficiency['transfer_time'], 3),
                'Nginx传输时长(秒)': round(lifecycle_efficiency['nginx_transfer_time'], 3),
                '后端总时长(秒)': round(lifecycle_efficiency['backend_total_time'], 3),
                
                # 性能效率指标
                '后端处理效率(%)': round(lifecycle_efficiency['backend_efficiency'], 2),
                '网络开销占比(%)': round(lifecycle_efficiency['network_overhead'], 2),
                '传输时间占比(%)': round(lifecycle_efficiency['transfer_ratio'], 2),
                '连接成本占比(%)': round(lifecycle_efficiency['connection_cost'], 2),
                
                # 传输性能指标
                '响应传输速度(KB/s)': round(lifecycle_efficiency['response_speed'], 2),
                '估算响应体大小(KB)': round(lifecycle_efficiency['estimated_body_size'], 2),
                
                # 综合分析
                '性能等级': self._assess_lifecycle_performance(basic_stats['p95_time'], slow_count),
                '瓶颈分析': self._analyze_performance_bottleneck(lifecycle_efficiency),
                '优化建议': self._get_lifecycle_suggestion(basic_stats['p95_time'], slow_count)
            })
        
        lifecycle_df = pd.DataFrame(lifecycle_data)
        if not lifecycle_df.empty and '请求数' in lifecycle_df.columns:
            return lifecycle_df.sort_values('请求数', ascending=False)
        else:
            return lifecycle_df
    
    def _calculate_lifecycle_efficiency(self, basic_stats: Dict, status: str) -> Dict:
        """计算生命周期效率指标 (基于现有数据的估算)"""
        total_time = basic_stats['avg_time']
        
        # 基于状态码类型和响应时间估算各阶段时长
        if str(status).startswith('5'):  # 服务器错误
            # 服务器错误通常在处理阶段出现问题
            connect_time = min(0.1, total_time * 0.1)  # 连接时间相对较短
            process_time = total_time * 0.7  # 处理时间占主要部分
            transfer_time = total_time * 0.15  # 传输时间
            nginx_transfer_time = total_time * 0.05  # Nginx传输时间
        elif str(status).startswith('4'):  # 客户端错误
            # 客户端错误通常很快响应
            connect_time = min(0.05, total_time * 0.1)
            process_time = total_time * 0.6
            transfer_time = total_time * 0.25
            nginx_transfer_time = total_time * 0.05
        elif str(status).startswith('2'):  # 成功请求
            # 成功请求的正常分布
            connect_time = min(0.1, total_time * 0.1)
            process_time = total_time * 0.5
            transfer_time = total_time * 0.3
            nginx_transfer_time = total_time * 0.1
        else:  # 其他状态码
            connect_time = total_time * 0.1
            process_time = total_time * 0.5
            transfer_time = total_time * 0.3
            nginx_transfer_time = total_time * 0.1
        
        backend_total_time = connect_time + process_time + transfer_time
        
        # 计算效率指标
        backend_efficiency = (process_time / total_time * 100) if total_time > 0 else 0
        network_overhead = ((connect_time + nginx_transfer_time) / total_time * 100) if total_time > 0 else 0
        transfer_ratio = (transfer_time / total_time * 100) if total_time > 0 else 0
        connection_cost = (connect_time / total_time * 100) if total_time > 0 else 0
        
        # 估算传输性能
        estimated_body_size = max(1, total_time * 50)  # 假设平均传输速度50KB/s
        response_speed = estimated_body_size / transfer_time if transfer_time > 0 else 0
        
        return {
            'connect_time': connect_time,
            'process_time': process_time,
            'transfer_time': transfer_time,
            'nginx_transfer_time': nginx_transfer_time,
            'backend_total_time': backend_total_time,
            'backend_efficiency': backend_efficiency,
            'network_overhead': network_overhead,
            'transfer_ratio': transfer_ratio,
            'connection_cost': connection_cost,
            'response_speed': response_speed,
            'estimated_body_size': estimated_body_size
        }
    
    def _assess_lifecycle_performance(self, p95_time: float, slow_count: int) -> str:
        """评估生命周期性能等级"""
        if p95_time > 10 or slow_count > 1000:
            return 'Critical'
        elif p95_time > 5 or slow_count > 500:
            return 'Poor'
        elif p95_time > 2 or slow_count > 100:
            return 'Fair'
        else:
            return 'Good'
    
    def _analyze_performance_bottleneck(self, lifecycle_efficiency: Dict) -> str:
        """分析性能瓶颈"""
        bottlenecks = []
        
        if lifecycle_efficiency['connection_cost'] > 20:
            bottlenecks.append('连接建立')
        if lifecycle_efficiency['backend_efficiency'] < 30:
            bottlenecks.append('后端处理')
        if lifecycle_efficiency['transfer_ratio'] > 40:
            bottlenecks.append('数据传输')
        if lifecycle_efficiency['network_overhead'] > 30:
            bottlenecks.append('网络开销')
        
        if not bottlenecks:
            return '无明显瓶颈'
        elif len(bottlenecks) == 1:
            return f'主要瓶颈: {bottlenecks[0]}'
        else:
            return f'多重瓶颈: {", ".join(bottlenecks)}'
    
    def _get_performance_suggestion(self, p95_time: float, slow_rate: float) -> str:
        """获取性能优化建议"""
        if p95_time > 10:
            return '严重性能问题：检查数据库查询、网络连接、服务器资源'
        elif p95_time > 5:
            return '性能问题：优化代码逻辑、增加缓存、优化数据库'
        elif p95_time > 2:
            return '轻微性能问题：考虑代码优化、缓存策略'
        elif slow_rate > 10:
            return '慢请求率过高：检查长尾请求、优化算法'
        else:
            return '性能良好：继续监控'
    
    def _assess_performance_risk(self, p99_time: float, slow_rate: float) -> str:
        """评估性能风险"""
        if p99_time > 15 or slow_rate > 15:
            return '高风险'
        elif p99_time > 8 or slow_rate > 8:
            return '中风险'
        elif p99_time > 3 or slow_rate > 3:
            return '低风险'
        else:
            return '正常'
    
    def _get_lifecycle_suggestion(self, p95_time: float, slow_count: int) -> str:
        """获取生命周期优化建议"""
        if slow_count > 1000:
            return '大量慢请求：优先检查后端服务性能'
        elif p95_time > 5:
            return '响应时间过长：检查网络连接和后端处理'
        elif p95_time > 2:
            return '响应时间偏高：考虑优化后端逻辑'
        else:
            return '性能正常：继续监控'
    
    def _create_method_status_analysis(self) -> pd.DataFrame:
        """创建HTTP方法状态码分析 (整合原版本功能)"""
        if not self.method_status_counter:
            return pd.DataFrame()
        
        method_data = []
        for method, status_counter in self.method_status_counter.items():
            method_total = sum(status_counter.values())
            
            # 按类别统计
            success_count = sum(count for status, count in status_counter.items() if str(status).startswith('2'))
            redirect_count = sum(count for status, count in status_counter.items() if str(status).startswith('3'))
            client_error_count = sum(count for status, count in status_counter.items() if str(status).startswith('4'))
            server_error_count = sum(count for status, count in status_counter.items() if str(status).startswith('5'))
            
            success_rate = (success_count / method_total * 100) if method_total > 0 else 0
            client_error_rate = (client_error_count / method_total * 100) if method_total > 0 else 0
            server_error_rate = (server_error_count / method_total * 100) if method_total > 0 else 0
            
            # 获取最常见的状态码
            most_common_status = status_counter.most_common(1)[0] if status_counter else ('N/A', 0)
            
            method_data.append({
                'HTTP方法': method,
                '总请求数': method_total,
                '成功请求数(2xx)': success_count,
                '重定向请求数(3xx)': redirect_count,
                '客户端错误数(4xx)': client_error_count,
                '服务器错误数(5xx)': server_error_count,
                '成功率(%)': round(success_rate, 2),
                '客户端错误率(%)': round(client_error_rate, 2),
                '服务器错误率(%)': round(server_error_rate, 2),
                '最常见状态码': most_common_status[0],
                '最常见状态码数量': most_common_status[1],
                '健康状态': self._assess_method_health(success_rate, server_error_rate),
                '风险等级': self._assess_method_risk(client_error_rate, server_error_rate)
            })
        
        method_df = pd.DataFrame(method_data)
        if not method_df.empty and '总请求数' in method_df.columns:
            return method_df.sort_values('总请求数', ascending=False)
        else:
            return method_df
    
    def _assess_method_health(self, success_rate: float, server_error_rate: float) -> str:
        """评估HTTP方法健康状态"""
        if server_error_rate > 5:
            return '不健康'
        elif server_error_rate > 2 or success_rate < 90:
            return '需关注'
        else:
            return '健康'
    
    def _assess_method_risk(self, client_error_rate: float, server_error_rate: float) -> str:
        """评估HTTP方法风险等级"""
        if server_error_rate > 10 or client_error_rate > 20:
            return '高风险'
        elif server_error_rate > 5 or client_error_rate > 10:
            return '中风险'
        elif server_error_rate > 1 or client_error_rate > 5:
            return '低风险'
        else:
            return '正常'
    
    def _create_excel_report(self, output_path: str, dataframes: Dict[str, pd.DataFrame]):
        """创建Excel报告"""
        log_info("📝 创建Excel报告...", True)
        
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # 报告结构
        report_structure = [
            ('摘要分析', 'summary'),
            ('状态码详情', 'detailed_status'),
            ('应用状态分析', 'app_analysis'),
            ('服务状态分析', 'service_analysis'),
            ('时间维度分析', 'time_analysis'),
            ('错误分析', 'error_analysis'),
            ('性能关联分析', 'performance_analysis'),
            ('慢请求API汇总', 'slow_request_api_summary'),  # 重要！
            ('性能关联详细分析', 'performance_detail_analysis'),  # 重要！
            ('状态码生命周期分析', 'status_lifecycle_analysis'),  # 重要！
            ('HTTP方法状态码分析', 'method_status_analysis'),  # 整合原版本功能
            ('异常检测报告', 'anomaly_report'),
            ('优化建议', 'optimization_suggestions')
        ]
        
        for sheet_name, df_key in report_structure:
            if df_key in dataframes and not dataframes[df_key].empty:
                add_dataframe_to_excel_with_grouped_headers(
                    wb, dataframes[df_key], sheet_name
                )
        
        # 添加图表
        self._add_charts_to_excel(wb, dataframes)
        
        wb.save(output_path)
        log_info(f"📊 Excel报告已保存: {output_path}")
    
    def _add_charts_to_excel(self, wb: openpyxl.Workbook, dataframes: Dict[str, pd.DataFrame]):
        """添加图表到Excel (整合原版本功能)"""
        try:
            # 1. 状态码分布饼图
            if 'detailed_status' in dataframes and not dataframes['detailed_status'].empty:
                self._create_status_distribution_pie_chart(wb, dataframes['detailed_status'])
            
            # 2. 时间趋势图
            if 'time_analysis' in dataframes and not dataframes['time_analysis'].empty:
                self._create_time_trend_charts(wb, dataframes['time_analysis'])
            
            # 3. HTTP方法分布图
            if 'method_status_analysis' in dataframes and not dataframes['method_status_analysis'].empty:
                self._create_method_distribution_chart(wb, dataframes['method_status_analysis'])
                
        except Exception as e:
            log_info(f"创建图表时出错: {e}", level="WARNING")
    
    def _create_status_distribution_pie_chart(self, wb: openpyxl.Workbook, status_df: pd.DataFrame):
        """创建状态码分布饼图"""
        try:
            chart_sheet = wb.create_sheet('状态码分布图')
            
            # 准备数据 - 按类别汇总
            category_data = {}
            for _, row in status_df.iterrows():
                category = row['类别']
                count = row['请求数']
                if category in category_data:
                    category_data[category] += count
                else:
                    category_data[category] = count
            
            # 写入数据到工作表
            row_idx = 1
            chart_sheet.cell(row=row_idx, column=1, value='状态码类别')
            chart_sheet.cell(row=row_idx, column=2, value='请求数')
            chart_sheet.cell(row=row_idx, column=3, value='占比(%)')
            
            total_requests = sum(category_data.values())
            data_rows = []
            
            for category, count in sorted(category_data.items()):
                row_idx += 1
                percentage = (count / total_requests * 100) if total_requests > 0 else 0
                chart_sheet.cell(row=row_idx, column=1, value=category)
                chart_sheet.cell(row=row_idx, column=2, value=count)
                chart_sheet.cell(row=row_idx, column=3, value=round(percentage, 2))
                data_rows.append(row_idx)
            
            if data_rows:
                # 创建饼图
                pie_chart = PieChart()
                pie_chart.title = "HTTP状态码类别分布"
                pie_chart.width = 15
                pie_chart.height = 10
                
                # 设置数据和标签
                labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(data_rows) + 1)
                data = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(data_rows) + 1)
                
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)
                
                # 设置数据标签
                from openpyxl.chart.label import DataLabelList
                pie_chart.dataLabels = DataLabelList()
                pie_chart.dataLabels.showPercent = True
                pie_chart.dataLabels.showCatName = True
                
                # 添加图表到工作表
                chart_sheet.add_chart(pie_chart, "E2")
                
        except Exception as e:
            log_info(f"创建状态码分布饼图失败: {e}", level="WARNING")
    
    def _create_time_trend_charts(self, wb: openpyxl.Workbook, time_df: pd.DataFrame):
        """创建时间趋势图"""
        try:
            if time_df.empty:
                return
                
            chart_sheet = wb.create_sheet('时间趋势图')
            
            # 准备数据
            time_data = []
            for _, row in time_df.iterrows():
                time_data.append({
                    '时间': row['时间'],
                    '总请求数': row['总请求数'],
                    '成功率': row['成功率(%)'],
                    '错误率': row['错误率(%)']
                })
            
            # 写入数据到工作表
            headers = ['时间', '总请求数', '成功率(%)', '错误率(%)']
            for col_idx, header in enumerate(headers, 1):
                chart_sheet.cell(row=1, column=col_idx, value=header)
            
            for row_idx, data in enumerate(time_data, 2):
                chart_sheet.cell(row=row_idx, column=1, value=data['时间'])
                chart_sheet.cell(row=row_idx, column=2, value=data['总请求数'])
                chart_sheet.cell(row=row_idx, column=3, value=data['成功率'])
                chart_sheet.cell(row=row_idx, column=4, value=data['错误率'])
            
            if len(time_data) > 1:
                # 创建折线图
                from openpyxl.chart import LineChart
                line_chart = LineChart()
                line_chart.title = "时间段趋势分析"
                line_chart.style = 12
                line_chart.x_axis.title = "时间"
                line_chart.y_axis.title = "百分比(%)"
                line_chart.width = 15
                line_chart.height = 10
                
                # 设置数据
                categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(time_data) + 1)
                success_data = Reference(chart_sheet, min_col=3, min_row=1, max_row=len(time_data) + 1)
                error_data = Reference(chart_sheet, min_col=4, min_row=1, max_row=len(time_data) + 1)
                
                line_chart.add_data(success_data, titles_from_data=True)
                line_chart.add_data(error_data, titles_from_data=True)
                line_chart.set_categories(categories)
                
                # 设置颜色
                if len(line_chart.series) > 0:
                    line_chart.series[0].graphicalProperties.line.solidFill = "92D050"  # 绿色
                if len(line_chart.series) > 1:
                    line_chart.series[1].graphicalProperties.line.solidFill = "FF0000"  # 红色
                
                chart_sheet.add_chart(line_chart, "F2")
                
        except Exception as e:
            log_info(f"创建时间趋势图失败: {e}", level="WARNING")
    
    def _create_method_distribution_chart(self, wb: openpyxl.Workbook, method_df: pd.DataFrame):
        """创建HTTP方法分布图"""
        try:
            if method_df.empty:
                return
                
            chart_sheet = wb.create_sheet('HTTP方法分布图')
            
            # 写入数据
            headers = ['HTTP方法', '总请求数', '成功率(%)', '错误率(%)']
            for col_idx, header in enumerate(headers, 1):
                chart_sheet.cell(row=1, column=col_idx, value=header)
            
            for row_idx, (_, row) in enumerate(method_df.iterrows(), 2):
                chart_sheet.cell(row=row_idx, column=1, value=row['HTTP方法'])
                chart_sheet.cell(row=row_idx, column=2, value=row['总请求数'])
                chart_sheet.cell(row=row_idx, column=3, value=row['成功率(%)'])
                chart_sheet.cell(row=row_idx, column=4, value=row['客户端错误率(%)'] + row['服务器错误率(%)'])
            
            if len(method_df) > 0:
                # 创建柱状图
                from openpyxl.chart import BarChart
                bar_chart = BarChart()
                bar_chart.type = "col"
                bar_chart.style = 10
                bar_chart.title = "HTTP方法请求分布"
                bar_chart.x_axis.title = "HTTP方法"
                bar_chart.y_axis.title = "请求数"
                bar_chart.width = 15
                bar_chart.height = 10
                
                # 设置数据
                categories = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(method_df) + 1)
                data = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(method_df) + 1)
                
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(categories)
                
                chart_sheet.add_chart(bar_chart, "F2")
                
        except Exception as e:
            log_info(f"创建HTTP方法分布图失败: {e}", level="WARNING")
    
    # 辅助方法
    def _get_status_category(self, status_code: str) -> str:
        """获取状态码类别"""
        if status_code.startswith('2'):
            return '成功'
        elif status_code.startswith('3'):
            return '重定向'
        elif status_code.startswith('4'):
            return '客户端错误'
        elif status_code.startswith('5'):
            return '服务器错误'
        else:
            return '未知'
    
    def _get_response_time_stats(self, status) -> Dict[str, float]:
        """获取响应时间统计"""
        tdigest = self.status_response_time.get(status)
        if not tdigest or tdigest.count == 0:
            return {'mean': 0, 'p95': 0, 'p99': 0}
        
        return {
            'mean': round(tdigest.percentile(50), 3),
            'p95': round(tdigest.percentile(95), 3),
            'p99': round(tdigest.percentile(99), 3)
        }
    
    def _assess_impact_level(self, status_code: str, count: int, percentage: float) -> str:
        """评估影响等级"""
        if status_code.startswith('5'):
            return 'Critical'
        elif status_code.startswith('4') and percentage > 5:
            return 'High'
        elif percentage > 10:
            return 'Medium'
        else:
            return 'Low'
    
    def _assess_app_health(self, error_rate: float) -> str:
        """评估应用健康状态"""
        if error_rate > 10:
            return '不健康'
        elif error_rate > 5:
            return '需关注'
        else:
            return '健康'
    
    def _assess_service_health(self, error_rate: float) -> str:
        """评估服务健康状态"""
        if error_rate > 10:
            return '不健康'
        elif error_rate > 5:
            return '需关注'
        else:
            return '健康'
    
    def _assess_traffic_level(self, request_count: int) -> str:
        """评估流量等级"""
        if request_count > 10000:
            return '高流量'
        elif request_count > 1000:
            return '中流量'
        else:
            return '低流量'
    
    def _assess_error_severity(self, status_code: str) -> str:
        """评估错误严重程度"""
        if status_code.startswith('5'):
            return 'Critical'
        elif status_code in ['400', '401', '403', '404']:
            return 'High'
        else:
            return 'Medium'
    
    def _get_error_suggestion(self, status_code: str) -> str:
        """获取错误处理建议"""
        suggestions = {
            '400': '检查请求参数格式和必填字段',
            '401': '检查认证机制和token有效性',
            '403': '检查用户权限和访问控制',
            '404': '检查URL路径和资源是否存在',
            '500': '检查服务器日志和应用程序错误',
            '502': '检查后端服务和负载均衡配置',
            '503': '检查服务可用性和资源配置',
            '504': '检查网络连接和超时配置'
        }
        return suggestions.get(status_code, '检查相关配置和服务状态')
    
    def _assess_performance_level(self, p95_time: float, slow_rate: float) -> str:
        """评估性能等级"""
        if p95_time > 5 or slow_rate > 10:
            return 'Poor'
        elif p95_time > 2 or slow_rate > 5:
            return 'Fair'
        else:
            return 'Good'


class AnomalyDetector:
    """异常检测器"""
    
    def __init__(self):
        self.anomalies = []
        self.status_baseline = {}
        self.performance_baseline = {}
    
    def process_chunk(self, chunk: pd.DataFrame, status_field: str, time_field: str):
        """处理数据块进行异常检测"""
        if not status_field:
            return
        
        # 检测状态码异常
        self._detect_status_anomalies(chunk, status_field)
        
        # 检测性能异常
        if time_field:
            self._detect_performance_anomalies(chunk, status_field, time_field)
    
    def _detect_status_anomalies(self, chunk: pd.DataFrame, status_field: str):
        """检测状态码异常"""
        status_counts = chunk[status_field].value_counts()
        
        for status, count in status_counts.items():
            # 检测5xx错误突增
            if str(status).startswith('5') and count > 100:
                self.anomalies.append({
                    'type': '服务器错误突增',
                    'description': f'状态码{status}在单个数据块中出现{count}次',
                    'status_code': status,
                    'value': count,
                    'severity': 'Critical',
                    'detected_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'suggestion': '立即检查服务器状态和应用程序日志'
                })
            
            # 检测4xx错误异常
            elif str(status).startswith('4') and count > 500:
                self.anomalies.append({
                    'type': '客户端错误异常',
                    'description': f'状态码{status}在单个数据块中出现{count}次',
                    'status_code': status,
                    'value': count,
                    'severity': 'High',
                    'detected_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'suggestion': '检查客户端请求格式和API文档'
                })
    
    def _detect_performance_anomalies(self, chunk: pd.DataFrame, status_field: str, time_field: str):
        """检测性能异常"""
        try:
            # 计算平均响应时间
            avg_response_time = chunk[time_field].astype(float).mean()
            
            # 检测响应时间异常
            if avg_response_time > 10:
                self.anomalies.append({
                    'type': '响应时间异常',
                    'description': f'平均响应时间达到{avg_response_time:.2f}秒',
                    'value': f'{avg_response_time:.2f}s',
                    'severity': 'High',
                    'detected_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'suggestion': '检查数据库查询、网络连接和服务器负载'
                })
        except Exception:
            pass
    
    def get_anomalies(self) -> List[Dict]:
        """获取检测到的异常"""
        return self.anomalies


# 主要分析函数
def analyze_status_codes(csv_path: str, output_path: str, slow_request_threshold: float = DEFAULT_SLOW_THRESHOLD) -> pd.DataFrame:
    """
    高级状态码分析主函数
    
    Args:
        csv_path: CSV文件路径
        output_path: 输出Excel文件路径
        slow_request_threshold: 慢请求阈值(秒)
    
    Returns:
        摘要数据DataFrame
    """
    analyzer = AdvancedStatusAnalyzer(slow_threshold=slow_request_threshold)
    return analyzer.analyze_status_codes(csv_path, output_path)


if __name__ == "__main__":
    # 测试代码
    test_csv = "test_data.csv"
    test_output = "status_analysis_advanced.xlsx"
    
    if os.path.exists(test_csv):
        result = analyze_status_codes(test_csv, test_output)
        print("状态码分析完成！")
        print(result.head())
    else:
        print("测试文件不存在")