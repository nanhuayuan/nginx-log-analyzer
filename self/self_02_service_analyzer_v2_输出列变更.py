import gc
import numpy as np
import pandas as pd
from openpyxl import Workbook

from self_00_04_excel_processor import add_dataframe_to_excel_with_grouped_headers, format_excel_sheet
from self_00_01_constants import DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD, TIME_METRICS, SIZE_METRICS, PERCENTILES, \
    TIME_METRICS_MAPPING


def analyze_service_performance(csv_path, output_path, success_codes=None, slow_threshold=DEFAULT_SLOW_THRESHOLD):
    """分析服务性能并创建Excel报告"""
    if success_codes is None:
        from self_00_01_constants import DEFAULT_SUCCESS_CODES
        success_codes = DEFAULT_SUCCESS_CODES

    # 初始化统计数据
    chunk_size = DEFAULT_CHUNK_SIZE
    service_stats = {}
    app_stats = {}

    total_requests = 0
    success_requests = 0
    success_codes = [str(code) for code in success_codes]

    # 分块读取和处理CSV文件
    for chunk_idx, chunk in enumerate(pd.read_csv(csv_path, chunksize=chunk_size)):
        chunk_rows = len(chunk)
        total_requests += chunk_rows

        # 处理服务和应用统计
        for service, group in chunk.groupby('service_name'):
            if not service or service == '':
                service = 'unknown'

            # 初始化服务统计数据
            if service not in service_stats:
                initialize_stats(service_stats, service, group)
            service_stats[service]['total_requests'] += len(group)

            # 处理应用统计数据
            app_name = group['app_name'].iloc[0] if not group['app_name'].empty else "unknown"
            if app_name not in app_stats:
                app_stats[app_name] = {'app_name': app_name, 'total_requests': 0, 'success_requests': 0,
                                       'slow_requests_count': 0}
                initialize_metrics_stats(app_stats[app_name])
            app_stats[app_name]['total_requests'] += len(group)

        # 处理成功请求
        successful_requests = chunk[chunk['status'].astype(str).isin(success_codes)]
        success_count = len(successful_requests)
        success_requests += success_count

        # 更新服务级别的成功请求统计
        for service, group in successful_requests.groupby('service_name'):
            if not service or service == '':
                service = 'unknown'

            if service not in service_stats:
                initialize_stats(service_stats, service, group)

            service_stats[service]['success_requests'] += len(group)

            # 更新时间和大小统计
            update_metrics_stats(service_stats[service], group)

            # 更新慢请求统计
            request_times = pd.to_numeric(group['request_time'], errors='coerce')
            slow_requests_count = (request_times > slow_threshold).sum()
            service_stats[service]['slow_requests_count'] += slow_requests_count

            # 更新应用级别统计
            app_name = group['app_name'].iloc[0] if not group['app_name'].empty else "unknown"
            app_stats[app_name]['success_requests'] += len(group)
            update_metrics_stats(app_stats[app_name], group)
            app_stats[app_name]['slow_requests_count'] += slow_requests_count

        # 内存管理
        if (chunk_idx + 1) % 5 == 0:
            del chunk
            gc.collect()

    # 计算最终统计结果
    service_results = calculate_stats_results(service_stats, success_requests, slow_threshold, stats_type='service')
    app_results = calculate_stats_results(app_stats, success_requests, slow_threshold, stats_type='app')

    # 创建Excel报告
    create_performance_excel(service_results, app_results, output_path, total_requests, success_requests)

    return service_results.head(5) if not service_results.empty else pd.DataFrame()


def initialize_stats(stats_dict, key, group):
    """初始化统计数据结构"""
    app_name = group['app_name'].iloc[0] if not group['app_name'].empty else "unknown"

    stats = {
        'service_name': key,
        'app_name': app_name,
        'total_requests': 0,
        'success_requests': 0,
        'slow_requests_count': 0,
    }

    initialize_metrics_stats(stats)
    stats_dict[key] = stats


def initialize_metrics_stats(stats):
    """初始化指标统计数据"""
    for metric in TIME_METRICS:
        if metric == 'upstream_connect_phase':
            continue
        stats[f'{metric}_min'] = float('inf')
        stats[f'{metric}_max'] = 0
        stats[f'{metric}_total'] = 0
        stats[f'{metric}_values'] = []

    for metric in SIZE_METRICS:
        stats[f'{metric}_min'] = float('inf')
        stats[f'{metric}_max'] = 0
        stats[f'{metric}_total'] = 0
        stats[f'{metric}_values'] = []


def update_metrics_stats(stats, group):
    """更新指标统计数据"""
    # 更新时间指标
    for metric in TIME_METRICS:
        if metric == 'upstream_connect_phase':
            continue
        times = pd.to_numeric(group[metric], errors='coerce')
        update_stat_values(stats, metric, times)

    # 更新大小指标
    for metric in SIZE_METRICS:
        sizes = pd.to_numeric(group[metric], errors='coerce')
        update_stat_values(stats, metric, sizes)


def update_stat_values(stats, field, value_series):
    """更新统计指标的值"""
    valid_values = value_series.dropna()
    if len(valid_values) == 0:
        return

    min_value = valid_values.min()
    max_value = valid_values.max()
    total_value = valid_values.sum()

    min_field = f"{field}_min"
    max_field = f"{field}_max"
    total_field = f"{field}_total"
    values_field = f"{field}_values"

    if min_field not in stats or max_field not in stats or total_field not in stats or values_field not in stats:
        return

    stats[min_field] = min(stats[min_field], min_value) if stats[min_field] != float('inf') else min_value
    stats[max_field] = max(stats[max_field], max_value)
    stats[total_field] += total_value
    stats[values_field].extend(valid_values.tolist())


def calculate_stats_results(stats_dict, success_requests, slow_threshold, stats_type='service'):
    """计算指标结果"""
    results = []

    for key, stats in stats_dict.items():
        if stats['success_requests'] == 0:
            continue

        result = {
            f'{stats_type}_name': key,
        }

        # 服务类型需要额外显示应用名称
        if stats_type == 'service':
            result['app_name'] = stats['app_name']

        # 计算基本指标
        result.update({
            'total_requests': stats['total_requests'],
            'request_count': stats['success_requests'],
            'request_percentage': round(stats['success_requests'] / success_requests * 100,
                                        2) if success_requests > 0 else 0,
            'success_rate': round(stats['success_requests'] / stats['total_requests'] * 100, 2) if stats[
                                                                                                       'total_requests'] > 0 else 0,
            'slow_requests_count': stats['slow_requests_count'],
            'slow_requests_percentage': round(stats['slow_requests_count'] / stats['success_requests'] * 100, 2) if
            stats['success_requests'] > 0 else 0,
        })

        # 计算时间指标
        for metric in TIME_METRICS:
            if metric == 'upstream_connect_phase':
                continue

            values = np.array(stats[f'{metric}_values'])
            if len(values) > 0:
                result[f'avg_{metric}'] = stats[f'{metric}_total'] / stats['success_requests']
                result[f'min_{metric}'] = stats[f'{metric}_min'] if stats[f'{metric}_min'] != float('inf') else 0
                result[f'max_{metric}'] = stats[f'{metric}_max']
                result[f'median_{metric}'] = np.median(values)
                for percentile in PERCENTILES:
                    result[f'p{percentile}_{metric}'] = np.percentile(values, percentile)
            else:
                result[f'avg_{metric}'] = 0
                result[f'min_{metric}'] = 0
                result[f'max_{metric}'] = 0
                result[f'median_{metric}'] = 0
                for percentile in PERCENTILES:
                    result[f'p{percentile}_{metric}'] = 0

        # 计算大小指标
        for metric in SIZE_METRICS:
            values = np.array(stats[f'{metric}_values'])
            if len(values) > 0:
                kb_factor = 1024
                result[f'avg_{metric}_kb'] = stats[f'{metric}_total'] / stats['success_requests'] / kb_factor
                result[f'min_{metric}_kb'] = stats[f'{metric}_min'] / kb_factor if stats[f'{metric}_min'] != float(
                    'inf') else 0
                result[f'max_{metric}_kb'] = stats[f'{metric}_max'] / kb_factor
                result[f'median_{metric}_kb'] = np.median(values) / kb_factor
                for percentile in PERCENTILES:
                    result[f'p{percentile}_{metric}_kb'] = np.percentile(values, percentile) / kb_factor
            else:
                result[f'avg_{metric}_kb'] = 0
                result[f'min_{metric}_kb'] = 0
                result[f'max_{metric}_kb'] = 0
                result[f'median_{metric}_kb'] = 0
                for percentile in PERCENTILES:
                    result[f'p{percentile}_{metric}_kb'] = 0

        # 清理内存 - 统计值不再需要
        for key in stats:
            if key.endswith('_values'):
                stats[key] = None

        results.append(result)

    if not results:
        return pd.DataFrame()

    # 创建DataFrame并排序
    results_df = pd.DataFrame(results)
    if not results_df.empty and 'avg_request_time' in results_df.columns:
        results_df = results_df.sort_values(by='avg_request_time', ascending=False)

    # 应用列名映射
    column_mapping = create_column_mapping(stats_type)
    results_df = results_df.rename(columns=column_mapping)

    return results_df


def create_column_mapping(stats_type='service'):
    """创建列名映射"""
    column_mapping = {
        f'{stats_type}_name': '服务名称' if stats_type == 'service' else '应用名称',
        'app_name': '应用名称',
        'total_requests': '接口请求总数',
        'request_count': '成功请求数',
        'request_percentage': '占总请求比例(%)',
        'success_rate': '成功率(%)',
        'slow_requests_count': '慢请求数',
        'slow_requests_percentage': '慢请求占比(%)',
    }

    # 时间指标映射
    for metric in TIME_METRICS:
        if metric == 'upstream_connect_phase':
            continue

        display_name = TIME_METRICS_MAPPING.get(metric, metric)
        column_mapping[f'avg_{metric}'] = f'平均{display_name}(秒)'
        column_mapping[f'min_{metric}'] = f'最小{display_name}(秒)'
        column_mapping[f'max_{metric}'] = f'最大{display_name}(秒)'
        column_mapping[f'median_{metric}'] = f'中位数{display_name}(秒)'
        for percentile in PERCENTILES:
            column_mapping[f'p{percentile}_{metric}'] = f'P{percentile}{display_name}(秒)'

    # 大小指标映射
    for metric in SIZE_METRICS:
        display_name = '响应体大小' if metric == 'body_bytes_sent' else '总发送字节'
        column_mapping[f'avg_{metric}_kb'] = f'平均{display_name}(KB)'
        column_mapping[f'min_{metric}_kb'] = f'最小{display_name}(KB)'
        column_mapping[f'max_{metric}_kb'] = f'最大{display_name}(KB)'
        column_mapping[f'median_{metric}_kb'] = f'中位数{display_name}(KB)'
        for percentile in PERCENTILES:
            column_mapping[f'p{percentile}_{metric}_kb'] = f'P{percentile}{display_name}(KB)'

    return column_mapping


def create_header_groups():
    """创建表头分组"""
    header_groups = {
        '基本信息': ['服务名称', '应用名称'],
        '请求统计': ['接口请求总数', '成功请求数', '占总请求比例(%)', '成功率(%)'],
        '慢请求': ['慢请求数', '慢请求占比(%)'],
    }

    # 时间指标分组
    for metric in TIME_METRICS:
        if metric == 'upstream_connect_phase':
            continue

        display_name = TIME_METRICS_MAPPING.get(metric, metric)
        header_groups[f'{display_name}(秒)'] = [
            f'平均', f'最小', f'最大', f'中位数'
        ]
        for percentile in PERCENTILES:
            header_groups[f'{display_name}(秒)'].append(f'P{percentile}')

    # 大小指标分组
    for metric in SIZE_METRICS:
        display_name = '响应体大小' if metric == 'body_bytes_sent' else '总发送字节'
        header_groups[f'{display_name}(KB)'] = [
            f'平均', f'最小', f'最大', f'中位数'
        ]
        for percentile in PERCENTILES:
            header_groups[f'{display_name}(KB)'].append(f'P{percentile}')

    return header_groups


def create_performance_excel(service_results, app_results, output_path, total_requests, success_requests):
    """创建性能分析Excel报告"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 创建服务分析表
    header_groups = create_header_groups()
    add_dataframe_to_excel_with_grouped_headers(wb, service_results, '服务分析', header_groups)

    # 创建应用分析表（去除应用名称列）
    app_header_groups = create_header_groups()
    if '基本信息' in app_header_groups and '应用名称' in app_header_groups['基本信息']:
        app_header_groups['基本信息'] = ['应用名称']
    add_dataframe_to_excel_with_grouped_headers(wb, app_results, '应用分析', app_header_groups)

    # 创建整体分析表
    create_overall_analysis_sheet(wb, service_results, app_results, total_requests, success_requests)

    wb.save(output_path)


def create_overall_analysis_sheet(wb, service_results, app_results, total_requests, success_requests):
    """创建整体分析表"""
    ws = wb.create_sheet(title='整体服务请求分析')

    # 计算整体指标
    all_request_times = []
    for _, stats in service_results.iterrows():
        if '平均请求耗时(秒)' in stats:
            request_count = stats['成功请求数']
            avg_time = stats['平均请求耗时(秒)']
            all_request_times.extend([avg_time] * int(request_count))

    avg_response_time = round(sum(all_request_times) / len(all_request_times), 3) if all_request_times else 0
    median_response_time = round(np.median(all_request_times), 3) if all_request_times else 0

    # 基本统计信息
    overall_stats = [
        ['总请求数', total_requests],
        ['成功请求数', success_requests],
        ['成功率(%)', round(success_requests / total_requests * 100, 2) if total_requests > 0 else 0],
        ['服务数量', len(service_results)],
        ['应用数量', len(app_results)],
        ['平均响应时间(秒)', avg_response_time],
        ['中位数响应时间(秒)', median_response_time],
    ]

    # 添加百分位统计
    if all_request_times:
        for percentile in PERCENTILES:
            p_value = round(np.percentile(all_request_times, percentile), 3)
            overall_stats.append([f'P{percentile}响应时间(秒)', p_value])

    # 填充表格
    for row_idx, (label, value) in enumerate(overall_stats, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    # 添加Top服务部分
    add_top_items_section(ws, service_results, len(overall_stats) + 2, "请求量最多的10个服务", '服务名称')

    # 如果有时间数据，添加最慢服务部分
    time_col = '平均请求耗时(秒)' if '平均请求耗时(秒)' in service_results.columns else '平均request_time(秒)'
    if time_col in service_results.columns:
        sorted_by_time = service_results.sort_values(by=time_col, ascending=False)
        add_top_items_section(ws, sorted_by_time, len(overall_stats) + 15, "响应时间最长的10个服务", '服务名称')

    # 添加Top应用部分
    add_top_items_section(ws, app_results, len(overall_stats) + 28, "请求量最多的10个应用", '应用名称', include_service=False)

    # 设置列宽和格式
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    format_excel_sheet(ws)


def add_top_items_section(worksheet, df, start_row, title, name_col, include_service=True):
    """添加Top项目部分"""
    from openpyxl.styles import Font

    worksheet.cell(row=start_row, column=1, value=title).font = Font(bold=True)

    # 设置表头
    if include_service:
        top_headers = ['服务名称', '应用名称', '接口请求总数', '成功请求数', '占总请求比例(%)', '平均响应时间(秒)']
    else:
        top_headers = ['应用名称', '接口请求总数', '成功请求数', '占总请求比例(%)', '平均响应时间(秒)']

    for col_idx, header in enumerate(top_headers, start=1):
        worksheet.cell(row=start_row + 1, column=col_idx, value=header)

    # 获取时间列
    time_col = '平均请求耗时(秒)' if '平均请求耗时(秒)' in df.columns else '平均request_time(秒)'
    if time_col not in df.columns:
        time_col = None

    # 取前10条数据
    top_items = df.head(10)

    # 填充数据
    for row_idx, item in enumerate(top_items.itertuples(), start=start_row + 2):
        col_offset = 0

        if include_service:
            worksheet.cell(row=row_idx, column=1, value=getattr(item, '服务名称', ''))
            worksheet.cell(row=row_idx, column=2, value=getattr(item, '应用名称', ''))
            col_offset = 2
        else:
            worksheet.cell(row=row_idx, column=1, value=getattr(item, '应用名称', ''))
            col_offset = 1

        worksheet.cell(row=row_idx, column=col_offset + 1, value=getattr(item, '接口请求总数', 0))
        worksheet.cell(row=row_idx, column=col_offset + 2, value=getattr(item, '成功请求数', 0))
        worksheet.cell(row=row_idx, column=col_offset + 3, value=getattr(item, '占总请求比例(%)', 0))

        if time_col:
            worksheet.cell(row=row_idx, column=col_offset + 4, value=getattr(item, time_col, 0))