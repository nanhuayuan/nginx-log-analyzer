"""
高级时间维度分析器 - 集成优化算法和完整功能
基于已优化模块经验，支持40G+数据处理，提供完整的时间维度统计

核心优化:
1. 基于T-Digest的分位数计算(P50/P95/P99)
2. 连接数统计(新建/并发/活跃连接数)
3. 统一时间维度计算逻辑(按完成时间分组)
4. 内存高效的流式处理
5. 智能采样和预聚合
6. 完整的输出列设计

时间维度定义:
- 成功请求总数: 完成时间在[T, T+N)内的2xx/3xx请求
- 总请求量: 到达时间在[T, T+N)内的所有请求
- 新建连接数: 到达时间在[T, T+N)内的请求数
- 并发连接数: 到达时间<T+N且完成时间≥T+N的请求数  
- 活跃连接数: 到达时间≤T+N且完成时间≥T的请求数

Author: Claude Code (Advanced Time Dimension Analyzer)
Date: 2025-07-20
"""

import gc
import math
import os
import time
import numpy as np
import pandas as pd
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font

from self_00_04_excel_processor import (
    add_dataframe_to_excel_with_grouped_headers,
    create_line_chart,
    format_excel_sheet
)
from self_00_01_constants import (
    DEFAULT_CHUNK_SIZE, DEFAULT_SLOW_THRESHOLD, 
    TIME_METRICS, SIZE_METRICS, HIGHLIGHT_FILL
)
from self_00_02_utils import log_info, get_distribution_stats
from self_00_05_sampling_algorithms import (
    TDigest, ReservoirSampler, CountMinSketch, HyperLogLog, 
    StratifiedSampler, AdaptiveSampler
)

# 核心指标配置 - 基于老版本高价值指标扩展
CORE_TIME_METRICS = [
    'total_request_duration',    # 请求总时长
    'upstream_response_time',    # 后端响应时长  
    'upstream_header_time',      # 后端处理时长
    'upstream_connect_time',     # 后端连接时长
]

# 阶段时间指标 (老版本高价值)
PHASE_TIME_METRICS = [
    'backend_connect_phase',     # 后端连接阶段
    'backend_process_phase',     # 后端处理阶段
    'backend_transfer_phase',    # 后端传输阶段
    'nginx_transfer_phase'       # Nginx传输阶段
]

# 组合时间指标 (老版本高价值)
COMPOSITE_TIME_METRICS = [
    'backend_total_phase',       # 后端总阶段
    'network_phase',             # 网络传输阶段
    'processing_phase',          # 纯处理阶段
    'transfer_phase'             # 纯传输阶段
]

# 传输大小指标 (老版本高价值)
SIZE_METRICS = [
    'response_body_size_kb',     # 响应体大小
    'total_bytes_sent_kb'        # 总传输大小
]

# 效率比率指标 (老版本高价值)
EFFICIENCY_METRICS = [
    'backend_efficiency',        # 后端处理效率
    'network_overhead',          # 网络开销占比
    'transfer_ratio',            # 传输时间占比
    'connection_cost_ratio'      # 连接成本占比
]

# 传输速度指标 (老版本高价值)
SPEED_METRICS = [
    'response_transfer_speed',   # 响应传输速度
    'total_transfer_speed',      # 总传输速度
    'nginx_transfer_speed'       # Nginx传输速度
]

# 所有指标汇总
ALL_METRICS = CORE_TIME_METRICS + PHASE_TIME_METRICS + COMPOSITE_TIME_METRICS + SIZE_METRICS + EFFICIENCY_METRICS + SPEED_METRICS

# 时间维度配置
TIME_DIMENSIONS = {
    'daily': {'seconds': 86400, 'format': '%Y-%m-%d'},
    'hourly': {'seconds': 3600, 'format': '%Y-%m-%d %H:00'},
    'minute': {'seconds': 60, 'format': '%Y-%m-%d %H:%M'},
    'second': {'seconds': 1, 'format': '%Y-%m-%d %H:%M:%S'}
}

# 分位数配置
PERCENTILES = [50, 95, 99]


class AdvancedTimeDimensionAnalyzer:
    """
    高级时间维度分析器
    使用多种采样算法提供准确和高效的时间维度分析
    """
    
    def __init__(self, slow_threshold=DEFAULT_SLOW_THRESHOLD):
        self.slow_threshold = slow_threshold
        
        # 初始化统计容器
        self.stats = {}
        self.time_samplers = {}
        self.ip_counters = {}
        self.all_requests = []  # 存储所有请求用于连接数计算
        
        for dimension in TIME_DIMENSIONS.keys():
            self.stats[dimension] = defaultdict(lambda: defaultdict(int))
            self.time_samplers[dimension] = defaultdict(lambda: {
                metric: TDigest(compression=100) for metric in ALL_METRICS
            })
            self.ip_counters[dimension] = defaultdict(lambda: HyperLogLog(precision=12))
    
    def process_chunk(self, chunk: pd.DataFrame) -> None:
        """处理数据块"""
        if chunk.empty:
            return
            
        # 预处理时间字段
        chunk = self._preprocess_time_fields(chunk)
        
        # 收集请求数据用于连接数计算（优化：只保留时间字段）
        if 'arrival_timestamp' in chunk.columns and 'timestamp' in chunk.columns:
            request_data = chunk[['arrival_timestamp', 'timestamp']].copy()
            request_data = request_data.dropna()
            if not request_data.empty:
                self.all_requests.append(request_data)
        
        # 创建状态掩码
        success_mask = self._create_success_mask(chunk)
        slow_mask = self._create_slow_mask(chunk)
        error_4xx_mask, error_5xx_mask = self._create_error_masks(chunk)
        
        # 处理各个时间维度
        for dimension in TIME_DIMENSIONS.keys():
            self._process_dimension(chunk, dimension, success_mask, slow_mask, error_4xx_mask, error_5xx_mask)
    
    def _preprocess_time_fields(self, chunk: pd.DataFrame) -> pd.DataFrame:
        """预处理时间字段"""
        chunk = chunk.copy()
        
        # 打印列名以调试
        log_info(f"CSV列名: {list(chunk.columns)}")
        
        # 尝试多种可能的时间字段名
        time_field_candidates = ['time', 'timestamp', 'raw_time', 'datetime']
        arrival_time_candidates = ['arrival_time', 'arrival_timestamp', 'request_time']
        
        completion_time_col = None
        arrival_time_col = None
        
        # 查找完成时间字段
        for candidate in time_field_candidates:
            if candidate in chunk.columns:
                completion_time_col = candidate
                log_info(f"使用完成时间字段: {candidate}")
                break
        
        # 查找到达时间字段  
        for candidate in arrival_time_candidates:
            if candidate in chunk.columns:
                arrival_time_col = candidate
                log_info(f"使用到达时间字段: {candidate}")
                break
        
        if not completion_time_col:
            log_info("警告: 未找到完成时间字段")
            return chunk
        
        # 转换时间戳
        timestamp_candidates = ['timestamp', 'arrival_timestamp', 'raw_timestamp']
        for col in timestamp_candidates:
            if col in chunk.columns:
                chunk[col] = pd.to_numeric(chunk[col], errors='coerce')
        
        # 转换并创建时间维度字段（基于完成时间）
        try:
            dt = pd.to_datetime(chunk[completion_time_col], errors='coerce')
            
            # 检查转换是否成功
            valid_times = dt.notna().sum()
            log_info(f"有效时间记录: {valid_times}/{len(chunk)}")
            
            if valid_times > 0:
                chunk['completion_daily'] = dt.dt.strftime('%Y-%m-%d')
                chunk['completion_hourly'] = dt.dt.strftime('%Y-%m-%d %H:00')
                chunk['completion_minute'] = dt.dt.strftime('%Y-%m-%d %H:%M')
                chunk['completion_second'] = dt.dt.strftime('%Y-%m-%d %H:%M:%S')
                
                # 显示样例数据
                first_valid_idx = dt.first_valid_index()
                if first_valid_idx is not None:
                    log_info(f"时间维度样例 - daily: {chunk.loc[first_valid_idx, 'completion_daily']}")
                    log_info(f"时间维度样例 - hourly: {chunk.loc[first_valid_idx, 'completion_hourly']}")
            else:
                log_info(f"警告: {completion_time_col} 字段无有效时间数据")
        except Exception as e:
            log_info(f"完成时间处理错误: {e}")
        
        # 处理到达时间（如果存在）
        if arrival_time_col:
            try:
                dt_arrival = pd.to_datetime(chunk[arrival_time_col], errors='coerce')
                valid_arrival_times = dt_arrival.notna().sum()
                log_info(f"有效到达时间记录: {valid_arrival_times}/{len(chunk)}")
                
                if valid_arrival_times > 0:
                    chunk['arrival_daily'] = dt_arrival.dt.strftime('%Y-%m-%d')
                    chunk['arrival_hourly'] = dt_arrival.dt.strftime('%Y-%m-%d %H:00')
                    chunk['arrival_minute'] = dt_arrival.dt.strftime('%Y-%m-%d %H:%M')
                    chunk['arrival_second'] = dt_arrival.dt.strftime('%Y-%m-%d %H:%M:%S')
            except Exception as e:
                log_info(f"到达时间处理错误: {e}")
        
        # 转换数值字段
        for metric in ALL_METRICS:
            if metric in chunk.columns:
                chunk[metric] = pd.to_numeric(chunk[metric], errors='coerce')
        
        return chunk
    
    def _create_success_mask(self, chunk: pd.DataFrame) -> np.ndarray:
        """创建成功请求掩码"""
        if 'status' in chunk.columns:
            status = pd.to_numeric(chunk['status'], errors='coerce')
            return ((status >= 200) & (status < 400)).values
        return np.ones(len(chunk), dtype=bool)
    
    def _create_slow_mask(self, chunk: pd.DataFrame) -> np.ndarray:
        """创建慢请求掩码"""
        if 'total_request_duration' in chunk.columns:
            duration = pd.to_numeric(chunk['total_request_duration'], errors='coerce')
            return (duration > self.slow_threshold).values
        return np.zeros(len(chunk), dtype=bool)
    
    def _create_error_masks(self, chunk: pd.DataFrame) -> tuple:
        """创建错误请求掩码"""
        if 'status' in chunk.columns:
            status = pd.to_numeric(chunk['status'], errors='coerce')
            error_4xx_mask = ((status >= 400) & (status < 500)).values
            error_5xx_mask = (status >= 500).values
            return error_4xx_mask, error_5xx_mask
        return np.zeros(len(chunk), dtype=bool), np.zeros(len(chunk), dtype=bool)
    
    def _process_dimension(self, chunk: pd.DataFrame, dimension: str, 
                          success_mask: np.ndarray, slow_mask: np.ndarray,
                          error_4xx_mask: np.ndarray, error_5xx_mask: np.ndarray) -> None:
        """处理单个时间维度"""
        completion_col = f'completion_{dimension}'
        arrival_col = f'arrival_{dimension}'
        
        if completion_col not in chunk.columns:
            log_info(f"跳过 {dimension} 维度: 缺少 {completion_col} 列")
            return
        
        # 检查非空值
        valid_times = chunk[completion_col].notna().sum()
        log_info(f"{dimension} 维度: {valid_times} 个有效时间值")
        
        if valid_times == 0:
            log_info(f"跳过 {dimension} 维度: 无有效时间数据")
            return
        
        # 基于完成时间的统计
        completion_groups = chunk.groupby(completion_col, sort=False)
        group_count = 0
        
        for time_key, group in completion_groups:
            if pd.isna(time_key):
                continue
                
            group_count += 1
            indices = group.index
            stats = self.stats[dimension][time_key]
            
            # 过滤有效索引，避免越界
            valid_indices = [i for i in indices if i < len(success_mask)]
            
            # 基础统计
            stats['total_requests'] += len(group)
            if valid_indices:
                stats['success_requests'] += success_mask[valid_indices].sum()
                stats['slow_requests'] += slow_mask[valid_indices].sum()
                
                # 错误统计 (新增)
                stats['error_4xx_requests'] += error_4xx_mask[valid_indices].sum()
                stats['error_5xx_requests'] += error_5xx_mask[valid_indices].sum()
            
            # 计算连接数指标
            self._calculate_connection_metrics(group, time_key, dimension, stats)
            
            # 更新时间指标采样器
            self._update_time_samplers(group, time_key, dimension)
            
            # 更新IP计数器
            if 'client_ip' in group.columns:
                for ip in group['client_ip'].dropna():
                    self.ip_counters[dimension][time_key].add(str(ip))
        
        log_info(f"{dimension} 维度处理完成: {group_count} 个时间组")
    
    def _calculate_connection_metrics(self, group: pd.DataFrame, time_key: str, 
                                    dimension: str, stats: Dict) -> None:
        """计算连接数指标 - 基于当前组的简化版本，最终会在calculate_derived_metrics中重新计算"""
        # 暂时使用组大小作为基础统计，真正的连接数计算在后面
        group_size = len(group)
        stats['new_connections'] = group_size
        stats['concurrent_connections'] = 0  # 稍后重新计算
        stats['active_connections'] = group_size  # 稍后重新计算
    
    def _update_time_samplers(self, group: pd.DataFrame, time_key: str, dimension: str) -> None:
        """更新时间指标采样器"""
        samplers = self.time_samplers[dimension][time_key]
        
        for metric in ALL_METRICS:
            if metric in group.columns:
                values = group[metric].dropna()
                for value in values:
                    # 时间指标：负值表示没有upstream，用0代替；其他保留原值
                    if metric in ['upstream_response_time', 'upstream_header_time', 'upstream_connect_time']:
                        clean_value = max(0, float(value))  # 负值变为0
                        samplers[metric].add(clean_value)
                    # 大小和速度指标：只接受非负值
                    elif metric.endswith('_kb') or metric.endswith('_speed'):
                        if value >= 0:
                            samplers[metric].add(float(value))
                    # 其他指标：接受所有有限值
                    else:
                        if pd.notna(value) and not math.isinf(float(value)):
                            samplers[metric].add(float(value))
    
    def calculate_derived_metrics(self) -> Dict:
        """计算衍生指标"""
        results = {}
        
        # 先计算正确的连接数指标
        log_info("计算连接数指标...")
        self._calculate_accurate_connection_metrics()
        
        for dimension in TIME_DIMENSIONS.keys():
            results[dimension] = {}
            window_seconds = TIME_DIMENSIONS[dimension]['seconds']
            
            # 调试输出
            total_time_keys = len(self.stats[dimension])
            log_info(f"  {dimension} 维度: {total_time_keys} 个时间键")
            
            for time_key, stats in self.stats[dimension].items():
                derived = {}
                
                # 基础比率计算
                total = stats.get('total_requests', 0)
                success = stats.get('success_requests', 0)
                slow = stats.get('slow_requests', 0)
                error_4xx = stats.get('error_4xx_requests', 0)
                error_5xx = stats.get('error_5xx_requests', 0)
                
                derived['success_rate'] = (success / total * 100) if total > 0 else 0
                derived['slow_rate'] = (slow / total * 100) if total > 0 else 0
                derived['error_4xx_rate'] = (error_4xx / total * 100) if total > 0 else 0
                derived['error_5xx_rate'] = (error_5xx / total * 100) if total > 0 else 0
                derived['total_error_rate'] = ((error_4xx + error_5xx) / total * 100) if total > 0 else 0
                derived['qps'] = success / window_seconds
                
                # 分位数计算
                percentiles = {}
                samplers = self.time_samplers[dimension][time_key]
                for metric in ALL_METRICS:
                    percentiles[metric] = {}
                    sampler = samplers[metric]
                    if sampler.count > 0:
                        for p in PERCENTILES:
                            percentiles[metric][f'P{p}'] = sampler.percentile(p)
                
                derived['percentiles'] = percentiles
                
                # 独立IP数
                derived['unique_ips'] = self.ip_counters[dimension][time_key].cardinality()
                
                # 合并到结果
                results[dimension][time_key] = {**stats, **derived}
        
        # 计算环比变化趋势
        self._calculate_trend_analysis(results)
        
        # 计算异常检测评分
        self._calculate_anomaly_detection(results)
        
        # 计算综合效率指数
        self._calculate_efficiency_index(results)
        
        # 调试输出最终结果
        for dimension in results:
            log_info(f"  最终 {dimension} 维度: {len(results[dimension])} 个时间组")
        
        return results
    
    def _calculate_trend_analysis(self, results: Dict) -> None:
        """计算环比变化趋势"""
        for dimension in results:
            time_keys = sorted(results[dimension].keys())
            
            for i, time_key in enumerate(time_keys):
                if i == 0:
                    # 第一个时间点没有前一个时间点作比较
                    results[dimension][time_key]['qps_change'] = 0
                    results[dimension][time_key]['avg_response_time_change'] = 0
                    results[dimension][time_key]['error_rate_change'] = 0
                    continue
                
                current_stats = results[dimension][time_key]
                prev_time_key = time_keys[i-1]
                prev_stats = results[dimension][prev_time_key]
                
                # QPS环比变化
                current_qps = current_stats.get('qps', 0)
                prev_qps = prev_stats.get('qps', 0)
                qps_change = ((current_qps - prev_qps) / prev_qps * 100) if prev_qps > 0 else 0
                
                # 平均响应时间环比变化 (使用P50作为平均值)
                current_percentiles = current_stats.get('percentiles', {})
                prev_percentiles = prev_stats.get('percentiles', {})
                current_avg_time = current_percentiles.get('total_request_duration', {}).get('P50', 0)
                prev_avg_time = prev_percentiles.get('total_request_duration', {}).get('P50', 0)
                time_change = ((current_avg_time - prev_avg_time) / prev_avg_time * 100) if prev_avg_time > 0 else 0
                
                # 错误率环比变化
                current_error_rate = current_stats.get('total_error_rate', 0)
                prev_error_rate = prev_stats.get('total_error_rate', 0)
                error_change = ((current_error_rate - prev_error_rate) / prev_error_rate * 100) if prev_error_rate > 0 else 0
                
                # 存储变化趋势
                results[dimension][time_key]['qps_change'] = round(qps_change, 2)
                results[dimension][time_key]['avg_response_time_change'] = round(time_change, 2)
                results[dimension][time_key]['error_rate_change'] = round(error_change, 2)
    
    def _calculate_anomaly_detection(self, results: Dict) -> None:
        """计算异常检测评分"""
        import numpy as np
        
        for dimension in results:
            # 收集各指标的数据用于统计分析
            metrics_data = {
                'qps': [],
                'total_error_rate': [],
                'slow_rate': [],
                'avg_response_time': []  # 使用P50作为平均响应时间
            }
            
            # 收集所有时间点的数据
            for time_key, stats in results[dimension].items():
                metrics_data['qps'].append(stats.get('qps', 0))
                metrics_data['total_error_rate'].append(stats.get('total_error_rate', 0))
                metrics_data['slow_rate'].append(stats.get('slow_rate', 0))
                
                # 获取P50作为平均响应时间
                percentiles = stats.get('percentiles', {})
                avg_time = percentiles.get('total_request_duration', {}).get('P50', 0)
                metrics_data['avg_response_time'].append(avg_time)
            
            # 计算各指标的统计信息（均值和标准差）
            metrics_stats = {}
            for metric, values in metrics_data.items():
                if values and len(values) > 1:
                    mean_val = np.mean(values)
                    std_val = np.std(values)
                    metrics_stats[metric] = {'mean': mean_val, 'std': std_val}
                else:
                    metrics_stats[metric] = {'mean': 0, 'std': 0}
            
            # 为每个时间点计算异常评分
            for time_key, stats in results[dimension].items():
                anomaly_scores = []
                
                # QPS异常评分 (使用Z-score)
                qps = stats.get('qps', 0)
                qps_stats = metrics_stats['qps']
                if qps_stats['std'] > 0:
                    qps_zscore = abs(qps - qps_stats['mean']) / qps_stats['std']
                    qps_anomaly = min(qps_zscore / 3.0, 1.0)  # 标准化到0-1，3-sigma规则
                else:
                    qps_anomaly = 0
                anomaly_scores.append(qps_anomaly)
                
                # 错误率异常评分
                error_rate = stats.get('total_error_rate', 0)
                error_stats = metrics_stats['total_error_rate']
                if error_stats['std'] > 0:
                    error_zscore = abs(error_rate - error_stats['mean']) / error_stats['std']
                    error_anomaly = min(error_zscore / 3.0, 1.0)
                else:
                    error_anomaly = 0
                # 错误率异常权重更高
                anomaly_scores.append(error_anomaly * 1.5)
                
                # 慢请求率异常评分
                slow_rate = stats.get('slow_rate', 0)
                slow_stats = metrics_stats['slow_rate']
                if slow_stats['std'] > 0:
                    slow_zscore = abs(slow_rate - slow_stats['mean']) / slow_stats['std']
                    slow_anomaly = min(slow_zscore / 3.0, 1.0)
                else:
                    slow_anomaly = 0
                anomaly_scores.append(slow_anomaly)
                
                # 响应时间异常评分
                percentiles = stats.get('percentiles', {})
                avg_time = percentiles.get('total_request_duration', {}).get('P50', 0)
                time_stats = metrics_stats['avg_response_time']
                if time_stats['std'] > 0:
                    time_zscore = abs(avg_time - time_stats['mean']) / time_stats['std']
                    time_anomaly = min(time_zscore / 3.0, 1.0)
                else:
                    time_anomaly = 0
                anomaly_scores.append(time_anomaly)
                
                # 计算综合异常评分（加权平均）
                if anomaly_scores:
                    comprehensive_anomaly = np.mean(anomaly_scores)
                    # 转换为0-100分制
                    anomaly_score = round(comprehensive_anomaly * 100, 1)
                else:
                    anomaly_score = 0
                
                # 存储异常评分
                results[dimension][time_key]['anomaly_score'] = anomaly_score
                
                # 添加异常等级分类
                if anomaly_score >= 80:
                    anomaly_level = "严重异常"
                elif anomaly_score >= 60:
                    anomaly_level = "中度异常"
                elif anomaly_score >= 40:
                    anomaly_level = "轻微异常"
                else:
                    anomaly_level = "正常"
                
                results[dimension][time_key]['anomaly_level'] = anomaly_level
    
    def _calculate_efficiency_index(self, results: Dict) -> None:
        """计算综合效率指数"""
        log_info("开始计算综合效率指数...")
        for dimension in results:
            # 收集各指标的数据用于归一化
            all_qps = []
            all_response_times = []
            all_error_rates = []
            all_slow_rates = []
            
            # 收集所有数据
            for time_key, stats in results[dimension].items():
                all_qps.append(stats.get('qps', 0))
                all_error_rates.append(stats.get('total_error_rate', 0))
                all_slow_rates.append(stats.get('slow_rate', 0))
                
                # 使用P50作为响应时间代表值
                percentiles = stats.get('percentiles', {})
                response_time = percentiles.get('total_request_duration', {}).get('P50', 0)
                all_response_times.append(response_time)
            
            log_info(f"{dimension} 维度效率指数数据收集: QPS={len(all_qps)}, 错误率={len(all_error_rates)}, 慢请求率={len(all_slow_rates)}, 响应时间={len(all_response_times)}")
            
            # 计算指标的最值用于归一化（避免除零错误）
            max_qps = max(all_qps) if all_qps and max(all_qps) > 0 else 1
            max_response_time = max(all_response_times) if all_response_times and max(all_response_times) > 0 else 1
            max_error_rate = max(all_error_rates) if all_error_rates and max(all_error_rates) > 0 else 1
            max_slow_rate = max(all_slow_rates) if all_slow_rates and max(all_slow_rates) > 0 else 1
            
            # 为每个时间点计算效率指数
            for time_key, stats in results[dimension].items():
                # 获取原始指标值
                qps = stats.get('qps', 0)
                error_rate = stats.get('total_error_rate', 0)
                slow_rate = stats.get('slow_rate', 0)
                percentiles = stats.get('percentiles', {})
                response_time = percentiles.get('total_request_duration', {}).get('P50', 0)
                
                # 计算各维度得分 (0-100分制)
                # 1. 吞吐量得分 (QPS越高越好)
                throughput_score = (qps / max_qps) * 100 if max_qps > 0 else 0
                
                # 2. 响应时间得分 (响应时间越低越好)
                if max_response_time > 0:
                    response_score = max(0, 100 - (response_time / max_response_time) * 100)
                else:
                    response_score = 100
                
                # 3. 可靠性得分 (错误率越低越好)
                if max_error_rate > 1:  # 只有当存在真实的错误率差异时才进行相对计算
                    reliability_score = max(0, 100 - (error_rate / max_error_rate) * 100)
                else:
                    # 如果所有错误率都很低或为0，则根据绝对错误率给分
                    reliability_score = max(0, 100 - error_rate * 10)  # 每1%错误率扣10分
                
                # 4. 性能稳定性得分 (慢请求率越低越好)
                if max_slow_rate > 1:  # 只有当存在真实的慢请求率差异时才进行相对计算
                    stability_score = max(0, 100 - (slow_rate / max_slow_rate) * 100)
                else:
                    # 如果所有慢请求率都很低或为0，则根据绝对慢请求率给分
                    stability_score = max(0, 100 - slow_rate * 5)  # 每1%慢请求率扣5分
                
                # 5. 异常健康度得分 (基于异常评分)
                anomaly_score = stats.get('anomaly_score', 0)
                health_score = max(0, 100 - anomaly_score)
                
                # 计算加权综合效率指数
                # 权重分配：吞吐量(25%) + 响应时间(25%) + 可靠性(20%) + 稳定性(15%) + 健康度(15%)
                efficiency_index = (
                    throughput_score * 0.25 +
                    response_score * 0.25 +
                    reliability_score * 0.20 +
                    stability_score * 0.15 +
                    health_score * 0.15
                )
                
                # 效率等级分类
                if efficiency_index >= 90:
                    efficiency_level = "优秀"
                elif efficiency_index >= 80:
                    efficiency_level = "良好"
                elif efficiency_index >= 70:
                    efficiency_level = "一般"
                elif efficiency_index >= 60:
                    efficiency_level = "较差"
                else:
                    efficiency_level = "很差"
                
                # 存储效率指数
                results[dimension][time_key]['efficiency_index'] = round(efficiency_index, 1)
                results[dimension][time_key]['efficiency_level'] = efficiency_level
                
                # 存储子指标得分 (用于详细分析)
                throughput_final = round(throughput_score, 1)
                response_final = round(response_score, 1)
                reliability_final = round(reliability_score, 1)
                stability_final = round(stability_score, 1)
                health_final = round(health_score, 1)
                
                results[dimension][time_key]['throughput_score'] = throughput_final
                results[dimension][time_key]['response_score'] = response_final
                results[dimension][time_key]['reliability_score'] = reliability_final
                results[dimension][time_key]['stability_score'] = stability_final
                results[dimension][time_key]['health_score'] = health_final
                
                # 调试输出
                if dimension == 'daily':  # 只在日维度输出调试信息避免过多日志
                    log_info(f"效率指数计算完成 {time_key}: 吞吐量={throughput_final}, 响应时间={response_final}, 可靠性={reliability_final}, 稳定性={stability_final}, 健康度={health_final}")
    
    def _calculate_accurate_connection_metrics(self) -> None:
        """基于所有请求数据重新计算准确的连接数指标"""
        if not self.all_requests:
            log_info("无请求数据，跳过连接数计算")
            return
        
        # 合并所有请求数据
        try:
            import pandas as pd
            all_data = pd.concat(self.all_requests, ignore_index=True)
            log_info(f"合并请求数据: {len(all_data)} 条记录")
        except Exception as e:
            log_info(f"合并请求数据失败: {e}")
            return
        
        # 为每个维度和时间键重新计算连接数
        for dimension in TIME_DIMENSIONS.keys():
            window_seconds = TIME_DIMENSIONS[dimension]['seconds']
            
            for time_key in self.stats[dimension].keys():
                try:
                    # 解析时间窗口
                    window_start = pd.to_datetime(time_key)
                    window_end = window_start + timedelta(seconds=window_seconds)
                    window_start_ts = window_start.timestamp()
                    window_end_ts = window_end.timestamp()
                    
                    # 计算新建连接数：到达时间在窗口内的请求数
                    new_conn_mask = ((all_data['arrival_timestamp'] >= window_start_ts) & 
                                   (all_data['arrival_timestamp'] < window_end_ts))
                    new_connections = new_conn_mask.sum()
                    
                    # 计算并发连接数：到达时间<窗口结束且完成时间≥窗口结束的请求数
                    concurrent_mask = ((all_data['arrival_timestamp'] < window_end_ts) & 
                                     (all_data['timestamp'] >= window_end_ts))
                    concurrent_connections = concurrent_mask.sum()
                    
                    # 计算活跃连接数：到达时间≤窗口结束且完成时间≥窗口开始的请求数  
                    active_mask = ((all_data['arrival_timestamp'] <= window_end_ts) & 
                                 (all_data['timestamp'] >= window_start_ts))
                    active_connections = active_mask.sum()
                    
                    # 更新统计
                    stats = self.stats[dimension][time_key]
                    stats['new_connections'] = new_connections
                    stats['concurrent_connections'] = concurrent_connections
                    stats['active_connections'] = active_connections
                    
                except Exception as e:
                    log_info(f"计算 {dimension} {time_key} 连接数失败: {e}")
                    continue
    
    def create_output_dataframe(self, dimension: str, results: Dict) -> pd.DataFrame:
        """创建输出DataFrame - 集成老版本所有高价值指标"""
        if dimension not in results or not results[dimension]:
            return pd.DataFrame()
        
        data = []
        time_label = self._get_time_label(dimension)
        
        for time_key in sorted(results[dimension].keys()):
            stats = results[dimension][time_key]
            percentiles = stats.get('percentiles', {})
            
            # 基础统计指标 (与老版本完全一致)
            row = {
                time_label: time_key,
                '总请求数': stats.get('total_requests', 0),
                '成功请求数': stats.get('success_requests', 0),
                '慢请求数': stats.get('slow_requests', 0),
                '慢请求占比(%)': round(stats.get('slow_rate', 0), 2),
                '新建连接数': stats.get('new_connections', 0),
                '并发连接数': stats.get('concurrent_connections', 0),
                '活跃连接数': stats.get('active_connections', 0),
                'QPS': round(stats.get('qps', 0), 3),
                
                # 错误分析指标 (新增高价值)
                '4xx错误数': stats.get('error_4xx_requests', 0),
                '4xx错误率(%)': round(stats.get('error_4xx_rate', 0), 2),
                '5xx错误数': stats.get('error_5xx_requests', 0), 
                '5xx错误率(%)': round(stats.get('error_5xx_rate', 0), 2),
                '总错误率(%)': round(stats.get('total_error_rate', 0), 2),
            }
            
            # 基础时间指标平均值 (老版本高价值)
            for metric in CORE_TIME_METRICS:
                metric_name = self._get_metric_display_name(metric)
                avg_value = percentiles.get(metric, {}).get('P50', 0)  # 用P50作为平均值
                row[f'平均{metric_name}(秒)'] = round(avg_value, 4)
            
            # 阶段时间指标平均值 (老版本高价值)
            for metric in PHASE_TIME_METRICS:
                metric_name = self._get_metric_display_name(metric)
                avg_value = percentiles.get(metric, {}).get('P50', 0)
                row[f'平均{metric_name}(秒)'] = round(avg_value, 4)
            
            # 组合时间指标平均值 (老版本高价值)
            for metric in COMPOSITE_TIME_METRICS:
                metric_name = self._get_metric_display_name(metric)
                avg_value = percentiles.get(metric, {}).get('P50', 0)
                row[f'平均{metric_name}(秒)'] = round(avg_value, 4)
            
            # 传输大小指标 (老版本高价值)
            for metric in SIZE_METRICS:
                metric_name = self._get_metric_display_name(metric)
                avg_value = percentiles.get(metric, {}).get('P50', 0)
                row[f'平均{metric_name}(KB)'] = round(avg_value, 2)
            
            # 效率比率指标 (老版本高价值)
            for metric in EFFICIENCY_METRICS:
                metric_name = self._get_metric_display_name(metric)
                avg_value = percentiles.get(metric, {}).get('P50', 0)
                row[f'平均{metric_name}(%)'] = round(avg_value, 2)
            
            # 传输速度指标 (老版本高价值)
            for metric in SPEED_METRICS:
                metric_name = self._get_metric_display_name(metric)
                avg_value = percentiles.get(metric, {}).get('P50', 0)
                row[f'平均{metric_name}(KB/s)'] = round(avg_value, 2)
            
            # 关键性能分位数指标 (P95/P99) - 新增高价值
            key_metrics_for_percentiles = ['total_request_duration', 'upstream_response_time', 'upstream_header_time']
            for metric in key_metrics_for_percentiles:
                metric_name = self._get_metric_display_name(metric)
                p95_value = percentiles.get(metric, {}).get('P95', 0)
                p99_value = percentiles.get(metric, {}).get('P99', 0)
                row[f'{metric_name}P95(秒)'] = round(p95_value, 4)
                row[f'{metric_name}P99(秒)'] = round(p99_value, 4)
            
            # 环比变化趋势指标 (新增高价值)
            row['QPS环比变化(%)'] = round(stats.get('qps_change', 0), 2)
            row['响应时间环比变化(%)'] = round(stats.get('avg_response_time_change', 0), 2)
            row['错误率环比变化(%)'] = round(stats.get('error_rate_change', 0), 2)
            
            # 异常检测指标 (新增高价值)
            row['异常评分(0-100)'] = round(stats.get('anomaly_score', 0), 1)
            row['异常等级'] = stats.get('anomaly_level', '正常')
            
            # 综合效率指数 (新增高价值)
            row['综合效率指数(0-100)'] = round(stats.get('efficiency_index', 0), 1)
            row['效率等级'] = stats.get('efficiency_level', '未知')
            row['吞吐量得分'] = round(stats.get('throughput_score', 0), 1)
            row['响应时间得分'] = round(stats.get('response_score', 0), 1)
            row['可靠性得分'] = round(stats.get('reliability_score', 0), 1)
            row['稳定性得分'] = round(stats.get('stability_score', 0), 1)
            row['健康度得分'] = round(stats.get('health_score', 0), 1)
            
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _get_time_label(self, dimension: str) -> str:
        """获取时间标签"""
        labels = {
            'daily': '日期',
            'hourly': '小时',
            'minute': '分钟', 
            'second': '秒'
        }
        return labels.get(dimension, '时间')
    
    def _get_metric_display_name(self, metric: str) -> str:
        """获取指标显示名称 - 完整版本"""
        names = {
            # 基础时间指标
            'total_request_duration': '请求总时长',
            'upstream_response_time': '后端响应时长',
            'upstream_header_time': '后端处理时长', 
            'upstream_connect_time': '后端连接时长',
            
            # 阶段时间指标 (老版本高价值)
            'backend_connect_phase': '后端连接阶段',
            'backend_process_phase': '后端处理阶段',
            'backend_transfer_phase': '后端传输阶段',
            'nginx_transfer_phase': 'Nginx传输阶段',
            
            # 组合时间指标 (老版本高价值)
            'backend_total_phase': '后端总阶段',
            'network_phase': '网络传输阶段',
            'processing_phase': '纯处理阶段',
            'transfer_phase': '纯传输阶段',
            
            # 传输大小指标 (老版本高价值)
            'response_body_size_kb': '响应体大小',
            'total_bytes_sent_kb': '总传输大小',
            
            # 效率比率指标 (老版本高价值)
            'backend_efficiency': '后端处理效率',
            'network_overhead': '网络开销占比',
            'transfer_ratio': '传输时间占比',
            'connection_cost_ratio': '连接成本占比',
            
            # 传输速度指标 (老版本高价值)
            'response_transfer_speed': '响应传输速度',
            'total_transfer_speed': '总传输速度',
            'nginx_transfer_speed': 'Nginx传输速度'
        }
        return names.get(metric, metric)
    
    def create_header_groups(self, df: pd.DataFrame, time_label: str) -> Dict:
        """创建表头分组 - 包含所有新增指标"""
        return {
            '时间维度': [time_label],
            '基础统计': ['总请求数', '成功请求数', '慢请求数', '慢请求占比(%)'],
            '连接统计': ['新建连接数', '并发连接数', '活跃连接数', 'QPS'],
            '错误分析': ['4xx错误数', '4xx错误率(%)', '5xx错误数', '5xx错误率(%)', '总错误率(%)'],
            '基础时间指标': [col for col in df.columns if col.startswith('平均') and any(x in col for x in ['请求总时长', '后端响应时长', '后端处理时长', '后端连接时长']) and '(秒)' in col],
            '阶段时间指标': [col for col in df.columns if col.startswith('平均') and any(x in col for x in ['连接阶段', '处理阶段', '传输阶段']) and '(秒)' in col],
            '组合时间指标': [col for col in df.columns if col.startswith('平均') and any(x in col for x in ['总阶段', '网络传输阶段', '纯处理阶段', '纯传输阶段']) and '(秒)' in col],
            '传输大小指标': [col for col in df.columns if col.startswith('平均') and '大小' in col and '(KB)' in col],
            '效率比率指标': [col for col in df.columns if col.startswith('平均') and ('效率' in col or '占比' in col) and '(%)' in col],
            '传输速度指标': [col for col in df.columns if col.startswith('平均') and '速度' in col and '(KB/s)' in col],
            '关键分位数指标': [col for col in df.columns if 'P95' in col or 'P99' in col],
            '趋势分析': ['QPS环比变化(%)', '响应时间环比变化(%)', '错误率环比变化(%)'],
            '异常检测': ['异常评分(0-100)', '异常等级'],
            '综合效率指数': ['综合效率指数(0-100)', '效率等级'],
            '效率子指标': ['吞吐量得分', '响应时间得分', '可靠性得分', '稳定性得分', '健康度得分']
        }


def analyze_time_dimension_advanced(csv_path: str, output_path: str, 
                                   specific_uri_list: Optional[List[str]] = None) -> str:
    """
    高级时间维度分析主函数
    
    Args:
        csv_path: CSV文件路径
        output_path: 输出Excel文件路径
        specific_uri_list: 特定URI列表(可选)
    
    Returns:
        输出文件路径
    """
    start_time = time.time()
    log_info("开始高级时间维度分析")
    
    # 准备输出文件名
    output_filename = _prepare_output_filename(output_path, specific_uri_list)
    
    # 初始化分析器
    analyzer = AdvancedTimeDimensionAnalyzer()
    
    # 流式处理数据
    total_records = _process_data_streaming(csv_path, analyzer, specific_uri_list)
    
    # 计算衍生指标
    log_info("计算衍生指标...")
    results = analyzer.calculate_derived_metrics()
    
    # 生成Excel报告
    log_info("生成Excel报告...")
    _create_excel_report(output_filename, analyzer, results, total_records)
    
    elapsed = time.time() - start_time
    log_info(f"高级时间维度分析完成，耗时: {elapsed:.2f}秒")
    log_info(f"报告已生成：{output_filename}")
    
    return output_filename


def _process_data_streaming(csv_path: str, analyzer: AdvancedTimeDimensionAnalyzer, 
                           specific_uri_list: Optional[List[str]] = None) -> int:
    """流式处理数据"""
    chunk_size = DEFAULT_CHUNK_SIZE
    total_records = 0
    processed_chunks = 0
    
    # URI过滤集合
    uri_set = None
    if specific_uri_list:
        uri_set = set(specific_uri_list) if isinstance(specific_uri_list, list) else {specific_uri_list}
        log_info(f"分析特定URI: {specific_uri_list}")
    else:
        log_info("分析所有请求")
    
    try:
        chunk_generator = pd.read_csv(csv_path, chunksize=chunk_size)
        
        for chunk in chunk_generator:
            processed_chunks += 1
            
            # URI过滤
            if uri_set and 'request_uri' in chunk.columns:
                chunk = chunk[chunk['request_uri'].isin(uri_set)]
                if chunk.empty:
                    continue
            
            # 处理数据块
            analyzer.process_chunk(chunk)
            total_records += len(chunk)
            
            # 内存管理
            del chunk
            if processed_chunks % 50 == 0:
                gc.collect()
                log_info(f"已处理 {processed_chunks} 个数据块, {total_records} 条记录")
    
    except Exception as e:
        log_info(f"数据处理错误: {e}")
        raise
    
    log_info(f"数据处理完成 - 总记录: {total_records}")
    return total_records


def _create_excel_report(output_path: str, analyzer: AdvancedTimeDimensionAnalyzer, 
                        results: Dict, total_records: int) -> None:
    """创建Excel报告"""
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 创建概览页
    _create_overview_sheet(wb, results, total_records)
    
    # 创建各维度分析页
    dimensions = [
        ('日期维度分析', 'daily'),
        ('小时维度分析', 'hourly'),
        ('分钟维度分析', 'minute'),
        ('秒级维度分析', 'second')
    ]
    
    for sheet_name, dimension in dimensions:
        if dimension in results and results[dimension]:
            log_info(f"创建工作表: {sheet_name} (维度: {dimension})")
            _create_dimension_sheet(wb, sheet_name, dimension, analyzer, results)
        else:
            log_info(f"跳过工作表: {sheet_name} (维度: {dimension}) - 无数据")
    
    # 保存文件
    wb.save(output_path)
    wb.close()


def _create_overview_sheet(wb: Workbook, results: Dict, total_records: int) -> None:
    """创建概览页"""
    ws = wb.create_sheet(title="概览")
    
    # 计算总体统计
    daily_results = results.get('daily', {})
    total_success = sum(stats.get('success_requests', 0) for stats in daily_results.values())
    total_slow = sum(stats.get('slow_requests', 0) for stats in daily_results.values())
    
    # 标题
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "高级时间维度分析报告"
    title_cell.font = Font(bold=True, size=16)
    
    # 统计数据
    overview_data = [
        ["指标名称", "数值", "单位", "说明"],
        ["总请求数", total_records, "个", "所有HTTP请求"],
        ["成功请求数", total_success, "个", "状态码2xx-3xx"],
        ["成功率", f"{total_success/total_records*100:.2f}" if total_records > 0 else "0", "%", "成功请求占比"],
        ["慢请求数", total_slow, "个", f"响应时间>{DEFAULT_SLOW_THRESHOLD}s"],
        ["慢请求率", f"{total_slow/total_records*100:.2f}" if total_records > 0 else "0", "%", "慢请求占比"],
        ["分析维度", "4个", "", "日/时/分/秒维度"],
        ["分位数", "P50/P95/P99", "", "响应时间分布"],
        ["连接指标", "新建/并发/活跃", "", "连接数统计"]
    ]
    
    # 写入数据
    for row_idx, row_data in enumerate(overview_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # 表头
                cell.font = Font(bold=True)
    
    format_excel_sheet(ws, has_grouped_header=False, header_end_row=3)


def _create_dimension_sheet(wb: Workbook, sheet_name: str, dimension: str, 
                           analyzer: AdvancedTimeDimensionAnalyzer, results: Dict) -> None:
    """创建维度分析页"""
    df = analyzer.create_output_dataframe(dimension, results)
    if df.empty:
        return
    
    time_label = analyzer._get_time_label(dimension)
    header_groups = analyzer.create_header_groups(df, time_label)
    
    # 添加到Excel
    ws = add_dataframe_to_excel_with_grouped_headers(wb, df, sheet_name, header_groups)
    
    # 添加图表
    if len(df) > 1 and len(df) <= 100:
        try:
            _add_dimension_charts(ws, df, sheet_name, time_label)
        except Exception as e:
            log_info(f"创建图表失败: {e}")


def _add_dimension_charts(ws, df: pd.DataFrame, sheet_name: str, time_label: str) -> None:
    """添加维度图表"""
    start_row = 4  # 跳过双行表头
    end_row = start_row + len(df) - 1
    
    # QPS趋势图
    create_line_chart(
        ws,
        min_row=start_row,
        max_row=end_row,
        title=f"{sheet_name}QPS趋势",
        x_title=time_label,
        y_title="QPS",
        y_cols=[7],  # QPS列
        chart_position="N5"
    )
    
    # 连接数分布图
    create_line_chart(
        ws,
        min_row=start_row,
        max_row=end_row,
        title=f"{sheet_name}连接数分布",
        x_title=time_label,
        y_title="连接数",
        y_cols=[8, 9, 10],  # 新建、并发、活跃连接数
        series_names=['新建连接数', '并发连接数', '活跃连接数'],
        chart_position="N20"
    )


def _prepare_output_filename(output_path: str, specific_uri_list: Optional[List[str]]) -> str:
    """准备输出文件名"""
    if not specific_uri_list:
        return output_path
    
    base_name = os.path.basename(output_path)
    name_parts = os.path.splitext(base_name)
    
    if isinstance(specific_uri_list, list):
        uri_identifier = specific_uri_list[0].replace('/', '_').replace('-', '_')
    else:
        uri_identifier = str(specific_uri_list).replace('/', '_').replace('-', '_')
    
    if len(uri_identifier) > 30:
        uri_identifier = uri_identifier[:30]
    
    new_filename = f"{name_parts[0]}_advanced_{uri_identifier}{name_parts[1]}"
    return os.path.join(os.path.dirname(output_path), new_filename)


# 主函数 - 保持向后兼容
def analyze_time_dimension(csv_path: str, output_path: str, 
                          specific_uri_list: Optional[List[str]] = None) -> str:
    """主分析函数 - 兼容原接口"""
    return analyze_time_dimension_advanced(csv_path, output_path, specific_uri_list)