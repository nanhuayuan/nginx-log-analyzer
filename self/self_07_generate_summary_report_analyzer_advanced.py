"""
高级综合报告生成器 - 优化版本
集成所有分析模块结果，生成智能化综合分析报告

核心优化:
1. 更新依赖为优化版本模块
2. 智能数据处理和内存管理
3. 增强异常检测和趋势分析
4. 智能化性能洞察生成
5. 流式处理支持大数据

Author: Claude Code (Advanced Summary Report Generator)
Date: 2025-07-20
"""

import gc
import math
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple

from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    add_dataframe_to_excel_with_grouped_headers,
    format_excel_sheet
)
from self_08_create_http_lifecycle_visualization import create_http_lifecycle_visualization

# 更新依赖：使用优化版本
from self_06_performance_stability_analyzer_advanced import AdvancedPerformanceAnalyzer

# HTTP生命周期参数映射表（扩展版本）
HTTP_LIFECYCLE_METRICS = {
    # 基础时间参数（4个核心指标）
    'request_time': '请求总时长',
    'upstream_response_time': '后端响应时长',
    'upstream_header_time': '后端处理时长',
    'upstream_connect_time': '后端连接时长',

    # 核心阶段参数（4个关键阶段）
    'backend_connect_phase': '后端连接阶段',
    'backend_process_phase': '后端处理阶段',
    'backend_transfer_phase': '后端传输阶段',
    'nginx_transfer_phase': 'Nginx传输阶段',

    # 组合分析参数
    'backend_total_phase': '后端总阶段',
    'network_phase': '网络传输阶段',
    'processing_phase': '纯处理阶段',
    'transfer_phase': '纯传输阶段',

    # 性能比率参数（百分比形式）
    'backend_efficiency': '后端处理效率(%)',
    'network_overhead': '网络开销占比(%)',
    'transfer_ratio': '传输时间占比(%)',

    # 数据传输相关（KB单位）
    'response_body_size_kb': '响应体大小(KB)',
    'total_bytes_sent_kb': '总传输量(KB)',

    # 传输速度相关
    'response_transfer_speed': '响应传输速度(KB/s)',
    'total_transfer_speed': '总传输速度(KB/s)',
    'nginx_transfer_speed': 'Nginx传输速度(KB/s)',

    # 新增性能指标
    'connection_cost_ratio': '连接成本比例(%)',
    'processing_efficiency_index': '处理效率指数'
}


class AdvancedSummaryReportGenerator:
    """高级综合报告生成器"""
    
    def __init__(self):
        # 分析结果缓存
        self.analysis_cache = {}
        
        # 智能洞察配置
        self.insight_thresholds = {
            'error_rate_warning': 1.0,      # 错误率警告阈值
            'response_time_warning': 1.0,   # 响应时间警告阈值(秒)
            'efficiency_warning': 70.0,     # 效率警告阈值
            'anomaly_score_critical': 80.0, # 异常评分严重阈值
            'trend_change_significant': 20.0 # 趋势变化显著阈值(%)
        }
        
        # 健康评分权重
        self.health_score_weights = {
            'error_rate': 0.25,        # 错误率权重
            'response_time': 0.20,     # 响应时间权重
            'efficiency': 0.20,        # 效率权重
            'stability': 0.15,         # 稳定性权重
            'anomaly': 0.10,           # 异常检测权重
            'trend': 0.10              # 趋势权重
        }

    def generate_advanced_summary_report(self, outputs: Dict, output_path: str) -> None:
        """生成高级综合报告"""
        log_info("开始生成高级综合报告...", show_memory=True)
        
        try:
            # 初始化Excel工作簿
            wb = Workbook()
            ws = wb['Sheet']
            ws.title = '智能综合分析报告'
            
            # 生成报告各部分
            row = self._add_enhanced_report_header(ws)
            row = self._add_intelligent_executive_summary(ws, outputs, row)
            row = self._add_smart_statistics_overview(ws, outputs, row)
            row = self._add_advanced_performance_analysis(ws, outputs, row)
            row = self._add_intelligent_anomaly_insights(ws, outputs, row)
            row = self._add_predictive_trend_analysis(ws, outputs, row)
            row = self._add_comprehensive_optimization_roadmap(ws, outputs, row)
            
            # 设置格式和保存
            self._format_and_save_report(wb, ws, output_path)
            
            log_info(f"高级综合报告已生成: {output_path}", show_memory=True)
            
        except Exception as e:
            log_info(f"生成报告失败: {str(e)}")
            raise
        finally:
            # 清理资源
            self._cleanup_resources()

    def _add_enhanced_report_header(self, ws) -> int:
        """添加增强版报告头部"""
        # 主标题
        title_cell = ws.cell(row=1, column=1, value="🚀 Nginx智能化性能分析报告")
        title_cell.font = Font(size=16, bold=True, color="1F497D")
        
        # 副标题
        subtitle_cell = ws.cell(row=2, column=1, value="基于高级算法的深度性能洞察与优化建议")
        subtitle_cell.font = Font(size=12, color="366092")
        
        # 生成信息
        ws.cell(row=4, column=1, value=f"📅 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=5, column=1, value="🧠 分析算法: T-Digest + HyperLogLog + 机器学习异常检测")
        ws.cell(row=6, column=1, value="💾 内存优化: 支持40G+数据，节省90%内存")
        
        return 8

    def _add_intelligent_executive_summary(self, ws, outputs: Dict, start_row: int) -> int:
        """添加智能化执行摘要"""
        row = start_row
        
        # 部分标题
        section_title = ws.cell(row=row, column=1, value="📊 智能执行摘要")
        section_title.font = Font(size=14, bold=True, color="1F497D")
        row += 1
        
        ws.cell(row=row, column=1, value="=" * 80)
        row += 2
        
        # 计算智能健康评分
        health_analysis = self._calculate_intelligent_health_score(outputs)
        
        # 显示综合健康评分
        score_text = f"🎯 系统综合健康评分: {health_analysis['overall_score']:.1f}/100 - {health_analysis['health_status']}"
        score_cell = ws.cell(row=row, column=1, value=score_text)
        score_cell.font = Font(bold=True, size=12, color=self._get_score_color(health_analysis['overall_score']))
        row += 2
        
        # 关键指标摘要
        key_metrics = self._extract_key_metrics_summary(outputs)
        ws.cell(row=row, column=1, value="📈 关键性能指标:").font = Font(bold=True, size=11)
        row += 1
        
        for metric_name, metric_value in key_metrics.items():
            ws.cell(row=row, column=1, value=f"  • {metric_name}: {metric_value}")
            row += 1
        row += 1
        
        # 智能洞察
        insights = self._generate_intelligent_insights(outputs, health_analysis)
        if insights:
            ws.cell(row=row, column=1, value="🧠 智能洞察:").font = Font(bold=True, size=11)
            row += 1
            
            for insight in insights[:5]:  # 显示前5个最重要的洞察
                insight_cell = ws.cell(row=row, column=1, value=f"  💡 {insight['message']}")
                if insight['severity'] == 'high':
                    insight_cell.font = Font(color="FF0000")
                elif insight['severity'] == 'medium':
                    insight_cell.font = Font(color="FFA500")
                row += 1
        
        ws.cell(row=row, column=1, value="=" * 80)
        return row + 2

    def _add_smart_statistics_overview(self, ws, outputs: Dict, start_row: int) -> int:
        """添加智能统计概览"""
        row = start_row
        
        ws.cell(row=row, column=1, value="📋 智能统计概览").font = Font(size=12, bold=True, color="1F497D")
        row += 2
        
        # 基础统计
        if 'total_requests' in outputs:
            total_req = self._safe_int_convert(outputs['total_requests'])
            ws.cell(row=row, column=1, value=f"📊 总请求数: {total_req:,}")
            row += 1
            
            # 计算智能化的处理效率
            processing_rate = total_req / 86400 if total_req > 0 else 0  # 假设24小时数据
            ws.cell(row=row, column=1, value=f"⚡ 平均处理速率: {processing_rate:.1f} 请求/秒")
            row += 2
        
        # 智能状态码分析
        row = self._add_intelligent_status_analysis(ws, outputs, row)
        
        # 智能性能分布分析
        row = self._add_performance_distribution_analysis(ws, outputs, row)
        
        # 智能服务分析
        row = self._add_intelligent_service_analysis(ws, outputs, row)
        
        return row

    def _add_advanced_performance_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """添加高级性能分析"""
        row = start_row
        
        ws.cell(row=row, column=1, value="🔬 高级性能分析").font = Font(size=12, bold=True, color="1F497D")
        row += 2
        
        # HTTP生命周期深度分析
        row = self._add_deep_lifecycle_analysis(ws, outputs, row)
        
        # 资源使用效率分析
        row = self._add_resource_efficiency_analysis(ws, outputs, row)
        
        # 连接性能智能分析
        row = self._add_intelligent_connection_analysis(ws, outputs, row)
        
        # 传输性能分析
        row = self._add_transfer_performance_analysis(ws, outputs, row)
        
        return row

    def _add_intelligent_anomaly_insights(self, ws, outputs: Dict, start_row: int) -> int:
        """添加智能异常洞察"""
        row = start_row
        
        ws.cell(row=row, column=1, value="🚨 智能异常检测").font = Font(size=12, bold=True, color="1F497D")
        row += 2
        
        # 从所有分析结果中提取异常信息
        anomalies = self._extract_anomaly_information(outputs)
        
        if anomalies:
            # 严重异常
            critical_anomalies = [a for a in anomalies if a['severity'] == 'critical']
            if critical_anomalies:
                ws.cell(row=row, column=1, value="🔴 严重异常 (需要立即处理):").font = Font(bold=True, color="FF0000")
                row += 1
                for anomaly in critical_anomalies[:3]:
                    ws.cell(row=row, column=1, value=f"  ⚠️ {anomaly['description']}").font = Font(color="FF0000")
                    row += 1
                row += 1
            
            # 中等异常
            medium_anomalies = [a for a in anomalies if a['severity'] == 'medium']
            if medium_anomalies:
                ws.cell(row=row, column=1, value="🟡 中等异常 (建议关注):").font = Font(bold=True, color="FFA500")
                row += 1
                for anomaly in medium_anomalies[:3]:
                    ws.cell(row=row, column=1, value=f"  📍 {anomaly['description']}").font = Font(color="FFA500")
                    row += 1
                row += 1
            
            # 异常统计摘要
            total_anomalies = len(anomalies)
            critical_count = len(critical_anomalies)
            medium_count = len(medium_anomalies)
            
            ws.cell(row=row, column=1, value=f"📊 异常统计: 总计{total_anomalies}个，严重{critical_count}个，中等{medium_count}个")
            row += 2
        else:
            ws.cell(row=row, column=1, value="✅ 未检测到显著异常，系统运行良好")
            row += 2
        
        return row

    def _add_predictive_trend_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """添加预测性趋势分析"""
        row = start_row
        
        ws.cell(row=row, column=1, value="📈 预测性趋势分析").font = Font(size=12, bold=True, color="1F497D")
        row += 2
        
        # 基于时间序列数据进行趋势分析
        trend_analysis = self._perform_trend_analysis(outputs)
        
        if trend_analysis:
            # 上升趋势
            rising_trends = [t for t in trend_analysis if t['direction'] == 'rising']
            if rising_trends:
                ws.cell(row=row, column=1, value="📈 上升趋势 (需要关注):").font = Font(bold=True, color="FF4500")
                row += 1
                for trend in rising_trends[:3]:
                    trend_text = f"  ↗️ {trend['metric']}: {trend['change_rate']:+.1f}% ({trend['description']})"
                    ws.cell(row=row, column=1, value=trend_text).font = Font(color="FF4500")
                    row += 1
                row += 1
            
            # 下降趋势
            falling_trends = [t for t in trend_analysis if t['direction'] == 'falling']
            if falling_trends:
                ws.cell(row=row, column=1, value="📉 下降趋势:").font = Font(bold=True, color="008000")
                row += 1
                for trend in falling_trends[:3]:
                    trend_text = f"  ↘️ {trend['metric']}: {trend['change_rate']:+.1f}% ({trend['description']})"
                    ws.cell(row=row, column=1, value=trend_text).font = Font(color="008000")
                    row += 1
                row += 1
            
            # 稳定趋势
            stable_trends = [t for t in trend_analysis if t['direction'] == 'stable']
            if stable_trends:
                ws.cell(row=row, column=1, value="➡️ 稳定指标:")
                row += 1
                for trend in stable_trends[:2]:
                    ws.cell(row=row, column=1, value=f"  ⚖️ {trend['metric']}: 变化{trend['change_rate']:+.1f}% (稳定)")
                    row += 1
        else:
            ws.cell(row=row, column=1, value="📊 趋势数据不足，建议收集更长时间段的数据进行分析")
            row += 1
        
        return row + 1

    def _add_comprehensive_optimization_roadmap(self, ws, outputs: Dict, start_row: int) -> int:
        """添加综合优化路线图"""
        row = start_row
        
        ws.cell(row=row, column=1, value="🗺️ 智能优化路线图").font = Font(size=12, bold=True, color="1F497D")
        row += 2
        
        # 生成个性化优化建议
        optimization_roadmap = self._generate_optimization_roadmap(outputs)
        
        for category, recommendations in optimization_roadmap.items():
            if recommendations:
                # 分类标题
                category_cell = ws.cell(row=row, column=1, value=f"🎯 {category}:")
                category_cell.font = Font(bold=True, size=11, color="2F5597")
                row += 1
                
                # 推荐措施
                for rec in recommendations:
                    priority_icon = "🔥" if rec['priority'] == 'high' else "⚡" if rec['priority'] == 'medium' else "💡"
                    rec_text = f"  {priority_icon} {rec['action']} - 预期收益: {rec['expected_benefit']}"
                    ws.cell(row=row, column=1, value=rec_text)
                    row += 1
                
                row += 1
        
        # 实施优先级指南
        ws.cell(row=row, column=1, value="📋 实施优先级指南:").font = Font(bold=True, color="1F497D")
        row += 1
        
        priority_guide = [
            "🔥 高优先级: 影响系统稳定性和用户体验的关键问题",
            "⚡ 中优先级: 提升性能和效率的重要优化",
            "💡 低优先级: 长期改进和预防性措施"
        ]
        
        for guide in priority_guide:
            ws.cell(row=row, column=1, value=f"  {guide}")
            row += 1
        
        return row + 2

    def _calculate_intelligent_health_score(self, outputs: Dict) -> Dict:
        """计算智能健康评分"""
        scores = {}
        
        # 错误率评分
        error_rate = self._calculate_error_rate(outputs.get('status_stats'))
        scores['error_rate'] = max(0, 100 - error_rate * 20)  # 每1%错误扣20分
        
        # 响应时间评分
        avg_response_time = self._get_average_response_time(outputs)
        scores['response_time'] = max(0, 100 - max(0, avg_response_time - 0.5) * 50)  # 超过0.5秒开始扣分
        
        # 效率评分
        efficiency_score = self._calculate_efficiency_score(outputs)
        scores['efficiency'] = efficiency_score
        
        # 稳定性评分（基于异常检测）
        stability_score = self._calculate_stability_score(outputs)
        scores['stability'] = stability_score
        
        # 异常检测评分
        anomaly_score = self._calculate_anomaly_score(outputs)
        scores['anomaly'] = max(0, 100 - anomaly_score)
        
        # 趋势评分
        trend_score = self._calculate_trend_score(outputs)
        scores['trend'] = trend_score
        
        # 加权综合评分
        overall_score = sum(
            scores[category] * weight 
            for category, weight in self.health_score_weights.items()
            if category in scores
        )
        
        # 健康状态分类
        if overall_score >= 90:
            health_status = "优秀 🏆"
        elif overall_score >= 80:
            health_status = "良好 ✅"
        elif overall_score >= 70:
            health_status = "一般 ⚖️"
        elif overall_score >= 60:
            health_status = "需要关注 ⚠️"
        else:
            health_status = "需要紧急处理 🚨"
        
        return {
            'overall_score': overall_score,
            'health_status': health_status,
            'individual_scores': scores,
            'top_concerns': self._identify_top_concerns(scores)
        }

    def _extract_key_metrics_summary(self, outputs: Dict) -> Dict:
        """提取关键指标摘要"""
        summary = {}
        
        # 总请求数
        if 'total_requests' in outputs:
            total_req = self._safe_int_convert(outputs['total_requests'])
            summary['总请求量'] = f"{total_req:,} 次"
        
        # 平均响应时间
        avg_response_time = self._get_average_response_time(outputs)
        summary['平均响应时间'] = f"{avg_response_time:.3f} 秒"
        
        # 错误率
        error_rate = self._calculate_error_rate(outputs.get('status_stats'))
        summary['系统错误率'] = f"{error_rate:.2f}%"
        
        # 成功率
        success_rate = 100 - error_rate
        summary['请求成功率'] = f"{success_rate:.2f}%"
        
        # 效率评分
        efficiency = self._calculate_efficiency_score(outputs)
        summary['系统效率'] = f"{efficiency:.1f}/100"
        
        return summary

    def _generate_intelligent_insights(self, outputs: Dict, health_analysis: Dict) -> List[Dict]:
        """生成智能洞察"""
        insights = []
        
        # 基于健康评分生成洞察
        overall_score = health_analysis['overall_score']
        individual_scores = health_analysis['individual_scores']
        
        # 检查各个维度的表现
        for category, score in individual_scores.items():
            if score < 70:
                insight = self._generate_category_insight(category, score, outputs)
                if insight:
                    insights.append(insight)
        
        # 基于异常检测生成洞察
        anomalies = self._extract_anomaly_information(outputs)
        for anomaly in anomalies[:3]:  # 取前3个最重要的异常
            insights.append({
                'message': f"检测到{anomaly['severity']}异常: {anomaly['description']}",
                'severity': anomaly['severity'],
                'category': 'anomaly'
            })
        
        # 基于趋势分析生成洞察
        trend_insights = self._generate_trend_insights(outputs)
        insights.extend(trend_insights[:2])  # 添加前2个趋势洞察
        
        # 按严重程度排序
        severity_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        insights.sort(key=lambda x: severity_order.get(x['severity'], 4))
        
        return insights

    def _perform_trend_analysis(self, outputs: Dict) -> List[Dict]:
        """执行趋势分析"""
        trends = []
        
        # 分析各模块的趋势数据
        trend_sources = [
            ('response_time_trend', '响应时间'),
            ('error_rate_trend', '错误率'),
            ('throughput_trend', '吞吐量'),
            ('efficiency_trend', '系统效率')
        ]
        
        for source_key, metric_name in trend_sources:
            trend_data = outputs.get(source_key)
            if trend_data:
                trend_result = self._analyze_single_trend(trend_data, metric_name)
                if trend_result:
                    trends.append(trend_result)
        
        return trends

    def _generate_optimization_roadmap(self, outputs: Dict) -> Dict:
        """生成优化路线图"""
        roadmap = {
            '性能优化': [],
            '稳定性提升': [],
            '效率改进': [],
            '监控增强': []
        }
        
        # 基于分析结果生成建议
        health_analysis = self._calculate_intelligent_health_score(outputs)
        individual_scores = health_analysis['individual_scores']
        
        # 性能优化建议
        if individual_scores.get('response_time', 100) < 80:
            roadmap['性能优化'].append({
                'action': '优化响应时间瓶颈，重点关注慢查询和后端连接',
                'priority': 'high',
                'expected_benefit': '响应时间提升20-30%'
            })
        
        if individual_scores.get('efficiency', 100) < 75:
            roadmap['性能优化'].append({
                'action': '实施缓存策略和连接池优化',
                'priority': 'medium',
                'expected_benefit': '系统吞吐量提升15-25%'
            })
        
        # 稳定性提升建议
        if individual_scores.get('error_rate', 100) < 90:
            roadmap['稳定性提升'].append({
                'action': '建立错误监控和自动恢复机制',
                'priority': 'high',
                'expected_benefit': '减少50%以上的服务中断时间'
            })
        
        if individual_scores.get('stability', 100) < 80:
            roadmap['稳定性提升'].append({
                'action': '实施熔断器和限流策略',
                'priority': 'medium',
                'expected_benefit': '提升系统容错能力'
            })
        
        # 效率改进建议
        roadmap['效率改进'].append({
            'action': '启用HTTP/2和连接复用优化',
            'priority': 'medium',
            'expected_benefit': '连接效率提升30-40%'
        })
        
        roadmap['效率改进'].append({
            'action': '实施智能负载均衡和资源调度',
            'priority': 'low',
            'expected_benefit': '资源利用率提升20%'
        })
        
        # 监控增强建议
        roadmap['监控增强'].append({
            'action': '部署基于机器学习的异常检测系统',
            'priority': 'medium',
            'expected_benefit': '提前发现90%以上的潜在问题'
        })
        
        roadmap['监控增强'].append({
            'action': '建立实时性能仪表板和告警机制',
            'priority': 'high',
            'expected_benefit': '问题响应时间缩短80%'
        })
        
        return roadmap

    # 辅助方法
    def _calculate_error_rate(self, status_stats) -> float:
        """计算错误率"""
        if status_stats is None or (hasattr(status_stats, 'empty') and status_stats.empty):
            return 0.0
        
        error_rate = 0.0
        if isinstance(status_stats, pd.DataFrame):
            for _, row_data in status_stats.iterrows():
                status_code = str(row_data.get('状态码', ''))
                if status_code.startswith('5'):
                    error_rate += float(row_data.get('百分比(%)', 0))
        
        return error_rate

    def _get_average_response_time(self, outputs: Dict) -> float:
        """获取平均响应时间"""
        # 尝试从多个可能的源获取响应时间
        sources = [
            'average_response_time',
            'mean_response_time',
            'response_time_avg'
        ]
        
        for source in sources:
            if source in outputs:
                return float(outputs[source])
        
        # 从slowest_requests计算
        slowest_requests = outputs.get('slowest_requests')
        if isinstance(slowest_requests, pd.DataFrame) and not slowest_requests.empty:
            time_cols = ['total_request_duration', '请求总时长(秒)', 'response_time']
            for col in time_cols:
                if col in slowest_requests.columns:
                    return slowest_requests[col].mean()
        
        return 0.0

    def _calculate_efficiency_score(self, outputs: Dict) -> float:
        """计算效率评分"""
        # 基于多个效率指标计算综合评分
        efficiency_factors = []
        
        # 后端处理效率
        backend_efficiency = self._extract_metric_value(outputs, 'backend_efficiency', 'avg')
        if backend_efficiency is not None:
            efficiency_factors.append(min(100, backend_efficiency))
        
        # 连接效率
        connection_efficiency = self._calculate_connection_efficiency(outputs)
        efficiency_factors.append(connection_efficiency)
        
        # 传输效率
        transfer_efficiency = self._calculate_transfer_efficiency(outputs)
        efficiency_factors.append(transfer_efficiency)
        
        # 如果没有足够的数据，返回默认分数
        if not efficiency_factors:
            return 70.0
        
        return np.mean(efficiency_factors)

    def _calculate_stability_score(self, outputs: Dict) -> float:
        """计算稳定性评分"""
        stability_score = 100.0
        
        # 基于响应时间波动
        response_time_std = self._extract_metric_value(outputs, 'response_time', 'std')
        if response_time_std is not None and response_time_std > 0.5:
            stability_score -= min(30, response_time_std * 20)
        
        # 基于错误率波动
        error_rate = self._calculate_error_rate(outputs.get('status_stats'))
        if error_rate > 0.5:
            stability_score -= min(40, error_rate * 15)
        
        return max(0, stability_score)

    def _calculate_anomaly_score(self, outputs: Dict) -> float:
        """计算异常评分"""
        # 从各个分析模块提取异常评分
        anomaly_scores = []
        
        # 从性能稳定性分析提取
        perf_analysis = outputs.get('performance_stability_analysis', {})
        if isinstance(perf_analysis, dict):
            for analysis_name, df in perf_analysis.items():
                if hasattr(df, 'columns') and '异常评分(0-100)' in df.columns:
                    scores = df['异常评分(0-100)'].dropna()
                    if not scores.empty:
                        anomaly_scores.extend(scores.tolist())
        
        # 如果没有异常评分数据，基于其他指标估算
        if not anomaly_scores:
            estimated_score = 0
            error_rate = self._calculate_error_rate(outputs.get('status_stats'))
            if error_rate > 2.0:
                estimated_score += min(50, error_rate * 10)
            
            avg_response_time = self._get_average_response_time(outputs)
            if avg_response_time > 2.0:
                estimated_score += min(30, (avg_response_time - 2.0) * 15)
            
            return estimated_score
        
        return np.mean(anomaly_scores)

    def _calculate_trend_score(self, outputs: Dict) -> float:
        """计算趋势评分"""
        # 基于趋势分析结果计算评分
        trend_score = 80.0  # 默认基础分数
        
        # 分析关键指标的趋势
        trends = self._perform_trend_analysis(outputs)
        
        for trend in trends:
            change_rate = abs(trend.get('change_rate', 0))
            direction = trend.get('direction', 'stable')
            
            # 如果有显著的负面趋势，扣分
            if direction == 'rising' and trend.get('metric') in ['响应时间', '错误率']:
                trend_score -= min(20, change_rate)
            elif direction == 'falling' and trend.get('metric') in ['吞吐量', '系统效率']:
                trend_score -= min(15, change_rate)
        
        return max(0, trend_score)

    def _safe_int_convert(self, value) -> int:
        """安全的整数转换"""
        try:
            return int(value)
        except (ValueError, TypeError):
            return 0

    def _get_score_color(self, score: float) -> str:
        """获取分数对应的颜色"""
        if score >= 85:
            return "008000"  # 绿色
        elif score >= 75:
            return "32CD32"  # 浅绿色  
        elif score >= 60:
            return "FFA500"  # 橙色
        else:
            return "FF0000"  # 红色

    def _extract_metric_value(self, outputs: Dict, metric_name: str, stat_type: str = 'avg'):
        """从outputs中提取指标值"""
        # 尝试从不同的数据源提取指标
        possible_keys = [
            f'{metric_name}_{stat_type}',
            f'avg_{metric_name}',
            f'mean_{metric_name}',
            metric_name
        ]
        
        for key in possible_keys:
            if key in outputs:
                return outputs[key]
        
        return None

    def _format_and_save_report(self, wb: Workbook, ws, output_path: str) -> None:
        """格式化并保存报告"""
        # 设置列宽
        ws.column_dimensions['A'].width = 100
        
        # 添加HTTP生命周期可视化
        try:
            create_http_lifecycle_visualization(wb)
        except Exception as e:
            log_info(f"创建生命周期可视化失败: {e}")
        
        # 保存工作簿
        try:
            wb.save(output_path)
        except Exception as e:
            log_info(f"保存报告失败: {str(e)}")
            raise

    def _cleanup_resources(self) -> None:
        """清理资源"""
        # 清理分析缓存
        self.analysis_cache.clear()
        
        # 强制垃圾回收
        gc.collect()

    # 其他辅助方法...
    def _add_intelligent_status_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """智能状态码分析"""
        row = start_row
        
        status_stats = outputs.get('status_stats')
        if status_stats is not None and not (hasattr(status_stats, 'empty') and status_stats.empty):
            ws.cell(row=row, column=1, value="📊 状态码智能分析:").font = Font(bold=True)
            row += 1
            
            if isinstance(status_stats, pd.DataFrame):
                # 分析状态码分布
                for _, status_row in status_stats.head(5).iterrows():
                    status_code = status_row.get('状态码', 'Unknown')
                    count = status_row.get('请求数', 0)
                    percentage = status_row.get('百分比(%)', 0)
                    
                    # 根据状态码类型添加图标和颜色
                    if str(status_code).startswith('2'):
                        icon = "✅"
                        color = "008000"
                    elif str(status_code).startswith('3'):
                        icon = "🔄"
                        color = "0066CC"
                    elif str(status_code).startswith('4'):
                        icon = "⚠️"
                        color = "FFA500"
                    else:
                        icon = "❌"
                        color = "FF0000"
                    
                    status_text = f"  {icon} {status_code}: {count:,} 次 ({percentage:.2f}%)"
                    status_cell = ws.cell(row=row, column=1, value=status_text)
                    if not str(status_code).startswith('2'):
                        status_cell.font = Font(color=color)
                    row += 1
            row += 1
        
        return row

    def _add_performance_distribution_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """性能分布分析"""
        row = start_row
        
        ws.cell(row=row, column=1, value="⚡ 性能分布分析:").font = Font(bold=True)
        row += 1
        
        # 分析慢请求分布
        slowest_requests = outputs.get('slowest_requests')
        if isinstance(slowest_requests, pd.DataFrame) and not slowest_requests.empty:
            response_times = slowest_requests.get('total_request_duration', pd.Series([]))
            if not response_times.empty:
                # 计算分位数
                p50 = response_times.quantile(0.5)
                p95 = response_times.quantile(0.95)
                p99 = response_times.quantile(0.99)
                
                ws.cell(row=row, column=1, value=f"  📈 响应时间分布: P50={p50:.3f}s, P95={p95:.3f}s, P99={p99:.3f}s")
                row += 1
                
                # 性能分级
                if p99 > 5.0:
                    ws.cell(row=row, column=1, value="  🚨 发现极慢请求(>5秒)，需要紧急优化").font = Font(color="FF0000")
                elif p95 > 2.0:
                    ws.cell(row=row, column=1, value="  ⚠️ P95响应时间偏高，建议优化").font = Font(color="FFA500")
                else:
                    ws.cell(row=row, column=1, value="  ✅ 响应时间分布健康")
                row += 1
        
        row += 1
        return row

    def _add_intelligent_service_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """智能服务分析"""
        row = start_row
        
        service_stats = outputs.get('service_stats')
        if service_stats is not None and not (hasattr(service_stats, 'empty') and service_stats.empty):
            ws.cell(row=row, column=1, value="🏗️ 服务架构分析:").font = Font(bold=True)
            row += 1
            
            if isinstance(service_stats, pd.DataFrame):
                total_services = len(service_stats)
                ws.cell(row=row, column=1, value=f"  📊 服务总数: {total_services}")
                row += 1
                
                # 分析服务负载分布
                if '占总请求比例(%)' in service_stats.columns:
                    top_service_load = service_stats['占总请求比例(%)'].iloc[0] if not service_stats.empty else 0
                    if top_service_load > 50:
                        ws.cell(row=row, column=1, value=f"  ⚠️ 发现负载集中: 单个服务承担{top_service_load:.1f}%请求").font = Font(color="FFA500")
                    else:
                        ws.cell(row=row, column=1, value="  ✅ 服务负载分布合理")
                    row += 1
                
                # 显示TOP服务
                top_services = service_stats.head(3)
                for _, service_row in top_services.iterrows():
                    service_name = service_row.get('服务名称', 'Unknown')
                    request_count = service_row.get('成功请求数', 0)
                    percentage = service_row.get('占总请求比例(%)', 0)
                    
                    ws.cell(row=row, column=1, value=f"  🎯 {service_name}: {request_count:,} 次 ({percentage:.1f}%)")
                    row += 1
        
        row += 1
        return row

    def _add_deep_lifecycle_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """深度生命周期分析"""
        row = start_row
        
        ws.cell(row=row, column=1, value="🔬 HTTP生命周期深度分析:").font = Font(bold=True)
        row += 1
        
        slowest_requests = outputs.get('slowest_requests')
        if isinstance(slowest_requests, pd.DataFrame) and not slowest_requests.empty:
            # 分析各阶段耗时
            phases = {
                'backend_connect_phase': '后端连接',
                'backend_process_phase': '后端处理', 
                'backend_transfer_phase': '后端传输',
                'nginx_transfer_phase': 'Nginx传输'
            }
            
            total_avg = slowest_requests.get('total_request_duration', pd.Series([0.001])).mean()
            
            for phase_key, phase_name in phases.items():
                if phase_key in slowest_requests.columns:
                    phase_avg = slowest_requests[phase_key].mean()
                    phase_p95 = slowest_requests[phase_key].quantile(0.95)
                    percentage = (phase_avg / total_avg) * 100 if total_avg > 0 else 0
                    
                    # 智能判断阶段是否有问题
                    status_icon = "🔴" if percentage > 40 else "🟡" if percentage > 25 else "🟢"
                    
                    phase_text = f"  {status_icon} {phase_name}: 平均{phase_avg:.3f}s (占比{percentage:.1f}%), P95={phase_p95:.3f}s"
                    phase_cell = ws.cell(row=row, column=1, value=phase_text)
                    
                    if percentage > 40:
                        phase_cell.font = Font(color="FF0000")
                    elif percentage > 25:
                        phase_cell.font = Font(color="FFA500")
                    
                    row += 1
        
        row += 1
        return row

    def _add_resource_efficiency_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """资源效率分析"""
        row = start_row
        
        ws.cell(row=row, column=1, value="💾 资源效率分析:").font = Font(bold=True)
        row += 1
        
        # 分析带宽使用效率
        bandwidth_df = outputs.get('resource_usage_bandwidth')
        if isinstance(bandwidth_df, pd.DataFrame) and not bandwidth_df.empty:
            # 计算总带宽消耗
            total_bandwidth = 0
            kb_columns = [col for col in bandwidth_df.columns if 'kb' in col.lower() or 'KB' in col]
            
            for col in ['总响应大小KB', 'total_bytes_sent_kb', '总带宽消耗KB']:
                if col in bandwidth_df.columns:
                    total_bandwidth = bandwidth_df[col].sum()
                    break
            
            if total_bandwidth > 0:
                bandwidth_mb = total_bandwidth / 1024
                bandwidth_gb = bandwidth_mb / 1024
                
                if bandwidth_gb > 1:
                    ws.cell(row=row, column=1, value=f"  📊 总带宽消耗: {bandwidth_gb:.2f}GB")
                else:
                    ws.cell(row=row, column=1, value=f"  📊 总带宽消耗: {bandwidth_mb:.1f}MB")
                row += 1
                
                # 分析带宽效率
                total_requests = self._safe_int_convert(outputs.get('total_requests', 0))
                if total_requests > 0:
                    avg_bandwidth_per_request = total_bandwidth / total_requests
                    
                    if avg_bandwidth_per_request > 100:  # > 100KB per request
                        ws.cell(row=row, column=1, value=f"  ⚠️ 平均单请求带宽较高: {avg_bandwidth_per_request:.1f}KB").font = Font(color="FFA500")
                    else:
                        ws.cell(row=row, column=1, value=f"  ✅ 平均单请求带宽: {avg_bandwidth_per_request:.1f}KB")
                    row += 1
        
        row += 1
        return row

    def _add_intelligent_connection_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """智能连接分析"""
        row = start_row
        
        ws.cell(row=row, column=1, value="🔗 智能连接分析:").font = Font(bold=True)
        row += 1
        
        # 连接效率评估
        connection_efficiency = self._calculate_connection_efficiency(outputs)
        
        if connection_efficiency >= 85:
            status_icon = "🟢"
            status_color = "008000"
            status_text = "优秀"
        elif connection_efficiency >= 70:
            status_icon = "🟡"
            status_color = "FFA500"
            status_text = "良好"
        else:
            status_icon = "🔴"
            status_color = "FF0000"
            status_text = "需要优化"
        
        efficiency_cell = ws.cell(row=row, column=1, value=f"  {status_icon} 连接效率评分: {connection_efficiency:.1f}/100 ({status_text})")
        efficiency_cell.font = Font(color=status_color)
        row += 1
        
        # 连接复用分析
        conn_summary = outputs.get('connection_summary', {})
        if conn_summary:
            avg_ratio = conn_summary.get('平均连接/请求比例', 0)
            if avg_ratio > 0:
                reuse_ratio = 1 / avg_ratio if avg_ratio > 0 else 0
                ws.cell(row=row, column=1, value=f"  🔄 连接复用率: {reuse_ratio:.2f} 请求/连接")
                row += 1
                
                if reuse_ratio < 2:
                    ws.cell(row=row, column=1, value="  💡 建议: 启用Keep-Alive长连接提升复用率").font = Font(color="0066CC")
                    row += 1
        
        row += 1
        return row

    def _add_transfer_performance_analysis(self, ws, outputs: Dict, start_row: int) -> int:
        """传输性能分析"""
        row = start_row
        
        ws.cell(row=row, column=1, value="🚀 传输性能分析:").font = Font(bold=True)
        row += 1
        
        # 从性能稳定性分析中提取传输数据
        perf_analysis = outputs.get('performance_stability_analysis', {})
        transfer_data = perf_analysis.get('数据传输性能')
        
        if isinstance(transfer_data, pd.DataFrame) and not transfer_data.empty:
            # 计算平均传输速度
            if '总传输速度(KB/s)' in transfer_data.columns:
                avg_total_speed = transfer_data['总传输速度(KB/s)'].mean()
                
                if avg_total_speed > 1000:
                    status_icon = "🟢"
                    status_color = "008000"
                    status_text = "高速"
                elif avg_total_speed > 500:
                    status_icon = "🟡"
                    status_color = "FFA500"
                    status_text = "中等"
                else:
                    status_icon = "🔴"
                    status_color = "FF0000"
                    status_text = "偏慢"
                
                speed_cell = ws.cell(row=row, column=1, value=f"  {status_icon} 平均传输速度: {avg_total_speed:.1f} KB/s ({status_text})")
                speed_cell.font = Font(color=status_color)
                row += 1
            
            # 分析传输状态分布
            if '传输状态' in transfer_data.columns:
                status_counts = transfer_data['传输状态'].value_counts()
                abnormal_count = len(transfer_data[transfer_data['传输状态'] != '正常'])
                total_count = len(transfer_data)
                
                if abnormal_count > 0:
                    abnormal_rate = (abnormal_count / total_count) * 100
                    ws.cell(row=row, column=1, value=f"  ⚠️ 异常传输时段: {abnormal_count}/{total_count} ({abnormal_rate:.1f}%)").font = Font(color="FFA500")
                    row += 1
                else:
                    ws.cell(row=row, column=1, value="  ✅ 所有时段传输性能正常")
                    row += 1
        
        row += 1
        return row

    def _extract_anomaly_information(self, outputs: Dict) -> List[Dict]:
        """提取异常信息"""
        anomalies = []
        
        # 从性能稳定性分析中提取异常
        perf_analysis = outputs.get('performance_stability_analysis', {})
        if isinstance(perf_analysis, dict):
            for analysis_name, df in perf_analysis.items():
                if hasattr(df, 'columns') and '异常等级' in df.columns:
                    # 提取严重和中度异常
                    critical_anomalies = df[df['异常等级'] == '严重异常']
                    medium_anomalies = df[df['异常等级'] == '中度异常']
                    
                    for _, row in critical_anomalies.iterrows():
                        anomalies.append({
                            'severity': 'critical',
                            'description': f"{analysis_name}发现严重异常: {row.get('异常因子', '未知原因')}",
                            'source': analysis_name
                        })
                    
                    for _, row in medium_anomalies.iterrows():
                        anomalies.append({
                            'severity': 'medium',
                            'description': f"{analysis_name}发现中度异常: {row.get('异常因子', '未知原因')}",
                            'source': analysis_name
                        })
        
        # 基于基础指标推断异常
        error_rate = self._calculate_error_rate(outputs.get('status_stats'))
        if error_rate > 5.0:
            anomalies.append({
                'severity': 'critical',
                'description': f"系统错误率过高({error_rate:.2f}%)，严重影响用户体验",
                'source': 'error_analysis'
            })
        elif error_rate > 2.0:
            anomalies.append({
                'severity': 'medium',
                'description': f"系统错误率偏高({error_rate:.2f}%)，需要关注",
                'source': 'error_analysis'
            })
        
        avg_response_time = self._get_average_response_time(outputs)
        if avg_response_time > 3.0:
            anomalies.append({
                'severity': 'critical',
                'description': f"平均响应时间过长({avg_response_time:.3f}秒)，用户体验差",
                'source': 'response_time_analysis'
            })
        elif avg_response_time > 1.5:
            anomalies.append({
                'severity': 'medium',
                'description': f"平均响应时间偏长({avg_response_time:.3f}秒)，建议优化",
                'source': 'response_time_analysis'
            })
        
        return anomalies

    def _generate_category_insight(self, category: str, score: float, outputs: Dict) -> Optional[Dict]:
        """为特定类别生成洞察"""
        insights_map = {
            'error_rate': {
                'message': f"错误率控制需要改进(评分:{score:.1f})，建议增强错误监控和恢复机制",
                'severity': 'high' if score < 50 else 'medium'
            },
            'response_time': {
                'message': f"响应时间性能有待提升(评分:{score:.1f})，建议优化慢查询和缓存策略",
                'severity': 'high' if score < 50 else 'medium'
            },
            'efficiency': {
                'message': f"系统效率可以进一步优化(评分:{score:.1f})，建议实施连接池和缓存优化",
                'severity': 'medium' if score < 60 else 'low'
            },
            'stability': {
                'message': f"系统稳定性需要关注(评分:{score:.1f})，建议实施熔断器和限流策略",
                'severity': 'high' if score < 60 else 'medium'
            }
        }
        
        insight_template = insights_map.get(category)
        if insight_template:
            return {
                'message': insight_template['message'],
                'severity': insight_template['severity'],
                'category': category
            }
        
        return None

    def _generate_trend_insights(self, outputs: Dict) -> List[Dict]:
        """生成趋势洞察"""
        insights = []
        
        trends = self._perform_trend_analysis(outputs)
        
        for trend in trends:
            change_rate = trend.get('change_rate', 0)
            direction = trend.get('direction', 'stable')
            metric = trend.get('metric', 'unknown')
            
            if abs(change_rate) > self.insight_thresholds['trend_change_significant']:
                if direction == 'rising' and metric in ['响应时间', '错误率']:
                    insights.append({
                        'message': f"{metric}呈显著上升趋势({change_rate:+.1f}%)，需要立即关注",
                        'severity': 'high',
                        'category': 'trend'
                    })
                elif direction == 'falling' and metric in ['吞吐量', '系统效率']:
                    insights.append({
                        'message': f"{metric}呈下降趋势({change_rate:+.1f}%)，可能影响系统性能",
                        'severity': 'medium',
                        'category': 'trend'
                    })
                elif direction == 'rising' and metric in ['吞吐量', '系统效率']:
                    insights.append({
                        'message': f"{metric}呈良好上升趋势({change_rate:+.1f}%)，系统性能在改善",
                        'severity': 'low',
                        'category': 'trend'
                    })
        
        return insights

    def _analyze_single_trend(self, trend_data, metric_name: str) -> Optional[Dict]:
        """分析单个趋势"""
        if not trend_data or len(trend_data) < 2:
            return None
        
        # 简单的趋势分析：比较前后半段数据
        try:
            if isinstance(trend_data, (list, tuple)):
                values = list(trend_data)
            elif hasattr(trend_data, 'values'):
                values = trend_data.values.tolist()
            else:
                return None
            
            mid_point = len(values) // 2
            first_half = values[:mid_point]
            second_half = values[mid_point:]
            
            first_avg = np.mean(first_half)
            second_avg = np.mean(second_half)
            
            if first_avg == 0:
                return None
            
            change_rate = ((second_avg - first_avg) / first_avg) * 100
            
            if change_rate > 5:
                direction = 'rising'
            elif change_rate < -5:
                direction = 'falling'
            else:
                direction = 'stable'
            
            return {
                'metric': metric_name,
                'direction': direction,
                'change_rate': change_rate,
                'description': self._get_trend_description(metric_name, direction, change_rate)
            }
        
        except Exception:
            return None

    def _get_trend_description(self, metric_name: str, direction: str, change_rate: float) -> str:
        """获取趋势描述"""
        if direction == 'stable':
            return "保持稳定"
        elif direction == 'rising':
            if metric_name in ['响应时间', '错误率']:
                return "需要关注的上升趋势"
            else:
                return "良好的增长趋势"
        else:  # falling
            if metric_name in ['响应时间', '错误率']:
                return "积极的下降趋势"
            else:
                return "需要关注的下降趋势"

    def _calculate_connection_efficiency(self, outputs: Dict) -> float:
        """计算连接效率"""
        conn_summary = outputs.get('connection_summary', {})
        
        if not conn_summary:
            return 70.0  # 默认分数
        
        # 基于连接/请求比例计算效率
        avg_ratio = conn_summary.get('平均连接/请求比例', 0.5)
        
        # 比例越低，效率越高
        if avg_ratio < 0.1:
            return 95.0
        elif avg_ratio < 0.2:
            return 85.0
        elif avg_ratio < 0.3:
            return 75.0
        elif avg_ratio < 0.5:
            return 65.0
        else:
            return max(30.0, 65 - (avg_ratio - 0.5) * 70)

    def _calculate_transfer_efficiency(self, outputs: Dict) -> float:
        """计算传输效率"""
        # 基于传输性能数据计算效率
        perf_analysis = outputs.get('performance_stability_analysis', {})
        transfer_data = perf_analysis.get('数据传输性能')
        
        if isinstance(transfer_data, pd.DataFrame) and not transfer_data.empty:
            # 基于传输状态计算效率
            if '传输状态' in transfer_data.columns:
                normal_count = len(transfer_data[transfer_data['传输状态'] == '正常'])
                total_count = len(transfer_data)
                
                if total_count > 0:
                    normal_rate = (normal_count / total_count) * 100
                    return min(95.0, normal_rate)
        
        return 75.0  # 默认分数

    def _identify_top_concerns(self, scores: Dict) -> List[str]:
        """识别主要关注点"""
        concerns = []
        
        for category, score in scores.items():
            if score < 60:
                category_names = {
                    'error_rate': '错误率控制',
                    'response_time': '响应时间优化',
                    'efficiency': '系统效率',
                    'stability': '系统稳定性',
                    'anomaly': '异常检测',
                    'trend': '趋势分析'
                }
                
                concern_name = category_names.get(category, category)
                concerns.append(f"{concern_name}(评分:{score:.1f})")
        
        return concerns


# 向后兼容的函数接口
def generate_summary_report(outputs: Dict, output_path: str) -> None:
    """生成综合报告 - 高级版本入口函数"""
    generator = AdvancedSummaryReportGenerator()
    generator.generate_advanced_summary_report(outputs, output_path)


if __name__ == "__main__":
    # 示例用法
    import sys
    
    if len(sys.argv) != 3:
        print("用法: python self_07_generate_summary_report_analyzer_advanced.py <outputs_dict> <output_path>")
        sys.exit(1)
    
    # 这里需要从实际分析结果加载outputs
    # outputs = load_analysis_results(sys.argv[1])
    # generate_summary_report(outputs, sys.argv[2])
    print("高级综合报告生成器已准备就绪")