import gc
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, Reference

from self_00_01_constants import DEFAULT_SLOW_THRESHOLD, DEFAULT_CHUNK_SIZE
from self_00_02_utils import log_info
from self_00_04_excel_processor import (
    format_excel_sheet,
    add_dataframe_to_excel_with_grouped_headers,
    create_pie_chart
)


def analyze_slow_requests(csv_path, output_path, threshold=DEFAULT_SLOW_THRESHOLD):
    """分析慢请求API，支持新增的性能指标和传输速度分析"""
    log_info(f"开始分析慢请求API (响应时间 > {threshold}秒)...", show_memory=True)

    # 优化后的字段映射：统一使用CSV优化后的列名
    column_mapping = {
        # 基础信息列
        'log_source_file': '来源文件',
        'application_name': '应用名称',
        'service_name': '服务名称',
        'raw_time': '请求时间',
        'http_method': '请求方法',
        'request_full_uri': '请求URI',
        'response_status_code': '状态码',

        # 核心时间指标（统一为"时长"术语）
        'total_request_duration': '请求总时长(秒)',
        'upstream_connect_time': '后端连接时长(秒)',
        'upstream_header_time': '后端处理时长(秒)',
        'upstream_response_time': '后端响应时长(秒)',

        # 阶段分析指标
        'backend_connect_phase': '后端连接阶段(秒)',
        'backend_process_phase': '后端处理阶段(秒)',
        'backend_transfer_phase': '后端传输阶段(秒)',
        'nginx_transfer_phase': 'Nginx传输阶段(秒)',

        # 组合分析指标
        'backend_total_phase': '后端总阶段(秒)',
        'network_phase': '网络传输阶段(秒)',
        'processing_phase': '纯处理阶段(秒)',
        'transfer_phase': '纯传输阶段(秒)',

        # 性能效率指标（百分比）
        'backend_efficiency': '后端处理效率(%)',
        'network_overhead': '网络开销占比(%)',
        'transfer_ratio': '传输时间占比(%)',

        # 数据传输指标（已统一为KB）
        'response_body_size_kb': '响应体大小(KB)',
        'total_bytes_sent_kb': '总传输大小(KB)',

        # 新增传输速度指标（具有重要分析价值）
        'response_transfer_speed': '响应体传输速度(KB/s)',
        'total_transfer_speed': '总传输速度(KB/s)',
        'nginx_transfer_speed': 'Nginx传输速度(KB/s)',

        # 新增性能分析指标
        'connection_cost_ratio': '连接成本占比(%)',
        'processing_efficiency_index': '处理效率指数'
    }

    # 核心分析指标分组
    time_metrics = [
        '请求总时长(秒)', '后端连接时长(秒)', '后端处理时长(秒)', '后端响应时长(秒)'
    ]

    phase_metrics = [
        '后端连接阶段(秒)', '后端处理阶段(秒)', '后端传输阶段(秒)', 'Nginx传输阶段(秒)',
        '后端总阶段(秒)', '网络传输阶段(秒)', '纯处理阶段(秒)', '纯传输阶段(秒)'
    ]

    efficiency_metrics = [
        '后端处理效率(%)', '网络开销占比(%)', '传输时间占比(%)',
        '连接成本占比(%)', '处理效率指数'
    ]

    transfer_metrics = [
        '响应体大小(KB)', '总传输大小(KB)', '响应体传输速度(KB/s)',
        '总传输速度(KB/s)', 'Nginx传输速度(KB/s)'
    ]

    all_metrics = time_metrics + phase_metrics + efficiency_metrics + transfer_metrics

    # 优化后的数据处理，减少内存使用
    chunk_size = max(DEFAULT_CHUNK_SIZE // 4, 10000)  # 减小chunk大小
    api_total_requests = {}
    total_processed = 0
    total_slow_requests = 0
    
    # 限制慢请求存储数量，避免OOM
    MAX_SLOW_REQUESTS = 50000  # 最多保存5万条慢请求
    
    # 使用临时文件存储慢请求数据
    import tempfile
    import os
    temp_dir = tempfile.mkdtemp()
    temp_slow_file = os.path.join(temp_dir, 'slow_requests.csv')
    
    # 第一遍：收集API总请求数和慢请求统计
    log_info("第一遍扫描：统计API请求数")
    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        chunk_size_actual = len(chunk)
        total_processed += chunk_size_actual

        # 统计API总请求数
        api_counts = chunk['request_full_uri'].value_counts()
        for api, count in api_counts.items():
            api_total_requests[api] = api_total_requests.get(api, 0) + count

        # 转换关键字段
        if 'total_request_duration' in chunk.columns:
            duration_col = pd.to_numeric(chunk['total_request_duration'], errors='coerce')
            slow_count = (duration_col > threshold).sum()
            total_slow_requests += slow_count

        del chunk
        if total_processed % 100000 == 0:
            gc.collect()
            log_info(f"已扫描 {total_processed:,} 条记录")

    log_info(f"第一遍扫描完成，总记录数：{total_processed:,}，预估慢请求数：{total_slow_requests:,}")
    
    # 第二遍：筛选和保存慢请求
    log_info("第二遍扫描：筛选慢请求")
    slow_requests_saved = 0
    header_written = False
    
    for chunk in pd.read_csv(csv_path, chunksize=chunk_size):
        # 数据类型转换（只转换必要列）
        numeric_columns = [
            'total_request_duration', 'upstream_connect_time', 'upstream_header_time', 'upstream_response_time',
            'backend_connect_phase', 'backend_process_phase', 'backend_transfer_phase', 'nginx_transfer_phase',
            'backend_total_phase', 'network_phase', 'processing_phase', 'transfer_phase',
            'backend_efficiency', 'network_overhead', 'transfer_ratio',
            'response_body_size_kb', 'total_bytes_sent_kb',
            'response_transfer_speed', 'total_transfer_speed', 'nginx_transfer_speed',
            'connection_cost_ratio', 'processing_efficiency_index'
        ]

        for col in numeric_columns:
            if col in chunk.columns:
                chunk.loc[:, col] = pd.to_numeric(chunk[col], errors='coerce')

        # 筛选慢请求
        slow_chunk = chunk[chunk['total_request_duration'] > threshold].copy()

        if not slow_chunk.empty and slow_requests_saved < MAX_SLOW_REQUESTS:
            # 限制保存数量
            remaining_slots = MAX_SLOW_REQUESTS - slow_requests_saved
            if len(slow_chunk) > remaining_slots:
                # 按响应时间倒序排序，保留最慢的请求
                slow_chunk = slow_chunk.nlargest(remaining_slots, 'total_request_duration')
                log_info(f"慢请求数量达到上限，仅保留最慢的 {remaining_slots} 条")
            
            # 选择需要的列
            selected_columns = [
                'log_source_file', 'application_name', 'service_name', 'raw_time', 'http_method',
                'request_full_uri', 'response_status_code'
            ] + [col for col in numeric_columns if col in slow_chunk.columns]

            # 确保所有列都存在
            for col in selected_columns:
                if col not in slow_chunk.columns:
                    slow_chunk[col] = None

            # 保存到临时文件
            slow_chunk[selected_columns].to_csv(temp_slow_file, mode='a', header=not header_written, index=False)
            header_written = True
            slow_requests_saved += len(slow_chunk)
            
            log_info(f"已保存 {len(slow_chunk):,} 个慢请求，累计 {slow_requests_saved:,}")
            
            if slow_requests_saved >= MAX_SLOW_REQUESTS:
                log_info(f"慢请求保存已达到上限 {MAX_SLOW_REQUESTS:,}，停止收集")
                break

        del chunk, slow_chunk
        gc.collect()

    log_info(f"数据处理完成，共处理 {total_processed:,} 条记录")

    if slow_requests_saved == 0:
        log_info(f"没有发现超过 {threshold} 秒的慢请求", level="WARNING")
        # 清理临时文件
        try:
            os.remove(temp_slow_file)
            os.rmdir(temp_dir)
        except:
            pass
        return pd.DataFrame()

    # 从临时文件读取慢请求数据
    log_info(f"从临时文件读取 {slow_requests_saved:,} 个慢请求数据")
    slow_df = pd.read_csv(temp_slow_file)
    
    # 清理临时文件
    try:
        os.remove(temp_slow_file)
        os.rmdir(temp_dir)
    except Exception as e:
        log_info(f"清理临时文件失败: {e}", level="WARNING")
    
    log_info(f"成功加载 {len(slow_df):,} 个慢请求，正在准备分析结果")

    # 应用列名映射
    slow_df = slow_df.rename(columns=column_mapping)

    # 按总响应时长排序
    slow_df = slow_df.sort_values(by='请求总时长(秒)', ascending=False)

    # 创建Excel输出
    log_info(f"创建Excel输出文件: {output_path}")
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 定义分组表头
    header_groups = {
        '基础信息': ['来源文件', '应用名称', '服务名称', '请求时间', '请求方法', '请求URI', '状态码'],
        '核心时间指标': [col for col in time_metrics if col in slow_df.columns],
        '阶段分析指标': [col for col in phase_metrics if col in slow_df.columns],
        '性能效率指标': [col for col in efficiency_metrics if col in slow_df.columns],
        '传输分析指标': [col for col in transfer_metrics if col in slow_df.columns]
    }

    # 添加慢请求列表工作表
    add_dataframe_to_excel_with_grouped_headers(wb, slow_df, '慢请求列表', header_groups=header_groups)

    # 生成API汇总统计
    if not slow_df.empty:
        api_stats = _generate_api_summary_stats(slow_df, api_total_requests, slow_requests_saved, all_metrics)

        # API汇总表头分组
        base_cols = ['请求URI', '请求总数', '慢请求次数', 'API内慢请求占比(%)', '全局慢请求占比(%)']
        api_header_groups = {
            '基础信息': base_cols,
            '核心时间统计': [f'{col}_{stat}' for col in time_metrics for stat in ['平均', '中位数', 'P90', 'P95', 'P99'] if
                       f'{col}_{stat}' in api_stats.columns],
            '阶段分析统计': [f'{col}_{stat}' for col in phase_metrics for stat in ['平均', '中位数', 'P90', 'P95', 'P99'] if
                       f'{col}_{stat}' in api_stats.columns],
            '效率指标统计': [f'{col}_{stat}' for col in efficiency_metrics for stat in ['平均', '中位数', 'P90', 'P95', 'P99'] if
                       f'{col}_{stat}' in api_stats.columns],
            '传输指标统计': [f'{col}_{stat}' for col in transfer_metrics for stat in ['平均', '中位数', 'P90', 'P95', 'P99'] if
                       f'{col}_{stat}' in api_stats.columns]
        }

        add_dataframe_to_excel_with_grouped_headers(wb, api_stats, '慢请求API汇总', header_groups=api_header_groups)

        # 创建性能分析工作表
        _create_performance_analysis_sheet(wb, slow_df)

        # 创建传输效率分析工作表（新增）
        _create_transfer_efficiency_sheet(wb, slow_df)

    wb.save(output_path)
    log_info(f"慢请求分析完成，结果已保存至: {output_path}", show_memory=True)
    return slow_df.head(10)


def _generate_api_summary_stats(slow_df, api_total_requests, total_slow_requests, metrics):
    """生成API汇总统计数据"""
    stat_funcs = {
        '平均': np.mean,
        '中位数': np.median,
        '最小': np.min,
        '最大': np.max,
        'P90': lambda x: np.percentile(x, 90),
        'P95': lambda x: np.percentile(x, 95),
        'P99': lambda x: np.percentile(x, 99),
    }

    records = []
    for uri, group in slow_df.groupby('请求URI'):
        record = {
            '请求URI': uri,
            '慢请求次数': len(group),
            '请求总数': api_total_requests.get(uri, 0)
        }

        # 计算各项指标统计
        for metric in metrics:
            if metric in group.columns:
                values = group[metric].dropna().values
                if len(values) > 0:
                    for stat_name, func in stat_funcs.items():
                        record[f'{metric}_{stat_name}'] = func(values)

        records.append(record)

    api_stats = pd.DataFrame.from_records(records)

    # 计算占比
    api_stats['API内慢请求占比(%)'] = api_stats.apply(
        lambda row: (row['慢请求次数'] / row['请求总数'] * 100) if row['请求总数'] > 0 else 0,
        axis=1
    ).round(2)

    api_stats['全局慢请求占比(%)'] = (api_stats['慢请求次数'] / total_slow_requests * 100).round(2)

    return api_stats.sort_values(by='慢请求次数', ascending=False)


def _create_performance_analysis_sheet(wb, slow_df):
    """创建性能分析工作表"""
    log_info("创建性能分析工作表")
    ws_perf = wb.create_sheet(title='性能分析')

    row = 1
    ws_perf.cell(row=row, column=1, value="慢请求性能深度分析").font = openpyxl.styles.Font(bold=True, size=14)
    row += 2

    # 阶段耗时分析
    phase_columns = {
        '后端连接阶段(秒)': '后端连接',
        '后端处理阶段(秒)': '后端处理',
        '后端传输阶段(秒)': '后端传输',
        'Nginx传输阶段(秒)': 'Nginx传输'
    }

    headers = ['阶段', '平均耗时(秒)', '占总耗时比例(%)', '中位数(秒)', 'P90(秒)', 'P95(秒)', 'P99(秒)']
    for col_idx, header in enumerate(headers, start=1):
        ws_perf.cell(row=row, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)

    total_avg_time = slow_df['请求总时长(秒)'].mean()
    phase_data = []

    row += 1
    for col, name in phase_columns.items():
        if col in slow_df.columns:
            values = slow_df[col].dropna()
            if values.empty:
                continue

            avg = values.mean()
            pct = (avg / total_avg_time * 100) if total_avg_time > 0 else 0
            phase_data.append((name, avg))

            ws_perf.cell(row=row, column=1, value=name)
            ws_perf.cell(row=row, column=2, value=round(avg, 4))
            ws_perf.cell(row=row, column=3, value=round(pct, 2))
            ws_perf.cell(row=row, column=4, value=round(values.median(), 4))
            ws_perf.cell(row=row, column=5, value=round(np.percentile(values, 90), 4))
            ws_perf.cell(row=row, column=6, value=round(np.percentile(values, 95), 4))
            ws_perf.cell(row=row, column=7, value=round(np.percentile(values, 99), 4))
            row += 1

    # 性能效率指标分析
    row += 2
    ws_perf.cell(row=row, column=1, value="性能效率指标分析").font = openpyxl.styles.Font(bold=True, size=12)
    row += 1

    efficiency_metrics = {
        '后端处理效率(%)': '后端处理效率',
        '网络开销占比(%)': '网络开销占比',
        '传输时间占比(%)': '传输时间占比',
        '连接成本占比(%)': '连接成本占比'
    }

    for col_idx, header in enumerate(['指标', '平均值', '中位数', '标准差', 'P90', 'P95'], start=1):
        ws_perf.cell(row=row, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)

    row += 1
    for col, name in efficiency_metrics.items():
        if col in slow_df.columns:
            values = slow_df[col].dropna()
            if values.empty:
                continue

            ws_perf.cell(row=row, column=1, value=name)
            ws_perf.cell(row=row, column=2, value=round(values.mean(), 2))
            ws_perf.cell(row=row, column=3, value=round(values.median(), 2))
            ws_perf.cell(row=row, column=4, value=round(values.std(), 2))
            ws_perf.cell(row=row, column=5, value=round(np.percentile(values, 90), 2))
            ws_perf.cell(row=row, column=6, value=round(np.percentile(values, 95), 2))
            row += 1

    # 添加饼图
    if phase_data:
        row += 2
        chart_start_row = row
        ws_perf.cell(row=row, column=1, value="阶段耗时占比").font = openpyxl.styles.Font(bold=True)
        row += 1

        for i, (name, value) in enumerate(phase_data):
            ws_perf.cell(row=row + i, column=1, value=name)
            ws_perf.cell(row=row + i, column=2, value=value)

        create_pie_chart(ws_perf, "各阶段耗时占比",
                         data_start_row=row,
                         data_end_row=row + len(phase_data) - 1,
                         labels_col=1,
                         values_col=2,
                         position="D" + str(chart_start_row))

    format_excel_sheet(ws_perf)


def _create_transfer_efficiency_sheet(wb, slow_df):
    """创建传输效率分析工作表（新增分析维度）"""
    log_info("创建传输效率分析工作表")
    ws_transfer = wb.create_sheet(title='传输效率分析')

    row = 1
    ws_transfer.cell(row=row, column=1, value="传输效率深度分析").font = openpyxl.styles.Font(bold=True, size=14)
    row += 2

    # 传输速度分析
    ws_transfer.cell(row=row, column=1, value="传输速度统计 (KB/s)").font = openpyxl.styles.Font(bold=True, size=12)
    row += 1

    speed_metrics = {
        '响应体传输速度(KB/s)': '响应体传输速度',
        '总传输速度(KB/s)': '总传输速度',
        'Nginx传输速度(KB/s)': 'Nginx传输速度'
    }

    headers = ['传输类型', '平均速度', '中位数速度', '最小速度', '最大速度', 'P90', 'P95']
    for col_idx, header in enumerate(headers, start=1):
        ws_transfer.cell(row=row, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)

    row += 1
    speed_data = []
    for col, name in speed_metrics.items():
        if col in slow_df.columns:
            values = slow_df[col].dropna()
            if values.empty:
                continue

            avg_speed = values.mean()
            speed_data.append((name, avg_speed))

            ws_transfer.cell(row=row, column=1, value=name)
            ws_transfer.cell(row=row, column=2, value=round(avg_speed, 2))
            ws_transfer.cell(row=row, column=3, value=round(values.median(), 2))
            ws_transfer.cell(row=row, column=4, value=round(values.min(), 2))
            ws_transfer.cell(row=row, column=5, value=round(values.max(), 2))
            ws_transfer.cell(row=row, column=6, value=round(np.percentile(values, 90), 2))
            ws_transfer.cell(row=row, column=7, value=round(np.percentile(values, 95), 2))
            row += 1

    # 数据量分析
    row += 2
    ws_transfer.cell(row=row, column=1, value="数据量统计 (KB)").font = openpyxl.styles.Font(bold=True, size=12)
    row += 1

    size_metrics = {
        '响应体大小(KB)': '响应体大小',
        '总传输大小(KB)': '总传输大小'
    }

    headers = ['数据类型', '平均大小', '中位数大小', '最小大小', '最大大小', 'P90', 'P95']
    for col_idx, header in enumerate(headers, start=1):
        ws_transfer.cell(row=row, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)

    row += 1
    for col, name in size_metrics.items():
        if col in slow_df.columns:
            values = slow_df[col].dropna()
            if values.empty:
                continue

            ws_transfer.cell(row=row, column=1, value=name)
            ws_transfer.cell(row=row, column=2, value=round(values.mean(), 2))
            ws_transfer.cell(row=row, column=3, value=round(values.median(), 2))
            ws_transfer.cell(row=row, column=4, value=round(values.min(), 2))
            ws_transfer.cell(row=row, column=5, value=round(values.max(), 2))
            ws_transfer.cell(row=row, column=6, value=round(np.percentile(values, 90), 2))
            ws_transfer.cell(row=row, column=7, value=round(np.percentile(values, 95), 2))
            row += 1

    # 传输效率相关性分析
    row += 2
    ws_transfer.cell(row=row, column=1, value="传输效率洞察").font = openpyxl.styles.Font(bold=True, size=12)
    row += 1

    # 计算传输效率阈值建议
    if '总传输速度(KB/s)' in slow_df.columns:
        speed_values = slow_df['总传输速度(KB/s)'].dropna()
        if not speed_values.empty:
            p25_speed = np.percentile(speed_values, 25)
            p75_speed = np.percentile(speed_values, 75)

            ws_transfer.cell(row=row, column=1, value="传输速度健康阈值建议:")
            row += 1
            ws_transfer.cell(row=row, column=1, value=f"• 优秀传输速度: > {p75_speed:.2f} KB/s")
            row += 1
            ws_transfer.cell(row=row, column=1, value=f"• 需要关注传输速度: < {p25_speed:.2f} KB/s")
            row += 1

    # 添加传输速度饼图
    if speed_data:
        row += 2
        chart_start_row = row
        ws_transfer.cell(row=row, column=1, value="传输速度对比").font = openpyxl.styles.Font(bold=True)
        row += 1

        for i, (name, value) in enumerate(speed_data):
            ws_transfer.cell(row=row + i, column=1, value=name)
            ws_transfer.cell(row=row + i, column=2, value=value)

        create_pie_chart(ws_transfer, "传输速度对比",
                         data_start_row=row,
                         data_end_row=row + len(speed_data) - 1,
                         labels_col=1,
                         values_col=2,
                         position="D" + str(chart_start_row))

    format_excel_sheet(ws_transfer)