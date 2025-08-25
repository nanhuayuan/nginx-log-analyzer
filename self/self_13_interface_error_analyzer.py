import gc
import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from self_00_01_constants import (
    DEFAULT_CHUNK_SIZE, EXCEL_MAX_ROWS,
    HEADER_FILL, ERROR_FILL, WARNING_FILL, SUCCESS_FILL,
    STATUS_DESCRIPTIONS
)
from self_00_02_utils import log_info
from self_00_04_excel_processor import add_dataframe_to_excel_with_grouped_headers


def analyze_interface_errors(csv_path, output_path, slow_request_threshold=3.0, 
                            error_threshold=0.05, min_requests=10):
    """
    接口错误分析主函数
    
    参数:
    - csv_path: CSV数据文件路径
    - output_path: 输出Excel文件路径  
    - slow_request_threshold: 慢请求阈值(秒)
    - error_threshold: 错误率阈值(默认5%)
    - min_requests: 最小请求数阈值(过滤低流量接口)
    """
    
    log_info("开始接口错误分析...", True)
    start_time = datetime.now()
    
    # 数据收集器
    collectors = _initialize_collectors()
    
    # 统计变量
    stats = {
        'total_requests': 0,
        'total_error_requests': 0,
        'total_interfaces': 0,
        'error_interfaces': 0,
        'chunks_processed': 0
    }
    
    # 分块处理数据
    chunk_size = DEFAULT_CHUNK_SIZE
    reader = pd.read_csv(csv_path, chunksize=chunk_size)
    
    for chunk in reader:
        stats['chunks_processed'] += 1
        chunk_rows = len(chunk)
        stats['total_requests'] += chunk_rows
        
        # 处理当前数据块
        _process_chunk(chunk, collectors, stats, slow_request_threshold, error_threshold)
        
        # 内存管理和进度报告
        if stats['chunks_processed'] % 5 == 0:
            elapsed = (datetime.now() - start_time).total_seconds()
            log_info(f"已处理 {stats['chunks_processed']} 个数据块, "
                    f"{stats['total_requests']} 条记录, 耗时: {elapsed:.2f}秒", show_memory=True)
            del chunk
            gc.collect()
    
    # 后处理：过滤低流量接口并计算派生指标
    _post_process_data(collectors, stats, min_requests, error_threshold)
    
    # 生成分析报告
    log_info("生成接口错误分析报告...", True)
    dataframes = _generate_analysis_dataframes(collectors, stats, slow_request_threshold)
    
    # 创建Excel报告
    _create_excel_report(output_path, dataframes)
    
    # 生成图表
    _create_charts(output_path, dataframes)
    
    end_time = datetime.now()
    log_info(f"接口错误分析完成，耗时: {(end_time - start_time).total_seconds():.2f} 秒", True)
    log_info(f"分析报告已保存至: {output_path}", True)
    
    # 返回关键统计信息
    return _generate_summary_stats(collectors, stats)


def _initialize_collectors():
    """初始化数据收集器"""
    return {
        # 接口错误统计
        'interface_stats': defaultdict(lambda: {
            'total_requests': 0,
            'error_requests': 0,
            'status_codes': Counter(),
            'error_codes': Counter(),
            'response_times': [],
            'error_response_times': [],
            'slow_requests': 0,
            'clients': set(),
            'applications': set(),
            'services': set(),
            'upstream_servers': set(),
            'first_error_time': None,
            'last_error_time': None,
            'error_time_distribution': defaultdict(int)  # 按小时统计错误分布
        }),
        
        # 错误状态码详情
        'error_details': defaultdict(list),
        
        # 时间维度错误统计
        'error_by_time': defaultdict(lambda: defaultdict(int)),  # {time: {interface: count}}
        
        # 上游服务错误统计
        'upstream_errors': defaultdict(lambda: {
            'total_errors': 0,
            'interfaces': set(),
            'error_codes': Counter(),
            'avg_connect_time': [],
            'avg_response_time': []
        }),
        
        # 错误影响面统计
        'impact_analysis': {
            'error_clients': set(),
            'error_applications': set(), 
            'error_services': set(),
            'total_clients': set(),
            'total_applications': set(),
            'total_services': set()
        },
        
        # 错误时间集中度分析
        'error_time_clusters': defaultdict(list),  # {interface: [error_times]}
        
        # 关键错误类型统计
        'critical_errors': {
            '499': defaultdict(int),  # 客户端取消
            '502': defaultdict(int),  # 网关错误
            '503': defaultdict(int),  # 服务不可用
            '504': defaultdict(int),  # 网关超时
            '500': defaultdict(int)   # 服务器内部错误
        }
    }


def _process_chunk(chunk, collectors, stats, slow_request_threshold, error_threshold):
    """处理单个数据块"""
    
    # 字段映射
    status_field = 'response_status_code'
    path_field = 'request_path'
    time_field = 'raw_time'
    duration_field = 'total_request_duration'
    client_ip_field = 'client_ip_address'
    app_field = 'application_name'
    service_field = 'service_name'
    upstream_field = 'upstream_server_address'
    upstream_connect_field = 'upstream_connect_time'
    upstream_response_field = 'upstream_response_time'
    
    for _, row in chunk.iterrows():
        interface = row.get(path_field, 'unknown')
        status_code = str(row.get(status_field, ''))
        request_time = float(row.get(duration_field, 0) or 0)
        timestamp_str = row.get(time_field, '')
        client_ip = row.get(client_ip_field, '')
        application = row.get(app_field, '')
        service = row.get(service_field, '')
        upstream = row.get(upstream_field, '')
        
        # 更新接口统计
        interface_stat = collectors['interface_stats'][interface]
        interface_stat['total_requests'] += 1
        interface_stat['status_codes'][status_code] += 1
        interface_stat['response_times'].append(request_time)
        interface_stat['clients'].add(client_ip)
        interface_stat['applications'].add(application)
        interface_stat['services'].add(service)
        if upstream:
            interface_stat['upstream_servers'].add(upstream)
        
        # 慢请求统计
        if request_time > slow_request_threshold:
            interface_stat['slow_requests'] += 1
        
        # 更新全局影响面统计
        impact = collectors['impact_analysis']
        impact['total_clients'].add(client_ip)
        impact['total_applications'].add(application)  
        impact['total_services'].add(service)
        
        # 错误请求处理
        is_error = status_code.startswith(('4', '5'))
        if is_error:
            stats['total_error_requests'] += 1
            interface_stat['error_requests'] += 1
            interface_stat['error_codes'][status_code] += 1
            interface_stat['error_response_times'].append(request_time)
            
            # 记录错误时间
            if timestamp_str:
                try:
                    error_time = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                    if not interface_stat['first_error_time']:
                        interface_stat['first_error_time'] = error_time
                    interface_stat['last_error_time'] = error_time
                    
                    # 错误时间分布（按小时）
                    hour_key = error_time.strftime('%H')
                    interface_stat['error_time_distribution'][hour_key] += 1
                    
                    # 错误时间集中度
                    collectors['error_time_clusters'][interface].append(error_time)
                    
                    # 时间维度错误统计
                    time_key = error_time.strftime('%Y-%m-%d %H')
                    collectors['error_by_time'][time_key][interface] += 1
                except:
                    pass
            
            # 更新影响面统计
            impact['error_clients'].add(client_ip)
            impact['error_applications'].add(application)
            impact['error_services'].add(service)
            
            # 上游服务错误统计
            if upstream:
                upstream_stat = collectors['upstream_errors'][upstream]
                upstream_stat['total_errors'] += 1
                upstream_stat['interfaces'].add(interface)
                upstream_stat['error_codes'][status_code] += 1
                
                # 上游连接时间
                upstream_connect = float(row.get(upstream_connect_field, 0) or 0)
                upstream_response = float(row.get(upstream_response_field, 0) or 0)
                if upstream_connect > 0:
                    upstream_stat['avg_connect_time'].append(upstream_connect)
                if upstream_response > 0:
                    upstream_stat['avg_response_time'].append(upstream_response)
            
            # 关键错误类型统计
            if status_code in collectors['critical_errors']:
                collectors['critical_errors'][status_code][interface] += 1
            
            # 错误详情收集(限制数量避免内存问题)
            if len(collectors['error_details'][interface]) < 100:
                error_detail = {
                    'time': timestamp_str,
                    'status_code': status_code,
                    'response_time': request_time,
                    'client_ip': client_ip,
                    'application': application,
                    'service': service,
                    'upstream': upstream,
                    'request_path': interface
                }
                collectors['error_details'][interface].append(error_detail)


def _post_process_data(collectors, stats, min_requests, error_threshold):
    """后处理：过滤和计算派生指标"""
    
    # 过滤低流量接口
    interfaces_to_remove = []
    for interface, stat in collectors['interface_stats'].items():
        if stat['total_requests'] < min_requests:
            interfaces_to_remove.append(interface)
    
    for interface in interfaces_to_remove:
        del collectors['interface_stats'][interface]
    
    # 统计错误接口数量
    stats['total_interfaces'] = len(collectors['interface_stats'])
    for interface, stat in collectors['interface_stats'].items():
        error_rate = (stat['error_requests'] / stat['total_requests'] * 100) if stat['total_requests'] > 0 else 0
        if error_rate >= error_threshold * 100:
            stats['error_interfaces'] += 1


def _generate_analysis_dataframes(collectors, stats, slow_request_threshold):
    """生成各类分析数据表"""
    dataframes = {}
    
    # 1. 摘要统计
    dataframes['summary_df'] = _create_summary_dataframe(stats, collectors)
    
    # 2. 接口错误统计
    dataframes['interface_errors_df'] = _create_interface_errors_dataframe(collectors['interface_stats'])
    
    # 3. Top错误接口排行榜  
    dataframes['top_error_interfaces_df'] = _create_top_error_interfaces_dataframe(collectors['interface_stats'])
    
    # 4. 关键错误类型分析
    dataframes['critical_errors_df'] = _create_critical_errors_dataframe(collectors['critical_errors'])
    
    # 5. 上游服务错误分析
    dataframes['upstream_errors_df'] = _create_upstream_errors_dataframe(collectors['upstream_errors'])
    
    # 6. 错误时间分析
    dataframes['error_time_analysis_df'] = _create_error_time_analysis_dataframe(collectors['error_by_time'])
    
    # 7. 错误影响面分析
    dataframes['impact_analysis_df'] = _create_impact_analysis_dataframe(collectors['impact_analysis'])
    
    # 8. 错误详情表
    dataframes['error_details_df'] = _create_error_details_dataframe(collectors['error_details'])
    
    # 9. 错误时间集中度分析
    dataframes['error_clusters_df'] = _create_error_clusters_dataframe(collectors['error_time_clusters'])
    
    return dataframes


def _create_summary_dataframe(stats, collectors):
    """创建摘要统计数据表"""
    total_requests = stats['total_requests']
    total_errors = stats['total_error_requests']
    error_rate = (total_errors / total_requests * 100) if total_requests > 0 else 0
    
    impact = collectors['impact_analysis']
    
    summary_data = [
        {'指标': '总请求数', '值': total_requests},
        {'指标': '总错误请求数', '值': total_errors},
        {'指标': '全局错误率(%)', '值': round(error_rate, 2)},
        {'指标': '总接口数', '值': stats['total_interfaces']},
        {'指标': '错误接口数', '值': stats['error_interfaces']},
        {'指标': '错误接口占比(%)', '值': round((stats['error_interfaces'] / stats['total_interfaces'] * 100), 2) if stats['total_interfaces'] > 0 else 0},
        {'指标': '受影响客户端数', '值': len(impact['error_clients'])},
        {'指标': '受影响应用数', '值': len(impact['error_applications'])},
        {'指标': '受影响服务数', '值': len(impact['error_services'])},
        {'指标': '客户端错误影响率(%)', '值': round((len(impact['error_clients']) / len(impact['total_clients']) * 100), 2) if len(impact['total_clients']) > 0 else 0}
    ]
    
    return pd.DataFrame(summary_data)


def _create_interface_errors_dataframe(interface_stats):
    """创建接口错误统计表"""
    interface_data = []
    
    for interface, stat in interface_stats.items():
        total_requests = stat['total_requests']
        error_requests = stat['error_requests']
        error_rate = (error_requests / total_requests * 100) if total_requests > 0 else 0
        
        # 响应时间统计
        avg_response_time = np.mean(stat['response_times']) if stat['response_times'] else 0
        avg_error_response_time = np.mean(stat['error_response_times']) if stat['error_response_times'] else 0
        
        # 主要错误类型
        top_error_code = stat['error_codes'].most_common(1)
        main_error_code = top_error_code[0][0] if top_error_code else ''
        main_error_count = top_error_code[0][1] if top_error_code else 0
        
        # 错误持续时间
        error_duration = 0
        if stat['first_error_time'] and stat['last_error_time']:
            error_duration = (stat['last_error_time'] - stat['first_error_time']).total_seconds() / 3600  # 小时
        
        interface_data.append({
            '接口路径': interface,
            '总请求数': total_requests,
            '错误请求数': error_requests,
            '错误率(%)': round(error_rate, 2),
            '平均响应时间(秒)': round(avg_response_time, 3),
            '平均错误响应时间(秒)': round(avg_error_response_time, 3),
            '慢请求数': stat['slow_requests'],
            '主要错误码': main_error_code,
            '主要错误码次数': main_error_count,
            '错误持续时间(小时)': round(error_duration, 2),
            '受影响客户端数': len(stat['clients']),
            '受影响应用数': len(stat['applications']),
            '受影响服务数': len(stat['services']),
            '上游服务数': len(stat['upstream_servers']),
            '首次错误时间': stat['first_error_time'].strftime('%Y-%m-%d %H:%M:%S') if stat['first_error_time'] else '',
            '最后错误时间': stat['last_error_time'].strftime('%Y-%m-%d %H:%M:%S') if stat['last_error_time'] else ''
        })
    
    interface_df = pd.DataFrame(interface_data)
    return interface_df.sort_values(by='错误请求数', ascending=False) if not interface_df.empty else interface_df


def _create_top_error_interfaces_dataframe(interface_stats):
    """创建Top错误接口排行榜"""
    # 按不同维度排序
    interface_list = []
    
    for interface, stat in interface_stats.items():
        total_requests = stat['total_requests']
        error_requests = stat['error_requests']
        error_rate = (error_requests / total_requests * 100) if total_requests > 0 else 0
        
        interface_list.append({
            'interface': interface,
            'error_count': error_requests,
            'error_rate': error_rate,
            'total_requests': total_requests,
            'impact_clients': len(stat['clients'])
        })
    
    # Top 10错误次数
    top_by_count = sorted(interface_list, key=lambda x: x['error_count'], reverse=True)[:10]
    # Top 10错误率(过滤低流量)
    top_by_rate = sorted([x for x in interface_list if x['total_requests'] >= 100], 
                        key=lambda x: x['error_rate'], reverse=True)[:10]
    # Top 10影响面
    top_by_impact = sorted(interface_list, key=lambda x: x['impact_clients'], reverse=True)[:10]
    
    top_data = []
    
    # 合并数据
    max_len = max(len(top_by_count), len(top_by_rate), len(top_by_impact))
    for i in range(max_len):
        row = {'排名': i + 1}
        
        if i < len(top_by_count):
            row.update({
                '错误次数-接口': top_by_count[i]['interface'],
                '错误次数-数量': top_by_count[i]['error_count']
            })
        
        if i < len(top_by_rate):
            row.update({
                '错误率-接口': top_by_rate[i]['interface'],
                '错误率-百分比': round(top_by_rate[i]['error_rate'], 2)
            })
            
        if i < len(top_by_impact):
            row.update({
                '影响面-接口': top_by_impact[i]['interface'],
                '影响面-客户端数': top_by_impact[i]['impact_clients']
            })
        
        top_data.append(row)
    
    return pd.DataFrame(top_data)


def _create_critical_errors_dataframe(critical_errors):
    """创建关键错误类型分析表"""
    critical_data = []
    
    for error_code, interfaces in critical_errors.items():
        if not interfaces:
            continue
            
        error_name = STATUS_DESCRIPTIONS.get(error_code, f'错误码{error_code}')
        total_count = sum(interfaces.values())
        affected_interfaces = len(interfaces)
        
        # Top 5受影响接口
        top_interfaces = sorted(interfaces.items(), key=lambda x: x[1], reverse=True)[:5]
        top_interface_list = [f"{iface}({count}次)" for iface, count in top_interfaces]
        
        critical_data.append({
            '错误码': error_code,
            '错误描述': error_name,
            '总错误次数': total_count,
            '受影响接口数': affected_interfaces,
            'Top5受影响接口': '; '.join(top_interface_list)
        })
    
    critical_df = pd.DataFrame(critical_data)
    return critical_df.sort_values(by='总错误次数', ascending=False) if not critical_df.empty else critical_df


def _create_upstream_errors_dataframe(upstream_errors):
    """创建上游服务错误分析表"""
    upstream_data = []
    
    for upstream, stat in upstream_errors.items():
        if stat['total_errors'] == 0:
            continue
            
        avg_connect_time = np.mean(stat['avg_connect_time']) if stat['avg_connect_time'] else 0
        avg_response_time = np.mean(stat['avg_response_time']) if stat['avg_response_time'] else 0
        
        # 主要错误类型
        top_error = stat['error_codes'].most_common(1)
        main_error = top_error[0][0] if top_error else ''
        main_error_count = top_error[0][1] if top_error else 0
        
        upstream_data.append({
            '上游服务地址': upstream,
            '总错误次数': stat['total_errors'],
            '受影响接口数': len(stat['interfaces']),
            '主要错误码': main_error,
            '主要错误码次数': main_error_count,
            '平均连接时间(秒)': round(avg_connect_time, 3),
            '平均响应时间(秒)': round(avg_response_time, 3),
            '受影响接口列表': '; '.join(list(stat['interfaces'])[:5])  # 显示前5个
        })
    
    upstream_df = pd.DataFrame(upstream_data)
    return upstream_df.sort_values(by='总错误次数', ascending=False) if not upstream_df.empty else upstream_df


def _create_error_time_analysis_dataframe(error_by_time):
    """创建错误时间分析表"""
    time_data = []
    
    for time_key, interfaces in error_by_time.items():
        total_errors = sum(interfaces.values())
        error_interfaces = len(interfaces)
        
        # Top 3错误接口
        top_interfaces = sorted(interfaces.items(), key=lambda x: x[1], reverse=True)[:3]
        top_interface_list = [f"{iface}({count})" for iface, count in top_interfaces]
        
        time_data.append({
            '时间': time_key,
            '总错误数': total_errors,
            '错误接口数': error_interfaces,
            'Top3错误接口': '; '.join(top_interface_list)
        })
    
    time_df = pd.DataFrame(time_data)
    return time_df.sort_values(by='总错误数', ascending=False) if not time_df.empty else time_df


def _create_impact_analysis_dataframe(impact_analysis):
    """创建错误影响面分析表"""
    impact_data = [
        {
            '维度': '客户端',
            '总数量': len(impact_analysis['total_clients']),
            '受影响数量': len(impact_analysis['error_clients']),
            '影响率(%)': round((len(impact_analysis['error_clients']) / len(impact_analysis['total_clients']) * 100), 2) if len(impact_analysis['total_clients']) > 0 else 0
        },
        {
            '维度': '应用',
            '总数量': len(impact_analysis['total_applications']),
            '受影响数量': len(impact_analysis['error_applications']),
            '影响率(%)': round((len(impact_analysis['error_applications']) / len(impact_analysis['total_applications']) * 100), 2) if len(impact_analysis['total_applications']) > 0 else 0
        },
        {
            '维度': '服务',
            '总数量': len(impact_analysis['total_services']),
            '受影响数量': len(impact_analysis['error_services']),
            '影响率(%)': round((len(impact_analysis['error_services']) / len(impact_analysis['total_services']) * 100), 2) if len(impact_analysis['total_services']) > 0 else 0
        }
    ]
    
    return pd.DataFrame(impact_data)


def _create_error_details_dataframe(error_details):
    """创建错误详情表"""
    details_data = []
    
    for interface, details in error_details.items():
        for detail in details[:50]:  # 限制每个接口最多显示50条
            details_data.append({
                '接口路径': interface,
                '错误时间': detail['time'],
                '错误状态码': detail['status_code'],
                '响应时间(秒)': detail['response_time'],
                '客户端IP': detail['client_ip'],
                '应用名称': detail['application'],
                '服务名称': detail['service'],
                '上游服务': detail['upstream']
            })
    
    details_df = pd.DataFrame(details_data)
    return details_df.sort_values(by=['接口路径', '错误时间'], ascending=[True, False]) if not details_df.empty else details_df


def _create_error_clusters_dataframe(error_time_clusters):
    """创建错误时间集中度分析表"""
    cluster_data = []
    
    for interface, error_times in error_time_clusters.items():
        if len(error_times) < 2:
            continue
            
        error_times.sort()
        
        # 计算错误时间间隔
        intervals = []
        for i in range(1, len(error_times)):
            interval = (error_times[i] - error_times[i-1]).total_seconds()
            intervals.append(interval)
        
        # 错误集中度分析
        avg_interval = np.mean(intervals) if intervals else 0
        min_interval = min(intervals) if intervals else 0
        max_interval = max(intervals) if intervals else 0
        
        # 错误爆发识别(间隔小于5分钟认为是爆发)
        burst_count = sum(1 for interval in intervals if interval < 300)
        
        cluster_data.append({
            '接口路径': interface,
            '总错误次数': len(error_times),
            '错误时间跨度(小时)': round((error_times[-1] - error_times[0]).total_seconds() / 3600, 2),
            '平均间隔(秒)': round(avg_interval, 1),
            '最短间隔(秒)': round(min_interval, 1),
            '最长间隔(秒)': round(max_interval, 1),
            '错误爆发次数': burst_count,
            '首次错误': error_times[0].strftime('%Y-%m-%d %H:%M:%S'),
            '最后错误': error_times[-1].strftime('%Y-%m-%d %H:%M:%S')
        })
    
    cluster_df = pd.DataFrame(cluster_data)
    return cluster_df.sort_values(by='错误爆发次数', ascending=False) if not cluster_df.empty else cluster_df


def _create_excel_report(output_path, dataframes):
    """创建Excel报告"""
    log_info("创建Excel工作簿...", True)
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 定义工作表顺序和分组标题
    sheets_config = [
        ('摘要统计', dataframes['summary_df'], None),
        
        ('接口错误统计', dataframes['interface_errors_df'], {
            '基本信息': ['接口路径', '总请求数', '错误请求数', '错误率(%)', '慢请求数'],
            '响应时间分析': ['平均响应时间(秒)', '平均错误响应时间(秒)'],
            '错误特征': ['主要错误码', '主要错误码次数', '错误持续时间(小时)'],
            '影响面分析': ['受影响客户端数', '受影响应用数', '受影响服务数', '上游服务数'],
            '时间分析': ['首次错误时间', '最后错误时间']
        }),
        
        ('Top错误接口排行', dataframes['top_error_interfaces_df'], {
            '排行信息': ['排名'],
            '错误次数排行': ['错误次数-接口', '错误次数-数量'],
            '错误率排行': ['错误率-接口', '错误率-百分比'],
            '影响面排行': ['影响面-接口', '影响面-客户端数']
        }),
        
        ('关键错误类型', dataframes['critical_errors_df'], None),
        ('上游服务错误', dataframes['upstream_errors_df'], None),
        ('错误时间分析', dataframes['error_time_analysis_df'], None),
        ('错误影响面', dataframes['impact_analysis_df'], None),
        ('错误时间集中度', dataframes['error_clusters_df'], None),
        ('错误详情', dataframes['error_details_df'], None)
    ]
    
    for sheet_name, df, header_groups in sheets_config:
        if not df.empty:
            add_dataframe_to_excel_with_grouped_headers(
                wb, df, sheet_name, header_groups=header_groups
            )
    
    wb.save(output_path)


def _create_charts(output_path, dataframes):
    """创建图表(简化版)"""
    # 为了保持代码简洁，这里省略了详细的图表创建代码
    # 实际使用中可以添加错误趋势图、接口错误分布图等
    pass


def _generate_summary_stats(collectors, stats):
    """生成摘要统计信息"""
    interface_stats = collectors['interface_stats']
    
    # 找出Top 5错误接口
    top_error_interfaces = sorted(
        [(iface, stat['error_requests']) for iface, stat in interface_stats.items()],
        key=lambda x: x[1], reverse=True
    )[:5]
    
    return {
        'total_requests': stats['total_requests'],
        'total_errors': stats['total_error_requests'],
        'error_rate': (stats['total_error_requests'] / stats['total_requests'] * 100) if stats['total_requests'] > 0 else 0,
        'error_interfaces': stats['error_interfaces'],
        'top_error_interfaces': top_error_interfaces,
        'affected_clients': len(collectors['impact_analysis']['error_clients']),
        'affected_applications': len(collectors['impact_analysis']['error_applications']),
        'affected_services': len(collectors['impact_analysis']['error_services'])
    }


# 主函数入口
if __name__ == "__main__":
    # 测试用例
    csv_path = "../data/demo/processed_logs.csv"
    output_path = "./13_接口错误分析.xlsx"
    result = analyze_interface_errors(csv_path, output_path)
    print("接口错误分析完成")