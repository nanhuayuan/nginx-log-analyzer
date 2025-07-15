#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Nginx监控报告生成系统 - 优化版
基于夜莺(Nightingale)监控数据生成全面的系统运行报告

主要特性:
- 现代化GUI界面设计
- 灵活的时间范围选择
- 单集群精准监控
- 数据与展示分离
- 结构化Excel报表
- 可配置的指标查询
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import datetime
import json
import logging
import os
import sys
from dataclasses import dataclass
from enum import Enum
from typing import Dict, List, Tuple, Optional, Any
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
import threading
import queue
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('monitoring_report.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


# 数据模型定义
@dataclass
class MetricResult:
    """监控指标结果"""
    value: float
    unit: str = ""
    timestamp: str = ""
    status: str = "normal"


@dataclass
class ConnectionConfig:
    """连接配置"""
    url: str
    username: str
    password: str
    datasource_ids: List[str]


class ReportType(Enum):
    """报告类型"""
    CUSTOM = "自定义时间段"


# 数字格式化工具
class NumberFormatter:
    """数字格式化工具类"""
    
    @staticmethod
    def format_for_display(value: float, for_excel: bool = False) -> any:
        """
        格式化数字显示
        Args:
            value: 数值
            for_excel: 是否用于Excel（True返回数值，False返回格式化字符串）
        Returns:
            格式化后的值
        """
        if not isinstance(value, (int, float)):
            return value
        
        if for_excel:
            # Excel中保持数值类型，便于计算
            return value
        else:
            # 文本报告中格式化显示
            if value == 0:
                return "0"
            elif value == int(value):
                return str(int(value))
            elif abs(value) < 0.01:
                return f"{value:.4f}".rstrip('0').rstrip('.')
            elif abs(value) < 1:
                return f"{value:.3f}".rstrip('0').rstrip('.')
            elif abs(value) < 100:
                return f"{value:.2f}".rstrip('0').rstrip('.')
            else:
                return f"{value:.1f}".rstrip('0').rstrip('.')
    
    @staticmethod
    def format_calculation_result(value: float, for_excel: bool = False) -> any:
        """
        格式化计算结果（保留2位小数）
        Args:
            value: 计算结果
            for_excel: 是否用于Excel
        Returns:
            格式化后的值
        """
        if not isinstance(value, (int, float)):
            return value
        
        if for_excel:
            # Excel中保持数值类型，设置2位小数
            return round(value, 2)
        else:
            # 文本报告中格式化显示
            if value == 0:
                return "0"
            elif value == int(value):
                return str(int(value))
            else:
                return f"{value:.2f}".rstrip('0').rstrip('.')


# 指标配置管理 - 单独模块，便于维护
class MetricsConfig:
    """监控指标配置管理"""

    @staticmethod
    def get_business_metrics() -> Dict[str, Dict[str, Any]]:
        """业务核心指标"""
        return {
            "availability": {
                "query": 'count(nginx_active{{cluster="{cluster}"}} > 0) / count(nginx_active{{cluster="{cluster}"}}) * 100',
                "type": "instant",
                "unit": "%",
                "name": "可用性",
                "thresholds": {"critical": 99.0, "warning": 99.9}
            },
            "total_requests": {
                "query": 'sum(increase(nginx_requests{{cluster="{cluster}"}}[{duration}]))',
                "query_fallback": 'sum(deriv(nginx_requests{{cluster="{cluster}"}}[{duration}]) * {duration_seconds})',
                "type": "instant",
                "unit": "万次",
                "name": "总请求数",
                "scale": 10000,
                "ghost_counter_threshold": 1000000000  # 10亿请求阈值
            },
            "qps_peak": {
                "query": 'sum(rate(nginx_requests{{cluster="{cluster}"}}[2m]))',
                "query_fallback": 'sum(deriv(nginx_requests{{cluster="{cluster}"}}[5m]))',
                "type": "range",
                "unit": "req/s",
                "name": "QPS峰值",
                "ghost_counter_threshold": 1000000  # 100万QPS明显异常
            },
            "bandwidth_in_peak": {
                "query": 'sum(rate(net_bytes_recv{{cluster="{cluster}"}}[2m]) * 8) / 1000000000',
                "query_fallback": 'sum(deriv(net_bytes_recv{{cluster="{cluster}"}}[5m]) * 8) / 1000000000',
                "type": "range",
                "unit": "Gbps",
                "name": "下行带宽峰值",
                "ghost_counter_threshold": 1000  # 1000Gbps明显异常
            },
            "bandwidth_out_peak": {
                "query": 'sum(rate(net_bytes_sent{{cluster="{cluster}"}}[2m]) * 8) / 1000000000',
                "query_fallback": 'sum(deriv(net_bytes_sent{{cluster="{cluster}"}}[5m]) * 8) / 1000000000',
                "type": "range",
                "unit": "Gbps",
                "name": "上行带宽峰值",
                "ghost_counter_threshold": 1000  # 1000Gbps明显异常
            },
            "bandwidth_total_peak": {
                "query": 'sum(rate(net_bytes_recv{{cluster="{cluster}"}}[2m]) * 8 + rate(net_bytes_sent{{cluster="{cluster}"}}[2m]) * 8) / 1000000000',
                "query_fallback": 'sum(deriv(net_bytes_recv{{cluster="{cluster}"}}[5m]) * 8 + deriv(net_bytes_sent{{cluster="{cluster}"}}[5m]) * 8) / 1000000000',
                "type": "range",
                "unit": "Gbps",
                "name": "总带宽峰值",
                "ghost_counter_threshold": 1000  # 1000Gbps明显异常
            },
            "traffic_in_total": {
                "query": 'sum(increase(net_bytes_recv{{cluster="{cluster}"}}[{duration}])) / 1024 / 1024 / 1024',
                "query_fallback": 'sum(deriv(net_bytes_recv{{cluster="{cluster}"}}[{duration}]) * {duration_seconds}) / 1024 / 1024 / 1024',
                "type": "instant",
                "unit": "GB",
                "name": "下行总流量",
                "ghost_counter_threshold": 10000  # 10TB明显异常
            },
            "traffic_out_total": {
                "query": 'sum(increase(net_bytes_sent{{cluster="{cluster}"}}[{duration}])) / 1024 / 1024 / 1024',
                "query_fallback": 'sum(deriv(net_bytes_sent{{cluster="{cluster}"}}[{duration}]) * {duration_seconds}) / 1024 / 1024 / 1024',
                "type": "instant",
                "unit": "GB",
                "name": "上行总流量",
                "ghost_counter_threshold": 10000  # 10TB明显异常
            },
            "traffic_total": {
                "query": 'sum(increase(net_bytes_recv{{cluster="{cluster}"}}[{duration}]) + increase(net_bytes_sent{{cluster="{cluster}"}}[{duration}])) / 1024 / 1024 / 1024',
                "query_fallback": 'sum(deriv(net_bytes_recv{{cluster="{cluster}"}}[{duration}]) * {duration_seconds} + deriv(net_bytes_sent{{cluster="{cluster}"}}[{duration}]) * {duration_seconds}) / 1024 / 1024 / 1024',
                "type": "instant",
                "unit": "GB",
                "name": "总流量",
                "ghost_counter_threshold": 10000  # 10TB明显异常
            }
        }

    @staticmethod
    def get_system_metrics() -> Dict[str, Dict[str, Any]]:
        """系统资源指标"""
        return {
            "cpu_usage_peak": {
                "query": 'avg(cpu_usage_active{{cluster="{cluster}"}})',
                "type": "range",
                "unit": "%",
                "name": "CPU使用率峰值",
                "thresholds": {"critical": 85, "warning": 70}
            },
            "memory_usage_peak": {
                "query": 'avg(mem_used_percent{{cluster="{cluster}"}})',
                "type": "range", 
                "unit": "%",
                "name": "内存使用率峰值",
                "thresholds": {"critical": 90, "warning": 75}
            },
            "system_load_peak": {
                "query": 'avg(system_load5{{cluster="{cluster}"}})',
                "type": "range",
                "unit": "",
                "name": "系统负载峰值"
            },
            "node_count": {
                "query": 'count(up{{cluster="{cluster}", job=~".*node.*"}} == 1 or system_uptime{{cluster="{cluster}"}} > 0)',
                "type": "instant",
                "unit": "个",
                "name": "在线节点数"
            }
        }

    @staticmethod
    def get_connection_metrics() -> Dict[str, Dict[str, Any]]:
        """连接状态指标"""
        return {
            "conn_active_peak": {
                "query": 'sum(nginx_active{{cluster="{cluster}"}})',
                "type": "range",
                "unit": "个",
                "name": "活跃连接峰值"
            },
            "conn_waiting_peak": {
                "query": 'sum(nginx_waiting{{cluster="{cluster}"}})',
                "type": "range",
                "unit": "个",
                "name": "等待连接峰值"
            },
            "conn_reading_peak": {
                "query": 'sum(nginx_reading{{cluster="{cluster}"}})',
                "type": "range",
                "unit": "个",
                "name": "读取连接峰值"
            },
            "conn_writing_peak": {
                "query": 'sum(nginx_writing{{cluster="{cluster}"}})',
                "type": "range",
                "unit": "个",
                "name": "写入连接峰值"
            }
        }
    
    @staticmethod
    def get_anomaly_metrics() -> Dict[str, Dict[str, Any]]:
        """异常事件指标"""
        return {
            "oom_events": {
                "query": 'sum(increase(kernel_vmstat_oom_kill{{cluster="{cluster}"}}[{duration}]))',
                "type": "instant",
                "unit": "次", 
                "name": "OOM事件",
                "thresholds": {"critical": 1}
            },
            "high_cpu_nodes": {
                "query": 'count(cpu_usage_active{{cluster="{cluster}"}} > {cpu_threshold})',
                "type": "instant",
                "unit": "个",
                "name": "高CPU节点数",
                "thresholds": {"warning": 1}
            },
            "network_errors": {
                "query": 'sum(increase(net_err_in{{cluster="{cluster}"}}[{duration}]) + increase(net_err_out{{cluster="{cluster}"}}[{duration}]))',
                "type": "instant", 
                "unit": "个",
                "name": "网络错误数",
                "thresholds": {"critical": 1000, "warning": 100}
            }
        }


# 夜莺客户端 - 专注于数据获取
class NightingaleClient:
    """夜莺监控API客户端"""
    
    def __init__(self, base_url: str, username: str, password: str, datasource_id: str = "1"):
        self.base_url = base_url.rstrip('/')
        self.datasource_id = datasource_id
        self.auth = HTTPBasicAuth(username, password)
        self.session = requests.Session()
        self.session.auth = self.auth
        self.session.headers.update({
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        })
        self.session.timeout = 30
    
    def test_connection(self) -> bool:
        """测试连接"""
        try:
            result = self.query_instant("up", int(datetime.datetime.now().timestamp()))
            return result and "data" in result
        except Exception as e:
            logger.error(f"连接测试失败: {e}")
            return False
    
    def query_instant(self, query: str, timestamp: int) -> Dict:
        """即时查询"""
        url = f"{self.base_url}/api/n9e/proxy/{self.datasource_id}/api/v1/query"
        params = {
            'query': query,
            'time': timestamp
        }
        
        try:
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            logger.error(f"即时查询失败: {query[:50]}... - {e}")
            return {}
    
    def query_range(self, query: str, start_time: int, end_time: int, step: str = "1m") -> Dict:
        """范围查询"""
        url = f"{self.base_url}/api/n9e/proxy/{self.datasource_id}/api/v1/query_range"
        params = {
            'query': query,
            'start': start_time,
            'end': end_time,
            'step': step
        }
        
        try:
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            logger.error(f"范围查询失败: {query[:50]}... - {e}")
            return {}


# 数据提取器 - 处理Prometheus查询结果
class MetricsExtractor:
    """监控数据提取器"""
    
    @staticmethod
    def extract_max_with_time(result: Dict) -> Tuple[float, str]:
        """提取最大值和时间"""
        max_value = 0.0
        max_time = "N/A"
        
        for series in result.get("data", {}).get("result", []):
            for timestamp, value in series.get("values", []):
                try:
                    val = float(value)
                    if val > max_value:
                        max_value = val
                        # 只显示到分钟，因为秒通常是00
                        max_time = datetime.datetime.fromtimestamp(timestamp).strftime("%m-%d %H:%M")
                except (ValueError, TypeError):
                    continue
        
        return max_value, max_time
    
    @staticmethod
    def extract_max_with_time_for_excel(result: Dict) -> Tuple[float, str]:
        """提取最大值和时间（Excel格式：只包含时分秒）"""
        max_value = 0.0
        max_time = "N/A"
        
        for series in result.get("data", {}).get("result", []):
            for timestamp, value in series.get("values", []):
                try:
                    val = float(value)
                    if val > max_value:
                        max_value = val
                        # Excel格式只显示到分钟
                        max_time = datetime.datetime.fromtimestamp(timestamp).strftime("%H:%M")
                except (ValueError, TypeError):
                    continue
        
        return max_value, max_time
    
    @staticmethod
    def extract_instant_value(result: Dict) -> float:
        """提取即时值"""
        for series in result.get("data", {}).get("result", []):
            try:
                return float(series["value"][1])
            except (KeyError, ValueError, TypeError, IndexError):
                continue
        return 0.0
    
    @staticmethod
    def extract_avg(result: Dict) -> float:
        """提取平均值"""
        total_sum = 0.0
        count = 0
        
        for series in result.get("data", {}).get("result", []):
            for timestamp, value in series.get("values", []):
                try:
                    total_sum += float(value)
                    count += 1
                except (ValueError, TypeError):
                    continue
        
        return total_sum / count if count > 0 else 0.0


# 数据收集器 - 核心业务逻辑
class DataCollector:
    """监控数据收集器"""
    
    def __init__(self, client: NightingaleClient, cpu_threshold: float = 85, memory_threshold: float = 90):
        self.client = client
        self.extractor = MetricsExtractor()
        self.cpu_threshold = cpu_threshold
        self.memory_threshold = memory_threshold
        self.metrics_config = MetricsConfig()
    
    def collect_daily_metrics(self, cluster: str, start_time: int, end_time: int, excel_format: bool = False) -> Dict[str, Any]:
        """收集单日指标数据"""
        duration = f"{end_time - start_time}s"
        duration_seconds = end_time - start_time
        
        # 合并所有指标配置
        all_metrics = {}
        all_metrics.update(self.metrics_config.get_business_metrics())
        all_metrics.update(self.metrics_config.get_system_metrics())
        all_metrics.update(self.metrics_config.get_connection_metrics())
        all_metrics.update(self.metrics_config.get_anomaly_metrics())
        
        daily_data = {}
        
        for metric_name, config in all_metrics.items():
            try:
                # 构建查询语句
                query = config["query"].format(
                    cluster=cluster,
                    duration=duration,
                    duration_seconds=duration_seconds,
                    cpu_threshold=self.cpu_threshold,
                    memory_threshold=self.memory_threshold
                )
                
                # 执行查询
                if config["type"] == "instant":
                    result = self.client.query_instant(query, end_time)
                    value = self.extractor.extract_instant_value(result)
                    peak_time = ""
                else:
                    result = self.client.query_range(query, start_time, end_time)
                    if excel_format:
                        value, peak_time = self.extractor.extract_max_with_time_for_excel(result)
                    else:
                        value, peak_time = self.extractor.extract_max_with_time(result)
                
                # 检测幽灵计数器重置
                ghost_threshold = config.get("ghost_counter_threshold")
                used_fallback = False
                original_value = value
                corrected_value = value
                
                if ghost_threshold and value > ghost_threshold and "query_fallback" in config:
                    logger.warning(f"检测到指标 {metric_name} 可能存在幽灵计数器重置 (值: {value}, 阈值: {ghost_threshold})，使用deriv()函数重新计算")
                    
                    # 使用fallback查询（deriv函数）
                    fallback_query = config["query_fallback"].format(
                        cluster=cluster,
                        duration=duration,
                        duration_seconds=duration_seconds,
                        cpu_threshold=self.cpu_threshold,
                        memory_threshold=self.memory_threshold
                    )
                    
                    if config["type"] == "instant":
                        fallback_result = self.client.query_instant(fallback_query, end_time)
                        fallback_value = self.extractor.extract_instant_value(fallback_result)
                        fallback_peak_time = ""
                    else:
                        fallback_result = self.client.query_range(fallback_query, start_time, end_time)
                        if excel_format:
                            fallback_value, fallback_peak_time = self.extractor.extract_max_with_time_for_excel(fallback_result)
                        else:
                            fallback_value, fallback_peak_time = self.extractor.extract_max_with_time(fallback_result)
                    
                    # 如果fallback结果更合理，使用fallback结果
                    if fallback_value < ghost_threshold and fallback_value > 0:
                        corrected_value = fallback_value
                        value = fallback_value
                        peak_time = fallback_peak_time
                        used_fallback = True
                        logger.info(f"使用deriv()计算的 {metric_name} 结果: {value}")
                
                # 应用单位换算
                if "scale" in config:
                    value = value / config["scale"]
                    if used_fallback:
                        # 同时记录原始值（换算前）和修正值（换算前）用于异常报告
                        original_value_scaled = original_value / config["scale"] if "scale" in config else original_value
                        corrected_value_scaled = corrected_value / config["scale"] if "scale" in config else corrected_value
                    else:
                        original_value_scaled = value
                        corrected_value_scaled = value
                else:
                    original_value_scaled = original_value
                    corrected_value_scaled = corrected_value
                
                daily_data[metric_name] = {
                    "value": max(0, value),  # 确保非负
                    "unit": config["unit"],
                    "name": config["name"],
                    "peak_time": peak_time,
                    "status": self._get_status(value, config.get("thresholds", {})),
                    "used_fallback": used_fallback,
                    "original_value": original_value_scaled if used_fallback else None,
                    "corrected_value": corrected_value_scaled if used_fallback else None
                }
                
            except Exception as e:
                logger.error(f"收集指标 {metric_name} 失败: {e}")
                daily_data[metric_name] = {
                    "value": 0.0,
                    "unit": config["unit"],
                    "name": config["name"],
                    "peak_time": "",
                    "status": "error",
                    "used_fallback": False,
                    "original_value": None,
                    "corrected_value": None
                }
        
        return daily_data
    
    def collect_period_data(self, cluster: str, start_date: datetime.date, end_date: datetime.date) -> List[Dict]:
        """收集时间段内的每日数据"""
        period_data = []
        current_date = start_date
        
        while current_date <= end_date:
            # 计算当日时间范围
            day_start = datetime.datetime.combine(current_date, datetime.time(0, 0, 0))
            day_end = datetime.datetime.combine(current_date, datetime.time(23, 59, 59))
            
            start_timestamp = int(day_start.timestamp())
            end_timestamp = int(day_end.timestamp())
            
            logger.info(f"收集 {current_date} 的数据...")
            
            # 收集当日数据
            daily_metrics = self.collect_daily_metrics(cluster, start_timestamp, end_timestamp)
            
            # 添加日期信息
            daily_metrics["date"] = current_date.strftime("%Y-%m-%d")
            period_data.append(daily_metrics)
            
            current_date += datetime.timedelta(days=1)
        
        return period_data

    def collect_datetime_data_for_excel(self, cluster: str, start_datetime: datetime.datetime, end_datetime: datetime.datetime) -> \
    List[Dict]:
        """收集指定时间段的数据（Excel格式，时间只显示时分秒）"""
        logger.info(f"收集 {start_datetime} 至 {end_datetime} 的数据 (Excel格式)...")

        # 判断是否跨越多个自然天
        start_date = start_datetime.date()
        end_date = end_datetime.date()
        
        # 如果跨越多个自然天，按天分别统计
        if start_date != end_date:
            logger.info(f"时间段跨越多个自然天，将按天分别统计...")
            period_data = []
            current_date = start_date
            
            while current_date <= end_date:
                # 计算当天的实际时间范围
                if current_date == start_date:
                    # 第一天：从开始时间到当天结束
                    day_start = start_datetime
                    day_end = datetime.datetime.combine(current_date, datetime.time(23, 59, 59))
                elif current_date == end_date:
                    # 最后一天：从当天开始到结束时间
                    day_start = datetime.datetime.combine(current_date, datetime.time(0, 0, 0))
                    day_end = end_datetime
                else:
                    # 中间天：完整一天
                    day_start = datetime.datetime.combine(current_date, datetime.time(0, 0, 0))
                    day_end = datetime.datetime.combine(current_date, datetime.time(23, 59, 59))
                
                start_timestamp = int(day_start.timestamp())
                end_timestamp = int(day_end.timestamp())
                
                logger.info(f"收集 {current_date} ({day_start.strftime('%H:%M:%S')} - {day_end.strftime('%H:%M:%S')}) 的数据...")
                
                # 收集当天数据（Excel格式）
                daily_metrics = self.collect_daily_metrics(cluster, start_timestamp, end_timestamp, excel_format=True)
                daily_metrics["date"] = current_date.strftime("%Y-%m-%d")
                period_data.append(daily_metrics)
                
                current_date += datetime.timedelta(days=1)
            
            return period_data
        else:
            # 单天内的时间段，仍按一天输出（Excel格式）
            logger.info(f"时间段在同一天内，按单天统计...")
            start_timestamp = int(start_datetime.timestamp())
            end_timestamp = int(end_datetime.timestamp())

            # 收集数据（Excel格式）
            metrics = self.collect_daily_metrics(cluster, start_timestamp, end_timestamp, excel_format=True)
            
            # 使用开始日期作为日期标识
            metrics["date"] = start_date.strftime("%Y-%m-%d")
            
            return [metrics]

    def collect_datetime_data(self, cluster: str, start_datetime: datetime.datetime, end_datetime: datetime.datetime) -> \
    List[Dict]:
        """收集指定时间段的数据（支持精确到秒）"""
        logger.info(f"收集 {start_datetime} 至 {end_datetime} 的数据...")

        # 判断是否跨越多个自然天
        start_date = start_datetime.date()
        end_date = end_datetime.date()
        
        # 如果跨越多个自然天，按天分别统计
        if start_date != end_date:
            logger.info(f"时间段跨越多个自然天，将按天分别统计...")
            period_data = []
            current_date = start_date
            
            while current_date <= end_date:
                # 计算当天的实际时间范围
                if current_date == start_date:
                    # 第一天：从开始时间到当天结束
                    day_start = start_datetime
                    day_end = datetime.datetime.combine(current_date, datetime.time(23, 59, 59))
                elif current_date == end_date:
                    # 最后一天：从当天开始到结束时间
                    day_start = datetime.datetime.combine(current_date, datetime.time(0, 0, 0))
                    day_end = end_datetime
                else:
                    # 中间天：完整一天
                    day_start = datetime.datetime.combine(current_date, datetime.time(0, 0, 0))
                    day_end = datetime.datetime.combine(current_date, datetime.time(23, 59, 59))
                
                start_timestamp = int(day_start.timestamp())
                end_timestamp = int(day_end.timestamp())
                
                logger.info(f"收集 {current_date} ({day_start.strftime('%H:%M:%S')} - {day_end.strftime('%H:%M:%S')}) 的数据...")
                
                # 收集当天数据
                daily_metrics = self.collect_daily_metrics(cluster, start_timestamp, end_timestamp)
                daily_metrics["date"] = current_date.strftime("%Y-%m-%d")
                period_data.append(daily_metrics)
                
                current_date += datetime.timedelta(days=1)
            
            return period_data
        else:
            # 单天内的时间段，仍按一天输出
            logger.info(f"时间段在同一天内，按单天统计...")
            start_timestamp = int(start_datetime.timestamp())
            end_timestamp = int(end_datetime.timestamp())

            # 收集数据
            metrics = self.collect_daily_metrics(cluster, start_timestamp, end_timestamp)
            
            # 使用开始日期作为日期标识
            metrics["date"] = start_date.strftime("%Y-%m-%d")
            
            return [metrics]

    def collect_comparison_data(self, cluster: str, current_start: int, current_end: int) -> Dict:
        """收集对比数据（环比）"""
        duration = current_end - current_start
        prev_start = current_start - duration
        prev_end = current_start

        logger.info(f"收集环比对比数据...")

        # 收集上一周期数据
        prev_metrics = self.collect_daily_metrics(cluster, prev_start, prev_end)

        return prev_metrics

    def _get_status(self, value: float, thresholds: Dict[str, float]) -> str:
        """根据阈值判断状态"""
        if not thresholds:
            return "normal"
        
        if "critical" in thresholds and value >= thresholds["critical"]:
            return "critical"
        elif "warning" in thresholds and value >= thresholds["warning"]:
            return "warning"
        else:
            return "normal"


# 报告生成器 - 负责格式化输出
class ReportGenerator:
    """报告生成器"""
    
    @staticmethod
    def generate_text_report(period_data: List[Dict], cluster: str, start_date: str, end_date: str) -> str:
        """生成文本报告"""
        if not period_data:
            return "无数据"
        
        # 计算汇总统计
        summary_stats = ReportGenerator._calculate_summary_stats(period_data)
        
        # 生成报告头部
        lines = [
            "=" * 80,
            f"🏢 【{cluster}】集群系统运行报告",
            f"📅 时间范围: {start_date} 至 {end_date}",
            f"📊 数据天数: {len(period_data)} 天",
            f"🕒 生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "=" * 80,
            "",
            "【📋 执行摘要】",
            ""
        ]
        
        # 添加关键指标摘要
        lines.extend([
            f"🎯 平均可用性: {NumberFormatter.format_for_display(summary_stats.get('availability_avg', 0))}%",
            f"📊 总请求数: {NumberFormatter.format_for_display(summary_stats.get('total_requests_sum', 0))}万次",
            f"⚡ QPS峰值: {NumberFormatter.format_for_display(summary_stats.get('qps_peak_max', 0))} req/s ({summary_stats.get('qps_peak_time', 'N/A')})",
            f"🌐 总带宽峰值: {NumberFormatter.format_for_display(summary_stats.get('bandwidth_total_peak_max', 0))} Gbps ({summary_stats.get('bandwidth_total_peak_time', 'N/A')})",
            f"📦 总流量: {NumberFormatter.format_for_display(summary_stats.get('traffic_total_sum', 0))} GB",
            f"💻 CPU峰值: {NumberFormatter.format_for_display(summary_stats.get('cpu_usage_peak_max', 0))}% ({summary_stats.get('cpu_usage_peak_time', 'N/A')})",
            f"🧠 内存峰值: {NumberFormatter.format_for_display(summary_stats.get('memory_usage_peak_max', 0))}% ({summary_stats.get('memory_usage_peak_time', 'N/A')})",
            f"🔗 活跃连接峰值: {NumberFormatter.format_for_display(summary_stats.get('conn_active_peak_max', 0))}个 ({summary_stats.get('conn_active_peak_time', 'N/A')})",
            f"⏱️ 等待连接峰值: {NumberFormatter.format_for_display(summary_stats.get('conn_waiting_peak_max', 0))}个 ({summary_stats.get('conn_waiting_peak_time', 'N/A')})",
            "",
            "【📈 详细数据】",
            ""
        ])
        
        # 添加每日数据表格
        lines.append("| 日期       | 可用性(%) | 请求数(万) | QPS峰值 | 总带宽峰值(Gbps) | 流量(GB) | CPU峰值(%) | 内存峰值(%) | 活跃连接峰值 | 等待连接峰值 |")
        lines.append("|------------|-----------|------------|---------|----------------|----------|------------|-------------|------------|------------|")
        
        for daily_data in period_data:
            date = daily_data.get("date", "N/A")
            availability = daily_data.get("availability", {}).get("value", 0)
            requests = daily_data.get("total_requests", {}).get("value", 0)
            qps = daily_data.get("qps_peak", {}).get("value", 0)
            bandwidth = daily_data.get("bandwidth_total_peak", {}).get("value", 0)
            traffic = daily_data.get("traffic_total", {}).get("value", 0)
            cpu = daily_data.get("cpu_usage_peak", {}).get("value", 0)
            memory = daily_data.get("memory_usage_peak", {}).get("value", 0)
            conn_active = daily_data.get("conn_active_peak", {}).get("value", 0)
            conn_waiting = daily_data.get("conn_waiting_peak", {}).get("value", 0)
            
            lines.append(f"| {date} | {NumberFormatter.format_for_display(availability):>7} | {NumberFormatter.format_for_display(requests):>10} | {NumberFormatter.format_for_display(qps):>7} | {NumberFormatter.format_for_display(bandwidth):>14} | {NumberFormatter.format_for_display(traffic):>8} | {NumberFormatter.format_for_display(cpu):>10} | {NumberFormatter.format_for_display(memory):>11} | {NumberFormatter.format_for_display(conn_active):>10} | {NumberFormatter.format_for_display(conn_waiting):>10} |")
        
        # 添加智能分析和异常检测
        anomalies = ReportGenerator._detect_anomalies(summary_stats, period_data)
        if anomalies:
            lines.extend([
                "",
                "【⚠️ 异常检测】",
                ""
            ])
            for anomaly in anomalies:
                lines.append(f"  ⚠️ {anomaly}")
            lines.append("")
        
        # 添加关键性能指标趋势
        trends = ReportGenerator._analyze_trends(period_data)
        if trends:
            lines.extend([
                "【📈 性能趋势】",
                ""
            ])
            for trend in trends:
                lines.append(f"  📊 {trend}")
            lines.append("")
        
        # 添加运维建议
        recommendations = ReportGenerator._generate_recommendations(summary_stats, period_data)
        if any(recommendations.values()):
            lines.extend([
                "【💡 运维建议】",
                ""
            ])
            for category, recs in recommendations.items():
                if recs:
                    lines.append(f"🔍 {category}:")
                    for rec in recs:
                        lines.append(f"  • {rec}")
                    lines.append("")
        
        lines.extend([
            "=" * 80,
            ""
        ])
        
        return "\n".join(lines)
    
    @staticmethod
    def generate_excel_report(period_data: List[Dict], cluster: str, start_date: str, end_date: str) -> openpyxl.Workbook:
        """生成Excel报告"""
        wb = openpyxl.Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建明细数据表
        detail_ws = wb.create_sheet("每日明细数据")
        
        # 构建明细数据
        if period_data:
            # 获取所有指标名称
            first_day = period_data[0]
            metric_names = [key for key in first_day.keys() if key not in ["date", "time_range"]]
            
            # 构建2行表头
            # 第一行：主要指标名称（跨列）
            # 第二行：具体数值列和峰值时间列
            
            first_row = ["日期"]
            second_row = [""]
            
            col_index = 2  # 从第2列开始（第1列是日期）
            
            for metric_name in metric_names:
                metric_data = first_day.get(metric_name, {})
                if isinstance(metric_data, dict):
                    name = metric_data.get("name", metric_name)
                    unit = metric_data.get("unit", "")
                    
                    # 检查是否有峰值时间
                    has_peak_time = bool(metric_data.get("peak_time"))
                    
                    if has_peak_time:
                        # 有峰值时间的指标占2列
                        first_row.extend([name, ""])  # 第一行跨2列
                        # 简化表头，去掉冗余的"峰值"字样和时间格式说明
                        if unit:
                            second_row.extend([f"数值({unit})", "时间"])
                        else:
                            second_row.extend(["数值", "时间"])
                        col_index += 2
                    else:
                        # 没有峰值时间的指标占1列
                        display_name = f"{name}({unit})" if unit else name
                        first_row.append(display_name)
                        second_row.append("")
                        col_index += 1
            
            # 添加表头
            detail_ws.append(first_row)
            detail_ws.append(second_row)
            
            # 合并第一行的多列标题
            col_index = 2
            for metric_name in metric_names:
                metric_data = first_day.get(metric_name, {})
                if isinstance(metric_data, dict):
                    has_peak_time = bool(metric_data.get("peak_time"))
                    
                    if has_peak_time:
                        # 合并2列
                        start_col = openpyxl.utils.get_column_letter(col_index)
                        end_col = openpyxl.utils.get_column_letter(col_index + 1)
                        detail_ws.merge_cells(f"{start_col}1:{end_col}1")
                        col_index += 2
                    else:
                        # 合并当前列的两行
                        col_letter = openpyxl.utils.get_column_letter(col_index)
                        detail_ws.merge_cells(f"{col_letter}1:{col_letter}2")
                        col_index += 1
            
            # 合并日期列的两行
            detail_ws.merge_cells("A1:A2")
            
            # 填充数据
            for daily_data in period_data:
                row = [daily_data.get("date", "")]

                for metric_name in metric_names:
                    metric_data = daily_data.get(metric_name, {})
                    if isinstance(metric_data, dict):
                        value = metric_data.get("value", 0)
                        peak_time = metric_data.get("peak_time", "")
                        
                        # Excel中保持数值类型，便于计算
                        formatted_value = NumberFormatter.format_for_display(value, for_excel=True)
                        
                        # 检查是否有峰值时间
                        if peak_time:
                            # 有峰值时间的指标：添加数值和时间
                            row.extend([formatted_value, peak_time])
                        else:
                            # 没有峰值时间的指标：只添加数值
                            row.append(formatted_value)

                detail_ws.append(row)
            
            # 添加统计行
            detail_ws.append([])  # 空行
            
            # 计算统计数据 - 只对数值列进行统计
            data_start_row = 3  # 数据从第3行开始（前两行是表头）
            data_end_row = len(period_data) + 2  # 数据结束行
            
            for i, stat_name in enumerate(["最大值", "最小值", "平均值", "总和"]):
                stat_row = [stat_name]
                
                # 遍历所有列，只对数值列计算统计
                col_index = 2
                for metric_name in metric_names:
                    metric_data = first_day.get(metric_name, {})
                    if isinstance(metric_data, dict):
                        has_peak_time = bool(metric_data.get("peak_time"))
                        
                        # 对数值列计算统计
                        col_letter = openpyxl.utils.get_column_letter(col_index)
                        
                        if stat_name == "最大值":
                            formula = f"=MAX({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                        elif stat_name == "最小值":
                            formula = f"=MIN({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                        elif stat_name == "平均值":
                            formula = f"=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                        else:  # 总和
                            formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                        
                        stat_row.append(formula)
                        
                        if has_peak_time:
                            # 峰值时间列不进行统计，留空
                            stat_row.append("")
                            col_index += 2
                        else:
                            col_index += 1
                
                detail_ws.append(stat_row)
                
                # 为统计行设置特殊样式
                current_row = detail_ws.max_row
                for col_num in range(1, detail_ws.max_column + 1):
                    try:
                        cell = detail_ws.cell(row=current_row, column=col_num)
                        if col_num == 1:  # 统计名称列
                            cell.font = Font(bold=True, size=10, color="FFFFFF")
                            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                        else:  # 数据列
                            cell.font = Font(bold=True, size=10, name="Consolas")
                            if i == 0:  # 最大值行用红色背景
                                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                            elif i == 3:  # 总和行用绿色背景
                                cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
                            else:  # 其他统计行用浅蓝色背景
                                cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                            # 为计算结果设置数值格式，保留2位小数
                            cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    except:
                        continue
        
        # 设置明细表样式
        ReportGenerator._apply_excel_styles(detail_ws)
        
        return wb
    
    @staticmethod
    def _calculate_summary_stats(period_data: List[Dict]) -> Dict:
        """计算汇总统计"""
        if not period_data:
            return {}
        
        summary = {}
        
        # 获取所有指标名称
        first_day = period_data[0]
        metric_names = [key for key in first_day.keys() if key not in ["date", "time_range"]]
        
        for metric_name in metric_names:
            values = []
            peak_times = []
            unit = ""
            name = metric_name
            max_value = 0
            max_peak_time = "N/A"
            
            for daily_data in period_data:
                metric_data = daily_data.get(metric_name, {})
                if isinstance(metric_data, dict):
                    value = metric_data.get("value", 0)
                    peak_time = metric_data.get("peak_time", "")
                    values.append(value)
                    peak_times.append(peak_time)
                    
                    # 记录最大值和对应的峰值时间
                    if value > max_value:
                        max_value = value
                        max_peak_time = peak_time if peak_time else "N/A"
                    
                    if not unit:
                        unit = metric_data.get("unit", "")
                    if name == metric_name:
                        name = metric_data.get("name", metric_name)
            
            if values:
                summary[f"{metric_name}_max"] = max(values)
                summary[f"{metric_name}_min"] = min(values)
                summary[f"{metric_name}_avg"] = sum(values) / len(values)
                summary[f"{metric_name}_sum"] = sum(values)
                summary[f"{metric_name}_time"] = max_peak_time  # 峰值对应的时间
                
                summary[metric_name] = {
                    "max": max(values),
                    "min": min(values),
                    "avg": sum(values) / len(values),
                    "sum": sum(values),
                    "unit": unit,
                    "name": name,
                    "peak_time": max_peak_time
                }
        
        return summary
    
    @staticmethod
    def _detect_anomalies(summary_stats: Dict, period_data: List[Dict]) -> List[str]:
        """检测异常情况"""
        anomalies = []
        
        # 检测幽灵计数器重置
        for daily_data in period_data:
            for metric_name, metric_data in daily_data.items():
                if isinstance(metric_data, dict) and metric_data.get("used_fallback", False):
                    original_val = metric_data.get("original_value", 0)
                    corrected_val = metric_data.get("corrected_value", 0)
                    unit = metric_data.get("unit", "")
                    metric_name_display = metric_data.get("name", metric_name)
                    
                    # 格式化原始值和修正值
                    original_display = NumberFormatter.format_for_display(original_val)
                    corrected_display = NumberFormatter.format_for_display(corrected_val)
                    
                    anomalies.append(f"{daily_data.get('date', 'N/A')} - {metric_name_display} 检测到幽灵计数器重置，原始值({original_display} {unit}) → 修正值({corrected_display} {unit})，已使用deriv()函数修正")
        
        # 检测异常高的资源使用率
        cpu_max = summary_stats.get('cpu_usage_peak_max', 0)
        memory_max = summary_stats.get('memory_usage_peak_max', 0)
        bandwidth_max = summary_stats.get('bandwidth_total_peak_max', 0)
        
        if cpu_max > 90:
            cpu_time = summary_stats.get('cpu_usage_peak_time', 'N/A')
            anomalies.append(f"CPU使用率峰值异常 - 达到 {NumberFormatter.format_for_display(cpu_max)}% ({cpu_time})，存在性能瓶颈风险")
        if memory_max > 95:
            memory_time = summary_stats.get('memory_usage_peak_time', 'N/A')
            anomalies.append(f"内存使用率峰值异常 - 达到 {NumberFormatter.format_for_display(memory_max)}% ({memory_time})，存在OOM风险")
        if bandwidth_max > 800:  # 800Gbps作为异常阈值
            bandwidth_time = summary_stats.get('bandwidth_total_peak_time', 'N/A')
            anomalies.append(f"带宽使用峰值异常 - 达到 {NumberFormatter.format_for_display(bandwidth_max)} Gbps ({bandwidth_time})，可能存在流量异常")
        
        # 检测可用性问题
        availability_avg = summary_stats.get('availability_avg', 100)
        if availability_avg < 99.9:
            availability_deficit = 99.9 - availability_avg
            anomalies.append(f"可用性SLA异常 - 平均可用性为 {NumberFormatter.format_for_display(availability_avg)}%，低于99.9%SLA要求 {NumberFormatter.format_for_display(availability_deficit)}%")
        
        return anomalies
    
    @staticmethod
    def _analyze_trends(period_data: List[Dict]) -> List[str]:
        """分析性能趋势"""
        if len(period_data) < 2:
            return []
        
        trends = []
        
        # 分析QPS趋势
        qps_values = [day.get("qps_peak", {}).get("value", 0) for day in period_data]
        if len(qps_values) >= 2:
            qps_start = qps_values[0]
            qps_end = qps_values[-1]
            qps_trend = qps_end - qps_start
            qps_change_pct = (qps_trend / qps_start * 100) if qps_start > 0 else 0
            if abs(qps_change_pct) > 20:
                trend_desc = "上升" if qps_trend > 0 else "下降"
                start_display = NumberFormatter.format_for_display(qps_start)
                end_display = NumberFormatter.format_for_display(qps_end)
                trends.append(f"QPS峰值呈{trend_desc}趋势，从 {start_display} req/s → {end_display} req/s，变化幅度 {qps_change_pct:+.1f}%")
        
        # 分析带宽趋势
        bandwidth_values = [day.get("bandwidth_total_peak", {}).get("value", 0) for day in period_data]
        if len(bandwidth_values) >= 2:
            bw_start = bandwidth_values[0]
            bw_end = bandwidth_values[-1]
            bw_trend = bw_end - bw_start
            bw_change_pct = (bw_trend / bw_start * 100) if bw_start > 0 else 0
            if abs(bw_change_pct) > 20:
                trend_desc = "上升" if bw_trend > 0 else "下降"
                start_display = NumberFormatter.format_for_display(bw_start)
                end_display = NumberFormatter.format_for_display(bw_end)
                trends.append(f"带宽使用呈{trend_desc}趋势，从 {start_display} Gbps → {end_display} Gbps，变化幅度 {bw_change_pct:+.1f}%")
        
        # 分析资源使用趋势
        cpu_values = [day.get("cpu_usage_peak", {}).get("value", 0) for day in period_data]
        if len(cpu_values) >= 2:
            cpu_start = cpu_values[0]
            cpu_end = cpu_values[-1]
            cpu_trend = cpu_end - cpu_start
            if abs(cpu_trend) > 10:
                trend_desc = "上升" if cpu_trend > 0 else "下降"
                start_display = NumberFormatter.format_for_display(cpu_start)
                end_display = NumberFormatter.format_for_display(cpu_end)
                trends.append(f"CPU使用率呈{trend_desc}趋势，从 {start_display}% → {end_display}%，变化 {cpu_trend:+.1f}%")
        
        return trends
    
    @staticmethod
    def _generate_recommendations(summary_stats: Dict, period_data: List[Dict]) -> Dict[str, List[str]]:
        """生成运维建议"""
        recommendations = {
            "性能优化": [],
            "资源管理": [],
            "容量规划": [],
            "运维建议": []
        }
        
        # 性能相关建议
        cpu_avg = summary_stats.get("cpu_usage_peak_avg", 0)
        if cpu_avg > 70:
            recommendations["性能优化"].append(f"CPU平均使用率达到{NumberFormatter.format_for_display(cpu_avg)}%，建议优化应用性能")
        
        memory_avg = summary_stats.get("memory_usage_peak_avg", 0)
        if memory_avg > 75:
            recommendations["性能优化"].append(f"内存平均使用率达到{NumberFormatter.format_for_display(memory_avg)}%，建议优化内存使用")
        
        # 容量规划建议
        qps_max = summary_stats.get("qps_peak_max", 0)
        if qps_max > 1000:
            recommendations["容量规划"].append(f"QPS峰值达到{NumberFormatter.format_for_display(qps_max)}，建议评估系统容量")
        
        bandwidth_max = summary_stats.get("bandwidth_peak_max", 0)
        if bandwidth_max > 1:
            recommendations["容量规划"].append(f"带宽峰值达到{NumberFormatter.format_for_display(bandwidth_max)}Gbps，建议关注网络容量")
        
        # 运维建议
        availability_avg = summary_stats.get("availability_avg", 0)
        if availability_avg < 99.9:
            recommendations["运维建议"].append(f"平均可用性为{NumberFormatter.format_for_display(availability_avg)}%，建议提升系统稳定性")
        
        # 如果没有问题，给出积极建议
        if not any(recommendations.values()):
            recommendations["运维建议"].append("系统运行状态良好，建议继续保持当前运维水平")
            recommendations["容量规划"].append("当前资源使用合理，可制定长期容量规划")
        
        return recommendations

    @staticmethod
    def _apply_excel_styles(worksheet):
        """应用Excel样式"""
        try:
            # 设置表头样式（前两行）
            for row_num in [1, 2]:
                for col_num in range(1, worksheet.max_column + 1):
                    try:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        if cell.value:  # 只为有内容的单元格设置样式
                            cell.font = Font(bold=True, color="FFFFFF", size=11)
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            # 表头支持自动换行和居中对齐
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    except:
                        continue

            # 为数据行设置样式
            for row_num in range(3, worksheet.max_row + 1):  # 跳过前两行表头
                # 添加交替行颜色以提高可读性
                is_even_row = (row_num % 2 == 0)
                row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") if is_even_row else None
                
                for col_num in range(1, worksheet.max_column + 1):
                    try:
                        cell = worksheet.cell(row=row_num, column=col_num)
                        
                        # 设置交替行背景色
                        if row_fill:
                            cell.fill = row_fill
                        
                        if col_num == 1:  # 日期列
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(size=10)
                        elif isinstance(cell.value, (int, float)):  # 数值列
                            # 优化数值格式：整数不显示小数，小数保持适当精度
                            if cell.value == 0:
                                cell.number_format = '0'  # 0不显示小数
                            elif cell.value == int(cell.value):
                                cell.number_format = '0'  # 整数不显示小数
                            elif abs(cell.value) < 0.01:
                                cell.number_format = '0.0000'
                            elif abs(cell.value) < 1:
                                cell.number_format = '0.000'
                            elif abs(cell.value) < 100:
                                cell.number_format = '0.00'
                            elif abs(cell.value) < 10000:
                                cell.number_format = '#,##0.0'  # 千分位分隔符
                            else:
                                cell.number_format = '#,##0'    # 大数值用千分位，不显示小数
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.font = Font(size=10, name="Consolas")  # 使用等宽字体提高数字可读性
                        elif isinstance(cell.value, str) and ":" in str(cell.value):  # 时间列
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(size=10, color="0066CC")  # 时间显示为蓝色
                        else:  # 其他列
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(size=10)
                    except:
                        continue

            # 智能调整列宽 - 优化版
            for col_num in range(1, worksheet.max_column + 1):
                column_letter = openpyxl.utils.get_column_letter(col_num)
                try:
                    max_length = 0
                    is_time_column = False
                    is_number_column = False
                    header_length = 0
                    
                    # 检查列的内容类型和最大长度
                    for row_num in range(1, worksheet.max_row + 1):
                        try:
                            cell = worksheet.cell(row=row_num, column=col_num)
                            if cell.value:
                                cell_str = str(cell.value)
                                
                                # 记录表头长度（前两行）
                                if row_num <= 2:
                                    header_length = max(header_length, len(cell_str))
                                
                                # 判断列类型
                                if ":" in cell_str and len(cell_str) <= 8:  # 时间列（HH:MM格式）
                                    is_time_column = True
                                elif row_num > 2 and isinstance(cell.value, (int, float)):  # 数字列（排除表头）
                                    is_number_column = True
                                
                                # 记录最大长度
                                if len(cell_str) > max_length:
                                    max_length = len(cell_str)
                        except:
                            continue
                    
                    # 根据列类型设置合适的宽度
                    if col_num == 1:  # 日期列
                        adjusted_width = 11
                    elif is_time_column:  # 时间列（HH:MM格式）
                        adjusted_width = 7  # 时间列较窄
                    elif is_number_column:  # 数字列
                        # 数字列基于内容长度，但考虑表头
                        content_width = max(max_length, header_length)
                        adjusted_width = min(content_width + 1, 12)  # 数字列最大12
                    else:
                        # 其他列基于内容长度，但优先考虑表头
                        content_width = max(max_length, header_length)
                        if content_width <= 8:
                            adjusted_width = content_width + 1
                        elif content_width <= 15:
                            adjusted_width = content_width + 2
                        else:
                            adjusted_width = min(content_width + 2, 20)  # 最大20
                    
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                except:
                    # 如果出错，使用默认宽度
                    worksheet.column_dimensions[column_letter].width = 10
            
            # 设置行高 - 优化表头高度以适应换行
            worksheet.row_dimensions[1].height = 35  # 第一行表头高度增加，适应换行
            worksheet.row_dimensions[2].height = 30  # 第二行表头
            
            # 为数据行设置适当的行高
            for row_num in range(3, worksheet.max_row + 1):
                worksheet.row_dimensions[row_num].height = 20
            
            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    try:
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
                    except:
                        continue
            
            # 冻结首行和首列
            worksheet.freeze_panes = "B3"  # 冻结前两行表头和第一列日期
            
        except Exception as e:
            # 如果样式设置失败，记录错误但不影响主流程
            logger.warning(f"Excel样式设置出现问题: {e}")


# GUI界面 - 现代化设计
class NginxMonitoringGUI:
    """Nginx监控系统GUI界面"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Nginx监控报告生成系统 - 优化版")
        self.root.geometry("900x750")
        self.root.resizable(True, True)
        
        # 配置样式
        self.setup_styles()
        
        # 初始化变量
        self.init_variables()
        
        # 创建界面
        self.create_widgets()
        
        # 状态
        self.client = None
        self.collector = None
        self.is_generating = False
    
    def setup_styles(self):
        """配置现代化界面样式"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # 配置颜色主题
        style.configure('Title.TLabel', font=('微软雅黑', 14, 'bold'), foreground='#2c3e50')
        style.configure('Section.TLabel', font=('微软雅黑', 11, 'bold'), foreground='#34495e')
        style.configure('Success.TLabel', foreground='#27ae60', font=('微软雅黑', 9))
        style.configure('Error.TLabel', foreground='#e74c3c', font=('微软雅黑', 9))
        style.configure('Info.TLabel', foreground='#3498db', font=('微软雅黑', 9))
        
        # 按钮样式
        style.configure('Action.TButton', font=('微软雅黑', 10, 'bold'))
        style.map('Action.TButton',
                 background=[('active', '#3498db'), ('pressed', '#2980b9')])

    def init_variables(self):
        """初始化界面变量"""
        # 连接配置
        self.url_var = tk.StringVar(value="http://localhost:17000")
        self.username_var = tk.StringVar(value="root")
        self.password_var = tk.StringVar(value="")
        self.datasource_id_var = tk.StringVar(value="1")

        # 监控配置
        self.cluster_var = tk.StringVar(value="default")

        # 时间配置 - 支持精确到秒
        yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
        self.start_datetime_var = tk.StringVar(value=yesterday.strftime("%Y-%m-%d %H:%M:%S"))
        self.end_datetime_var = tk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        # 快捷选择类型
        self.quick_select_var = tk.StringVar(value="自定义")

        # 阈值配置
        self.cpu_threshold_var = tk.DoubleVar(value=85.0)
        self.memory_threshold_var = tk.DoubleVar(value=90.0)

        # 输出配置
        self.output_text_var = tk.BooleanVar(value=True)
        self.output_excel_var = tk.BooleanVar(value=True)
        self.output_path_var = tk.StringVar(value="./reports/")

        # 状态
        self.status_var = tk.StringVar(value="就绪")
        self.connection_status_var = tk.StringVar(value="未连接")
    def create_widgets(self):
        """创建界面组件"""
        # 主容器
        main_container = ttk.Frame(self.root)
        main_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # 标题
        title_frame = ttk.Frame(main_container)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = ttk.Label(title_frame, text="🚀 Nginx监控报告生成系统", style='Title.TLabel')
        title_label.pack()
        
        subtitle_label = ttk.Label(title_frame, text="基于夜莺监控数据的智能报告生成工具", 
                                  font=('微软雅黑', 10), foreground='#7f8c8d')
        subtitle_label.pack(pady=(5, 0))
        
        # 创建notebook（标签页）
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # 连接配置页
        self.create_connection_tab()
        
        # 报告配置页
        self.create_report_tab()
        
        # 输出配置页
        self.create_output_tab()
        
        # 执行监控页
        self.create_execution_tab()
    
    def create_connection_tab(self):
        """创建连接配置标签页"""
        conn_frame = ttk.Frame(self.notebook)
        self.notebook.add(conn_frame, text="🔗 连接配置")
        
        # 主内容框架
        content_frame = ttk.Frame(conn_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 夜莺服务配置
        service_frame = ttk.LabelFrame(content_frame, text="夜莺服务配置", padding=15)
        service_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 服务器地址
        ttk.Label(service_frame, text="服务器地址:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky=tk.W, pady=8)
        url_entry = ttk.Entry(service_frame, textvariable=self.url_var, width=50, font=('微软雅黑', 10))
        url_entry.grid(row=0, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=8, padx=(10, 0))
        
        # 用户名
        ttk.Label(service_frame, text="用户名:", font=('微软雅黑', 10)).grid(row=1, column=0, sticky=tk.W, pady=8)
        username_entry = ttk.Entry(service_frame, textvariable=self.username_var, width=25, font=('微软雅黑', 10))
        username_entry.grid(row=1, column=1, sticky=tk.W, pady=8, padx=(10, 0))
        
        # 密码
        ttk.Label(service_frame, text="密码:", font=('微软雅黑', 10)).grid(row=2, column=0, sticky=tk.W, pady=8)
        password_entry = ttk.Entry(service_frame, textvariable=self.password_var, show="*", width=25, font=('微软雅黑', 10))
        password_entry.grid(row=2, column=1, sticky=tk.W, pady=8, padx=(10, 0))
        
        # 数据源ID
        ttk.Label(service_frame, text="数据源ID:", font=('微软雅黑', 10)).grid(row=3, column=0, sticky=tk.W, pady=8)
        datasource_entry = ttk.Entry(service_frame, textvariable=self.datasource_id_var, width=25, font=('微软雅黑', 10))
        datasource_entry.grid(row=3, column=1, sticky=tk.W, pady=8, padx=(10, 0))
        ttk.Label(service_frame, text="(通常为1或2)", font=('微软雅黑', 9), foreground='#7f8c8d').grid(row=3, column=2, sticky=tk.W, padx=(10, 0))
        
        # 连接按钮
        test_btn = ttk.Button(service_frame, text="🔧 测试连接", command=self.test_connection, style='Action.TButton')
        test_btn.grid(row=1, column=2, rowspan=3, padx=20, pady=8)
        
        # 连接状态
        status_frame = ttk.Frame(service_frame)
        status_frame.grid(row=4, column=1, columnspan=2, sticky=tk.W, pady=(10, 0))
        
        ttk.Label(status_frame, text="连接状态:", font=('微软雅黑', 10)).pack(side=tk.LEFT)
        self.connection_status_label = ttk.Label(status_frame, textvariable=self.connection_status_var, style='Error.TLabel')
        self.connection_status_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # 配置网格权重
        service_frame.columnconfigure(1, weight=1)
        
        # 使用说明
        help_frame = ttk.LabelFrame(content_frame, text="使用说明", padding=15)
        help_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        help_text = """📝 配置说明:

• 服务器地址: 夜莺监控系统的完整URL，例如: http://192.168.1.100:17000
• 用户名/密码: 夜莺系统的登录凭据
• 数据源ID: Prometheus数据源的ID，通常为1或2，可在夜莺管理界面查看
• 建议先点击"测试连接"确保连接正常

🔍 API接口说明:
• 即时查询: /api/n9e/proxy/{datasource_id}/api/v1/query
• 范围查询: /api/n9e/proxy/{datasource_id}/api/v1/query_range
• 支持Basic认证和Token认证

⚠️ 注意事项:
• 确保夜莺服务正常运行且API可访问
• 用户需要有Prometheus数据源的查询权限
• 数据源ID必须正确，否则查询将失败
• 网络连接超时时间为30秒"""
        
        help_label = ttk.Label(help_frame, text=help_text, font=('微软雅黑', 9), foreground='#555')
        help_label.pack(anchor=tk.W)

    def create_report_tab(self):
        """创建报告配置标签页"""
        report_frame = ttk.Frame(self.notebook)
        self.notebook.add(report_frame, text="📊 报告配置")

        content_frame = ttk.Frame(report_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # 快捷时间选择
        quick_frame = ttk.LabelFrame(content_frame, text="快捷时间选择", padding=15)
        quick_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(quick_frame, text="选择类型:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky=tk.W, pady=8)
        quick_combo = ttk.Combobox(quick_frame, textvariable=self.quick_select_var,
                                   values=["自定义", "最近1小时", "最近4小时", "最近12小时", "昨天", "最近7天",
                                           "最近30天"],
                                   state="readonly", width=15)
        quick_combo.grid(row=0, column=1, sticky=tk.W, pady=8, padx=(10, 0))
        quick_combo.bind('<<ComboboxSelected>>', self.on_quick_select_changed)

        ttk.Button(quick_frame, text="📅 应用", command=self.apply_quick_select).grid(row=0, column=2, padx=(10, 0))

        # 精确时间配置
        time_frame = ttk.LabelFrame(content_frame, text="精确时间配置", padding=15)
        time_frame.pack(fill=tk.X, pady=(0, 15))

        # 开始时间
        ttk.Label(time_frame, text="开始时间:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky=tk.W, pady=8)
        start_time_entry = ttk.Entry(time_frame, textvariable=self.start_datetime_var, width=25, font=('微软雅黑', 10))
        start_time_entry.grid(row=0, column=1, sticky=tk.W, pady=8, padx=(10, 0))
        ttk.Label(time_frame, text="(格式: YYYY-MM-DD HH:MM:SS)", font=('微软雅黑', 9), foreground='#7f8c8d').grid(
            row=0, column=2, sticky=tk.W, padx=(10, 0))

        # 结束时间
        ttk.Label(time_frame, text="结束时间:", font=('微软雅黑', 10)).grid(row=1, column=0, sticky=tk.W, pady=8)
        end_time_entry = ttk.Entry(time_frame, textvariable=self.end_datetime_var, width=25, font=('微软雅黑', 10))
        end_time_entry.grid(row=1, column=1, sticky=tk.W, pady=8, padx=(10, 0))
        ttk.Label(time_frame, text="(格式: YYYY-MM-DD HH:MM:SS)", font=('微软雅黑', 9), foreground='#7f8c8d').grid(
            row=1, column=2, sticky=tk.W, padx=(10, 0))

        # 时间范围信息显示
        self.time_info_label = ttk.Label(time_frame, text="", font=('微软雅黑', 9), foreground='#3498db')
        self.time_info_label.grid(row=2, column=1, columnspan=2, sticky=tk.W, pady=(5, 0))

        # 集群配置
        cluster_frame = ttk.LabelFrame(content_frame, text="集群配置", padding=15)
        cluster_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(cluster_frame, text="集群名称:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky=tk.W, pady=8)
        cluster_entry = ttk.Entry(cluster_frame, textvariable=self.cluster_var, width=30, font=('微软雅黑', 10))
        cluster_entry.grid(row=0, column=1, sticky=tk.W, pady=8, padx=(10, 0))
        ttk.Label(cluster_frame, text="(Prometheus集群标签值)", font=('微软雅黑', 9), foreground='#7f8c8d').grid(row=0,
                                                                                                                 column=2,
                                                                                                                 sticky=tk.W,
                                                                                                                 padx=(
                                                                                                                 10, 0))

        # 阈值配置
        threshold_frame = ttk.LabelFrame(content_frame, text="告警阈值配置", padding=15)
        threshold_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(threshold_frame, text="高CPU阈值(%):", font=('微软雅黑', 10)).grid(row=0, column=0, sticky=tk.W,
                                                                                     pady=8)
        cpu_spinbox = ttk.Spinbox(threshold_frame, from_=0, to=100, textvariable=self.cpu_threshold_var,
                                  increment=5, width=10, font=('微软雅黑', 10))
        cpu_spinbox.grid(row=0, column=1, sticky=tk.W, pady=8, padx=(10, 0))

        ttk.Label(threshold_frame, text="高内存阈值(%):", font=('微软雅黑', 10)).grid(row=0, column=2, sticky=tk.W,
                                                                                      pady=8, padx=(30, 0))
        memory_spinbox = ttk.Spinbox(threshold_frame, from_=0, to=100, textvariable=self.memory_threshold_var,
                                     increment=5, width=10, font=('微软雅黑', 10))
        memory_spinbox.grid(row=0, column=3, sticky=tk.W, pady=8, padx=(10, 0))

        # 绑定时间变化事件
        start_time_entry.bind('<KeyRelease>', self.update_time_info)
        end_time_entry.bind('<KeyRelease>', self.update_time_info)

        # 初始化时间信息
        self.update_time_info()

        # 指标说明（内容不变，省略...）

        # 指标说明
        metrics_frame = ttk.LabelFrame(content_frame, text="监控指标说明", padding=15)
        metrics_frame.pack(fill=tk.BOTH, expand=True)

        metrics_text = """📈 支持的监控指标:

🔹 业务指标:
  • 可用性(%): 服务在线率统计
  • 总请求数(万次): 统计时间段内的总请求量
  • QPS峰值(req/s): 每秒查询数峰值及发生时间
  • 带宽峰值(Gbps): 网络带宽使用峰值及发生时间
  • 总流量(GB): 统计时间段内的总流量

🔹 系统指标:
  • CPU使用率峰值(%): CPU使用率峰值及发生时间
  • 内存使用率峰值(%): 内存使用率峰值及发生时间
  • 系统负载峰值: 系统负载峰值及发生时间
  • 在线节点数(个): 当前在线的服务器节点数量

🔹 连接指标:
  • 活跃连接峰值(个): Nginx活跃连接数峰值
  • 等待连接峰值(个): Nginx等待连接数峰值

🔹 异常指标:
  • OOM事件(次): 内存溢出事件统计
  • 高CPU节点数(个): 超过阈值的节点数量
  • 网络错误数(个): 网络错误事件统计"""

        metrics_label = ttk.Label(metrics_frame, text=metrics_text, font=('微软雅黑', 9), foreground='#555')
        metrics_label.pack(anchor=tk.W)

    def create_output_tab(self):
        """创建输出配置标签页"""
        output_frame = ttk.Frame(self.notebook)
        self.notebook.add(output_frame, text="📁 输出配置")
        
        content_frame = ttk.Frame(output_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 输出格式配置
        format_frame = ttk.LabelFrame(content_frame, text="输出格式", padding=15)
        format_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(format_frame, text="选择输出格式:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky=tk.W, pady=8)
        
        format_options_frame = ttk.Frame(format_frame)
        format_options_frame.grid(row=0, column=1, sticky=tk.W, pady=8, padx=(20, 0))

        text_check = ttk.Checkbutton(format_options_frame, text="📄 文本报告(.txt)",
                                     variable=self.output_text_var)
        text_check.pack(anchor=tk.W, pady=2)

        excel_check = ttk.Checkbutton(format_options_frame, text="📊 Excel报表(.xlsx)",
                                      variable=self.output_excel_var)
        excel_check.pack(anchor=tk.W, pady=2)
        
        # 输出路径配置
        path_frame = ttk.LabelFrame(content_frame, text="保存路径", padding=15)
        path_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(path_frame, text="输出目录:", font=('微软雅黑', 10)).grid(row=0, column=0, sticky=tk.W, pady=8)
        
        path_input_frame = ttk.Frame(path_frame)
        path_input_frame.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=8, padx=(10, 0))
        path_input_frame.columnconfigure(0, weight=1)
        
        path_entry = ttk.Entry(path_input_frame, textvariable=self.output_path_var, font=('微软雅黑', 10))
        path_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        browse_btn = ttk.Button(path_input_frame, text="📂 浏览", command=self.browse_output_path)
        browse_btn.grid(row=0, column=1)
        
        path_frame.columnconfigure(1, weight=1)
        
        # 文件命名说明
        naming_frame = ttk.LabelFrame(content_frame, text="文件命名规则", padding=15)
        naming_frame.pack(fill=tk.X, pady=(0, 15))
        
        naming_text = """📝 文件命名格式:

• 文本报告: {集群名称}_监控报告_{开始日期}_{结束日期}_{生成时间}.txt
• Excel报表: {集群名称}_监控数据_{开始日期}_{结束日期}_{生成时间}.xlsx

📋 文件内容说明:

🔹 文本报告(.txt):
  • 执行摘要: 关键指标汇总
  • 每日明细: 表格形式展示每日数据
  • 运维建议: 基于数据分析的优化建议

🔹 Excel报表(.xlsx):
  • 每日明细数据: 包含所有指标的每日数据，列为指标，行为日期
  • 统计汇总: 包含最大值、最小值、平均值、总和的自动计算
  • 峰值时间: 记录各指标峰值出现的具体时间
  • 公式计算: 使用Excel公式自动计算统计数据

💡 特色功能:
  • 单位自动换算: 请求数显示为万次，带宽显示为Gbps
  • 峰值时间记录: 所有峰值数据都附带发生时间
  • 智能状态判断: 根据阈值自动判断指标状态"""
        
        naming_label = ttk.Label(naming_frame, text=naming_text, font=('微软雅黑', 9), foreground='#555')
        naming_label.pack(anchor=tk.W)
    
    def create_execution_tab(self):
        """创建执行监控标签页"""
        exec_frame = ttk.Frame(self.notebook)
        self.notebook.add(exec_frame, text="🚀 执行监控")
        
        content_frame = ttk.Frame(exec_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 操作按钮区域
        action_frame = ttk.LabelFrame(content_frame, text="报告生成", padding=15)
        action_frame.pack(fill=tk.X, pady=(0, 15))
        
        button_frame = ttk.Frame(action_frame)
        button_frame.pack()
        
        # 生成报告按钮
        self.generate_btn = ttk.Button(button_frame, text="🚀 生成报告", 
                                      command=self.generate_report_thread, style='Action.TButton')
        self.generate_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 停止按钮
        self.stop_btn = ttk.Button(button_frame, text="⏹️ 停止", 
                                  command=self.stop_generation, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # 清空日志按钮
        clear_btn = ttk.Button(button_frame, text="🗑️ 清空日志", command=self.clear_log)
        clear_btn.pack(side=tk.LEFT)
        
        # 状态显示区域
        status_frame = ttk.LabelFrame(content_frame, text="执行状态", padding=15)
        status_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 状态标签
        status_info_frame = ttk.Frame(status_frame)
        status_info_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(status_info_frame, text="当前状态:", font=('微软雅黑', 10)).pack(side=tk.LEFT)
        self.status_label = ttk.Label(status_info_frame, textvariable=self.status_var, 
                                     style='Info.TLabel', font=('微软雅黑', 10, 'bold'))
        self.status_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # 进度条
        self.progress = ttk.Progressbar(status_frame, mode='indeterminate', length=400)
        self.progress.pack(pady=(0, 10))
        
        # 日志显示区域
        log_frame = ttk.LabelFrame(content_frame, text="执行日志", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        # 创建文本框和滚动条
        log_container = ttk.Frame(log_frame)
        log_container.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(log_container, height=15, wrap=tk.WORD, font=('Consolas', 9))
        log_scrollbar = ttk.Scrollbar(log_container, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 配置日志文本样式
        self.log_text.tag_configure("INFO", foreground="#2980b9")
        self.log_text.tag_configure("ERROR", foreground="#e74c3c")
        self.log_text.tag_configure("SUCCESS", foreground="#27ae60")
        self.log_text.tag_configure("WARNING", foreground="#f39c12")

    def on_quick_select_changed(self, event=None):
        """快捷选择类型改变"""
        pass  # 仅用于绑定，实际逻辑在apply_quick_select中

    def apply_quick_select(self):
        """应用快捷时间选择"""
        selection = self.quick_select_var.get()
        now = datetime.datetime.now()

        if selection == "最近1小时":
            start_time = now - datetime.timedelta(hours=1)
            end_time = now
        elif selection == "最近4小时":
            start_time = now - datetime.timedelta(hours=4)
            end_time = now
        elif selection == "最近12小时":
            start_time = now - datetime.timedelta(hours=12)
            end_time = now
        elif selection == "昨天":
            yesterday = now - datetime.timedelta(days=1)
            start_time = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
            end_time = yesterday.replace(hour=23, minute=59, second=59, microsecond=0)
        elif selection == "最近7天":
            start_time = (now - datetime.timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_time = (now - datetime.timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=0)
        elif selection == "最近30天":
            start_time = (now - datetime.timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
            end_time = (now - datetime.timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=0)
        else:  # 自定义
            return

        self.start_datetime_var.set(start_time.strftime("%Y-%m-%d %H:%M:%S"))
        self.end_datetime_var.set(end_time.strftime("%Y-%m-%d %H:%M:%S"))
        self.update_time_info()

    def update_time_info(self, event=None):
        """更新时间范围信息显示"""
        try:
            start_str = self.start_datetime_var.get()
            end_str = self.end_datetime_var.get()

            start_time = datetime.datetime.strptime(start_str, "%Y-%m-%d %H:%M:%S")
            end_time = datetime.datetime.strptime(end_str, "%Y-%m-%d %H:%M:%S")

            duration = end_time - start_time

            if duration.total_seconds() <= 0:
                self.time_info_label.config(text="⚠️ 结束时间必须大于开始时间", foreground='#e74c3c')
            else:
                days = duration.days
                hours, remainder = divmod(duration.seconds, 3600)
                minutes, seconds = divmod(remainder, 60)

                duration_text = []
                if days > 0:
                    duration_text.append(f"{days}天")
                if hours > 0:
                    duration_text.append(f"{hours}小时")
                if minutes > 0:
                    duration_text.append(f"{minutes}分钟")
                if seconds > 0 and days == 0:
                    duration_text.append(f"{seconds}秒")

                duration_str = "".join(duration_text) if duration_text else "不足1秒"
                self.time_info_label.config(text=f"📊 时间跨度: {duration_str}", foreground='#3498db')

        except ValueError:
            self.time_info_label.config(text="⚠️ 时间格式错误", foreground='#e74c3c')

    def browse_output_path(self):
        """浏览输出路径"""
        path = filedialog.askdirectory(initialdir=self.output_path_var.get())
        if path:
            self.output_path_var.set(path)
    
    def test_connection(self):
        """测试连接"""
        try:
            self.log_message("🔄 正在测试连接...", "INFO")
            self.connection_status_var.set("连接中...")
            self.connection_status_label.configure(style='Info.TLabel')
            
            # 在后台线程中测试连接
            def test_worker():
                try:
                    client = NightingaleClient(
                        self.url_var.get().strip(),
                        self.username_var.get().strip(),
                        self.password_var.get().strip(),
                        self.datasource_id_var.get().strip()
                    )
                    
                    if client.test_connection():
                        self.root.after(0, self.on_connection_success)
                        self.client = client  # 保存客户端实例
                    else:
                        self.root.after(0, lambda: self.on_connection_error("连接失败或认证失败"))
                
                except Exception as e:
                    self.root.after(0, lambda: self.on_connection_error(str(e)))
            
            threading.Thread(target=test_worker, daemon=True).start()
            
        except Exception as e:
            self.on_connection_error(f"连接测试失败: {e}")
    
    def on_connection_success(self):
        """连接成功回调"""
        self.connection_status_var.set("连接成功")
        self.connection_status_label.configure(style='Success.TLabel')
        self.log_message("✅ 连接测试成功", "SUCCESS")
        messagebox.showinfo("连接测试", "连接测试成功!")
    
    def on_connection_error(self, error_msg):
        """连接错误回调"""
        self.connection_status_var.set("连接失败")
        self.connection_status_label.configure(style='Error.TLabel')
        self.log_message(f"❌ 连接测试失败: {error_msg}", "ERROR")
        messagebox.showerror("连接测试", f"连接测试失败:\n{error_msg}")
    
    def generate_report_thread(self):
        """在新线程中生成报告"""
        if self.is_generating:
            return
        
        thread = threading.Thread(target=self.generate_report, daemon=True)
        thread.start()
    
    def generate_report(self):
        """生成报告主流程"""
        try:
            # 验证输入
            if not self.validate_inputs():
                return
            
            self.is_generating = True
            self.generate_btn.config(state=tk.DISABLED)
            self.stop_btn.config(state=tk.NORMAL)
            self.progress.start()
            
            self.log_message("🚀 开始生成监控报告...", "INFO")
            self.status_var.set("正在生成报告...")

            # 解析参数
            cluster = self.cluster_var.get().strip()
            start_datetime = datetime.datetime.strptime(self.start_datetime_var.get(), "%Y-%m-%d %H:%M:%S")
            end_datetime = datetime.datetime.strptime(self.end_datetime_var.get(), "%Y-%m-%d %H:%M:%S")

            # 提取日期字符串用于文件命名
            start_date_str = self.start_datetime_var.get()[:10]  # 取日期部分 YYYY-MM-DD
            end_date_str = self.end_datetime_var.get()[:10]  # 取日期部分 YYYY-MM-DD

            # 创建数据收集器
            if not self.client:
                raise Exception("未连接到夜莺服务，请先测试连接")
            
            self.collector = DataCollector(
                self.client,
                self.cpu_threshold_var.get(),
                self.memory_threshold_var.get()
            )
            
            self.log_message(f"📊 收集时间范围: {start_datetime} 至 {end_datetime}", "INFO")
            self.log_message(f"🏢 目标集群: {cluster}", "INFO")
            
            # 收集数据（文本报告用）
            period_data = self.collector.collect_datetime_data(cluster, start_datetime, end_datetime)
            
            if not period_data:
                raise Exception("未收集到任何数据，请检查集群名称和时间范围")
            
            self.log_message(f"✅ 成功收集 {len(period_data)} 天的数据", "SUCCESS")
            
            # 为Excel报告单独收集数据（时间格式不同）
            period_data_excel = None
            if self.output_excel_var.get():
                self.log_message("📊 为Excel报告收集数据（时间格式优化）...", "INFO")
                period_data_excel = self.collector.collect_datetime_data_for_excel(cluster, start_datetime, end_datetime)
            
            # 生成报告
            output_dir = self.output_path_var.get().strip()
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                self.log_message(f"📁 创建输出目录: {output_dir}", "INFO")
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            saved_files = []
            
            # 生成文本报告
            if self.output_text_var.get():
                self.log_message("📝 生成文本报告...", "INFO")
                text_report = ReportGenerator.generate_text_report(
                    period_data, cluster,
                    start_date_str,
                    end_date_str
                )

                text_filename = f"{cluster}_监控报告_{start_date_str}_{end_date_str}_{timestamp}.txt"
                text_filepath = os.path.join(output_dir, text_filename)
                
                with open(text_filepath, 'w', encoding='utf-8') as f:
                    f.write(text_report)
                
                saved_files.append(text_filepath)
                self.log_message(f"✅ 文本报告已保存: {text_filepath}", "SUCCESS")
            
            # 生成Excel报告
            if self.output_excel_var.get() and period_data_excel:
                self.log_message("📊 生成Excel报告...", "INFO")
                excel_workbook = ReportGenerator.generate_excel_report(
                    period_data_excel, cluster,
                    start_date_str,
                    end_date_str
                )

                excel_filename = f"{cluster}_监控数据_{start_date_str}_{end_date_str}_{timestamp}.xlsx"
                excel_filepath = os.path.join(output_dir, excel_filename)
                
                excel_workbook.save(excel_filepath)
                saved_files.append(excel_filepath)
                self.log_message(f"✅ Excel报告已保存: {excel_filepath}", "SUCCESS")
            
            # 完成
            self.on_generation_complete(saved_files)
            
        except Exception as e:
            self.on_generation_error(str(e))
        
        finally:
            self.is_generating = False
            self.generate_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            self.progress.stop()
    
    def validate_inputs(self):
        """验证输入参数"""
        if not self.url_var.get().strip():
            messagebox.showerror("输入错误", "请输入夜莺服务器地址")
            return False
        
        if not self.username_var.get().strip():
            messagebox.showerror("输入错误", "请输入用户名")
            return False
        
        if not self.datasource_id_var.get().strip():
            messagebox.showerror("输入错误", "请输入数据源ID")
            return False
        
        if not self.cluster_var.get().strip():
            messagebox.showerror("输入错误", "请输入集群名称")
            return False
        
        # 验证日期格式
        try:
            start_date = datetime.datetime.strptime(self.start_datetime_var.get(), "%Y-%m-%d %H:%M:%S")
            end_date = datetime.datetime.strptime(self.end_datetime_var.get(), "%Y-%m-%d %H:%M:%S")
            
            if start_date > end_date:
                messagebox.showerror("输入错误", "开始日期不能大于结束日期")
                return False
            
            # 检查时间范围是否合理（不超过90天）
            if (end_date - start_date).days > 90:
                result = messagebox.askyesno("时间范围提醒", 
                                           "时间范围超过90天，可能需要较长时间处理，是否继续？")
                if not result:
                    return False

        except ValueError:
            messagebox.showerror("输入错误", "时间格式错误，请使用 YYYY-MM-DD HH:MM:SS 格式")
            return False
        
        if not self.output_text_var.get() and not self.output_excel_var.get():
            messagebox.showerror("输入错误", "请至少选择一种输出格式")
            return False
        
        return True
    
    def on_generation_complete(self, saved_files):
        """报告生成完成回调"""
        self.status_var.set("报告生成完成")
        self.log_message("🎉 报告生成完成!", "SUCCESS")
        
        # 显示完成对话框
        file_list = "\n".join([f"• {os.path.basename(f)}" for f in saved_files])
        message = f"监控报告生成完成!\n\n已保存文件:\n{file_list}\n\n保存位置: {self.output_path_var.get()}"
        
        result = messagebox.showinfo("生成完成", message)
        
        # 询问是否打开输出目录
        if messagebox.askyesno("打开目录", "是否打开输出目录查看文件？"):
            try:
                os.startfile(self.output_path_var.get())  # Windows
            except:
                try:
                    os.system(f'explorer "{self.output_path_var.get()}"')  # Windows备用
                except:
                    pass
    
    def on_generation_error(self, error_msg):
        """报告生成错误回调"""
        self.status_var.set("报告生成失败")
        self.log_message(f"❌ 报告生成失败: {error_msg}", "ERROR")
        messagebox.showerror("生成失败", f"报告生成失败:\n{error_msg}")
    
    def stop_generation(self):
        """停止报告生成"""
        self.status_var.set("操作已停止")
        self.log_message("⏹️ 用户停止操作", "WARNING")
        self.is_generating = False
    
    def clear_log(self):
        """清空日志"""
        self.log_text.delete(1.0, tk.END)
        self.log_message("🗑️ 日志已清空", "INFO")
    
    def log_message(self, message, level="INFO"):
        """添加日志消息"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        # 根据日志级别设置颜色
        tag = level
        self.log_text.insert(tk.END, log_entry, tag)
        self.log_text.see(tk.END)
        
        # 限制日志行数
        lines = self.log_text.get(1.0, tk.END).split('\n')
        if len(lines) > 200:
            self.log_text.delete(1.0, f"{len(lines) - 200}.0")
        
        # 更新界面
        self.root.update()


# 应用程序入口
def main():
    """主函数"""
    try:
        # 设置DPI感知（Windows高分辨率支持）
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass
        
        # 创建主窗口
        root = tk.Tk()
        
        # 设置窗口图标
        try:
            root.iconbitmap('nginx_monitor.ico')
        except:
            pass
        
        # 创建应用程序实例
        app = NginxMonitoringGUI(root)
        
        # 设置窗口关闭处理
        def on_closing():
            if app.is_generating:
                result = messagebox.askyesno("确认退出", "正在生成报告，确定要退出吗？")
                if not result:
                    return
            
            try:
                app.progress.stop()
                root.quit()
                root.destroy()
            except:
                pass
        
        root.protocol("WM_DELETE_WINDOW", on_closing)
        
        # 设置窗口最小尺寸
        root.minsize(800, 600)
        
        # 居中显示窗口
        center_window(root)
        
        # 记录启动日志
        app.log_message("🚀 Nginx监控报告生成系统已启动", "SUCCESS")
        app.log_message("📋 请先在连接配置页面配置夜莺服务连接", "INFO")
        
        # 启动主循环
        root.mainloop()
        
    except Exception as e:
        import traceback
        error_msg = f"启动失败: {e}\n\n详细错误:\n{traceback.format_exc()}"
        print(error_msg)
        
        try:
            messagebox.showerror("启动错误", error_msg)
        except:
            pass


def center_window(window):
    """将窗口居中显示"""
    try:
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        window.geometry(f"{width}x{height}+{x}+{y}")
    except Exception as e:
        logger.error(f"窗口居中失败: {e}")


def setup_logging():
    """设置日志配置"""
    log_dir = "./logs"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    
    log_filename = os.path.join(log_dir, f"nginx_monitoring_{datetime.datetime.now().strftime('%Y%m%d')}.log")
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    
    return logging.getLogger(__name__)


if __name__ == "__main__":
    """程序入口点"""
    
    # 设置全局异常处理
    def handle_exception(exc_type, exc_value, exc_traceback):
        if issubclass(exc_type, KeyboardInterrupt):
            print("\n程序被用户中断")
            sys.exit(0)
        
        import traceback
        error_msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_traceback))
        logger.error(f"未处理的异常: {error_msg}")
        
        try:
            with open("error.log", "a", encoding="utf-8") as f:
                f.write(f"\n[{datetime.datetime.now()}] 未处理的异常:\n{error_msg}\n")
        except:
            pass
    
    sys.excepthook = handle_exception
    
    try:
        # 检查Python版本
        if sys.version_info < (3, 6):
            print("此程序需要Python 3.6或更高版本")
            sys.exit(1)
        
        # 检查必要模块
        required_modules = [
            'tkinter', 'requests', 'pandas', 'openpyxl', 
            'datetime', 'json', 'threading'
        ]
        
        missing_modules = []
        for module in required_modules:
            try:
                __import__(module)
            except ImportError:
                missing_modules.append(module)
        
        if missing_modules:
            print(f"缺少必要的Python模块: {', '.join(missing_modules)}")
            print(f"请使用以下命令安装: pip install {' '.join(missing_modules)}")
            sys.exit(1)
        
        # 创建必要目录
        directories = ["./reports", "./logs", "./temp"]
        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)
        
        # 设置日志
        logger.info("=" * 50)
        logger.info("Nginx监控报告生成系统启动")
        logger.info("=" * 50)
        
        # 启动主程序
        main()
        
    except KeyboardInterrupt:
        print("\n程序被用户中断")
        logger.info("程序被用户中断")
        sys.exit(0)
    
    except Exception as e:
        error_msg = f"程序启动失败: {e}"
        print(error_msg)
        try:
            logger.error(error_msg, exc_info=True)
        except:
            pass
        sys.exit(1)
    
    finally:
        try:
            logger.info("程序正常退出")
            logger.info("=" * 50)
        except:
            pass


"""
🚀 Nginx监控报告生成系统 - 优化版使用说明

📋 主要特性:
1. ✅ 现代化GUI界面，标签页式设计
2. ✅ 灵活的时间范围选择（开始/结束日期）
3. ✅ 单集群精准监控
4. ✅ 数据与展示完全分离
5. ✅ 结构化Excel报表（列为指标，行为日期）
6. ✅ 可配置的监控指标和阈值
7. ✅ 智能单位换算和峰值时间记录

📊 支持的报表格式:
• 文本报告: 包含执行摘要、每日明细表格、运维建议
• Excel报表: 每日明细数据 + 统计汇总（最大、最小、平均、总和）

🔧 技术优化:
• 遵循单一职责原则，模块化设计
• 配置与逻辑分离，便于维护
• 异步处理，避免界面阻塞
• 完善的异常处理和日志记录
• 支持高DPI显示器

📝 使用步骤:
1. 在"连接配置"页面配置夜莺服务器信息
2. 在"报告配置"页面设置时间范围和集群
3. 在"输出配置"页面选择输出格式和路径
4. 在"执行监控"页面生成报告

💡 最佳实践:
• 建议时间范围不超过90天
• 确保集群名称与Prometheus标签一致
• 首次使用建议先测试连接
• 可使用快捷时间选择功能

⚠️ 注意事项:
• 仅支持Windows平台
• 需要Python 3.6+和相关依赖包
• 确保夜莺服务API可访问
• 大时间范围查询可能耗时较长
"""
            