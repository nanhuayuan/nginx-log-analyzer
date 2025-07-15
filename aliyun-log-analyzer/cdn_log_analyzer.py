#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CDN日志分析系统
用于分析阿里云CDN日志，统计第三方调用情况
作者：AI助手
"""

import gc
import os
import re
import time
from collections import defaultdict, Counter
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, PieChart, Reference
from urllib.parse import urlparse, parse_qs


class CDNLogParser:
    """CDN日志解析器"""
    
    def __init__(self):
        # 日志解析正则表达式
        self.log_pattern = re.compile(
            r'^\[(.*?)\]\s+(\S+)\s+(\S+)\s+(\d+)\s+"([^"]*)"\s+"([^"]*)"\s+(\d+)\s+(\d+)\s+(\d+)\s+([^"]*?)\s+"([^"]*)"\s+"([^"]*)"\s+(.*)$'
        )
        
        # 方法和URL解析
        self.method_url_pattern = re.compile(r'^(\w+)\s+(https?://[^?]+)(\?.*)?$')
        
    def parse_line(self, line):
        """解析单行日志"""
        line = line.strip()
        if not line:
            return None
            
        match = self.log_pattern.match(line)
        if not match:
            return None
            
        try:
            # 基础字段解析
            timestamp_str = match.group(1)
            client_ip = match.group(2)
            proxy_ip = match.group(3)
            response_time = int(match.group(4))
            referer = match.group(5)
            method_url = match.group(6)
            status_code = int(match.group(7))
            request_size = int(match.group(8))
            response_size = int(match.group(9))
            cache_status = match.group(10)
            user_agent = match.group(11)
            content_type = match.group(12)
            end_ip = match.group(13)
            
            # 解析时间
            timestamp = self._parse_timestamp(timestamp_str)
            if not timestamp:
                return None
                
            # 解析方法和URL
            method, full_url, uri, params = self._parse_method_url(method_url)
            
            return {
                'timestamp': timestamp,
                'client_ip': client_ip,
                'proxy_ip': proxy_ip,
                'response_time': response_time,
                'referer': referer,
                'method': method,
                'full_url': full_url,
                'uri': uri,
                'params': params,
                'status_code': status_code,
                'request_size': request_size,
                'response_size': response_size,
                'cache_status': cache_status,
                'user_agent': user_agent,
                'content_type': content_type,
                'end_ip': end_ip
            }
            
        except (ValueError, IndexError) as e:
            return None
    
    def _parse_timestamp(self, timestamp_str):
        """解析时间戳"""
        try:
            # 格式: 1/Jul/2025:21:47:26 +0800
            return datetime.strptime(timestamp_str, '%d/%b/%Y:%H:%M:%S %z')
        except ValueError:
            try:
                # 尝试其他格式
                return datetime.strptime(timestamp_str[:19], '%d/%b/%Y:%H:%M:%S')
            except ValueError:
                return None
    
    def _parse_method_url(self, method_url):
        """解析HTTP方法和URL"""
        match = self.method_url_pattern.match(method_url)
        if not match:
            return None, method_url, method_url, None
            
        method = match.group(1)
        base_url = match.group(2)
        params = match.group(3)
        
        # 提取URI（去掉域名）
        parsed_url = urlparse(base_url)
        uri = parsed_url.path
        
        full_url = base_url + (params if params else '')
        
        return method, full_url, uri, params


class CDNLogAnalyzer:
    """CDN日志分析器"""
    
    def __init__(self, target_ips=None, target_uris=None, slow_threshold=5000):
        self.parser = CDNLogParser()
        self.target_ips = set(target_ips) if target_ips else None
        self.target_uris = set(target_uris) if target_uris else None
        self.slow_threshold = slow_threshold  # 毫秒
        
        # 统计容器
        self.stats = {
            'daily': defaultdict(lambda: defaultdict(int)),
            'hourly': defaultdict(lambda: defaultdict(int)),
            'minute': defaultdict(lambda: defaultdict(int)),
            'second': defaultdict(lambda: defaultdict(int))
        }
        
        # 响应时间统计
        self.response_times = {
            'daily': defaultdict(list),
            'hourly': defaultdict(list),
            'minute': defaultdict(list),
            'second': defaultdict(list)
        }
        
        # 状态码统计
        self.status_codes = defaultdict(int)
        
        # User-Agent统计
        self.user_agents = defaultdict(int)
        
        # IP统计
        self.ip_stats = defaultdict(lambda: defaultdict(int))
        
        # URI统计
        self.uri_stats = defaultdict(lambda: defaultdict(int))
        
        # 处理计数
        self.total_records = 0
        self.processed_records = 0
        self.filtered_records = 0
        
    def get_time_keys(self, timestamp):
        """获取各时间维度的键"""
        return {
            'daily': timestamp.strftime('%Y-%m-%d'),
            'hourly': timestamp.strftime('%Y-%m-%d %H:00'),
            'minute': timestamp.strftime('%Y-%m-%d %H:%M'),
            'second': timestamp.strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def should_process_record(self, record):
        """判断是否需要处理该记录"""
        # IP过滤
        if self.target_ips and record['client_ip'] not in self.target_ips:
            return False
            
        # URI过滤
        if self.target_uris and record['uri'] not in self.target_uris:
            return False
            
        return True
    
    def process_record(self, record):
        """处理单条记录"""
        if not self.should_process_record(record):
            self.filtered_records += 1
            return
            
        self.processed_records += 1
        
        # 获取时间键
        time_keys = self.get_time_keys(record['timestamp'])
        
        # 基础统计
        status_code = record['status_code']
        response_time = record['response_time']
        is_success = 200 <= status_code < 400
        is_slow = response_time > self.slow_threshold
        
        # 更新各维度统计
        for dimension, time_key in time_keys.items():
            stats = self.stats[dimension][time_key]
            stats['total_requests'] += 1
            
            if is_success:
                stats['success_requests'] += 1
            if is_slow:
                stats['slow_requests'] += 1
                
            # 响应时间记录
            self.response_times[dimension][time_key].append(response_time)
        
        # 状态码统计
        self.status_codes[status_code] += 1
        
        # User-Agent统计
        ua = record['user_agent']
        if ua and ua != '-':
            # 提取主要的User-Agent信息
            ua_simplified = self._simplify_user_agent(ua)
            self.user_agents[ua_simplified] += 1
        
        # IP统计
        client_ip = record['client_ip']
        self.ip_stats[client_ip]['total_requests'] += 1
        if is_success:
            self.ip_stats[client_ip]['success_requests'] += 1
        
        # URI统计
        uri = record['uri']
        self.uri_stats[uri]['total_requests'] += 1
        if is_success:
            self.uri_stats[uri]['success_requests'] += 1
    
    def _simplify_user_agent(self, ua):
        """简化User-Agent字符串"""
        ua_lower = ua.lower()
        
        # 识别常见的工具和浏览器
        if 'hutool' in ua_lower:
            return 'Hutool'
        elif 'wst-sdk' in ua_lower:
            return 'WST-SDK'
        elif 'chrome' in ua_lower:
            return 'Chrome'
        elif 'firefox' in ua_lower:
            return 'Firefox'
        elif 'safari' in ua_lower and 'chrome' not in ua_lower:
            return 'Safari'
        elif 'curl' in ua_lower:
            return 'cURL'
        elif 'wget' in ua_lower:
            return 'wget'
        elif 'python' in ua_lower:
            return 'Python'
        elif 'java' in ua_lower:
            return 'Java'
        else:
            # 返回前50个字符
            return ua[:50] + ('...' if len(ua) > 50 else '')
    
    def process_file(self, file_path):
        """处理单个日志文件"""
        print(f"处理文件: {file_path}")
        
        file_records = 0
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line_no, line in enumerate(f, 1):
                    self.total_records += 1
                    file_records += 1
                    
                    record = self.parser.parse_line(line)
                    if record:
                        self.process_record(record)
                    
                    # 定期输出进度
                    if file_records % 10000 == 0:
                        print(f"  已处理 {file_records} 行")
                        
        except Exception as e:
            print(f"处理文件 {file_path} 时出错: {e}")
        
        print(f"文件 {file_path} 处理完成，共 {file_records} 行")
    
    def calculate_qps(self):
        """计算QPS"""
        window_seconds = {
            'daily': 86400,
            'hourly': 3600,
            'minute': 60,
            'second': 1
        }
        
        for dimension, seconds in window_seconds.items():
            for time_key, stats in self.stats[dimension].items():
                success_requests = stats.get('success_requests', 0)
                stats['qps'] = success_requests / seconds
    
    def calculate_averages(self):
        """计算平均响应时间"""
        avg_response_times = {}
        
        for dimension in self.stats.keys():
            avg_response_times[dimension] = {}
            for time_key in self.stats[dimension].keys():
                times = self.response_times[dimension][time_key]
                if times:
                    avg_response_times[dimension][time_key] = {
                        'avg_response_time': sum(times) / len(times),
                        'max_response_time': max(times),
                        'min_response_time': min(times),
                        'p95_response_time': self._percentile(times, 95),
                        'p99_response_time': self._percentile(times, 99)
                    }
                else:
                    avg_response_times[dimension][time_key] = {
                        'avg_response_time': 0,
                        'max_response_time': 0,
                        'min_response_time': 0,
                        'p95_response_time': 0,
                        'p99_response_time': 0
                    }
        
        return avg_response_times
    
    def _percentile(self, data, percentile):
        """计算百分位数"""
        if not data:
            return 0
        sorted_data = sorted(data)
        index = int(len(sorted_data) * percentile / 100)
        return sorted_data[min(index, len(sorted_data) - 1)]


class FileTimeChecker:
    """文件时间连续性检查器"""
    
    def __init__(self):
        # 时间格式解析器
        self.time_pattern = re.compile(r'(\d{4})_(\d{2})_(\d{2})_(\d{6})_(\d{6})')
    
    def extract_time_from_filename(self, filename):
        """从文件名提取时间"""
        match = self.time_pattern.search(filename)
        if not match:
            return None
            
        try:
            year, month, day, start_time, end_time = match.groups()
            
            # 解析开始时间
            start_hour = int(start_time[:2])
            start_minute = int(start_time[2:4])
            start_second = int(start_time[4:6])
            
            start_dt = datetime(
                int(year), int(month), int(day),
                start_hour, start_minute, start_second
            )
            
            # 解析结束时间
            end_hour = int(end_time[:2])
            end_minute = int(end_time[2:4])
            end_second = int(end_time[4:6])
            
            end_dt = datetime(
                int(year), int(month), int(day),
                end_hour, end_minute, end_second
            )
            
            return start_dt, end_dt
            
        except ValueError:
            return None
    
    def check_continuity(self, file_list):
        """检查文件时间连续性"""
        time_ranges = []
        invalid_files = []
        
        for file_path in file_list:
            filename = os.path.basename(file_path)
            time_range = self.extract_time_from_filename(filename)
            
            if time_range:
                time_ranges.append((file_path, time_range[0], time_range[1]))
            else:
                invalid_files.append(file_path)
        
        # 按开始时间排序
        time_ranges.sort(key=lambda x: x[1])
        
        # 检查连续性
        gaps = []
        if len(time_ranges) > 1:
            for i in range(len(time_ranges) - 1):
                current_end = time_ranges[i][2]
                next_start = time_ranges[i + 1][1]
                
                if current_end < next_start:
                    gap_duration = next_start - current_end
                    gaps.append((current_end, next_start, gap_duration))
        
        return {
            'total_files': len(file_list),
            'valid_files': len(time_ranges),
            'invalid_files': invalid_files,
            'time_ranges': time_ranges,
            'gaps': gaps,
            'continuous': len(gaps) == 0
        }


class ExcelReportGenerator:
    """Excel报告生成器"""
    
    def __init__(self, analyzer, avg_response_times, continuity_info):
        self.analyzer = analyzer
        self.avg_response_times = avg_response_times
        self.continuity_info = continuity_info
    
    def generate_report(self, output_path):
        """生成Excel报告"""
        wb = Workbook()
        
        # 删除默认工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # 创建各个工作表
        self._create_overview_sheet(wb)
        self._create_time_analysis_sheets(wb)
        self._create_status_code_sheet(wb)
        self._create_user_agent_sheet(wb)
        self._create_ip_analysis_sheet(wb)
        self._create_uri_analysis_sheet(wb)
        self._create_continuity_sheet(wb)
        
        # 保存文件
        wb.save(output_path)
        wb.close()
        print(f"Excel报告已生成: {output_path}")
    
    def _create_overview_sheet(self, wb):
        """创建概览页"""
        ws = wb.create_sheet(title="概览")
        
        # 标题
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "CDN日志分析报告"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        
        # 统计数据
        total_requests = self.analyzer.total_records
        processed_requests = self.analyzer.processed_records
        filtered_requests = self.analyzer.filtered_records
        
        overview_data = [
            ["指标名称", "数值", "单位", "说明"],
            ["总日志记录数", total_requests, "条", "所有日志记录"],
            ["已处理记录数", processed_requests, "条", "通过过滤条件的记录"],
            ["过滤记录数", filtered_requests, "条", "被过滤掉的记录"],
            ["处理成功率", f"{processed_requests/total_requests*100:.2f}" if total_requests > 0 else "0", "%", "处理记录占比"],
        ]
        
        # 添加状态码统计
        if self.analyzer.status_codes:
            success_codes = sum(count for code, count in self.analyzer.status_codes.items() if 200 <= code < 400)
            overview_data.extend([
                ["成功请求数", success_codes, "个", "状态码2xx-3xx"],
                ["成功率", f"{success_codes/processed_requests*100:.2f}" if processed_requests > 0 else "0", "%", "成功请求占比"],
            ])
        
        # 写入数据
        for row_idx, row_data in enumerate(overview_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # 表头
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
    
    def _create_time_analysis_sheets(self, wb):
        """创建时间维度分析页"""
        dimensions = [
            ('日维度分析', 'daily', '日期'),
            ('小时维度分析', 'hourly', '小时'),
            ('分钟维度分析', 'minute', '分钟'),
            ('秒维度分析', 'second', '秒')
        ]
        
        for sheet_name, dimension, time_label in dimensions:
            if self.analyzer.stats[dimension]:
                self._create_dimension_sheet(wb, sheet_name, dimension, time_label)
    
    def _create_dimension_sheet(self, wb, sheet_name, dimension, time_label):
        """创建维度分析页"""
        ws = wb.create_sheet(title=sheet_name)
        
        # 准备数据
        data = []
        stats_dict = self.analyzer.stats[dimension]
        avg_dict = self.avg_response_times[dimension]
        
        for time_key in sorted(stats_dict.keys()):
            stats = stats_dict[time_key]
            avg_stats = avg_dict.get(time_key, {})
            
            total = stats.get('total_requests', 0)
            success = stats.get('success_requests', 0)
            slow = stats.get('slow_requests', 0)
            qps = stats.get('qps', 0)
            
            success_rate = (success / total * 100) if total > 0 else 0
            slow_rate = (slow / total * 100) if total > 0 else 0
            
            row = [
                time_key,
                total,
                success,
                success_rate,
                slow,
                slow_rate,
                qps,
                avg_stats.get('avg_response_time', 0),
                avg_stats.get('max_response_time', 0),
                avg_stats.get('p95_response_time', 0),
                avg_stats.get('p99_response_time', 0)
            ]
            data.append(row)
        
        # 表头
        headers = [
            time_label, '总请求数', '成功请求数', '成功率(%)',
            '慢请求数', '慢请求率(%)', 'QPS',
            '平均响应时间(ms)', '最大响应时间(ms)', 'P95响应时间(ms)', 'P99响应时间(ms)'
        ]
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_status_code_sheet(self, wb):
        """创建状态码分析页"""
        ws = wb.create_sheet(title="状态码分析")
        
        # 表头
        ws.cell(row=1, column=1, value="状态码").font = Font(bold=True)
        ws.cell(row=1, column=2, value="请求数").font = Font(bold=True)
        ws.cell(row=1, column=3, value="占比(%)").font = Font(bold=True)
        
        # 数据
        total_requests = sum(self.analyzer.status_codes.values())
        for row_idx, (status_code, count) in enumerate(sorted(self.analyzer.status_codes.items()), 2):
            percentage = (count / total_requests * 100) if total_requests > 0 else 0
            ws.cell(row=row_idx, column=1, value=status_code)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{percentage:.2f}")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
    
    def _create_user_agent_sheet(self, wb):
        """创建User-Agent分析页"""
        ws = wb.create_sheet(title="UserAgent分析")
        
        # 表头
        ws.cell(row=1, column=1, value="User-Agent").font = Font(bold=True)
        ws.cell(row=1, column=2, value="请求数").font = Font(bold=True)
        ws.cell(row=1, column=3, value="占比(%)").font = Font(bold=True)
        
        # 数据（取前50个）
        total_requests = sum(self.analyzer.user_agents.values())
        sorted_agents = sorted(self.analyzer.user_agents.items(), key=lambda x: x[1], reverse=True)[:50]
        
        for row_idx, (user_agent, count) in enumerate(sorted_agents, 2):
            percentage = (count / total_requests * 100) if total_requests > 0 else 0
            ws.cell(row=row_idx, column=1, value=user_agent)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{percentage:.2f}")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
    
    def _create_ip_analysis_sheet(self, wb):
        """创建IP分析页"""
        ws = wb.create_sheet(title="IP分析")
        
        # 表头
        headers = ["客户端IP", "总请求数", "成功请求数", "成功率(%)", "失败请求数"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据（按请求数排序，取前100个）
        sorted_ips = sorted(self.analyzer.ip_stats.items(), key=lambda x: x[1]['total_requests'], reverse=True)[:100]
        
        for row_idx, (ip, stats) in enumerate(sorted_ips, 2):
            total = stats.get('total_requests', 0)
            success = stats.get('success_requests', 0)
            failed = total - success
            success_rate = (success / total * 100) if total > 0 else 0
            
            ws.cell(row=row_idx, column=1, value=ip)
            ws.cell(row=row_idx, column=2, value=total)
            ws.cell(row=row_idx, column=3, value=success)
            ws.cell(row=row_idx, column=4, value=f"{success_rate:.2f}")
            ws.cell(row=row_idx, column=5, value=failed)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def _create_uri_analysis_sheet(self, wb):
        """创建URI分析页"""
        ws = wb.create_sheet(title="URI分析")
        
        # 表头
        headers = ["URI路径", "总请求数", "成功请求数", "成功率(%)", "失败请求数"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据（按请求数排序，取前100个）
        sorted_uris = sorted(self.analyzer.uri_stats.items(), key=lambda x: x[1]['total_requests'], reverse=True)[:100]
        
        for row_idx, (uri, stats) in enumerate(sorted_uris, 2):
            total = stats.get('total_requests', 0)
            success = stats.get('success_requests', 0)
            failed = total - success
            success_rate = (success / total * 100) if total > 0 else 0
            
            ws.cell(row=row_idx, column=1, value=uri)
            ws.cell(row=row_idx, column=2, value=total)
            ws.cell(row=row_idx, column=3, value=success)
            ws.cell(row=row_idx, column=4, value=f"{success_rate:.2f}")
            ws.cell(row=row_idx, column=5, value=failed)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 60
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
    
    def _create_continuity_sheet(self, wb):
        """创建文件连续性分析页"""
        ws = wb.create_sheet(title="文件连续性检查")
        
        # 标题
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "日志文件连续性检查报告"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # 基本信息
        info_data = [
            ["检查项目", "结果", "说明"],
            ["总文件数", self.continuity_info['total_files'], "目录下所有文件"],
            ["有效文件数", self.continuity_info['valid_files'], "符合命名规范的文件"],
            ["无效文件数", len(self.continuity_info['invalid_files']), "不符合命名规范的文件"],
            ["时间连续性", "连续" if self.continuity_info['continuous'] else "不连续", "文件时间是否连续"],
            ["时间间隙数", len(self.continuity_info['gaps']), "发现的时间间隙数量"]
        ]
        
        # 写入基本信息
        for row_idx, row_data in enumerate(info_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # 表头
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 如果有时间间隙，显示详细信息
        if self.continuity_info['gaps']:
            ws.cell(row=10, column=1, value="时间间隙详情:").font = Font(bold=True)
            
            gap_headers = ["间隙开始时间", "间隙结束时间", "间隙时长"]
            for col_idx, header in enumerate(gap_headers, 1):
                cell = ws.cell(row=11, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            for row_idx, (start_time, end_time, duration) in enumerate(self.continuity_info['gaps'], 12):
                ws.cell(row=row_idx, column=1, value=start_time.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_idx, column=2, value=end_time.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_idx, column=3, value=str(duration))
        
        # 调整列宽
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 25


def get_all_log_files(directory):
    """获取目录下所有日志文件"""
    log_files = []
    directory = Path(directory)
    
    if not directory.exists():
        print(f"目录不存在: {directory}")
        return log_files
    
    # 获取所有文件（排除隐藏文件和目录）
    for file_path in directory.iterdir():
        if file_path.is_file() and not file_path.name.startswith('.'):
            log_files.append(str(file_path))
    
    return sorted(log_files)


def analyze_cdn_logs(
    log_directory,
    output_path=None,
    target_ips=None,
    target_uris=None,
    slow_threshold=5000
):
    """
    分析CDN日志主函数
    
    Args:
        log_directory: 日志文件目录
        output_path: 输出Excel文件路径
        target_ips: 目标IP列表，None表示分析所有IP
        target_uris: 目标URI列表，None表示分析所有URI
        slow_threshold: 慢请求阈值（毫秒）
    
    Returns:
        str: 输出文件路径
    """
    print("开始CDN日志分析...")
    print(f"日志目录: {log_directory}")
    print(f"目标IP: {target_ips if target_ips else '所有IP'}")
    print(f"目标URI: {target_uris if target_uris else '所有URI'}")
    print(f"慢请求阈值: {slow_threshold}ms")
    
    # 获取所有日志文件
    log_files = get_all_log_files(log_directory)
    if not log_files:
        print("未找到日志文件")
        return None
    
    print(f"找到 {len(log_files)} 个日志文件")
    
    # 检查文件连续性
    print("检查文件时间连续性...")
    checker = FileTimeChecker()
    continuity_info = checker.check_continuity(log_files)
    
    if not continuity_info['continuous']:
        print(f"警告: 发现 {len(continuity_info['gaps'])} 个时间间隙")
        for gap_start, gap_end, duration in continuity_info['gaps']:
            print(f"  间隙: {gap_start} 到 {gap_end}, 时长: {duration}")
    else:
        print("文件时间连续性检查通过")
    
    # 创建分析器
    analyzer = CDNLogAnalyzer(
        target_ips=target_ips,
        target_uris=target_uris,
        slow_threshold=slow_threshold
    )
    
    # 处理日志文件
    start_time = time.time()
    
    for file_path in log_files:
        analyzer.process_file(file_path)
        
        # 定期清理内存
        if analyzer.total_records % 100000 == 0:
            gc.collect()
    
    # 计算统计指标
    print("计算统计指标...")
    analyzer.calculate_qps()
    avg_response_times = analyzer.calculate_averages()
    
    elapsed_time = time.time() - start_time
    print(f"日志处理完成:")
    print(f"  总记录数: {analyzer.total_records}")
    print(f"  处理记录数: {analyzer.processed_records}")
    print(f"  过滤记录数: {analyzer.filtered_records}")
    print(f"  处理时间: {elapsed_time:.2f}秒")
    print(f"  处理速度: {analyzer.total_records / elapsed_time:.0f} 记录/秒")
    
    # 生成Excel报告
    if not output_path:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = f"cdn_analysis_report_{timestamp}.xlsx"
    
    print("生成Excel报告...")
    report_generator = ExcelReportGenerator(analyzer, avg_response_times, continuity_info)
    report_generator.generate_report(output_path)
    
    # 输出关键统计信息
    print("\n=== 关键统计信息 ===")
    
    # 状态码分布
    if analyzer.status_codes:
        print("状态码分布:")
        for status_code, count in sorted(analyzer.status_codes.items()):
            percentage = count / analyzer.processed_records * 100
            print(f"  {status_code}: {count} ({percentage:.2f}%)")
    
    # Top IP
    if analyzer.ip_stats:
        print("\nTop 10 IP:")
        sorted_ips = sorted(analyzer.ip_stats.items(), key=lambda x: x[1]['total_requests'], reverse=True)[:10]
        for ip, stats in sorted_ips:
            print(f"  {ip}: {stats['total_requests']} 请求")
    
    # Top URI
    if analyzer.uri_stats:
        print("\nTop 10 URI:")
        sorted_uris = sorted(analyzer.uri_stats.items(), key=lambda x: x[1]['total_requests'], reverse=True)[:10]
        for uri, stats in sorted_uris:
            print(f"  {uri}: {stats['total_requests']} 请求")
    
    # Top User-Agent
    if analyzer.user_agents:
        print("\nTop 10 User-Agent:")
        sorted_agents = sorted(analyzer.user_agents.items(), key=lambda x: x[1], reverse=True)[:10]
        for agent, count in sorted_agents:
            print(f"  {agent}: {count} 请求")
    
    return output_path


def main():
    """
    主函数 - 用于PyCharm调试
    """
    # 默认参数，方便调试
    default_log_directory = "D:/ob_智桂通/01-运维/高考专题/2025/阿里云全站加速数据/7月/日志数据/解压后日志数据-20250703"  # 修改为实际日志目录
    default_output_path = "cdn_analysis_report.xlsx"
    
    # 示例：分析特定IP的特定URI
    target_ips = [
        "222.84.157.38",  # 示例IP
        # "其他目标IP"
    ]
    
    target_uris = [
        "/scmp-gateway/wxxcx/prod-api/portals/affairsGuide/search",  # 示例URI
        # "其他目标URI"
    ]
    
    # 如果不需要过滤，设为None
    # target_ips = None
    # target_uris = None
    
    try:
        result = analyze_cdn_logs(
            log_directory=default_log_directory,
            output_path=default_output_path,
            target_ips=target_ips,
            target_uris=target_uris,
            slow_threshold=5000  # 5秒作为慢请求阈值
        )
        
        if result:
            print(f"\n分析完成！报告已保存到: {result}")
        else:
            print("分析失败！")
            
    except Exception as e:
        print(f"分析过程中出现错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()